import streamlit as st
import pandas as pd
import math
import io
import json
import datetime

st.set_page_config(page_title="System Projektowania", layout="wide")

st.title("🏭 Inżynieryjny Reaktor Procesowy & Logistyczny")
st.subheader("Kompletna Platforma Wymiarowania Linii, Reologii, Logistyki i Surowców")
st.markdown("---")

st.markdown("""
    <style>
    div[data-testid="stTabs"] {
        position: sticky;
        top: 2.875rem;
        background-color: white;
        z-index: 999;
        padding-top: 10px;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 0. STAŁE GLOBALNE (dawniej "magiczne liczby" rozsiane po kodzie)
# ==========================================
WORKING_DAYS_YEAR = 250          # dni robocze zakładu w roku
MONTHS_PER_YEAR = 12
OIL_FILL_FACTOR = 0.88           # współczynnik napełnienia silosu dla mediów olejowych
WATER_FILL_FACTOR = 1.00         # współczynnik napełnienia silosu dla wody DEMI
TANK_SAFETY_FILL = 0.85          # bufor bezpieczeństwa objętości silosu (85% pojemności nominalnej)
MAX_TANK_UTILIZATION_PCT = 85.0  # próg ostrzegawczy wykorzystania czasowego mieszalnika
MIN_TANK_VOLUME_M3 = 5.0         # minimalna pojemność mieszalnika akceptowana w fabryce
VELOCITY_MIN_MS = 0.5            # dolna granica prędkości przepływu w rurociągu
VELOCITY_MAX_MS = 2.5            # górna granica prędkości przepływu w rurociągu
LMTD_MIN_K = 15.0                # dolna granica "zdrowego" LMTD
LMTD_MAX_K = 60.0                # górna granica "zdrowego" LMTD
STEAM_LATENT_HEAT_KJKG = 2200.0  # ciepło skraplania pary nasyconej (~2 bar) [kJ/kg], wartość orientacyjna
G_ACCEL = 9.81
RAMPUP_YEARS = 5                 # horyzont symulacji rozruchu (Zakładka 2 + Zakładka 3)

# --- 1. BAZA DANYCH PROCESOWYCH I FIZYKOCHEMICZNYCH ---
# (Dawny zestaw z liniami marek - np. "Hydraulic Oils (Marka X)" - został
# usunięty. Jedyna taksonomia w apce to teraz 7 grup produktowych, patrz GENERIC_PORTFOLIO
# i GROUP_PHYSICAL_DEFAULTS niżej.)

PACK_CONFIGS = {
    "1l (Detal)": {"size_l": 1.0, "per_pallet": 480, "rate_szt_h": 2500},
    "4l (Karton)": {"size_l": 4.0, "per_pallet": 120, "rate_szt_h": 1200},
    "5l (Karton)": {"size_l": 5.0, "per_pallet": 96, "rate_szt_h": 1000},
    "10l (Kanister)": {"size_l": 10.0, "per_pallet": 40, "rate_szt_h": 600},
    "20l (Kanister)": {"size_l": 20.0, "per_pallet": 24, "rate_szt_h": 400},
    "60l (Beczka)": {"size_l": 60.0, "per_pallet": 9, "rate_szt_h": 150},
    "200l (Beczka)": {"size_l": 200.0, "per_pallet": 4, "rate_szt_h": 60},
    "1000l (IBC)": {"size_l": 1000.0, "per_pallet": 1, "rate_szt_h": 15},
    "Cysterna (luzem)": {"size_l": 24000.0, "per_pallet": 0, "rate_szt_h": 1.5},
    # per_pallet=0 to sentinel: produkt luzem w cysternie NIE jest paletyzowany - nie generuje
    # miejsc magazynowych/palet, tylko bezpośrednie wysyłki cysterną (patrz logika w Zakładce 4).
}

AGITATOR_TYPES = {
    "Turbinowe (Rushton)": {"laminar_C": 70.0, "turbulent_Ne": 5.0},
    "Łapowe / Płatowe": {"laminar_C": 50.0, "turbulent_Ne": 2.5},
    "Propelerowe (Śmigłowe)": {"laminar_C": 35.0, "turbulent_Ne": 0.8}
}

MEDIA_PROCESOWE = {
    "Woda technologiczna": {"cp": 4.19, "t_max": 95.0, "t_min": 5.0, "steam": False, "density_kg_m3": 1000.0},
    "Olej termiczny": {"cp": 2.00, "t_max": 300.0, "t_min": 40.0, "steam": False, "density_kg_m3": 850.0},
    "Para nasycona": {"cp": 2.15, "t_max": 180.0, "t_min": 100.0, "steam": True, "density_kg_m3": None}
}

# Katalog testów laboratoryjnych oznaczonych jako "QC" (zwolnienie szarży) w dostarczonej
# liście testów. Testy oznaczone wyłącznie jako "R&D" (np. korozja, pianotwórczość, EP/AW,
# wielkość cząstek, tribologia, czystość ISO 4406) są pominięte - nie leżą na ścieżce
# krytycznej standardowego zwolnienia partii produkcyjnej.
# Czasy trwania [min] to orientacyjne wartości domyślne oparte na typowej praktyce
# laboratoryjnej - EDYTOWALNE bezpośrednio w Zakładce 6 (VSM), bo rzeczywisty czas zależy
# od obciążenia laboratorium i wprawy technika.
QC_TEST_CATALOG = {
    "Lepkość kinematyczna @40°C (łaźnia ręczna)": {"duration_min": 40, "equipment": "Łaźnia wiskozymetryczna ręczna, 40°C", "count": 1},
    "Lepkość kinematyczna @40°C (półautomat)": {"duration_min": 25, "equipment": "Łaźnia wiskozymetryczna półautomatyczna, 40°C", "count": 1},
    "Lepkość kinematyczna @40°C (automat)": {"duration_min": 15, "equipment": "Wiskozymetr automatyczny, 40°C", "count": 1},
    "Lepkość kinematyczna @100°C (łaźnia ręczna)": {"duration_min": 60, "equipment": "Łaźnia wiskozymetryczna ręczna, 100°C", "count": 1},
    "Lepkość kinematyczna @100°C (automat)": {"duration_min": 20, "equipment": "Wiskozymetr automatyczny, 100°C", "count": 1},
    "Lepkość dynamiczna (ASTM D2669)": {"duration_min": 30, "equipment": "Wiskozymetr rotacyjny (ASTM D2669)", "count": 1},
    "Lepkość dynamiczna (ASTM D2983)": {"duration_min": 45, "equipment": "Wiskozymetr rotacyjny (ASTM D2983)", "count": 1},
    "Lepkość dynamiczna @-40°C": {"duration_min": 45, "equipment": "Wiskozymetr rotacyjny (niska temp.)", "count": 1},
    "Krzywa Stribecka": {"duration_min": 60, "equipment": "Reometr", "count": 1},
    "Barwa ASTM": {"duration_min": 5, "equipment": "Kolorymetr (ASTM D1500)", "count": 1},
    "Wielkość cząstek (Particle size)": {"duration_min": 15, "equipment": "Analizator wielkości cząstek", "count": 1},
    "Klasa czystości (Cleanliness code)": {"duration_min": 15, "equipment": "Licznik cząstek (ISO 4406)", "count": 1},
    "Gęstość": {"duration_min": 10, "equipment": "Hydrometr / densimetr", "count": 1},
    "Wskaźnik refrakcji": {"duration_min": 5, "equipment": "Refraktometr cyfrowy", "count": 1},
    "Temp. zapłonu - półautomat": {"duration_min": 45, "equipment": "Aparat do temp. zapłonu, półautomatyczny", "count": 1},
    "Temp. zapłonu - automat": {"duration_min": 25, "equipment": "Aparat do temp. zapłonu, automatyczny", "count": 1},
    "Krzywa chłodzenia (Cooling Curve)": {"duration_min": 30, "equipment": "Aparat do krzywej chłodzenia", "count": 1},
    "XRF": {"duration_min": 15, "equipment": "Spektrometr XRF", "count": 1},
    "Punkt aniliny": {"duration_min": 50, "equipment": "Aparat do punktu aniliny", "count": 1},
    "Zawartość ciał stałych": {"duration_min": 20, "equipment": "Wagosuszarka", "count": 1},
    "Przewodność": {"duration_min": 10, "equipment": "Konduktometr", "count": 1},
    "Zawartość wody (Karl Fischer)": {"duration_min": 20, "equipment": "Kulometr (metoda Karl Fischera)", "count": 1},
    "Spektroskopia FTIR": {"duration_min": 15, "equipment": "Spektrometr FT-IR", "count": 1},
    "pH": {"duration_min": 10, "equipment": "pH-metr stołowy", "count": 1},
    "Zasadowość (Alkalinity)": {"duration_min": 25, "equipment": "Automatyczny tytrator potencjometryczny", "count": 1},
    "Spektroskopia UV-Vis": {"duration_min": 20, "equipment": "Spektrometr UV/Vis", "count": 1},
    "Test korozji": {"duration_min": 1440, "equipment": "Komora klimatyczna (DIN 6270-2) — czas oczekiwania 24h+", "count": 1},
    "Pienienie (Foam)": {"duration_min": 30, "equipment": "Aparat do badania pienienia", "count": 1},
    "EP / AW (tarcie i zużycie)": {"duration_min": 90, "equipment": "Aparat do tarcia i zużycia (EP/AW)", "count": 1},
    "Tapping Torque Test": {"duration_min": 45, "equipment": "Trybometr", "count": 1},
    "Demulgowalność": {"duration_min": 45, "equipment": "Aparat do demulgowalności", "count": 1},
    "Pour Point": {"duration_min": 30, "equipment": "Aparat Pour Point", "count": 1},
}

# Prefiks/sufiks kolumn Excela definiujących, KTÓRE testy QC dotyczą KONKRETNEGO produktu (nie
# tylko całej linii jak w Zakładce 6/VSM) - ten sam wzorzec co "Opak: {nazwa} [%]" dla opakowań.
# Kolumna: "QC: {nazwa testu z QC_TEST_CATALOG} [x]", wartość "x"/1/TRUE = test dotyczy produktu.
QC_COL_PREFIX = "QC: "
QC_COL_SUFFIX = " [x]"


# Lista surowców receptury produktowej (Zakładka 1 / Receptury) - dozowanie w kg na tonę
# produktu gotowego [kg/t], kolejność zgodna z szablonem Excel wgrywanym przez użytkownika.
RECIPE_RAW_MATERIALS = [
    "Base Oil Group I [kg/t]", "Base Oil Group II [kg/t]", "Base Oil Group III [kg/t]",
    "Base Oil Group IV (PAO) [kg/t]", "Base Oil Group V (Estry/PAG) [kg/t]",
    "RRBO (Re-refined) [kg/t]", "Bio-Base Oil (Ester/Vegetable) [kg/t]",
    "Modyfikator Lepkości (VI Improver) [kg/t]", "Depresator (PPD) [kg/t]",
    "Dodatek Smarnościowy / Anti-wear (AW) [kg/t]", "Dodatek Wysokociśnieniowy (EP) [kg/t]",
    "Inhibitor Utleniania (AO) [kg/t]", "Inhibitor Korozji / Pasywator [kg/t]",
    "Detergenty (TBN Boosters) [kg/t]", "Dyspergatory Bezpopiołowe [kg/t]",
    "Dodatek Przeciwpienny (Antifoam) [kg/t]", "Modyfikator Tarcia (FM) [kg/t]",
    "Deemulgatory / Emulgatory [kg/t]", "Modyfikator Uszczelek (Seal Swell) [kg/t]",
    "Pakiet Silnikowy (PCMO/HDDO) [kg/t]", "Pakiet Przekładniowy (Gear) [kg/t]",
    "Pakiet Hydrauliczny (Hydraulic) [kg/t]", "Zagęszczacz Mydłowy (Li/Ca/Complex) [kg/t]",
    "Zagęszczacz Niemydłowy (Polyurea/Bentonit) [kg/t]", "Smar Stały: MoS2 / Grafit [kg/t]",
    "Smar Stały: PTFE / Boron Nitride [kg/t]", "Barwnik / Znacznik / Zapach [kg/t]",
    "Woda Demineralizowana [kg/t]", "Biocyd / Fungicyd [kg/t]",
]

# Grupy produktowe do wyboru (lista rozwijana w szablonie Excel + walidacja przy imporcie).
RECIPE_PRODUCT_GROUPS = ["Cleaners", "Engine Oils", "Glycols", "Greases", "Hydraulic Oils", "Watermiscibles", "Waxes",
                          "Preservative Oils", "Coolants", "Cutting Oils"]

# Domyślne właściwości fizykochemiczne i procesowe per grupa produktowa - używane do
# automatycznego zasilenia floty (Zakładka 2) danymi z wgranej receptury (Zakładka 1),
# tam gdzie receptura sama nie precyzuje danej wartości (np. materiał zbiornika, cp).
# Gęstość NIE jest tu potrzebna - ta zawsze pochodzi z konkretnego wiersza receptury.
GROUP_PHYSICAL_DEFAULTS = {
    "Cleaners": {"material": "Stal nierdzewna", "density": 1.01, "cp": 3.9, "oil_group": "Brak (Specjalistyczne)", "water_content": 0.85, "cycle_h": 4},
    "Engine Oils": {"material": "Stal zwykła", "density": 0.87, "cp": 2.1, "oil_group": "Syntetyczne (Gr. III/IV)", "water_content": 0.0, "cycle_h": 5},
    "Glycols": {"material": "Stal nierdzewna", "density": 1.05, "cp": 2.4, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.3, "cycle_h": 4},
    "Greases": {"material": "Stal zwykła", "density": 0.90, "cp": 2.0, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.0, "cycle_h": 6},
    "Hydraulic Oils": {"material": "Stal zwykła", "density": 0.88, "cp": 2.0, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.0, "cycle_h": 4},
    "Watermiscibles": {"material": "Stal nierdzewna", "density": 0.99, "cp": 3.8, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.65, "cycle_h": 6},
    "Waxes": {"material": "Stal zwykła", "density": 0.91, "cp": 2.2, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.0, "cycle_h": 5},
    "Preservative Oils": {"material": "Stal zwykła", "density": 0.85, "cp": 1.95, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.0, "cycle_h": 4},
    "Coolants": {"material": "Stal nierdzewna", "density": 1.02, "cp": 3.9, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.85, "cycle_h": 5},
    "Cutting Oils": {"material": "Stal zwykła", "density": 0.89, "cp": 1.95, "oil_group": "Mineralne (Gr. I/II)", "water_content": 0.0, "cycle_h": 5},
}

# Zestaw startowy active_portfolio - JEDYNA taksonomia w apce, wspólna dla trybu ręcznego i
# recepturowego: 7 grup produktowych z GROUP_PHYSICAL_DEFAULTS, bez osobnych, zaszytych w
# kodzie linii marek (dawne nazwy typu "Hydraulic Oils (Marka X)") - to
# było źródłem dublowania się nazewnictwa, gdy obok wbudowanych linii pojawiały się grupy
# z wgranej receptury o niemal identycznej nazwie.
GENERIC_PORTFOLIO = {
    name: {"material": d["material"], "density": d["density"], "cycle_h": d["cycle_h"],
           "cp": d["cp"], "oil_group": d["oil_group"], "water_content": d["water_content"]}
    for name, d in GROUP_PHYSICAL_DEFAULTS.items()
}

# Arkusz opakowań (opcjonalny, w tym samym pliku co receptury) - pozwala predefiniować
# typy opakowań i ich pojemności w Excelu; po wgraniu nadpisują/uzupełniają wbudowane
# domyślne wartości (PACK_CONFIGS), a dalej pozostają w pełni edytowalne w samej aplikacji.
PACKAGING_SHEET_NAME = "Opakowania"
PACKAGING_NAME_COL = "Nazwa Opakowania"
PACKAGING_SIZE_COL = "Pojemność [L]"
PACKAGING_PER_PALLET_COL = "Sztuk na Palecie"
PACKAGING_RATE_COL = "Wydajność 1 głowicy [kg/min]"
PACKAGING_NOZZLES_COL = "Liczba głowic (domyślna)"

QC_SHEET_NAME = "Badania Laboratoryjne"
QC_SHEET_TEST_NAME_COL = "Nazwa Testu"

DIRECT_RM_SHEET_NAME = "Zużycie Surowców (bez recept.)"
DIRECT_RM_GROUP_COL = "Grupa Produktowa (opcjonalnie - puste = ogólna krzywa rozruchu całego zakładu)"
DIRECT_RM_MATERIAL_COL = "Nazwa Surowca (dowolna, np. z dostawcą)"
DIRECT_RM_CATEGORY_COL = "Kategoria (opcjonalnie, do raportów zbiorczych)"
DIRECT_RM_ANNUAL_COL = "Roczne Zużycie Docelowe [tony]"
DIRECT_RM_TANK_ID_COL = "ID Zbiornika (opcjonalnie - te samo ID = wspólny zbiornik)"

SUPPLIER_SPLIT_SHEET_NAME = "Rozbicie Dostawców Surowców"
SUPPLIER_SPLIT_GROUP_COL = "Grupa Produktowa (opcjonalnie - puste = cały zakład)"
SUPPLIER_SPLIT_CATEGORY_COL = "Kategoria API"
SUPPLIER_SPLIT_SUPPLIER_COL = "Konkretna Baza/Dostawca"
SUPPLIER_SPLIT_PCT_COL = "Udział w tej kategorii [%]"
SUPPLIER_SPLIT_TECH_COL = "Technologia (opcjonalnie)"
SUPPLIER_SPLIT_TANK_ID_COL = "ID Zbiornika (opcjonalnie - te samo ID = wspólny zbiornik)"

RECIPE_GROUP_COL = "Grupa Produktowa"
RECIPE_PRODUCT_COL = "Produkt"
RECIPE_ANNUAL_COL = "Roczne Zapotrzebowanie Produktu [tony]"
RECIPE_SUM_COL = "Suma Udziałów Składników [kg]"
RECIPE_DENSITY_COL = "Gęstość 15°C [g/cm³]"
RECIPE_LOSS_COL = "Szacowane Straty Procesowe [%]"
RECIPE_RAW_DEMAND_COL = "Roczne Zapotrzebowanie Surowcowe [tony]"
RECIPE_NOTES_COL = "Uwagi Technologiczne / Status QA"

# Sposób pozyskania produktu: na starcie część produktów bywa IMPOROWANA (z innego zakładu/od
# dostawcy) zamiast produkowana lokalnie, a dopiero po jakimś czasie zaczyna się produkcja
# własna. Przejście modelowane jest jako NAGŁE, w jednym pełnym roku symulacji rozruchu
# (Zakładka 2/5) - przed rokiem przejścia produkt w 100% importowany, od tego roku w 100%
# produkowany lokalnie. "Nigdy" = produkt na stałe pozostaje importowany (nigdy nie trafia do
# floty mieszalników - typowy przypadek dla niszowego SKU, którego produkcja lokalna się nie
# opłaca nawet w dojrzałości).
RECIPE_SOURCING_COL = "Sposób Pozyskania"
RECIPE_SOURCING_OPTIONS = ["Produkcja własna", "Import"]
RECIPE_IMPORT_TRANSITION_COL = "Rok Przejścia na Produkcję Własną"
RECIPE_IMPORT_TRANSITION_OPTIONS = ["Rok 1", "Rok 2", "Rok 3", "Rok 4", "Rok 5", "Nigdy (stały import)", "Nigdy (bufor)"]
RECIPE_IMPORT_FREQ_COL = "Częstotliwość Dostawy Importowej [dni]"
RECIPE_IMPORT_LOT_COL = "Wielkość 1 Dostawy Importowej [tony]"
RECIPE_IMPORT_SAFETY_DAYS_COL = "Bufor Bezpieczeństwa Importu [dni]"

# Opcjonalna kolumna: kilka produktów może współdzielić JEDEN fizyczny mieszalnik (produkcja
# kampanijna) zamiast dostawać każdy swój dedykowany zbiornik. Puste = własny zbiornik (jak
# dotychczas); te samo ID w kilku wierszach TEJ SAMEJ grupy produktowej = wspólny zbiornik.
RECIPE_TANK_ID_COL = "ID Zbiornika (opcjonalnie - te samo ID = wspólny mieszalnik)"
RECIPE_SHARED_UTIL_COL = "Suma Wykorzystania Zbiornika Współdzielonego [%]"

# Wymiarowanie mieszalnika i szacowane wykorzystanie zdolności produkcyjnej - wpisywane
# bezpośrednio w Excelu, żeby dać natychmiastową (choć uproszczoną) informację zwrotną, zanim
# w ogóle dojdzie do konfiguracji floty w Zakładce 2/4 (tam wykorzystanie liczone jest w pełni,
# z rzeczywistą hydrauliką/bilansem cieplnym - to tutaj to szybki szacunek wstępny).
RECIPE_MIXER_VOL_COL = "Pojemność Mieszalnika [m³]"
RECIPE_CYCLE_COL = "Szacowany Cykl Szarży [h]"
RECIPE_AVAIL_HOURS_COL = "Dostępne Godziny Pracy / Rok [h]"
RECIPE_BATCH_MASS_COL = "Masa Szarży [kg]"
RECIPE_BATCHES_YEAR_COL = "Szarż / Rok"
RECIPE_UTILIZATION_COL = "Wykorzystanie Mieszalnika [%]"
RECIPE_UTILIZATION_WARN_PCT = 85.0  # spójne z MAX_TANK_UTILIZATION_PCT w Zakładce 2

# Rozbicie procentowe na typy opakowań - kolumny generowane dynamicznie z aktualnej listy
# opakowań (domyślne PACK_CONFIGS lub to, co użytkownik ma w arkuszu 'Opakowania').
RECIPE_PACK_SUM_COL = "Suma % Opakowań (kontrola)"

# Docelowa suma dozowania składników na tonę produktu - 1000 kg/t (1 tona), z tolerancją na
# zaokrąglenia ręcznego wpisywania receptur.
RECIPE_TARGET_SUM_KG = 1000.0
RECIPE_SUM_TOLERANCE_KG = 50.0


def recipe_pack_pct_col(pack_name):
    """Nazwa kolumny procentowego udziału danego opakowania w arkuszu receptur."""
    return f"Opak: {pack_name} [%]"


RAMPUP_YEAR_TARGET_SENTINEL = 9999  # "Docelowa (100%)" = pełna dojrzałość, po wszystkich przejściach z importu


def is_product_imported_in_year(sourcing, transition_label, year_idx):
    """
    Czy dany produkt jest w danym roku symulacji (year_idx, 0-based; RAMPUP_YEAR_TARGET_SENTINEL
    = widok docelowy/100%) nadal importowany, czy już produkowany lokalnie. Przejście jest
    NAGŁE (patrz komentarz przy RECIPE_SOURCING_COL) - przed rokiem przejścia 100% import, od
    tego roku 100% produkcja własna. "Nigdy" = zawsze import, nawet w widoku docelowym.
    """
    if sourcing != "Import":
        return False
    if transition_label in ("Nigdy (stały import)", "Nigdy (bufor)") or not transition_label:
        return True
    if year_idx == RAMPUP_YEAR_TARGET_SENTINEL:
        return False  # widok docelowy = pełna dojrzałość = po przejściu (chyba że "Nigdy", obsłużone wyżej)
    try:
        transition_year = int(transition_label.split(" ")[1])
    except (IndexError, ValueError):
        transition_year = 1
    return (year_idx + 1) < transition_year

# Informacja, czy dany surowiec nadaje się fizycznie/praktycznie do magazynowania
# w zbiorniku (płyn luzem) czy zawsze zostaje w beczkach/IBC/workach - niezależnie od
# wolumenu (ciała stałe, bardzo niskie dozowania, ograniczona trwałość po otwarciu itp.).
# To jest orientacyjna reguła inżynierska do podpowiedzi w aplikacji, nie sztywna norma -
# użytkownik może ją zignorować dla konkretnego przypadku.
RAW_MATERIAL_STORAGE_INFO = {
    "Base Oil Group I [kg/t]": {"bulk_eligible": True, "note": "Baza mineralna - standardowo magazynowana luzem w zbiornikach."},
    "Base Oil Group II [kg/t]": {"bulk_eligible": True, "note": "Baza mineralna - standardowo magazynowana luzem w zbiornikach."},
    "Base Oil Group III [kg/t]": {"bulk_eligible": True, "note": "Baza syntetyczna - standardowo magazynowana luzem w zbiornikach."},
    "Base Oil Group IV (PAO) [kg/t]": {"bulk_eligible": True, "note": "PAO - magazynowanie luzem możliwe, zwykle droższe medium więc warto pilnować rotacji."},
    "Base Oil Group V (Estry/PAG) [kg/t]": {"bulk_eligible": True, "note": "Estry/PAG - higroskopijne, magazynowanie luzem OK, ale zbiornik wymaga osuszania/inertyzacji i krótszej rotacji."},
    "RRBO (Re-refined) [kg/t]": {"bulk_eligible": True, "note": "Baza re-refinowana - magazynowanie luzem jak dla baz mineralnych."},
    "Bio-Base Oil (Ester/Vegetable) [kg/t]": {"bulk_eligible": True, "note": "Podatna na utlenianie - luzem możliwe, ale przy niższej rotacji zalecana krótsza autonomia zapasu."},
    "Modyfikator Lepkości (VI Improver) [kg/t]": {"bulk_eligible": True, "note": "Przy dużym wolumenie opłacalny zbiornik dedykowany; przy małym - beczki/IBC."},
    "Depresator (PPD) [kg/t]": {"bulk_eligible": True, "note": "Zwykle niskie dozowanie - zbiornik opłacalny dopiero przy dużym wolumenie."},
    "Dodatek Smarnościowy / Anti-wear (AW) [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny przy dużym wolumenie."},
    "Dodatek Wysokociśnieniowy (EP) [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny przy dużym wolumenie (typowo oleje przekładniowe)."},
    "Inhibitor Utleniania (AO) [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny przy dużym wolumenie."},
    "Inhibitor Korozji / Pasywator [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny przy dużym wolumenie."},
    "Detergenty (TBN Boosters) [kg/t]": {"bulk_eligible": True, "note": "Przy wysokowolumenowych olejach silnikowych zbiornik dedykowany jest standardem branżowym."},
    "Dyspergatory Bezpopiołowe [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny przy dużym wolumenie."},
    "Dodatek Przeciwpienny (Antifoam) [kg/t]": {"bulk_eligible": False, "note": "Bardzo niskie dozowanie (dziesiąte części %) - zawsze beczki/małe pojemniki, precyzja dozowania ważniejsza niż koszt logistyki."},
    "Modyfikator Tarcia (FM) [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny dopiero przy dużym wolumenie, zwykle beczki/IBC."},
    "Deemulgatory / Emulgatory [kg/t]": {"bulk_eligible": True, "note": "Zbiornik opłacalny przy dużym wolumenie."},
    "Modyfikator Uszczelek (Seal Swell) [kg/t]": {"bulk_eligible": False, "note": "Niskie dozowanie i wysoka cena jednostkowa - zwykle beczki/IBC nawet przy większych wolumenach."},
    "Pakiet Silnikowy (PCMO/HDDO) [kg/t]": {"bulk_eligible": True, "note": "Gotowy pakiet dodatków - przy wysokim wolumenie standardowo dedykowany zbiornik (dostawy cysternami)."},
    "Pakiet Przekładniowy (Gear) [kg/t]": {"bulk_eligible": True, "note": "Gotowy pakiet dodatków - zbiornik opłacalny przy dużym wolumenie."},
    "Pakiet Hydrauliczny (Hydraulic) [kg/t]": {"bulk_eligible": True, "note": "Gotowy pakiet dodatków - zbiornik opłacalny przy dużym wolumenie."},
    "Zagęszczacz Mydłowy (Li/Ca/Complex) [kg/t]": {"bulk_eligible": False, "note": "Pasta/proszek do produkcji smarów - magazynowany w beczkach/workach, nie w zbiorniku cieczy."},
    "Zagęszczacz Niemydłowy (Polyurea/Bentonit) [kg/t]": {"bulk_eligible": False, "note": "Ciało stałe/proszek - beczki/worki, nie zbiornik cieczy."},
    "Smar Stały: MoS2 / Grafit [kg/t]": {"bulk_eligible": False, "note": "Proszek stały - zawsze worki/beczki, nie nadaje się do zbiornika."},
    "Smar Stały: PTFE / Boron Nitride [kg/t]": {"bulk_eligible": False, "note": "Proszek stały - zawsze worki/beczki, nie nadaje się do zbiornika."},
    "Barwnik / Znacznik / Zapach [kg/t]": {"bulk_eligible": False, "note": "Śladowe dozowanie - zawsze małe pojemniki/beczki, zbiornik nieopłacalny przy żadnym wolumenie."},
    "Woda Demineralizowana [kg/t]": {"bulk_eligible": True, "note": "Zawsze warto trzymać w zbiorniku/cysternie - tani, wysokowolumenowy nośnik."},
    "Biocyd / Fungicyd [kg/t]": {"bulk_eligible": False, "note": "Ograniczona trwałość i niskie dozowanie - zalecane beczki/IBC z szybką rotacją, nie długie magazynowanie luzem."},
}

# Typy pojemników dla surowców NIE magazynowanych luzem (beczki/IBC/worki) - do przeliczenia
# rocznego zużycia [t/rok] na liczbę pojemników/palet/miejsc magazynowych w Zakładce 4, żeby
# te surowce trafiły do tego samego bilansu powierzchni magazynowej co wyroby gotowe (Zakładka
# 4) - w końcu wszystko, co nie trafia do zbiornika, musi stanąć w magazynie. Pojemność podana
# w kg (nie L), bo dla surowców pracujemy na masie, nie na objętości/gęstości.
RM_CONTAINER_TYPES = {
    "Beczka 200 kg (ciecz)": {"capacity_kg": 200.0, "per_pallet": 4},
    "IBC 1000 kg (ciecz)": {"capacity_kg": 1000.0, "per_pallet": 1},
    "Worek 25 kg (ciało stałe, na palecie)": {"capacity_kg": 25.0, "per_pallet": 40},
    "Big Bag 500 kg (ciało stałe)": {"capacity_kg": 500.0, "per_pallet": 1},
}


def default_rm_container_for(material_name, info):
    """Sensowna domyślna propozycja pojemnika na podstawie opisu surowca w RAW_MATERIAL_STORAGE_INFO."""
    note_lower = (info.get("note", "") + " " + material_name).lower()
    if "worki" in note_lower or "proszek" in note_lower or "pasta" in note_lower:
        return "Worek 25 kg (ciało stałe, na palecie)"
    return "Beczka 200 kg (ciecz)"


# Klucze session_state, które użytkownik faktycznie KONFIGURUJE (nie liczy się od nowa za każdym
# przebiegiem) - to jest zapisywane/wczytywane przy "Zapisz Projekt" / "Wczytaj Projekt". Wartości
# WYNIKOWE (raporty, PDF/Excel, symulacje) celowo pominięte - przeliczą się same z tych danych.
PROJECT_SAVE_KEYS = [
    "active_portfolio", "prod_dict", "confirmed_mixers", "mixer_tech_advanced_details",
    "tag_to_recipe_product", "tag_to_shared_members", "batch_time_components", "calculated_times",
    "confirmed_rm_tanks", "rm_tank_tech_details", "rm_storage_method_override", "rm_container_assignment",
    "confirmed_fg_buffer_tanks", "fg_buffer_tank_tech_details", "shared_pumps",
    "rampup_global_pct", "rampup_per_line_pct", "rampup_differentiate", "rampup_start_pct", "import_follows_rampup",
    "group_pricing", "pack_configs", "filling_lines_config", "opakowania_podzial",
    "qc_tests_by_product", "vsm_qc_config", "vsm_qc_queue_days", "vsm_oee", "direct_raw_materials",
    "qc_equipment_count_override", "supplier_splits", "custom_qc_tests",
    "tanker_capacity_t", "mixer_fill_factor", "days_of_stock_tab5", "max_single_tank_m3",
    "czas_skladowania_tab3", "cena_mwh_tab4", "cena_gazu_mwh", "sprawnosc_kotla_frac",
    "import_pallet_mass_kg", "capex_lump_sum", "boiler_capacity_installed_kw",
    "typ_kotla", "equipment_install_counts",
]


class _NumpyJsonEncoder(json.JSONEncoder):
    """Konwertuje typy numpy (int64/float64 z obliczeń pandas) na natywne typy Pythona - inaczej
    standardowy json.dumps rzuca błąd na tych typach, mimo że 'wyglądają' jak zwykłe liczby."""
    def default(self, obj):
        if hasattr(obj, "item"):
            return obj.item()
        return super().default(obj)


def export_project_bytes():
    """Buduje plik JSON do pobrania, zawierający cały skonfigurowany stan projektu (receptura,
    flota, zbiorniki, ceny, rozruch, itd.) - do wczytania później albo w innej sesji/przeglądarce,
    zamiast uzupełniać wszystko od nowa."""
    payload = {"_saved_at": datetime.datetime.now().isoformat(), "_app_version": "1.0"}
    for key in PROJECT_SAVE_KEYS:
        if key not in st.session_state:
            continue
        val = st.session_state[key]
        if isinstance(val, pd.DataFrame):
            payload[key] = {"__dataframe__": True, "data": val.to_dict(orient="split")}
        else:
            payload[key] = val
    if "recipes_df" in st.session_state and st.session_state.recipes_df is not None:
        payload["recipes_df"] = {"__dataframe__": True, "data": st.session_state.recipes_df.to_dict(orient="split")}
    return json.dumps(payload, cls=_NumpyJsonEncoder, ensure_ascii=False, indent=2).encode("utf-8")


def import_project_bytes(uploaded_bytes):
    """Wczytuje plik JSON zapisany przez export_project_bytes i odtwarza cały stan projektu.
    Zwraca (sukces: bool, komunikat: str)."""
    try:
        payload = json.loads(uploaded_bytes.decode("utf-8"))
    except Exception as exc:
        return False, f"Nie udało się odczytać pliku projektu: {exc}"

    restored_keys = []
    for key, val in payload.items():
        if key.startswith("_"):
            continue
        if isinstance(val, dict) and val.get("__dataframe__"):
            df_data = val["data"]
            st.session_state[key] = pd.DataFrame(df_data["data"], columns=df_data["columns"], index=df_data["index"])
        else:
            st.session_state[key] = val
        restored_keys.append(key)

    saved_at = payload.get("_saved_at", "nieznana data")
    return True, f"Wczytano projekt zapisany {saved_at} — przywrócono {len(restored_keys)} sekcji konfiguracji."


def generate_recipe_template_bytes():
    """
    Buduje w pamięci szablon Excel (openpyxl) do uzupełnienia przez użytkownika: grupa
    produktowa, nazwa produktu, roczne zapotrzebowanie na produkt [tony], wstępny dobór
    mieszalnika (pojemność/cykl/dostępne godziny -> wyliczone wykorzystanie zdolności
    produkcyjnej w %), dozowanie surowców [kg/t] (29 pozycji), gęstość, szacowane straty
    procesowe, wyliczone roczne zapotrzebowanie surowcowe, rozbicie procentowe na typy
    opakowań, oraz pole na uwagi technologiczne/status QA.
    Kolumny wyliczane (Masa Szarży, Szarż/Rok, Wykorzystanie Mieszalnika, Suma Udziałów
    Składników, Roczne Zapotrzebowanie Surowcowe, Suma % Opakowań) są formułami Excela (nie
    sztywnymi wartościami z Pythona), żeby przeliczały się po edycji w arkuszu.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Receptury"

    header_font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="DDEBF7")
    computed_fill = PatternFill("solid", fgColor="F2F2F2")
    example_font = Font(name="Arial", size=10, italic=True, color="0000FF")
    normal_font = Font(name="Arial", size=10)
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    try:
        current_pack_defaults = st.session_state.get("pack_configs", PACK_CONFIGS)
    except Exception:
        current_pack_defaults = PACK_CONFIGS
    pack_names = list(current_pack_defaults.keys())
    pack_pct_cols = [recipe_pack_pct_col(n) for n in pack_names]

    headers = ([RECIPE_GROUP_COL, RECIPE_PRODUCT_COL, RECIPE_ANNUAL_COL,
                RECIPE_SOURCING_COL, RECIPE_IMPORT_TRANSITION_COL, RECIPE_IMPORT_FREQ_COL,
                RECIPE_IMPORT_LOT_COL, RECIPE_IMPORT_SAFETY_DAYS_COL,
                RECIPE_MIXER_VOL_COL, RECIPE_CYCLE_COL, RECIPE_AVAIL_HOURS_COL,
                RECIPE_BATCH_MASS_COL, RECIPE_BATCHES_YEAR_COL, RECIPE_UTILIZATION_COL] +
               RECIPE_RAW_MATERIALS +
               [RECIPE_SUM_COL, RECIPE_DENSITY_COL, RECIPE_LOSS_COL, RECIPE_RAW_DEMAND_COL] +
               pack_pct_cols + [RECIPE_PACK_SUM_COL, RECIPE_TANK_ID_COL, RECIPE_SHARED_UTIL_COL, RECIPE_NOTES_COL])

    n_fixed_left = 3                      # Grupa, Produkt, Roczne Zapotrzebowanie Produktu
    sourcing_col = n_fixed_left + 1
    transition_col = sourcing_col + 1
    import_freq_col = transition_col + 1
    import_lot_col = import_freq_col + 1
    import_safety_col = import_lot_col + 1
    mixer_vol_col = import_safety_col + 1
    cycle_col = mixer_vol_col + 1
    avail_hours_col = cycle_col + 1
    batch_mass_col = avail_hours_col + 1
    batches_year_col = batch_mass_col + 1
    utilization_col = batches_year_col + 1
    n_materials = len(RECIPE_RAW_MATERIALS)
    first_mat_col = utilization_col + 1
    last_mat_col = first_mat_col + n_materials - 1
    sum_col = last_mat_col + 1
    density_col = sum_col + 1
    loss_col = density_col + 1
    raw_demand_col = loss_col + 1
    first_pack_col = raw_demand_col + 1
    last_pack_col = first_pack_col + len(pack_names) - 1
    pack_sum_col = last_pack_col + 1
    tank_id_col = pack_sum_col + 1
    shared_util_col = tank_id_col + 1
    notes_col = shared_util_col + 1

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
    ws.freeze_panes = get_column_letter(n_fixed_left + 1) + "2"
    ws.row_dimensions[1].height = 50

    ws_info = wb.create_sheet("Instrukcja")
    info_lines = [
        "INSTRUKCJA WYPEŁNIANIA:",
        "1. Nie zmieniaj nazw ani kolejności kolumn w arkuszu 'Receptury'.",
        f"2. '{RECIPE_GROUP_COL}' - wybierz z listy rozwijanej: {', '.join(RECIPE_PRODUCT_GROUPS)}.",
        f"3. '{RECIPE_ANNUAL_COL}' - roczny wolumen produkcji GOTOWEGO produktu (bez strat procesowych).",
        f"4. '{RECIPE_MIXER_VOL_COL}', '{RECIPE_CYCLE_COL}', '{RECIPE_AVAIL_HOURS_COL}' - wstępny dobór mieszalnika; "
        f"'{RECIPE_UTILIZATION_COL}' liczy się sama (formuła) i podświetla na czerwono powyżej {RECIPE_UTILIZATION_WARN_PCT:.0f}% "
        "(przeciążenie) - to szybki szacunek, w Zakładce 2/4 aplikacji policzysz to dokładnie z rzeczywistą hydrauliką.",
        "5. Kolumny surowcowe [kg/t] - ile kg danego surowca zużywa się na 1 tonę GOTOWEGO produktu.",
        f"6. '{RECIPE_SUM_COL}' liczy się sama (formuła) - powinna wynosić ok. 1000 kg (tolerancja +/-{RECIPE_SUM_TOLERANCE_KG:.0f} kg).",
        f"7. '{RECIPE_DENSITY_COL}' i '{RECIPE_LOSS_COL}' wpisz ręcznie dla każdego produktu.",
        f"8. '{RECIPE_RAW_DEMAND_COL}' liczy się sama (formuła) = zapotrzebowanie produktu / (1 - straty%)"
        " - to ILE SUROWCA trzeba faktycznie zakupić, uwzględniając straty procesowe.",
        f"9. Kolumny 'Opak: ... [%]' - opcjonalne, jeśli chcesz z góry rozbić produkt na opakowania; "
        f"'{RECIPE_PACK_SUM_COL}' liczy się sama i powinna wynosić 100%, jeśli używasz tego rozbicia (możesz zostawić 0, jeśli nie dotyczy).",
        "10. Puste komórki w kolumnach surowcowych/opakowaniowych są traktowane jako 0.",
        "11. Dodaj tyle wierszy produktów, ile potrzebujesz - przeciągnij formuły w dół.",
        "12. Przykładowe wiersze (niebieska kursywa) pokazują format - usuń je lub nadpisz własnymi danymi.",
        f"13. '{RECIPE_TANK_ID_COL}' - zostaw puste, jeśli każdy produkt ma mieć własny, dedykowany mieszalnik "
        "(domyślne zachowanie). Jeśli KILKA produktów TEJ SAMEJ grupy produktowej ma współdzielić JEDEN fizyczny "
        "mieszalnik (produkcja kampanijna - różne produkty na przemian na tym samym reaktorze), wpisz im "
        "IDENTYCZNY identyfikator, np. 'R-01'. Aplikacja policzy wtedy jeden zbiornik z łącznym wykorzystaniem "
        "czasowym (suma szarż x cykl każdego produktu), dobierając pojemność pod największą recepturę spośród nich. "
        "Ustaw też tę SAMĄ pojemność mieszalnika dla wszystkich wierszy ze wspólnym ID - to fizycznie jeden zbiornik.",
        f"14. '{RECIPE_SHARED_UTIL_COL}' liczy się sama (formuła) - dla wierszy ze wspólnym ID Zbiornika sumuje ich "
        "indywidualne wykorzystania (orientacyjnie, każde liczone tak, jakby miało cały zbiornik dla siebie), "
        f"żeby wykryć przeciążenie WSPÓLNEGO zbiornika już tutaj, przed wgraniem do aplikacji - podświetla się na "
        f"czerwono powyżej {RECIPE_UTILIZATION_WARN_PCT:.0f}%.",
        f"15. '{RECIPE_SOURCING_COL}' - 'Produkcja własna' (domyślne) albo 'Import'. Produkty importowane NIE trafiają "
        f"do floty mieszalników, dopóki nie nadejdzie ich '{RECIPE_IMPORT_TRANSITION_COL}' (Rok 1-5, zgodnie z 5-letnią "
        "symulacją rozruchu w Zakładce 2) - przejście jest NAGŁE, w jednym pełnym roku. Wybierz 'Nigdy (stały import)', "
        "jeśli produkt ma zawsze pozostać importowany (nigdy nie dostanie własnego mieszalnika, nawet w widoku "
        f"docelowym). '{RECIPE_IMPORT_FREQ_COL}', '{RECIPE_IMPORT_LOT_COL}' i '{RECIPE_IMPORT_SAFETY_DAYS_COL}' "
        "opisują rytm dostaw - używane w Zakładce 3 do wyliczenia miejsc magazynowych na produkt importowany "
        "(analogicznie do bufora surowców w Zakładce 4, ale liczone z rytmu dostaw, nie z cyklu produkcji).",
    ]
    for i, line in enumerate(info_lines, start=1):
        c = ws_info.cell(row=i, column=1, value=line)
        c.font = Font(bold=(i == 1), name="Arial", size=11)
    ws_info.column_dimensions["A"].width = 110

    example_rows = [
        {
            RECIPE_GROUP_COL: "Hydraulic Oils", RECIPE_PRODUCT_COL: "Przykład: Hydraulic Oil 46",
            RECIPE_ANNUAL_COL: 500, RECIPE_MIXER_VOL_COL: 15, RECIPE_CYCLE_COL: 4, RECIPE_AVAIL_HOURS_COL: 2000,
            RECIPE_DENSITY_COL: 0.876, RECIPE_LOSS_COL: 1.5,
            RECIPE_SOURCING_COL: "Import", RECIPE_IMPORT_TRANSITION_COL: "Rok 3",
            RECIPE_IMPORT_FREQ_COL: 21, RECIPE_IMPORT_LOT_COL: 20, RECIPE_IMPORT_SAFETY_DAYS_COL: 10,
            RECIPE_NOTES_COL: "Receptura referencyjna - status QA: zatwierdzona. Importowana do Roku 3, od Roku 3 produkcja własna.",
            "Base Oil Group II [kg/t]": 965, "Depresator (PPD) [kg/t]": 5,
            "Inhibitor Utleniania (AO) [kg/t]": 8, "Inhibitor Korozji / Pasywator [kg/t]": 3,
            "Dodatek Przeciwpienny (Antifoam) [kg/t]": 1, "Deemulgatory / Emulgatory [kg/t]": 18,
            recipe_pack_pct_col("5l (Karton)"): 30, recipe_pack_pct_col("200l (Beczka)"): 40,
            recipe_pack_pct_col("1000l (IBC)"): 30,
        },
        {
            RECIPE_GROUP_COL: "Engine Oils", RECIPE_PRODUCT_COL: "Przykład: Engine Oil 5W-30",
            RECIPE_ANNUAL_COL: 800, RECIPE_MIXER_VOL_COL: 20, RECIPE_CYCLE_COL: 5, RECIPE_AVAIL_HOURS_COL: 2000,
            RECIPE_DENSITY_COL: 0.855, RECIPE_LOSS_COL: 2.0, RECIPE_TANK_ID_COL: "R-01",
            RECIPE_NOTES_COL: "Receptura referencyjna - status QA: w walidacji. Współdzieli mieszalnik R-01 z Engine Oil 10W-40 poniżej.",
            "Base Oil Group III [kg/t]": 820, "Modyfikator Lepkości (VI Improver) [kg/t]": 80,
            "Depresator (PPD) [kg/t]": 3, "Dodatek Smarnościowy / Anti-wear (AW) [kg/t]": 10,
            "Inhibitor Utleniania (AO) [kg/t]": 12, "Inhibitor Korozji / Pasywator [kg/t]": 5,
            "Pakiet Silnikowy (PCMO/HDDO) [kg/t]": 68, "Dodatek Przeciwpienny (Antifoam) [kg/t]": 1,
            "Modyfikator Tarcia (FM) [kg/t]": 1,
            recipe_pack_pct_col("5l (Karton)"): 60, recipe_pack_pct_col("1000l (IBC)"): 40,
        },
        {
            RECIPE_GROUP_COL: "Engine Oils", RECIPE_PRODUCT_COL: "Przykład: Engine Oil 10W-40 (współdzielony mieszalnik)",
            RECIPE_ANNUAL_COL: 150, RECIPE_MIXER_VOL_COL: 20, RECIPE_CYCLE_COL: 5, RECIPE_AVAIL_HOURS_COL: 2000,
            RECIPE_DENSITY_COL: 0.86, RECIPE_LOSS_COL: 2.0, RECIPE_TANK_ID_COL: "R-01",
            RECIPE_NOTES_COL: "Współdzieli mieszalnik R-01 z Engine Oil 5W-30 powyżej (produkcja kampanijna - "
                               "ten sam reaktor, na przemian).",
            "Base Oil Group III [kg/t]": 850, "Modyfikator Lepkości (VI Improver) [kg/t]": 60,
            "Depresator (PPD) [kg/t]": 3, "Dodatek Smarnościowy / Anti-wear (AW) [kg/t]": 10,
            "Inhibitor Utleniania (AO) [kg/t]": 12, "Inhibitor Korozji / Pasywator [kg/t]": 5,
            "Pakiet Silnikowy (PCMO/HDDO) [kg/t]": 58, "Dodatek Przeciwpienny (Antifoam) [kg/t]": 1,
            "Modyfikator Tarcia (FM) [kg/t]": 1,
            recipe_pack_pct_col("5l (Karton)"): 60, recipe_pack_pct_col("1000l (IBC)"): 40,
        },
    ]

    start_data_row = 2
    n_blank_rows = 20
    total_rows = len(example_rows) + n_blank_rows
    annual_col_letter = get_column_letter(n_fixed_left)
    mixer_vol_letter = get_column_letter(mixer_vol_col)
    cycle_letter = get_column_letter(cycle_col)
    avail_hours_letter = get_column_letter(avail_hours_col)
    batch_mass_letter = get_column_letter(batch_mass_col)
    batches_year_letter = get_column_letter(batches_year_col)
    utilization_letter = get_column_letter(utilization_col)
    first_mat_letter = get_column_letter(first_mat_col)
    last_mat_letter = get_column_letter(last_mat_col)
    sum_col_letter = get_column_letter(sum_col)
    density_col_letter = get_column_letter(density_col)
    loss_col_letter = get_column_letter(loss_col)
    raw_demand_col_letter = get_column_letter(raw_demand_col)
    first_pack_letter = get_column_letter(first_pack_col) if pack_names else None
    last_pack_letter = get_column_letter(last_pack_col) if pack_names else None
    pack_sum_col_letter = get_column_letter(pack_sum_col)
    tank_id_col_letter = get_column_letter(tank_id_col)
    shared_util_col_letter = get_column_letter(shared_util_col)

    group_dv = DataValidation(type="list", formula1='"' + ",".join(RECIPE_PRODUCT_GROUPS) + '"', allow_blank=True)
    ws.add_data_validation(group_dv)
    sourcing_dv = DataValidation(type="list", formula1='"' + ",".join(RECIPE_SOURCING_OPTIONS) + '"', allow_blank=True)
    ws.add_data_validation(sourcing_dv)
    transition_dv = DataValidation(type="list", formula1='"' + ",".join(RECIPE_IMPORT_TRANSITION_OPTIONS) + '"', allow_blank=True)
    ws.add_data_validation(transition_dv)

    for r_offset in range(total_rows):
        row = start_data_row + r_offset
        is_example = r_offset < len(example_rows)
        font_to_use = example_font if is_example else normal_font

        if is_example:
            data = example_rows[r_offset]
            ws.cell(row=row, column=1, value=data.get(RECIPE_GROUP_COL, "")).font = font_to_use
            ws.cell(row=row, column=2, value=data.get(RECIPE_PRODUCT_COL, "")).font = font_to_use
            ws.cell(row=row, column=3, value=data.get(RECIPE_ANNUAL_COL, "")).font = font_to_use
            ws.cell(row=row, column=sourcing_col, value=data.get(RECIPE_SOURCING_COL, "Produkcja własna")).font = font_to_use
            ws.cell(row=row, column=transition_col, value=data.get(RECIPE_IMPORT_TRANSITION_COL, "")).font = font_to_use
            ws.cell(row=row, column=import_freq_col, value=data.get(RECIPE_IMPORT_FREQ_COL, "")).font = font_to_use
            ws.cell(row=row, column=import_lot_col, value=data.get(RECIPE_IMPORT_LOT_COL, "")).font = font_to_use
            ws.cell(row=row, column=import_safety_col, value=data.get(RECIPE_IMPORT_SAFETY_DAYS_COL, "")).font = font_to_use
            ws.cell(row=row, column=mixer_vol_col, value=data.get(RECIPE_MIXER_VOL_COL, "")).font = font_to_use
            ws.cell(row=row, column=cycle_col, value=data.get(RECIPE_CYCLE_COL, "")).font = font_to_use
            ws.cell(row=row, column=avail_hours_col, value=data.get(RECIPE_AVAIL_HOURS_COL, "")).font = font_to_use
            for m_idx, mat in enumerate(RECIPE_RAW_MATERIALS, start=first_mat_col):
                ws.cell(row=row, column=m_idx, value=data.get(mat, 0)).font = font_to_use
            ws.cell(row=row, column=density_col, value=data.get(RECIPE_DENSITY_COL, "")).font = font_to_use
            ws.cell(row=row, column=loss_col, value=data.get(RECIPE_LOSS_COL, "")).font = font_to_use
            for p_idx, pname in enumerate(pack_names, start=first_pack_col):
                ws.cell(row=row, column=p_idx, value=data.get(recipe_pack_pct_col(pname), 0)).font = font_to_use
            ws.cell(row=row, column=tank_id_col, value=data.get(RECIPE_TANK_ID_COL, "")).font = font_to_use
            ws.cell(row=row, column=notes_col, value=data.get(RECIPE_NOTES_COL, "")).font = font_to_use
        else:
            prod_num = r_offset - len(example_rows) + 1
            ws.cell(row=row, column=1).font = normal_font
            ws.cell(row=row, column=1).fill = input_fill
            ws.cell(row=row, column=2, value=f"Product {prod_num}").font = normal_font
            ws.cell(row=row, column=2).fill = input_fill
            for col_idx in [3, mixer_vol_col, cycle_col, avail_hours_col]:
                ws.cell(row=row, column=col_idx).font = normal_font
                ws.cell(row=row, column=col_idx).fill = input_fill
            ws.cell(row=row, column=sourcing_col, value="Produkcja własna").font = normal_font
            ws.cell(row=row, column=sourcing_col).fill = input_fill
            for col_idx in [transition_col, import_freq_col, import_lot_col, import_safety_col]:
                ws.cell(row=row, column=col_idx).font = normal_font
                ws.cell(row=row, column=col_idx).fill = input_fill
            for col_idx in range(first_mat_col, last_mat_col + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = normal_font
                cell.fill = input_fill
            ws.cell(row=row, column=density_col).fill = input_fill
            ws.cell(row=row, column=loss_col).fill = input_fill
            for col_idx in range(first_pack_col, last_pack_col + 1) if pack_names else []:
                cell = ws.cell(row=row, column=col_idx)
                cell.font = normal_font
                cell.fill = input_fill
            ws.cell(row=row, column=tank_id_col).fill = input_fill
            ws.cell(row=row, column=notes_col).fill = input_fill
            group_dv.add(f"A{row}")
            sourcing_dv.add(ws.cell(row=row, column=sourcing_col).coordinate)
            transition_dv.add(ws.cell(row=row, column=transition_col).coordinate)

        # Masa Szarży [kg] = pojemność [m3] * 1000 (L/m3) * gęstość [kg/L, liczbowo = g/cm3]
        batch_mass_formula = f"=IF(OR({mixer_vol_letter}{row}<=0,{density_col_letter}{row}<=0),0,{mixer_vol_letter}{row}*1000*{density_col_letter}{row})"
        bm_cell = ws.cell(row=row, column=batch_mass_col, value=batch_mass_formula)
        bm_cell.font = font_to_use
        bm_cell.number_format = '#,##0" kg"'
        bm_cell.fill = computed_fill

        # Szarż / Rok = (roczne zapotrzebowanie [t] * 1000) / masa szarży [kg]
        batches_year_formula = f"=IF({batch_mass_letter}{row}<=0,0,{annual_col_letter}{row}*1000/{batch_mass_letter}{row})"
        by_cell = ws.cell(row=row, column=batches_year_col, value=batches_year_formula)
        by_cell.font = font_to_use
        by_cell.number_format = '0.0'
        by_cell.fill = computed_fill

        # Wykorzystanie Mieszalnika [%] = Szarż/Rok * Cykl [h] / Dostępne godziny/rok * 100
        util_formula = f"=IF({avail_hours_letter}{row}<=0,0,{batches_year_letter}{row}*{cycle_letter}{row}/{avail_hours_letter}{row}*100)"
        util_cell = ws.cell(row=row, column=utilization_col, value=util_formula)
        util_cell.font = font_to_use
        util_cell.number_format = '0.0"%"'
        util_cell.fill = computed_fill

        # Suma wykorzystania dla wspólnego ID Zbiornika - jeśli kilka produktów współdzieli
        # jeden fizyczny mieszalnik, ich pojedyncze % wykorzystania (każdy liczony tak, jakby
        # miał zbiornik tylko dla siebie) trzeba zsumować, żeby zobaczyć REALNE łączne
        # obciążenie. Dla pustego ID (własny zbiornik) to po prostu to samo co Wykorzystanie.
        tank_id_range = f"${tank_id_col_letter}${start_data_row}:${tank_id_col_letter}${start_data_row + total_rows - 1}"
        util_range = f"${utilization_letter}${start_data_row}:${utilization_letter}${start_data_row + total_rows - 1}"
        shared_util_formula = (f'=IF({tank_id_col_letter}{row}="",{utilization_letter}{row},'
                                f'SUMIF({tank_id_range},{tank_id_col_letter}{row},{util_range}))')
        shared_cell = ws.cell(row=row, column=shared_util_col, value=shared_util_formula)
        shared_cell.font = font_to_use
        shared_cell.number_format = '0.0"%"'
        shared_cell.fill = computed_fill

        sum_formula = f"=SUM({first_mat_letter}{row}:{last_mat_letter}{row})"
        sum_cell = ws.cell(row=row, column=sum_col, value=sum_formula)
        sum_cell.font = font_to_use
        sum_cell.number_format = '0.0" kg"'
        sum_cell.fill = computed_fill

        raw_demand_formula = (f'=IF({loss_col_letter}{row}>=100,"błąd: straty>=100%",'
                               f'{annual_col_letter}{row}/(1-{loss_col_letter}{row}/100))')
        rd_cell = ws.cell(row=row, column=raw_demand_col, value=raw_demand_formula)
        rd_cell.font = font_to_use
        rd_cell.number_format = '0.00" t"'
        rd_cell.fill = computed_fill

        if pack_names:
            pack_sum_formula = f"=SUM({first_pack_letter}{row}:{last_pack_letter}{row})"
            ps_cell = ws.cell(row=row, column=pack_sum_col, value=pack_sum_formula)
            ps_cell.font = font_to_use
            ps_cell.number_format = '0.0"%"'
            ps_cell.fill = computed_fill

    red_fill = PatternFill("solid", fgColor="FFC7CE")
    rng = f"{sum_col_letter}{start_data_row}:{sum_col_letter}{start_data_row + total_rows - 1}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f"ABS({sum_col_letter}{start_data_row}-{RECIPE_TARGET_SUM_KG})>{RECIPE_SUM_TOLERANCE_KG}"], fill=red_fill)
    )

    util_rng = f"{utilization_letter}{start_data_row}:{utilization_letter}{start_data_row + total_rows - 1}"
    ws.conditional_formatting.add(util_rng, CellIsRule(operator="greaterThan", formula=[str(RECIPE_UTILIZATION_WARN_PCT)], fill=red_fill))

    shared_util_rng = f"{shared_util_col_letter}{start_data_row}:{shared_util_col_letter}{start_data_row + total_rows - 1}"
    ws.conditional_formatting.add(shared_util_rng, CellIsRule(operator="greaterThan", formula=[str(RECIPE_UTILIZATION_WARN_PCT)], fill=red_fill))

    if pack_names:
        pack_sum_rng = f"{pack_sum_col_letter}{start_data_row}:{pack_sum_col_letter}{start_data_row + total_rows - 1}"
        ws.conditional_formatting.add(
            pack_sum_rng,
            FormulaRule(formula=[f"AND({pack_sum_col_letter}{start_data_row}<>0,ABS({pack_sum_col_letter}{start_data_row}-100)>0.5)"], fill=red_fill)
        )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions[get_column_letter(sourcing_col)].width = 16
    ws.column_dimensions[get_column_letter(transition_col)].width = 18
    ws.column_dimensions[get_column_letter(import_freq_col)].width = 16
    ws.column_dimensions[get_column_letter(import_lot_col)].width = 16
    ws.column_dimensions[get_column_letter(import_safety_col)].width = 16
    for letter in [mixer_vol_letter, cycle_letter, avail_hours_letter, batch_mass_letter, batches_year_letter, utilization_letter]:
        ws.column_dimensions[letter].width = 14
    for col_idx in range(first_mat_col, last_mat_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.column_dimensions[sum_col_letter].width = 14
    ws.column_dimensions[density_col_letter].width = 12
    ws.column_dimensions[loss_col_letter].width = 12
    ws.column_dimensions[raw_demand_col_letter].width = 16
    if pack_names:
        for col_idx in range(first_pack_col, last_pack_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.column_dimensions[pack_sum_col_letter].width = 14
    ws.column_dimensions[get_column_letter(tank_id_col)].width = 16
    ws.column_dimensions[get_column_letter(shared_util_col)].width = 16
    ws.column_dimensions[get_column_letter(notes_col)].width = 30

    # --- Arkusz 'Opakowania' (opcjonalny) - predefiniowanie typów opakowań i pojemności ---
    ws_pack = wb.create_sheet(PACKAGING_SHEET_NAME)
    pack_headers = [PACKAGING_NAME_COL, PACKAGING_SIZE_COL, PACKAGING_PER_PALLET_COL, PACKAGING_RATE_COL, PACKAGING_NOZZLES_COL]
    for col_idx, h in enumerate(pack_headers, start=1):
        cell = ws_pack.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
    ws_pack.row_dimensions[1].height = 32

    pack_row = 2
    for name, cfg in current_pack_defaults.items():
        is_small_pack = "5l" in name.lower() or "1l" in name.lower() or "4l" in name.lower()
        default_nozzles = 4 if is_small_pack else 1
        default_rate_kg_min = 15.0 if is_small_pack else 60.0
        ws_pack.cell(row=pack_row, column=1, value=name).fill = input_fill
        ws_pack.cell(row=pack_row, column=2, value=cfg.get("size_l", 0)).fill = input_fill
        ws_pack.cell(row=pack_row, column=3, value=cfg.get("per_pallet", 0)).fill = input_fill
        ws_pack.cell(row=pack_row, column=4, value=default_rate_kg_min).fill = input_fill
        ws_pack.cell(row=pack_row, column=5, value=default_nozzles).fill = input_fill
        pack_row += 1
    for col, w in zip("ABCDE", [22, 16, 16, 20, 20]):
        ws_pack.column_dimensions[col].width = w

    # --- Arkusz 'Badania Laboratoryjne' (opcjonalny) - testy jako wiersze (z gotowego katalogu
    # aplikacji), przykładowe produkty jako kolumny do zaznaczenia "x". Dopisz kolejne kolumny z
    # DOKŁADNĄ nazwą produktu z arkusza 'Receptury', żeby przypisać im testy. ---
    ws_qc = wb.create_sheet(QC_SHEET_NAME)
    example_qc_products = ["Przykład: SAF", "Przykład: MX"]
    qc_headers = [QC_SHEET_TEST_NAME_COL, "Sprzęt", "Czas [min]"] + example_qc_products
    for col_idx, h in enumerate(qc_headers, start=1):
        cell = ws_qc.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
    ws_qc.row_dimensions[1].height = 32

    qc_row = 2
    for test_name, test_info in QC_TEST_CATALOG.items():
        ws_qc.cell(row=qc_row, column=1, value=test_name).fill = computed_fill
        ws_qc.cell(row=qc_row, column=2, value=test_info["equipment"]).fill = computed_fill
        ws_qc.cell(row=qc_row, column=3, value=test_info["duration_min"]).fill = computed_fill
        for pc in range(len(example_qc_products)):
            ws_qc.cell(row=qc_row, column=4 + pc, value="").fill = input_fill
        qc_row += 1
    # Przykład WŁASNEGO testu spoza wbudowanego katalogu - wystarczy wypełnić 'Sprzęt' i
    # 'Czas [min]', a aplikacja zarejestruje go jako nowy test przy wgraniu pliku. To dokładnie
    # Twój przypadek (drugi aparat do pienienia) - podmień na Twój rzeczywisty sprzęt/czas.
    ws_qc.cell(row=qc_row, column=1, value="XXX Foam Tester").fill = input_fill
    ws_qc.cell(row=qc_row, column=2, value="XXX Foam Tester").fill = input_fill
    ws_qc.cell(row=qc_row, column=3, value=30).fill = input_fill
    for pc in range(len(example_qc_products)):
        ws_qc.cell(row=qc_row, column=4 + pc, value="").fill = input_fill
    qc_row += 1
    for col, w in zip("AB", [40, 40]):
        ws_qc.column_dimensions[col].width = w
    ws_qc.column_dimensions["C"].width = 12
    for pc in range(len(example_qc_products)):
        ws_qc.column_dimensions[get_column_letter(4 + pc)].width = 18
    ws_qc.freeze_panes = "D2"

    # --- Arkusz 'Zużycie Surowców (bez receptury)' (opcjonalny) - dla surowców, których roczne
    # zużycie znasz wprost, bez ujawniania pełnej receptury (ochrona know-how). ---
    ws_direct_rm = wb.create_sheet(DIRECT_RM_SHEET_NAME)
    direct_rm_headers = [DIRECT_RM_GROUP_COL, DIRECT_RM_MATERIAL_COL, DIRECT_RM_CATEGORY_COL, DIRECT_RM_ANNUAL_COL, DIRECT_RM_TANK_ID_COL]
    for col_idx, h in enumerate(direct_rm_headers, start=1):
        cell = ws_direct_rm.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
    ws_direct_rm.row_dimensions[1].height = 32
    # Przykład: dwaj dostawcy tej samej bazy - osobne pozycje (własne śledzenie zużycia), ale
    # WSPÓLNE ID Zbiornika (rotacja dostawców w tym samym fizycznym zbiorniku R-01).
    example_direct_rm = [
        ("Hydraulic Oils", "Baza Grupy I - Dostawca X1", "Base Oil Group I", 700.0, "R-01"),
        ("Hydraulic Oils", "Baza Grupy I - Dostawca X2", "Base Oil Group I", 500.0, "R-01"),
        ("Engine Oils", "Base Oil Group III", "Base Oil Group III", 800.0, ""),
    ]
    for row_idx, (grp, mat, cat, val, tank_id) in enumerate(example_direct_rm, start=2):
        ws_direct_rm.cell(row=row_idx, column=1, value=grp).fill = input_fill
        ws_direct_rm.cell(row=row_idx, column=2, value=mat).fill = input_fill
        ws_direct_rm.cell(row=row_idx, column=3, value=cat).fill = input_fill
        ws_direct_rm.cell(row=row_idx, column=4, value=val).fill = input_fill
        ws_direct_rm.cell(row=row_idx, column=5, value=tank_id).fill = input_fill
    for col, w in zip("ABCDE", [22, 32, 26, 20, 30]):
        ws_direct_rm.column_dimensions[col].width = w

    # --- Arkusz 'Rozbicie Dostawców Surowców' (opcjonalny) - dzieli CAŁKOWITE zużycie jednej
    # kategorii (np. Base Oil Group II, policzone z pełnych receptur) na kilku konkretnych
    # dostawców/baz, żeby pokazać kilka mniejszych zbiorników zamiast jednego dużego. ---
    ws_supplier = wb.create_sheet(SUPPLIER_SPLIT_SHEET_NAME)
    supplier_headers = [SUPPLIER_SPLIT_GROUP_COL, SUPPLIER_SPLIT_CATEGORY_COL, SUPPLIER_SPLIT_SUPPLIER_COL,
                         SUPPLIER_SPLIT_PCT_COL, SUPPLIER_SPLIT_TECH_COL, SUPPLIER_SPLIT_TANK_ID_COL]
    for col_idx, h in enumerate(supplier_headers, start=1):
        cell = ws_supplier.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
    ws_supplier.row_dimensions[1].height = 32
    # Przykład: rozbicie "całego zakładu" (puste pole grupy) dla Base Oil Group II - 60/40 na
    # dwóch dostawców, sumujące się do 100%.
    example_supplier = [
        ("", "Base Oil Group II", "Baza Grupy II - Dostawca X1", 60.0, "", ""),
        ("", "Base Oil Group II", "Baza Grupy II - Dostawca X2", 40.0, "", ""),
    ]
    for row_idx, (grp, cat, sup, pct, tech, tank_id) in enumerate(example_supplier, start=2):
        ws_supplier.cell(row=row_idx, column=1, value=grp).fill = input_fill
        ws_supplier.cell(row=row_idx, column=2, value=cat).fill = input_fill
        ws_supplier.cell(row=row_idx, column=3, value=sup).fill = input_fill
        ws_supplier.cell(row=row_idx, column=4, value=pct).fill = input_fill
        ws_supplier.cell(row=row_idx, column=5, value=tech).fill = input_fill
        ws_supplier.cell(row=row_idx, column=6, value=tank_id).fill = input_fill
    for col, w in zip("ABCDEF", [26, 22, 26, 18, 22, 30]):
        ws_supplier.column_dimensions[col].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def parse_recipe_excel(uploaded_file):
    """
    Wczytuje wgrany plik Excel z arkusza 'Receptury' (lub pierwszego arkusza, jeśli nazwa
    inna) i zwraca (df_czysty, lista_bledow). Waliduje: obecność wymaganych kolumn, brak
    wartości ujemnych, brak pustego produktu/rocznego zapotrzebowania, poprawność grupy
    produktowej, oraz sumę dozowania surowców w tolerancji wokół 1000 kg/t. Kolumny
    wyliczane ('Suma Udziałów Składników', 'Roczne Zapotrzebowanie Surowcowe') są PRZELICZANE
    w Pythonie (nie czytane z formuł Excela), żeby działać niezależnie od tego, czy plik był
    wcześniej przeliczony przez Excel/LibreOffice.
    """
    errors = []
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_name = "Receptury" if "Receptury" in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name)
    except Exception as exc:
        return None, [f"Nie udało się odczytać pliku Excel: {exc}"]

    required_cols = [RECIPE_GROUP_COL, RECIPE_PRODUCT_COL, RECIPE_ANNUAL_COL] + RECIPE_RAW_MATERIALS + [
        RECIPE_DENSITY_COL, RECIPE_LOSS_COL]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        return None, [f"W pliku brakuje wymaganych kolumn: {', '.join(missing_cols)}. "
                       f"Pobierz i użyj oficjalnego szablonu, nie zmieniając nazw kolumn."]

    if RECIPE_NOTES_COL not in df.columns:
        df[RECIPE_NOTES_COL] = ""
    if RECIPE_TANK_ID_COL not in df.columns:
        df[RECIPE_TANK_ID_COL] = ""
    else:
        df[RECIPE_TANK_ID_COL] = df[RECIPE_TANK_ID_COL].fillna("").astype(str).str.strip()

    # Sposób pozyskania (Import/Produkcja własna) - opcjonalne kolumny, dla zgodności wstecznej
    # z plikami wygenerowanymi starszą wersją szablonu (bez tych kolumn -> wszystko "Produkcja
    # własna", zero zmiany zachowania).
    if RECIPE_SOURCING_COL not in df.columns:
        df[RECIPE_SOURCING_COL] = "Produkcja własna"
    else:
        df[RECIPE_SOURCING_COL] = df[RECIPE_SOURCING_COL].fillna("Produkcja własna").astype(str).str.strip()
        invalid_sourcing = df[~df[RECIPE_SOURCING_COL].isin(RECIPE_SOURCING_OPTIONS)]
        if not invalid_sourcing.empty:
            bad = invalid_sourcing[RECIPE_PRODUCT_COL].tolist()
            errors.append(f"Nieznany '{RECIPE_SOURCING_COL}' (musi być jednym z: {', '.join(RECIPE_SOURCING_OPTIONS)}) dla: "
                           f"{', '.join(map(str, bad))} - przyjęto 'Produkcja własna'.")
            df.loc[~df[RECIPE_SOURCING_COL].isin(RECIPE_SOURCING_OPTIONS), RECIPE_SOURCING_COL] = "Produkcja własna"
    if RECIPE_IMPORT_TRANSITION_COL not in df.columns:
        df[RECIPE_IMPORT_TRANSITION_COL] = ""
    else:
        df[RECIPE_IMPORT_TRANSITION_COL] = df[RECIPE_IMPORT_TRANSITION_COL].fillna("").astype(str).str.strip()
        import_mask = df[RECIPE_SOURCING_COL] == "Import"
        invalid_transition = df[import_mask & ~df[RECIPE_IMPORT_TRANSITION_COL].isin(RECIPE_IMPORT_TRANSITION_OPTIONS)]
        if not invalid_transition.empty:
            bad = invalid_transition[RECIPE_PRODUCT_COL].tolist()
            errors.append(f"Produkty importowane bez poprawnego '{RECIPE_IMPORT_TRANSITION_COL}' (jeden z: "
                           f"{', '.join(RECIPE_IMPORT_TRANSITION_OPTIONS)}): {', '.join(map(str, bad))} - przyjęto 'Rok 1'.")
            bad_idx = invalid_transition.index
            df.loc[bad_idx, RECIPE_IMPORT_TRANSITION_COL] = "Rok 1"
    for col in [RECIPE_IMPORT_FREQ_COL, RECIPE_IMPORT_LOT_COL, RECIPE_IMPORT_SAFETY_DAYS_COL]:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    df = df[df[RECIPE_PRODUCT_COL].notna()].copy()
    df = df[~df[RECIPE_PRODUCT_COL].astype(str).str.startswith("Przykład")].copy()

    if df.empty:
        return None, ["Plik nie zawiera żadnych wierszy produktów poza przykładami."]

    import_no_lot = df[(df[RECIPE_SOURCING_COL] == "Import") & (df[RECIPE_IMPORT_LOT_COL] <= 0)]
    if not import_no_lot.empty:
        bad = import_no_lot[RECIPE_PRODUCT_COL].tolist()
        errors.append(f"Produkty importowane bez podanej '{RECIPE_IMPORT_LOT_COL}' (>0): {', '.join(map(str, bad))} - "
                       f"bufor magazynowy importu w Zakładce 3 dla nich nie policzy się poprawnie, dopóki nie uzupełnisz tej wartości.")

    for mat in RECIPE_RAW_MATERIALS:
        df[mat] = pd.to_numeric(df[mat], errors="coerce").fillna(0.0)

    df[RECIPE_ANNUAL_COL] = pd.to_numeric(df[RECIPE_ANNUAL_COL], errors="coerce")
    missing_annual = df[df[RECIPE_ANNUAL_COL].isna()][RECIPE_PRODUCT_COL].tolist()
    if missing_annual:
        errors.append(f"Brak / niepoprawne roczne zapotrzebowanie dla: {', '.join(map(str, missing_annual))}.")
        df = df[df[RECIPE_ANNUAL_COL].notna()].copy()

    df[RECIPE_DENSITY_COL] = pd.to_numeric(df[RECIPE_DENSITY_COL], errors="coerce")
    missing_density = df[df[RECIPE_DENSITY_COL].isna()][RECIPE_PRODUCT_COL].tolist()
    if missing_density:
        errors.append(f"Brak / niepoprawna gęstość dla: {', '.join(map(str, missing_density))} - przyjęto 0.88 g/cm³.")
        df[RECIPE_DENSITY_COL] = df[RECIPE_DENSITY_COL].fillna(0.88)

    df[RECIPE_LOSS_COL] = pd.to_numeric(df[RECIPE_LOSS_COL], errors="coerce").fillna(0.0)
    invalid_loss = df[(df[RECIPE_LOSS_COL] < 0) | (df[RECIPE_LOSS_COL] >= 100)][RECIPE_PRODUCT_COL].tolist()
    if invalid_loss:
        errors.append(f"Nieprawidłowe straty procesowe (muszą być 0-99.9%) dla: {', '.join(map(str, invalid_loss))} - przyjęto 0%.")
        df.loc[(df[RECIPE_LOSS_COL] < 0) | (df[RECIPE_LOSS_COL] >= 100), RECIPE_LOSS_COL] = 0.0

    # Wymiarowanie mieszalnika (opcjonalne - pliki wygenerowane starszą wersją szablonu mogą
    # nie mieć tych kolumn; wtedy wykorzystanie po prostu nie jest liczone dla tych wierszy).
    for col in [RECIPE_MIXER_VOL_COL, RECIPE_CYCLE_COL, RECIPE_AVAIL_HOURS_COL]:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    # Rozbicie na opakowania (opcjonalne, dynamiczne - dowolna liczba kolumn 'Opak: ... [%]').
    pack_cols_found = [c for c in df.columns if c.startswith("Opak: ") and c.endswith(" [%]")]
    for c in pack_cols_found:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
        negative_pack = df[df[c] < 0][RECIPE_PRODUCT_COL].tolist()
        if negative_pack:
            errors.append(f"Ujemny udział opakowania '{c}' dla: {', '.join(map(str, negative_pack))} - przyjęto 0%.")
            df.loc[df[c] < 0, c] = 0.0
    if pack_cols_found:
        df[RECIPE_PACK_SUM_COL] = df[pack_cols_found].sum(axis=1)
        bad_pack_sum = df[(df[RECIPE_PACK_SUM_COL] != 0) & ((df[RECIPE_PACK_SUM_COL] - 100).abs() > 0.5)]
        if not bad_pack_sum.empty:
            details = ", ".join(f"{r[RECIPE_PRODUCT_COL]} ({r[RECIPE_PACK_SUM_COL]:.0f}%)" for _, r in bad_pack_sum.iterrows())
            errors.append(f"Suma % opakowań różni się od 100% dla: {details} - rozbicie na opakowania dla tych wierszy zignorowane "
                           f"(produkcja nadal zaliczona), popraw jeśli chcesz z niego korzystać.")
    else:
        df[RECIPE_PACK_SUM_COL] = 0.0

    # Testy QC per KONKRETNY produkt (opcjonalne, dynamiczne - dowolna liczba kolumn
    # 'QC: {nazwa testu} [x]'), analogicznie do opakowań. Wartość "x"/1/TRUE/"tak" = test dotyczy
    # tego produktu; jeśli produkt nie ma ŻADNEJ takiej kolumny wypełnionej, panel zwolnienia
    # nadal można skonfigurować per LINIA w Zakładce 6 (VSM) jako dotychczasowy fallback.
    qc_cols_found = [c for c in df.columns if c.startswith(QC_COL_PREFIX) and c.endswith(QC_COL_SUFFIX)]
    unknown_qc_tests = [c for c in qc_cols_found if c[len(QC_COL_PREFIX):-len(QC_COL_SUFFIX)] not in QC_TEST_CATALOG
                         and c[len(QC_COL_PREFIX):-len(QC_COL_SUFFIX)] not in st.session_state.get("custom_qc_tests", {})]
    if unknown_qc_tests:
        errors.append(f"Nieznane testy QC w kolumnach: {', '.join(unknown_qc_tests)} - sprawdź pisownię względem "
                       "katalogu testów (Zakładka 6, VSM). Kolumny te są ignorowane.")
        qc_cols_found = [c for c in qc_cols_found if c not in unknown_qc_tests]
    for c in qc_cols_found:
        df[c] = df[c].astype(str).str.strip().str.lower().isin(["x", "1", "true", "tak", "yes"])

    unknown_group_mask = ~df[RECIPE_GROUP_COL].astype(str).isin(RECIPE_PRODUCT_GROUPS)
    if unknown_group_mask.any():
        bad = df.loc[unknown_group_mask, RECIPE_PRODUCT_COL].tolist()
        errors.append(f"Nieznana/brakująca grupa produktowa (musi być jedną z: {', '.join(RECIPE_PRODUCT_GROUPS)}) dla: "
                       f"{', '.join(map(str, bad))}. Wiersze pominięte.")
        df = df[~unknown_group_mask].copy()

    negative_mask = (df[RECIPE_RAW_MATERIALS] < 0).any(axis=1)
    if negative_mask.any():
        bad = df.loc[negative_mask, RECIPE_PRODUCT_COL].tolist()
        errors.append(f"Ujemne wartości dozowania [kg/t] dla: {', '.join(map(str, bad))}.")
        df = df[~negative_mask].copy()

    df[RECIPE_SUM_COL] = df[RECIPE_RAW_MATERIALS].sum(axis=1)
    bad_sum_mask = (df[RECIPE_SUM_COL] - RECIPE_TARGET_SUM_KG).abs() > RECIPE_SUM_TOLERANCE_KG
    # Produkty importowane NA STAŁE ("Nigdy (stały import)" / "Nigdy (bufor)") mają celowo pusty/
    # zerowy skład - to nie surowce do produkcji, tylko gotowy import, więc nie sprawdzamy dla nich
    # sumy dozowania i NIE usuwamy tych wierszy (inaczej znikałyby z całej reszty obliczeń, w tym
    # z sizingu zbiornika buforowego w Zakładce 2).
    if RECIPE_SOURCING_COL in df.columns and RECIPE_IMPORT_TRANSITION_COL in df.columns:
        permanent_import_exempt = (df[RECIPE_SOURCING_COL] == "Import") & \
                                   (df[RECIPE_IMPORT_TRANSITION_COL].isin(["Nigdy (stały import)", "Nigdy (bufor)"]))
        bad_sum_mask = bad_sum_mask & ~permanent_import_exempt
    if bad_sum_mask.any():
        bad_rows = df.loc[bad_sum_mask, [RECIPE_PRODUCT_COL, RECIPE_SUM_COL]]
        details = ", ".join(f"{r[RECIPE_PRODUCT_COL]} ({r[RECIPE_SUM_COL]:.0f} kg/t)" for _, r in bad_rows.iterrows())
        errors.append(f"Suma dozowania surowców odbiega od 1000 kg/t o więcej niż {RECIPE_SUM_TOLERANCE_KG:.0f} kg dla: "
                       f"{details}. Te wiersze zostały pominięte.")
        df = df[~bad_sum_mask].copy()

    dup_mask = df[RECIPE_PRODUCT_COL].duplicated(keep=False)
    if dup_mask.any():
        dups = sorted(set(df.loc[dup_mask, RECIPE_PRODUCT_COL].astype(str)))
        errors.append(f"Zduplikowane nazwy produktów (potraktowane jako oddzielne pozycje): {', '.join(dups)}.")

    if df.empty:
        return None, errors

    # Roczne zapotrzebowanie surowcowe = zapotrzebowanie na GOTOWY produkt / (1 - straty%) -
    # ile surowca faktycznie trzeba zakupić, żeby po stratach procesowych uzyskać zakładany
    # wolumen produktu. Liczone w Pythonie, nie czytane z formuły Excela (patrz docstring).
    df[RECIPE_RAW_DEMAND_COL] = df[RECIPE_ANNUAL_COL] / (1.0 - df[RECIPE_LOSS_COL] / 100.0)

    # Masa Szarży [kg] = pojemność mieszalnika [m3] * 1000 (L/m3) * gęstość [kg/L] - liczone
    # w Pythonie tak samo jak w Excelu, z zabezpieczeniem przed dzieleniem przez zero.
    df[RECIPE_BATCH_MASS_COL] = df[RECIPE_MIXER_VOL_COL] * 1000.0 * df[RECIPE_DENSITY_COL]
    df[RECIPE_BATCHES_YEAR_COL] = 0.0
    has_batch_mass = df[RECIPE_BATCH_MASS_COL] > 0
    df.loc[has_batch_mass, RECIPE_BATCHES_YEAR_COL] = (
        df.loc[has_batch_mass, RECIPE_ANNUAL_COL] * 1000.0 / df.loc[has_batch_mass, RECIPE_BATCH_MASS_COL])

    df[RECIPE_UTILIZATION_COL] = 0.0
    has_avail_hours = df[RECIPE_AVAIL_HOURS_COL] > 0
    df.loc[has_avail_hours, RECIPE_UTILIZATION_COL] = (
        df.loc[has_avail_hours, RECIPE_BATCHES_YEAR_COL] * df.loc[has_avail_hours, RECIPE_CYCLE_COL]
        / df.loc[has_avail_hours, RECIPE_AVAIL_HOURS_COL] * 100.0)

    overloaded = df[df[RECIPE_UTILIZATION_COL] > RECIPE_UTILIZATION_WARN_PCT]
    if not overloaded.empty:
        details = ", ".join(f"{r[RECIPE_PRODUCT_COL]} ({r[RECIPE_UTILIZATION_COL]:.0f}%)" for _, r in overloaded.iterrows())
        errors.append(f"⚠️ Wykorzystanie mieszalnika powyżej {RECIPE_UTILIZATION_WARN_PCT:.0f}% (przeciążenie) dla: {details}.")

    return df.reset_index(drop=True), errors


def parse_packaging_excel(uploaded_file):
    """
    Wczytuje opcjonalny arkusz 'Opakowania' z tego samego pliku Excel co receptury.
    Zwraca (dict_opakowan, lista_bledow). dict_opakowan ma strukturę zgodną z PACK_CONFIGS
    (size_l, per_pallet, rate_szt_h) rozszerzoną o domyślne nozzles/speed_kg_min do
    prekonfiguracji sekcji rozlewu w Zakładce 3. Jeśli arkusz nie istnieje w pliku, zwraca
    (None, []) po cichu - to pole jest opcjonalne, nie każdy plik musi go zawierać.
    """
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as exc:
        return None, [f"Nie udało się odczytać pliku Excel (opakowania): {exc}"]

    if PACKAGING_SHEET_NAME not in xls.sheet_names:
        return None, []

    try:
        df = pd.read_excel(xls, sheet_name=PACKAGING_SHEET_NAME)
    except Exception as exc:
        return None, [f"Nie udało się odczytać arkusza '{PACKAGING_SHEET_NAME}': {exc}"]

    required_cols = [PACKAGING_NAME_COL, PACKAGING_SIZE_COL, PACKAGING_PER_PALLET_COL]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        return None, [f"Arkusz '{PACKAGING_SHEET_NAME}' istnieje, ale brakuje kolumn: {', '.join(missing_cols)}. Pominięto import opakowań."]

    df = df[df[PACKAGING_NAME_COL].notna()].copy()
    if df.empty:
        return None, []

    errors = []
    df[PACKAGING_SIZE_COL] = pd.to_numeric(df[PACKAGING_SIZE_COL], errors="coerce")
    df[PACKAGING_PER_PALLET_COL] = pd.to_numeric(df[PACKAGING_PER_PALLET_COL], errors="coerce")
    if PACKAGING_RATE_COL in df.columns:
        df[PACKAGING_RATE_COL] = pd.to_numeric(df[PACKAGING_RATE_COL], errors="coerce")
    if PACKAGING_NOZZLES_COL in df.columns:
        df[PACKAGING_NOZZLES_COL] = pd.to_numeric(df[PACKAGING_NOZZLES_COL], errors="coerce")

    bad_rows = df[df[PACKAGING_SIZE_COL].isna() | df[PACKAGING_PER_PALLET_COL].isna() |
                  (df[PACKAGING_SIZE_COL] <= 0) | (df[PACKAGING_PER_PALLET_COL] <= 0)]
    if not bad_rows.empty:
        bad_names = bad_rows[PACKAGING_NAME_COL].astype(str).tolist()
        errors.append(f"Pominięto opakowania z brakującą/niepoprawną pojemnością lub liczbą sztuk na palecie: {', '.join(bad_names)}.")
        df = df.drop(bad_rows.index)

    packaging_dict = {}
    filling_defaults = {}
    for _, row in df.iterrows():
        name = str(row[PACKAGING_NAME_COL])
        packaging_dict[name] = {
            "size_l": float(row[PACKAGING_SIZE_COL]),
            "per_pallet": int(row[PACKAGING_PER_PALLET_COL]),
            "rate_szt_h": 0,
        }
        rate = row.get(PACKAGING_RATE_COL) if PACKAGING_RATE_COL in df.columns else None
        nozzles = row.get(PACKAGING_NOZZLES_COL) if PACKAGING_NOZZLES_COL in df.columns else None
        filling_defaults[name] = {
            "speed_kg_min": float(rate) if pd.notna(rate) else 30.0,
            "nozzles": float(nozzles) if pd.notna(nozzles) else 1.0,
        }

    return {"pack_configs": packaging_dict, "filling_defaults": filling_defaults}, errors


def parse_qc_tests_excel(uploaded_file):
    """
    Wczytuje opcjonalny arkusz 'Badania Laboratoryjne' z tego samego pliku Excel co receptury.
    Struktura ODWROTNA niż receptura: TESTY jako WIERSZE (kolumna 'Nazwa Testu'), PRODUKTY jako
    KOLUMNY (nagłówek = dokładna nazwa produktu z arkusza 'Receptury') - "x" w komórce = ten test
    dotyczy tego produktu.

    WAŻNE: katalog testów NIE jest już zamknięty na wbudowaną listę (QC_TEST_CATALOG) - jeśli
    "Nazwa Testu" nie pasuje do żadnego znanego testu, ale wiersz ma wypełnione własne kolumny
    'Sprzęt' i 'Czas [min]', REJESTRUJEMY to jako NOWY, własny test (z Twoim czasem i sprzętem),
    zamiast odrzucać jako "nieznany". Tylko wiersz bez ŻADNEGO dopasowania i bez tych dwóch
    kolumn wypełnionych faktycznie się odrzuca.

    Zwraca (dict_produkt_na_liste_testów, dict_nowych_testów_do_zarejestrowania, lista_błędów).
    dict_nowych_testów ma format {nazwa_testu: {"duration_min", "equipment", "count": 1}} -
    scal go do session_state.custom_qc_tests w kodzie wywołującym. Jeśli arkusz nie istnieje,
    (None, {}, []).
    """
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as exc:
        return None, {}, [f"Nie udało się odczytać pliku Excel (badania laboratoryjne): {exc}"]

    if QC_SHEET_NAME not in xls.sheet_names:
        return None, {}, []

    try:
        df = pd.read_excel(xls, sheet_name=QC_SHEET_NAME)
    except Exception as exc:
        return None, {}, [f"Nie udało się odczytać arkusza '{QC_SHEET_NAME}': {exc}"]

    if QC_SHEET_TEST_NAME_COL not in df.columns:
        return None, {}, [f"Arkusz '{QC_SHEET_NAME}' istnieje, ale brakuje kolumny '{QC_SHEET_TEST_NAME_COL}'. Pominięto import badań."]

    df = df[df[QC_SHEET_TEST_NAME_COL].notna()].copy()
    if df.empty:
        return None, {}, []

    equipment_col = next((c for c in ["Sprzęt", "Equipment Description"] if c in df.columns), None)
    duration_col = next((c for c in ["Czas [min]", "Avr. Time (min)"] if c in df.columns), None)

    def _clean(val):
        return "" if pd.isna(val) else str(val).strip()

    errors = []
    new_tests = {}
    known_tests = set(QC_TEST_CATALOG.keys()) | set(st.session_state.get("custom_qc_tests", {}).keys())
    rows_to_drop = []
    for idx, row in df.iterrows():
        test_name = str(row[QC_SHEET_TEST_NAME_COL])
        if test_name in known_tests:
            continue
        equip_val = _clean(row.get(equipment_col, "")) if equipment_col else ""
        duration_val = row.get(duration_col, None) if duration_col else None
        duration_ok = duration_val is not None and not pd.isna(duration_val) and float(duration_val) > 0
        if equip_val and duration_ok:
            new_tests[test_name] = {"duration_min": float(duration_val), "equipment": equip_val, "count": 1}
            known_tests.add(test_name)  # żeby kolejne wiersze o tej samej nazwie już nie liczyły się jako nowe
        else:
            errors.append(f"Nieznany test '{test_name}' w arkuszu '{QC_SHEET_NAME}' - uzupełnij kolumny 'Sprzęt' i "
                           "'Czas [min]' dla tego wiersza, żeby zarejestrować go jako nowy własny test, albo popraw "
                           "pisownię, jeśli to literówka istniejącego testu. Wiersz pominięty.")
            rows_to_drop.append(idx)

    if rows_to_drop:
        df = df.drop(index=rows_to_drop)
    if df.empty:
        return None, new_tests, errors

    # Kolumny produktowe = wszystko poza opisowymi (Nazwa Testu, Sprzęt, Czas - te dwie ostatnie
    # teraz FAKTYCZNIE odczytywane wyżej, nie tylko pomijane).
    descriptive_cols = {QC_SHEET_TEST_NAME_COL, "No.", "Sprzęt", "Equipment Description", "Avr. Time (min)", "Czas [min]"}
    product_cols = [c for c in df.columns if c not in descriptive_cols]

    tests_by_product = {}
    for _, row in df.iterrows():
        test_name = str(row[QC_SHEET_TEST_NAME_COL])
        for prod_col in product_cols:
            val = str(row.get(prod_col, "")).strip().lower()
            if val in ("x", "1", "true", "tak", "yes"):
                existing = tests_by_product.setdefault(str(prod_col), [])
                if test_name not in existing:  # unikaj podwójnego liczenia tego samego testu,
                    existing.append(test_name)  # jeśli pojawia się na więcej niż jednym wierszu

    return tests_by_product, new_tests, errors


def parse_direct_raw_materials_excel(uploaded_file):
    """
    Wczytuje opcjonalny arkusz 'Zużycie Surowców (bez receptury)' - dla surowców, których roczne
    zużycie znasz wprost (np. "zużywamy X ton Bazy Grupy I od Dostawcy A rocznie"), ale NIE
    chcesz/możesz ujawnić pełnej receptury produktu z uwagi na ochronę know-how.

    Nazwa Surowca jest DOWOLNA (np. "Baza Grupy I - Dostawca X1", "Baza Grupy I - Dostawca X2" jako
    dwie ODRĘBNE pozycje - własne zbiorniki/cysterny/zużycie) - nie musi pasować do sztywnej listy
    kategorii z Zakładki 1. Kategoria (opcjonalna) łączy to z jedną z tych kategorii WYŁĄCZNIE do
    raportów zbiorczych - sam dobór zbiornika działa na konkretnej nazwie, nie na kategorii.

    ID Zbiornika (opcjonalne) - TA SAMA konwencja co w recepturze (ID Zbiornika): te samo ID w
    kilku wierszach = te surowce mają dzielić JEDEN fizyczny zbiornik (np. rotacja dostawców w
    tym samym zbiorniku), zamiast dostawać osobne zbiorniki każdy.

    Ochrona przed literówką: przy wielu dostawcach łatwo o "Baza I - Dostawca A" i "Baza I -
    dostawca a " (inna wielkość liter/spacje) wpisane przez pomyłkę jako DWIE różne pozycje,
    zamiast jednej - wykrywamy to (znormalizowane porównanie: małe litery + pojedyncze spacje) i
    ostrzegamy, zamiast po cichu liczyć je jako dwa osobne, niezamierzone surowce.

    Zwraca (lista_wpisów, lista_błędów), gdzie wpis to dict {"group","material","category",
    "tank_id","annual_t"}. Jeśli arkusz nie istnieje, ([], []).
    """
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as exc:
        return [], [f"Nie udało się odczytać pliku Excel (zużycie surowców bez receptury): {exc}"]

    if DIRECT_RM_SHEET_NAME not in xls.sheet_names:
        return [], []

    try:
        df = pd.read_excel(xls, sheet_name=DIRECT_RM_SHEET_NAME)
    except Exception as exc:
        return [], [f"Nie udało się odczytać arkusza '{DIRECT_RM_SHEET_NAME}': {exc}"]

    required_cols = {DIRECT_RM_GROUP_COL, DIRECT_RM_MATERIAL_COL, DIRECT_RM_ANNUAL_COL}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        return [], [f"Arkusz '{DIRECT_RM_SHEET_NAME}' istnieje, ale brakuje kolumn: {', '.join(missing_cols)}. Pominięto import."]

    df = df[df[DIRECT_RM_MATERIAL_COL].notna()].copy()
    if df.empty:
        return [], []

    errors = []
    known_categories = {m.replace(" [kg/t]", "") for m in RECIPE_RAW_MATERIALS}
    entries = []
    seen_normalized = {}  # znormalizowana nazwa -> oryginalna nazwa (pierwsze wystąpienie)

    def _normalize(name):
        return " ".join(name.lower().split())

    for _, row in df.iterrows():
        material_raw = str(row[DIRECT_RM_MATERIAL_COL]).strip()
        if not material_raw:
            continue

        norm = _normalize(material_raw)
        if norm in seen_normalized and seen_normalized[norm] != material_raw:
            errors.append(f"⚠️ Podobne nazwy surowca w arkuszu '{DIRECT_RM_SHEET_NAME}': '{material_raw}' i "
                           f"'{seen_normalized[norm]}' różnią się tylko wielkością liter/spacjami - to prawdopodobnie "
                           "literówka, nie dwa różne surowce. Ujednolić pisownię, żeby nie liczyły się osobno.")
        else:
            seen_normalized[norm] = material_raw

        group_val = row.get(DIRECT_RM_GROUP_COL, "")
        group_raw = "" if pd.isna(group_val) else str(group_val).strip()
        if group_raw and group_raw not in RECIPE_PRODUCT_GROUPS:
            errors.append(f"Nieznana grupa produktowa '{group_raw}' dla surowca '{material_raw}' w arkuszu "
                           f"'{DIRECT_RM_SHEET_NAME}' - sprawdź pisownię. Wiersz pominięty.")
            continue
        # Puste pole = OK - surowiec skaluje się wtedy wspólną, ogólną krzywą rozruchu całego
        # zakładu zamiast krzywej konkretnej linii (przydatne, gdy nie wiesz, do którego
        # produktu trafia dany surowiec - dokładnie scenariusz ochrony know-how).

        category_val = row.get(DIRECT_RM_CATEGORY_COL, "")
        category_raw = "" if pd.isna(category_val) else str(category_val).strip()
        if category_raw and category_raw not in known_categories:
            errors.append(f"Nieznana kategoria '{category_raw}' dla surowca '{material_raw}' - sprawdź pisownię "
                           "względem listy kategorii (Zakładka 1). Wiersz zaimportowany, ale bez kategorii "
                           "(nie trafi do zbiorczych raportów per kategoria).")
            category_raw = ""

        tank_id_val = row.get(DIRECT_RM_TANK_ID_COL, "")
        tank_id_raw = "" if pd.isna(tank_id_val) else str(tank_id_val).strip()

        annual_t = float(row.get(DIRECT_RM_ANNUAL_COL, 0) or 0)
        if annual_t <= 0:
            continue
        entries.append({"group": group_raw, "material": material_raw, "category": category_raw,
                         "tank_id": tank_id_raw, "annual_t": annual_t})

    return entries, errors


def parse_supplier_split_excel(uploaded_file):
    """
    Wczytuje opcjonalny arkusz 'Rozbicie Dostawców Surowców' - dla kategorii surowca (np. 'Base
    Oil Group II'), której CAŁKOWITE zużycie jest już policzone (z pełnych receptur i/lub z
    arkusza 'Zużycie Surowców bez receptury'), ale w rzeczywistości pochodzi z KILKU różnych
    konkretnych baz/dostawców - i chcesz to zobaczyć jako kilka mniejszych zbiorników zamiast
    jednego dużego.

    Rozbicie jest na poziomie KATEGORII (nie pojedynczego produktu) - jedna wspólna mieszanka
    dostawców stosowana do całego zużycia tej kategorii w całym zakładzie, chyba że w kolumnie
    'Grupa Produktowa' wskażesz konkretną linię - wtedy rozbicie dotyczy TYLKO zużycia tej
    kategorii pochodzącego z tamtej linii (reszta zakładu, jeśli też korzysta z tej kategorii,
    zostaje nierozbita, jako jedna pozycja). Zostaw puste, żeby rozbicie objęło cały zakład -
    to najszybsza droga do przetestowania wariantu (zmieniasz 2-3 liczby, nie dziesiątki wierszy).

    Zwraca (lista_wpisów, lista_błędów). Wpis: {"group" (lub "" dla całego zakładu), "category",
    "supplier", "pct", "tech", "tank_id"}.
    """
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as exc:
        return [], [f"Nie udało się odczytać pliku Excel (rozbicie dostawców surowców): {exc}"]

    if SUPPLIER_SPLIT_SHEET_NAME not in xls.sheet_names:
        return [], []

    try:
        df = pd.read_excel(xls, sheet_name=SUPPLIER_SPLIT_SHEET_NAME)
    except Exception as exc:
        return [], [f"Nie udało się odczytać arkusza '{SUPPLIER_SPLIT_SHEET_NAME}': {exc}"]

    required_cols = {SUPPLIER_SPLIT_CATEGORY_COL, SUPPLIER_SPLIT_SUPPLIER_COL, SUPPLIER_SPLIT_PCT_COL}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        return [], [f"Arkusz '{SUPPLIER_SPLIT_SHEET_NAME}' istnieje, ale brakuje kolumn: {', '.join(missing_cols)}. Pominięto import."]

    df = df[df[SUPPLIER_SPLIT_SUPPLIER_COL].notna()].copy()
    if df.empty:
        return [], []

    def _clean(val):
        return "" if pd.isna(val) else str(val).strip()

    errors = []
    known_categories = {m.replace(" [kg/t]", "") for m in RECIPE_RAW_MATERIALS}
    raw_entries = []
    for _, row in df.iterrows():
        category_raw = _clean(row.get(SUPPLIER_SPLIT_CATEGORY_COL, ""))
        if category_raw not in known_categories:
            errors.append(f"Nieznana kategoria API '{category_raw}' w arkuszu '{SUPPLIER_SPLIT_SHEET_NAME}' - sprawdź "
                           "pisownię względem listy surowców (Zakładka 1). Wiersz pominięty.")
            continue
        group_raw = _clean(row.get(SUPPLIER_SPLIT_GROUP_COL, ""))
        if group_raw and group_raw not in RECIPE_PRODUCT_GROUPS:
            errors.append(f"Nieznana grupa produktowa '{group_raw}' w arkuszu '{SUPPLIER_SPLIT_SHEET_NAME}' - sprawdź "
                           "pisownię. Wiersz pominięty.")
            continue
        supplier_raw = _clean(row.get(SUPPLIER_SPLIT_SUPPLIER_COL, ""))
        if not supplier_raw:
            continue
        pct_raw = float(row.get(SUPPLIER_SPLIT_PCT_COL, 0) or 0)
        if pct_raw <= 0:
            continue
        raw_entries.append({
            "group": group_raw, "category": category_raw, "supplier": supplier_raw, "pct": pct_raw,
            "tech": _clean(row.get(SUPPLIER_SPLIT_TECH_COL, "")), "tank_id": _clean(row.get(SUPPLIER_SPLIT_TANK_ID_COL, "")),
        })

    # Walidacja sumy % w obrębie każdej kombinacji (grupa, kategoria) - jeśli nie sumuje się do
    # ~100%, rozbicie dawałoby błędne (za małe lub za duże) zużycie per dostawca.
    sums = {}
    for e in raw_entries:
        key = (e["group"], e["category"])
        sums[key] = sums.get(key, 0.0) + e["pct"]
    for (grp, cat), total_pct in sums.items():
        if abs(total_pct - 100.0) > 1.0:
            scope_txt = f"linii '{grp}'" if grp else "całego zakładu"
            errors.append(f"⚠️ Rozbicie dostawców dla '{cat}' ({scope_txt}) sumuje się do {total_pct:.0f}%, nie 100% - "
                           "popraw udziały, inaczej zużycie per dostawca będzie błędnie przeskalowane.")

    return raw_entries, errors

# Cennik / Standardowa Instalacja (BOM) - osobny, aktualizowalny arkusz per grupa produktowa,
# z ceną jednostkową komponentu; źródłem treści jest zwykle istniejący P&ID danej instalacji
# standardowej (użytkownik przepisuje stamtąd listę urządzeń, nie generujemy P&ID w apce).
EQUIPMENT_SHEET_NAME = "Cennik Instalacji"
EQUIPMENT_GROUP_COL = "Grupa Produktowa"
EQUIPMENT_COMPONENT_COL = "Komponent"
EQUIPMENT_MODEL_COL = "Typ / Model"
EQUIPMENT_SUPPLIER_COL = "Dostawca"
EQUIPMENT_QTY_COL = "Ilość na instalację [szt]"
EQUIPMENT_UNIT_PRICE_COL = "Cena jednostkowa"
EQUIPMENT_CURRENCY_COL = "Waluta"
EQUIPMENT_NOTES_COL = "Uwagi"
EQUIPMENT_LINE_TOTAL_COL = "Wartość pozycji"


def generate_equipment_template_bytes():
    """
    Buduje w pamięci szablon Excel (openpyxl) do zdefiniowania cennika standardowej instalacji
    per grupa produktowa: komponent, typ/model, dostawca, ilość na instalację, cena
    jednostkowa, waluta, uwagi. 'Wartość pozycji' to formuła (ilość x cena), przeliczana po
    edycji w arkuszu. Wypełniony przykładowymi wierszami dla instalacji silnikowej (Borger
    PL200, czujniki E+H, zawory/elektrozawory Metalwork) jako wzór formatu.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = EQUIPMENT_SHEET_NAME

    header_font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="DDEBF7")
    computed_fill = PatternFill("solid", fgColor="F2F2F2")
    example_font = Font(name="Arial", size=10, italic=True, color="0000FF")
    normal_font = Font(name="Arial", size=10)
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [EQUIPMENT_GROUP_COL, EQUIPMENT_COMPONENT_COL, EQUIPMENT_MODEL_COL, EQUIPMENT_SUPPLIER_COL,
               EQUIPMENT_QTY_COL, EQUIPMENT_UNIT_PRICE_COL, EQUIPMENT_CURRENCY_COL, EQUIPMENT_LINE_TOTAL_COL,
               EQUIPMENT_NOTES_COL]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 32

    ws_info = wb.create_sheet("Instrukcja")
    info_lines = [
        "INSTRUKCJA WYPEŁNIANIA:",
        "1. Nie zmieniaj nazw ani kolejności kolumn.",
        f"2. '{EQUIPMENT_GROUP_COL}' - wybierz z listy rozwijanej: {', '.join(RECIPE_PRODUCT_GROUPS)}.",
        "3. Jeden wiersz = jeden typ komponentu standardowej instalacji dla danej grupy (np. pompa, czujnik, zawór) - "
        "przepisz listę ze swojego P&ID danej instalacji standardowej.",
        f"4. '{EQUIPMENT_QTY_COL}' - ile sztuk tego komponentu jest w JEDNEJ instalacji (np. 2 czujniki temperatury).",
        f"5. '{EQUIPMENT_LINE_TOTAL_COL}' liczy się sama (formuła) = ilość x cena jednostkowa.",
        f"6. Aktualizuj '{EQUIPMENT_UNIT_PRICE_COL}' wraz ze zmianami cen dostawców - to jest właśnie po to, żeby "
        "nie trzeba było grzebać w kodzie aplikacji przy każdej zmianie cennika.",
        "7. W aplikacji (Zakładka 5) podajesz, ile takich instalacji planujesz dla danej grupy - CAPEX przelicza się automatycznie.",
    ]
    for i, line in enumerate(info_lines, start=1):
        c = ws_info.cell(row=i, column=1, value=line)
        c.font = Font(bold=(i == 1), name="Arial", size=11)
    ws_info.column_dimensions["A"].width = 110

    example_rows = [
        {"grp": "Engine Oils", "comp": "Pompa rozładunkowa", "model": "Borger PL200", "sup": "Borger",
         "qty": 1, "price": 18500, "cur": "PLN", "notes": "Pompa wyporowa do rozładunku cystern"},
        {"grp": "Engine Oils", "comp": "Czujnik ciśnienia", "model": "Cerabar PMC21", "sup": "Endress+Hauser",
         "qty": 3, "price": 2400, "cur": "PLN", "notes": ""},
        {"grp": "Engine Oils", "comp": "Czujnik temperatury", "model": "TC10/iTHERM", "sup": "Endress+Hauser",
         "qty": 2, "price": 1800, "cur": "PLN", "notes": ""},
        {"grp": "Engine Oils", "comp": "Zawór z siłownikiem pneumatycznym", "model": "Standard DN50", "sup": "Metalwork",
         "qty": 4, "price": 3200, "cur": "PLN", "notes": ""},
        {"grp": "Engine Oils", "comp": "Elektrozawór sterujący", "model": "Standard 24VDC", "sup": "Metalwork",
         "qty": 4, "price": 650, "cur": "PLN", "notes": ""},
    ]

    group_dv = DataValidation(type="list", formula1='"' + ",".join(RECIPE_PRODUCT_GROUPS) + '"', allow_blank=True)
    ws.add_data_validation(group_dv)

    start_row = 2
    n_blank_rows = 30
    total_rows = len(example_rows) + n_blank_rows

    for r_offset in range(total_rows):
        row = start_row + r_offset
        is_example = r_offset < len(example_rows)
        font_to_use = example_font if is_example else normal_font

        if is_example:
            d = example_rows[r_offset]
            ws.cell(row=row, column=1, value=d["grp"]).font = font_to_use
            ws.cell(row=row, column=2, value=d["comp"]).font = font_to_use
            ws.cell(row=row, column=3, value=d["model"]).font = font_to_use
            ws.cell(row=row, column=4, value=d["sup"]).font = font_to_use
            ws.cell(row=row, column=5, value=d["qty"]).font = font_to_use
            ws.cell(row=row, column=6, value=d["price"]).font = font_to_use
            ws.cell(row=row, column=7, value=d["cur"]).font = font_to_use
            ws.cell(row=row, column=9, value=d["notes"]).font = font_to_use
        else:
            for col_idx in [1, 2, 3, 4, 5, 6, 7, 9]:
                cell = ws.cell(row=row, column=col_idx)
                cell.font = normal_font
                cell.fill = input_fill
            group_dv.add(f"A{row}")

        total_formula = f"=E{row}*F{row}"
        tot_cell = ws.cell(row=row, column=8, value=total_formula)
        tot_cell.font = font_to_use
        tot_cell.number_format = '#,##0.00'
        tot_cell.fill = computed_fill

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 30

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def parse_equipment_excel(uploaded_file):
    """
    Wczytuje plik Excel z arkusza 'Cennik Instalacji' (lub pierwszego arkusza) i zwraca
    (df_czysty, lista_bledow). Waliduje: wymagane kolumny, poprawność grupy produktowej,
    nieujemne ilości/ceny. 'Wartość pozycji' PRZELICZANA w Pythonie (nie czytana z formuły).
    """
    errors = []
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_name = EQUIPMENT_SHEET_NAME if EQUIPMENT_SHEET_NAME in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name)
    except Exception as exc:
        return None, [f"Nie udało się odczytać pliku Excel: {exc}"]

    required_cols = [EQUIPMENT_GROUP_COL, EQUIPMENT_COMPONENT_COL, EQUIPMENT_QTY_COL, EQUIPMENT_UNIT_PRICE_COL]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        return None, [f"W pliku brakuje wymaganych kolumn: {', '.join(missing_cols)}. "
                       f"Pobierz i użyj oficjalnego szablonu, nie zmieniając nazw kolumn."]

    for c in [EQUIPMENT_MODEL_COL, EQUIPMENT_SUPPLIER_COL, EQUIPMENT_CURRENCY_COL, EQUIPMENT_NOTES_COL]:
        if c not in df.columns:
            df[c] = ""

    df = df[df[EQUIPMENT_COMPONENT_COL].notna()].copy()
    if df.empty:
        return None, ["Plik nie zawiera żadnych wierszy komponentów."]

    df[EQUIPMENT_QTY_COL] = pd.to_numeric(df[EQUIPMENT_QTY_COL], errors="coerce")
    df[EQUIPMENT_UNIT_PRICE_COL] = pd.to_numeric(df[EQUIPMENT_UNIT_PRICE_COL], errors="coerce")

    bad_rows = df[df[EQUIPMENT_QTY_COL].isna() | df[EQUIPMENT_UNIT_PRICE_COL].isna() |
                  (df[EQUIPMENT_QTY_COL] < 0) | (df[EQUIPMENT_UNIT_PRICE_COL] < 0)]
    if not bad_rows.empty:
        bad_names = bad_rows[EQUIPMENT_COMPONENT_COL].astype(str).tolist()
        errors.append(f"Pominięto komponenty z brakującą/ujemną ilością lub ceną: {', '.join(bad_names)}.")
        df = df.drop(bad_rows.index)

    unknown_group_mask = ~df[EQUIPMENT_GROUP_COL].astype(str).isin(RECIPE_PRODUCT_GROUPS)
    if unknown_group_mask.any():
        bad = df.loc[unknown_group_mask, EQUIPMENT_COMPONENT_COL].tolist()
        errors.append(f"Nieznana/brakująca grupa produktowa (musi być jedną z: {', '.join(RECIPE_PRODUCT_GROUPS)}) dla: "
                       f"{', '.join(map(str, bad))}. Wiersze pominięte.")
        df = df[~unknown_group_mask].copy()

    if df.empty:
        return None, errors

    df[EQUIPMENT_LINE_TOTAL_COL] = df[EQUIPMENT_QTY_COL] * df[EQUIPMENT_UNIT_PRICE_COL]

    return df.reset_index(drop=True), errors


# ==========================================
# RAPORT 5-LETNI (PDF) - zbieranie danych i generowanie dokumentu
# ==========================================

def compute_rm_consumption_for_year(year_idx):
    """
    Zużycie surowców [t/rok] per materiał, TYLKO dla produktów aktualnie PRODUKOWANYCH (nie
    importowanych) w danym roku symulacji (year_idx, 0-based; RAMPUP_YEAR_TARGET_SENTINEL =
    widok docelowy/100%), skalowane krzywą rozruchu ich linii. Surowce potrzebne wyłącznie do
    produktów, które w danym roku są jeszcze importowane, są pomijane - nie ma sensu ich
    kupować/magazynować, skoro dany produkt nie jest jeszcze lokalnie wytwarzany. Jedno źródło
    prawdy używane zarówno przez Zakładkę 4 (silosy/beczki) jak i Zakładkę 4 (magazyn RM).
    """
    recipes_df_local = st.session_state.get("recipes_df")
    consumption = {mat: 0.0 for mat in RECIPE_RAW_MATERIALS}
    has_recipe = recipes_df_local is not None and not recipes_df_local.empty
    if has_recipe:
        has_sourcing = RECIPE_SOURCING_COL in recipes_df_local.columns
        for _, r in recipes_df_local.iterrows():
            if has_sourcing:
                is_imp = is_product_imported_in_year(r.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                                      r.get(RECIPE_IMPORT_TRANSITION_COL, ""), year_idx)
                if is_imp:
                    continue
            frac = get_rampup_fraction(r[RECIPE_GROUP_COL], year_idx) if year_idx != RAMPUP_YEAR_TARGET_SENTINEL else 1.0
            raw_demand_t_target = float(r.get(RECIPE_RAW_DEMAND_COL, 0) or 0)
            raw_demand_t_year = raw_demand_t_target * frac
            for mat in RECIPE_RAW_MATERIALS:
                dozowanie = float(r.get(mat, 0) or 0)
                consumption[mat] += raw_demand_t_year * (dozowanie / 1000.0)

    # Surowce zadeklarowane WPROST (bez pełnej receptury - ochrona know-how), z arkusza
    # 'Zużycie Surowców (bez receptury)' - dolicza się do zużycia z receptur powyżej (jeśli w
    # ogóle jakaś receptura jest wgrana - działa też SAMODZIELNIE, bez żadnej receptury), skalowane
    # tą samą krzywą rozruchu co linia, do której surowiec jest przypisany.
    for entry in st.session_state.get("direct_raw_materials", []):
        frac_direct = get_rampup_fraction(entry["group"], year_idx) if year_idx != RAMPUP_YEAR_TARGET_SENTINEL else 1.0
        consumption[entry["material"]] = consumption.get(entry["material"], 0.0) + entry["annual_t"] * frac_direct

    return consumption


def compute_rm_consumption_for_month(year_idx, month_idx):
    """Wariant MIESIĘCZNY compute_rm_consumption_for_year - zużycie [t/miesiąc] per materiał, z
    interpolacją liniową rozruchu (get_rampup_fraction_month) zamiast płaskiej wartości rocznej.
    Status importu sprawdzany per ROK (przejścia są roczne, nie ułamkowe - patrz
    is_product_imported_in_year), tylko skalowanie w obrębie roku jest miesięczne."""
    recipes_df_local = st.session_state.get("recipes_df")
    consumption = {mat: 0.0 for mat in RECIPE_RAW_MATERIALS}
    has_recipe = recipes_df_local is not None and not recipes_df_local.empty
    if has_recipe:
        has_sourcing = RECIPE_SOURCING_COL in recipes_df_local.columns
        for _, r in recipes_df_local.iterrows():
            if has_sourcing:
                is_imp = is_product_imported_in_year(r.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                                      r.get(RECIPE_IMPORT_TRANSITION_COL, ""), year_idx)
                if is_imp:
                    continue
            frac = get_rampup_fraction_month(r[RECIPE_GROUP_COL], year_idx, month_idx)
            raw_demand_t_target_month = float(r.get(RECIPE_RAW_DEMAND_COL, 0) or 0) / MONTHS_PER_YEAR
            raw_demand_t_month = raw_demand_t_target_month * frac
            for mat in RECIPE_RAW_MATERIALS:
                dozowanie = float(r.get(mat, 0) or 0)
                consumption[mat] += raw_demand_t_month * (dozowanie / 1000.0)

    for entry in st.session_state.get("direct_raw_materials", []):
        frac_direct_month = get_rampup_fraction_month(entry["group"], year_idx, month_idx)
        consumption[entry["material"]] = consumption.get(entry["material"], 0.0) + (entry["annual_t"] / MONTHS_PER_YEAR) * frac_direct_month

    return consumption


def compute_rm_category_consumption_by_group(year_idx, category_full_name):
    """
    Zużycie JEDNEJ konkretnej kategorii surowca (np. 'Base Oil Group II [kg/t]'), rozbite PER
    GRUPA PRODUKTOWA - potrzebne, żeby rozbicie dostawców zawężone do konkretnej linii
    (SUPPLIER_SPLIT_GROUP_COL) mogło zadziałać na WŁAŚCIWEJ części całkowitego zużycia tej
    kategorii, nie na całości. Uwzględnia pełne receptury ORAZ pasujące wpisy z arkusza
    'Zużycie Surowców (bez receptury)'.
    """
    result = {}
    recipes_df_local = st.session_state.get("recipes_df")
    if recipes_df_local is not None and not recipes_df_local.empty and category_full_name in recipes_df_local.columns:
        has_sourcing = RECIPE_SOURCING_COL in recipes_df_local.columns
        for _, r in recipes_df_local.iterrows():
            if has_sourcing:
                is_imp = is_product_imported_in_year(r.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                                      r.get(RECIPE_IMPORT_TRANSITION_COL, ""), year_idx)
                if is_imp:
                    continue
            frac = get_rampup_fraction(r[RECIPE_GROUP_COL], year_idx) if year_idx != RAMPUP_YEAR_TARGET_SENTINEL else 1.0
            raw_demand_t_year = float(r.get(RECIPE_RAW_DEMAND_COL, 0) or 0) * frac
            dozowanie = float(r.get(category_full_name, 0) or 0)
            result[r[RECIPE_GROUP_COL]] = result.get(r[RECIPE_GROUP_COL], 0.0) + raw_demand_t_year * (dozowanie / 1000.0)

    category_short = category_full_name.replace(" [kg/t]", "")
    for entry in st.session_state.get("direct_raw_materials", []):
        if entry.get("category") == category_short:
            frac_direct = get_rampup_fraction(entry["group"], year_idx) if year_idx != RAMPUP_YEAR_TARGET_SENTINEL else 1.0
            result[entry["group"]] = result.get(entry["group"], 0.0) + entry["annual_t"] * frac_direct

    return result


def apply_supplier_splits(consumption_dict, year_idx):
    """
    Rozbija pozycje kategorii (np. 'Base Oil Group II [kg/t]'), dla których zdefiniowano
    'Rozbicie Dostawców Surowców', na osobne pozycje per konkretny dostawca - każda dostaje
    WŁASNĄ kandydaturę na zbiornik zamiast jednej dużej, wspólnej. Rozbicie per LINIA (jeśli
    wskazana) dotyczy TYLKO zużycia tej kategorii z tamtej linii - reszta zakładu zostaje
    nierozbita. Rozbicie "całego zakładu" (puste pole grupy) dotyczy całości kategorii.
    Bezpieczny krok POST-PROCESSING na zwykłym dict {surowiec: t/rok}, analogicznie do
    collapse_shared_tank_materials - nie zmienia istniejącej logiki liczącej zużycie.
    """
    splits_by_category = {}
    for s in st.session_state.get("supplier_splits", []):
        splits_by_category.setdefault(s["category"], []).append(s)
    if not splits_by_category:
        return consumption_dict

    result = dict(consumption_dict)
    for category_short, splits in splits_by_category.items():
        category_full = category_short + " [kg/t]"
        total_category_t = result.get(category_full, 0.0)
        if total_category_t <= 0:
            continue

        group_specific = [s for s in splits if s["group"]]
        plant_wide = [s for s in splits if not s["group"]]

        if group_specific:
            by_group = compute_rm_category_consumption_by_group(year_idx, category_full)
            remaining_t = total_category_t
            for s in group_specific:
                group_t = by_group.get(s["group"], 0.0)
                supplier_t = group_t * (s["pct"] / 100.0)
                supplier_key = f"{category_short} - {s['supplier']}"
                result[supplier_key] = result.get(supplier_key, 0.0) + supplier_t
                remaining_t -= supplier_t
            result[category_full] = max(remaining_t, 0.0)
        elif plant_wide:
            del result[category_full]
            for s in plant_wide:
                supplier_key = f"{category_short} - {s['supplier']}"
                result[supplier_key] = result.get(supplier_key, 0.0) + total_category_t * (s["pct"] / 100.0)

    return result


def collapse_shared_tank_materials(consumption_dict):
    """
    Zwija pozycje z arkusza 'Zużycie Surowców (bez receptury)' LUB z 'Rozbicie Dostawców
    Surowców', które mają WSPÓLNE ID Zbiornika (ta sama konwencja co ID Zbiornika w recepturze -
    te samo ID = współdzielony fizyczny zbiornik), w JEDNĄ pozycję o zsumowanym zużyciu - zamiast
    każdej z osobna dostawać własną rekomendację zbiornika. Bezpieczne posunięcie: działa na
    zwykłym dict {surowiec: t/rok} PRZED wejściem do istniejącej pętli doboru zbiorników, więc
    sama pętla nie wymaga żadnych zmian. Materiały bez ID Zbiornika (albo spoza tych arkuszy)
    przechodzą bez zmian.
    """
    tank_groups = {}  # tank_id -> lista nazw materiałów z tym ID
    for entry in st.session_state.get("direct_raw_materials", []):
        if entry.get("tank_id"):
            tank_groups.setdefault(entry["tank_id"], []).append(entry["material"])
    for s in st.session_state.get("supplier_splits", []):
        if s.get("tank_id"):
            tank_groups.setdefault(s["tank_id"], []).append(f"{s['category']} - {s['supplier']}")

    collapsed = dict(consumption_dict)
    for tank_id, materials in tank_groups.items():
        materials_present = [m for m in materials if m in collapsed]
        if len(materials_present) < 2:
            continue  # tylko 1 materiał z tym ID w danych - nic do zwinięcia
        combined_t = sum(collapsed.pop(m) for m in materials_present)
        combined_name = f"🔗 Zbiornik {tank_id}: {' + '.join(materials_present)}"
        collapsed[combined_name] = combined_t

    return collapsed


def get_full_qc_catalog():
    """
    JEDNO ŹRÓDŁO PRAWDY dla pełnego katalogu testów QC - wbudowany katalog (QC_TEST_CATALOG)
    SCALONY z testami dopisanymi przez użytkownika w arkuszu 'Badania Laboratoryjne' (własne
    kolumny 'Sprzęt'/'Czas [min]', patrz parse_qc_tests_excel). Używaj tego wszędzie zamiast
    odwoływać się bezpośrednio do QC_TEST_CATALOG - inaczej własne testy użytkownika "znikają"
    (brak czasu/sprzętu) w miejscach, które nie wiedzą o custom_qc_tests.
    """
    return {**QC_TEST_CATALOG, **st.session_state.get("custom_qc_tests", {})}


def get_qc_tests_for_mixer(mixer):
    """
    JEDNO ŹRÓDŁO PRAWDY dla ustalenia, które testy QC dotyczą danego mieszalnika (i jego
    konkretnego produktu, jeśli przypisany) - PRIORYTET 1: arkusz 'Badania Laboratoryjne' per
    KONKRETNY PRODUKT; PRIORYTET 2: kolumny 'QC: {test} [x]' wprost w recepturze; PRIORYTET 3
    (fallback): panel zwolnienia per LINIA (Zakładka 6/VSM). Wcześniej ta sama logika była
    duplikowana osobno w widgecie porównawczym (Karta Maszyn) i na Dashboardzie.
    Zwraca (lista_nazw_testów, etykieta_źródła).
    """
    recipe_product = mixer.get("recipe_product")
    recipes_df_local = st.session_state.get("recipes_df")

    qc_tests_from_sheet = st.session_state.get("qc_tests_by_product", {}).get(recipe_product, []) if recipe_product else []
    if qc_tests_from_sheet:
        return qc_tests_from_sheet, f"arkusz 'Badania Laboratoryjne': {', '.join(qc_tests_from_sheet)}"

    if recipes_df_local is not None and not recipes_df_local.empty and recipe_product:
        qc_cols = [c for c in recipes_df_local.columns if c.startswith(QC_COL_PREFIX) and c.endswith(QC_COL_SUFFIX)]
        match = recipes_df_local[recipes_df_local[RECIPE_PRODUCT_COL] == recipe_product]
        if qc_cols and not match.empty:
            row0 = match.iloc[0]
            qc_tests_from_recipe = [c[len(QC_COL_PREFIX):-len(QC_COL_SUFFIX)] for c in qc_cols if bool(row0.get(c, False))]
            if qc_tests_from_recipe:
                return qc_tests_from_recipe, f"kolumny w recepturze (Zakładka 1): {', '.join(qc_tests_from_recipe)}"

    qc_cfg = st.session_state.get("vsm_qc_config", {}).get(mixer["product_family"], {
        "tests": ["Lepkość kinematyczna @40°C (półautomat)", "Barwa ASTM", "Temp. zapłonu - półautomat"]
    })
    qc_tests_used = qc_cfg.get("tests", [])
    return qc_tests_used, f"panel zwolnienia linii (Zakładka 6): {len(qc_tests_used)} testów"


def compute_filling_time_h(mass_kg, recipe_product, kat, mixer_tag, rho_linii, opakowania_podzial=None):
    """
    JEDNO ŹRÓDŁO PRAWDY dla czasu rozlewu/napełniania - działa na DOWOLNEJ masie podanej wprost
    (masa jednej szarży, miesiąca, roku - cokolwiek przekażesz), więc liczy się to samo, niezależnie
    od tego, gdzie w aplikacji jest wywoływane. Priorytet rozbicia na opakowania: 1) receptura per
    KONKRETNY produkt (Zakładka 1, kolumny 'Opak: ... [%]'), 2) ręczny podział per LINIA (panel
    boczny), znormalizowany do 100% w obu przypadkach. Ograniczenie przepływu: MNIEJSZE z (a) pompy
    TEGO KONKRETNEGO mieszalnika (Zakładka 2) i (b) wydajności linii nalewającej danego opakowania.
    Zwraca łączny czas [h] rozlewu tej masy, zsumowany po wszystkich typach opakowań z udziałem % > 0.
    """
    recipes_df_local = st.session_state.get("recipes_df")
    pack_pcts = None
    if recipes_df_local is not None and not recipes_df_local.empty and recipe_product:
        pack_cols = [c for c in recipes_df_local.columns if c.startswith("Opak: ") and c.endswith(" [%]")]
        match = recipes_df_local[recipes_df_local[RECIPE_PRODUCT_COL] == recipe_product]
        if pack_cols and not match.empty:
            row0 = match.iloc[0]
            pack_sum = sum(row0.get(c, 0) or 0 for c in pack_cols)
            if pack_sum > 0.5:
                pack_pcts = {c[len("Opak: "):-len(" [%]")]: (row0.get(c, 0) or 0) for c in pack_cols if (row0.get(c, 0) or 0) > 0}

    if pack_pcts is None:
        opakowania_podzial = opakowania_podzial or {}
        pack_pcts = {p: opakowania_podzial.get(f"pct_{kat}_{p}", 0.0) for p in st.session_state.pack_configs.keys()}

    pack_pcts_sum = sum(pack_pcts.values())
    if pack_pcts_sum > 0.5 and abs(pack_pcts_sum - 100.0) > 2.0:
        pack_pcts = {p: v * (100.0 / pack_pcts_sum) for p, v in pack_pcts.items()}

    tech_details = st.session_state.get("mixer_tech_advanced_details", {}).get(mixer_tag, {})
    q_pump_m3h = tech_details.get("pump_flow_m3h", 15.0)

    total_h = 0.0
    for p, udzial_pct in pack_pcts.items():
        if udzial_pct <= 0 or p not in st.session_state.pack_configs:
            continue
        mass_this_pack_kg = mass_kg * (udzial_pct / 100.0)
        pack_capacity_kg = st.session_state.pack_configs[p]["size_l"] * rho_linii
        if pack_capacity_kg <= 0:
            continue
        cfg_fill = st.session_state.filling_lines_config.get(p, default_filling_line_config(p))
        sekcja_nalewania_m3_h = (cfg_fill["nozzles"] * cfg_fill["speed_kg_min"] * 60.0) / (rho_linii * 1000.0)
        q_effective_flow_m3h = min(q_pump_m3h, sekcja_nalewania_m3_h)
        if q_effective_flow_m3h > 0:
            total_h += (mass_this_pack_kg / (rho_linii * 1000.0)) / q_effective_flow_m3h
    return total_h


def default_filling_line_config(pack_name):
    """
    JEDNO ŹRÓDŁO PRAWDY dla domyślnej konfiguracji linii nalewającej (głowice + prędkość/głowicę) -
    wcześniej te same domyślne wartości były wpisane osobno w 3 różnych miejscach kodu i ROZJECHAŁY
    SIĘ (30 vs 60 kg/min dla beczek, 1 vs 4 głowice) - to bezpośrednio powodowało nierealistycznie
    długie czasy rozlewu (np. 27h) dla dużych mieszalników, gdy akurat trafiał się mniej korzystny
    z tych dwóch domyślnych. Cysterna ma zupełnie inny rząd wielkości przepływu (ramię załadowcze,
    nie dysza) - bez tego rozróżnienia dziedziczyłaby domyślne dla beczek, dając fizycznie
    absurdalny czas napełnienia (~12h zamiast realnych 30-60 min).
    """
    if pack_name in st.session_state.get("pack_configs", {}) and st.session_state.pack_configs[pack_name].get("per_pallet") == 0:
        return {"nozzles": 1, "speed_kg_min": 800.0}  # ramię załadowcze cysterny, ~48 t/h
    if "5l" in pack_name.lower() or "1l" in pack_name.lower():
        return {"nozzles": 4, "speed_kg_min": 15.0}  # mała pakowarka wielogłowicowa (detal/karton)
    return {"nozzles": 2, "speed_kg_min": 60.0}  # standardowa dysza do beczek/kanistrów/IBC - 2 głowice
    # jako bardziej realistyczny punkt startowy dla większych partii niż pojedyncza dysza; i tak
    # zawsze warto zweryfikować/dostosować w Zakładce 4 do rzeczywistej linii nalewającej.


def default_filling_speed_kg_min(pack_name):
    """Zachowane dla zgodności - zwraca tylko prędkość z default_filling_line_config."""
    return default_filling_line_config(pack_name)["speed_kg_min"]


def round_visible(value, min_significant=2, max_decimals=8):
    """
    JEDNO ŹRÓDŁO PRAWDY dla zaokrąglania małych ilości (dozowanie surowców, tony/szarżę,
    zużycie) tak, żeby ŚLADOWE dodatki (barwniki, zapachy, biocydy, przeciwpienne - często
    0,0001-0,01 kg/t) NIGDY nie znikały jako '0' przy stałym zaokrągleniu do 2-3 miejsc po
    przecinku. Dla wartości >= 1 zaokrągla zwyczajnie do 2 miejsc (czytelność); dla mniejszych
    dobiera tyle miejsc po przecinku, żeby zachować min_significant cyfr znaczących.
    """
    if value == 0:
        return 0.0
    if abs(value) >= 1:
        return round(value, 2)
    decimals = min_significant - int(math.floor(math.log10(abs(value)))) - 1
    return round(value, min(max(decimals, 2), max_decimals))


def compute_rm_drummed_pallets_per_month(year_idx):
    """
    JEDNO ŹRÓDŁO PRAWDY dla palet RM w beczkach/IBC/workach [pal/mies] per materiał, danego
    roku - PRZEPŁYW (ile faktycznie produkujemy/zużywamy miesięcznie), NIE stan magazynowy.
    Wyklucza materiały skierowane do zbiorników (rm_storage_method_override). Używane przez
    Zakładkę 4 (Magazynowanie), Zakładkę 5 (Finanse, tabela rampup) i Dashboard - wcześniej ta
    sama logika była skopiowana osobno w każdym miejscu i z czasem się rozjechała.
    """
    rm_year_consumption = compute_rm_consumption_for_year(year_idx)
    rm_container_assignment = st.session_state.get("rm_container_assignment", {})
    rm_storage_method_override = st.session_state.get("rm_storage_method_override", {})
    pallets_per_material = {}
    for mat, ann_t in rm_year_consumption.items():
        if ann_t <= 0 or rm_storage_method_override.get(mat) == "Zbiornik (luzem)":
            continue
        container_name = rm_container_assignment.get(mat, "Beczka 200 kg (ciecz)")
        container_cfg = RM_CONTAINER_TYPES.get(container_name)
        if not container_cfg:
            continue
        monthly_kg = (ann_t * 1000.0) / MONTHS_PER_YEAR
        n_containers = math.ceil(monthly_kg / container_cfg["capacity_kg"]) if container_cfg["capacity_kg"] > 0 else 0
        pallets_per_material[mat] = math.ceil(n_containers / container_cfg["per_pallet"]) if container_cfg["per_pallet"] > 0 else 0
    return pallets_per_material


def compute_rm_drummed_positions_for_year(year_idx, days_of_stock_val):
    """
    JEDNO ŹRÓDŁO PRAWDY dla MIEJSC MAGAZYNOWYCH RM w beczkach [szt] - to INNA wielkość niż
    'palety/miesiąc' (compute_rm_drummed_pallets_per_month): tu przepływ miesięczny jest
    rozciągnięty na bufor dni zapasu (days_of_stock), dając liczbę POZYCJI, jakie faktycznie
    muszą stać w magazynie naraz - zgodnie z formułą kanoniczną z Zakładki 4 (Magazynowanie).
    """
    pallets_per_material = compute_rm_drummed_pallets_per_month(year_idx)
    dni_robocze_miesiac = WORKING_DAYS_YEAR / MONTHS_PER_YEAR
    total_positions = 0
    for mat, n_pallets_month in pallets_per_material.items():
        total_positions += math.ceil((n_pallets_month / dni_robocze_miesiac) * days_of_stock_val) if dni_robocze_miesiac > 0 else 0
    return total_positions


def compute_fg_buffer_positions_for_year(year_idx, days_of_stock_val):
    """
    Miejsca magazynowe FG [szt], TYM SAMYM modelem bufora dni zapasu co RM/Import (przepływ
    miesięczny × dni_zapasu / dni_robocze_miesiąc) - na poziomie ZAGREGOWANYM (cała flota razem),
    nie per SKU/opakowanie jak w Zakładce 4 (tamten szczegółowy rozkład zostaje osobno, jako
    źródło prawdy dla wymiarowania budynku per opakowanie). Tu chodzi o spójny, prosty model do
    porównań rok-do-roku w Dashboardzie, konsekwentnie stosujący założenia buforowe wszędzie.
    """
    total_palety_month_fg_target = st.session_state.get("total_palety_month_fg_target_report", 0.0)
    target_annual_t_fg = sum(m["annual_volume"] for m in st.session_state.get("confirmed_mixers", [])) / 1000.0
    if target_annual_t_fg <= 0:
        return 0
    year_tonnage_t = sum((m["annual_volume"] / 1000.0) * get_rampup_fraction(m["product_family"], year_idx)
                          for m in st.session_state.get("confirmed_mixers", []))
    frac_year_blended = year_tonnage_t / target_annual_t_fg
    monthly_pallets_year = total_palety_month_fg_target * frac_year_blended
    dni_robocze_miesiac = WORKING_DAYS_YEAR / MONTHS_PER_YEAR
    return math.ceil((monthly_pallets_year / dni_robocze_miesiac) * days_of_stock_val) if dni_robocze_miesiac > 0 else 0


def get_import_volume_fraction(product_family, year_idx_for_calc):
    """
    Ułamek (0-1) wolumenu produktu JESZCZE IMPORTOWANEGO w danym roku - JEDNO ŹRÓDŁO PRAWDY,
    respektujące przełącznik 'import_follows_rampup' (panel boczny). Domyślnie (True) import
    skaluje się tą samą krzywą rozruchu co produkcja własna (popyt klienta rośnie niezależnie
    od tego, kto go zaspokaja). Gdy False - import liczony zawsze jako pełny docelowy wolumen
    (100%), bo w rzeczywistości może mieć własny, niepowiązany harmonogram (np. stały kontrakt).
    """
    if year_idx_for_calc == RAMPUP_YEAR_TARGET_SENTINEL:
        return 1.0
    if not st.session_state.get("import_follows_rampup", True):
        return 1.0
    return get_rampup_fraction(product_family, year_idx_for_calc)


def compute_import_positions_for_year(year_idx_for_calc, import_pallet_mass_kg=None):
    """
    JEDNO ŹRÓDŁO PRAWDY dla miejsc paletowych produktów importowanych (jeszcze lub na stałe,
    z wyłączeniem 'Nigdy (bufor)' - te mają dedykowany zbiornik, nie paletę). Szczytowy zapas =
    wielkość 1 dostawy + bufor bezpieczeństwa. Używane przez Zakładkę 4 (Logistyka) i Dashboard.
    """
    if import_pallet_mass_kg is None:
        import_pallet_mass_kg = st.session_state.get("import_pallet_mass_kg", 800.0)
    recipes_df_local = st.session_state.get("recipes_df")
    total_positions = 0
    if recipes_df_local is None or recipes_df_local.empty or RECIPE_SOURCING_COL not in recipes_df_local.columns:
        return total_positions
    import_rows_df = recipes_df_local[
        (recipes_df_local[RECIPE_SOURCING_COL] == "Import") &
        (recipes_df_local.get(RECIPE_IMPORT_TRANSITION_COL, pd.Series(dtype=str)) != "Nigdy (bufor)")
    ]
    for _, r in import_rows_df.iterrows():
        if not is_product_imported_in_year(r.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                             r.get(RECIPE_IMPORT_TRANSITION_COL, ""), year_idx_for_calc):
            continue
        annual_t_target = float(r.get(RECIPE_ANNUAL_COL, 0) or 0)
        frac = get_import_volume_fraction(r[RECIPE_GROUP_COL], year_idx_for_calc)
        effective_annual_t = annual_t_target * frac
        daily_t = effective_annual_t / WORKING_DAYS_YEAR if WORKING_DAYS_YEAR > 0 else 0.0
        lot_t = float(r.get(RECIPE_IMPORT_LOT_COL, 0) or 0)
        safety_days = float(r.get(RECIPE_IMPORT_SAFETY_DAYS_COL, 0) or 0)
        peak_stock_t = lot_t + safety_days * daily_t
        total_positions += math.ceil((peak_stock_t * 1000.0) / import_pallet_mass_kg) if import_pallet_mass_kg > 0 else 0
    return total_positions


def compute_pdf_report_year_data():
    """
    Zbiera dane per rok symulacji rozruchu (1-5) potrzebne do raportu PDF: tonaż docelowy vs
    produkowany vs importowany, produkty, flota (kotły/mieszalniki), wykorzystanie magazynu i
    KPI energetyczne. Czyta WYŁĄCZNIE z session_state / już policzonych struktur (confirmed_mixers,
    recipes_df, stock_simulation_df z Zakładki 3, energy_kpi_rows_report z Zakładki 5) - nie
    przelicza niczego od nowa poza prostym rozbiciem produkcja/import per produkt, tą samą logiką
    co Zakładka 3 (is_product_imported_in_year). Zwraca None, jeśli flota nie jest jeszcze
    zatwierdzona (nic do zaraportowania).
    """
    mixers = st.session_state.get("confirmed_mixers", [])
    if not mixers:
        return None

    target_annual_t = sum(m["annual_volume"] for m in mixers) / 1000.0
    recipes_df = st.session_state.get("recipes_df")
    energy_rows = st.session_state.get("energy_kpi_rows_report", [])
    stock_df = st.session_state.get("stock_simulation_df")
    fg_capacity = st.session_state.get("fg_capacity_pallets_report", 0.0)
    has_recipes = recipes_df is not None and not recipes_df.empty and RECIPE_SOURCING_COL in recipes_df.columns

    years_data = []
    for i in range(RAMPUP_YEARS):
        year_tonnage_t = sum((m["annual_volume"] / 1000.0) * get_rampup_fraction(m["product_family"], i) for m in mixers)
        frac_year = (year_tonnage_t / target_annual_t) if target_annual_t > 0 else 0.0

        produced_t, imported_t = 0.0, 0.0
        product_rows = []
        if has_recipes:
            for _, r in recipes_df.iterrows():
                ann_t_target = float(r.get(RECIPE_ANNUAL_COL, 0) or 0)
                line_frac = get_rampup_fraction(r[RECIPE_GROUP_COL], i)
                eff_t = ann_t_target * line_frac
                is_imp = is_product_imported_in_year(r.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                                      r.get(RECIPE_IMPORT_TRANSITION_COL, ""), i)
                if is_imp:
                    imported_t += eff_t
                else:
                    produced_t += eff_t
                if eff_t > 0:
                    product_rows.append({"product": str(r[RECIPE_PRODUCT_COL]), "group": str(r[RECIPE_GROUP_COL]),
                                          "kg": eff_t * 1000.0, "mode": "Import" if is_imp else "Production"})
        else:
            produced_t = year_tonnage_t  # brak wgranej receptury - zakładamy 100% produkcja własna

        wh_util_pct, wh_value = None, None
        if stock_df is not None and not stock_df.empty:
            month_row_idx = i * 12 + 11  # ostatni miesiąc danego roku (indeks 0-based)
            if month_row_idx < len(stock_df):
                row = stock_df.iloc[month_row_idx]
                stock_pal = row["Stan magazynowy [pal]"]
                wh_util_pct = (stock_pal / fg_capacity * 100.0) if fg_capacity > 0 else None
                value_cols = [c for c in stock_df.columns if c.startswith("Wartość zapasu")]
                wh_value = row[value_cols[0]] if value_cols else None

        years_data.append({
            "year": i + 1, "target_pct": frac_year * 100.0, "total_t": year_tonnage_t,
            "produced_t": produced_t, "imported_t": imported_t,
            "products": sorted(product_rows, key=lambda x: -x["kg"])[:12],
            "wh_util_pct": wh_util_pct, "wh_value": wh_value,
            "energy": energy_rows[i] if i < len(energy_rows) else {},
        })

    return {"years": years_data, "target_annual_t": target_annual_t, "fleet": mixers, "has_recipes": has_recipes}


def _mpl_fig_to_png_bytes(fig):
    """Zapisuje figurę matplotlib do bajtów PNG (do osadzenia w PDF przez reportlab)."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.getvalue()


def generate_pdf_report_bytes(report_data, waluta_report):
    """
    Buduje dokument PDF (reportlab) po angielsku, z tabelami i wykresami (matplotlib -> PNG),
    podsumowujący 5-letnią symulację rozruchu: skala produkcji per rok, produkty, produkcja
    własna vs import, flota (mieszalniki), wykorzystanie magazynu, KPI energetyczne.
    Zwraca bajty PDF. Zgłasza ImportError, jeśli reportlab/matplotlib nie są zainstalowane -
    obsługiwane w UI (Zakładka 5) komunikatem z instrukcją dopisania do requirements.txt.
    """
    import os
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Domyślne czcionki reportlab (Helvetica) nie obsługują polskich znaków diakrytycznych -
    # rejestrujemy DejaVu Sans (dołączone do matplotlib, więc zawsze dostępne skoro matplotlib
    # jest już zależnością tej apki), żeby np. polskie nazwy produktów wpisane w recepturach
    # renderowały się poprawnie zamiast jako czarne kwadraty.
    try:
        mpl_font_dir = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
        pdfmetrics.registerFont(TTFont("DejaVuSans", os.path.join(mpl_font_dir, "DejaVuSans.ttf")))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", os.path.join(mpl_font_dir, "DejaVuSans-Bold.ttf")))
        font_regular, font_bold = "DejaVuSans", "DejaVuSans-Bold"
    except Exception:
        font_regular, font_bold = "Helvetica", "Helvetica-Bold"

    years = report_data["years"]
    target_annual_t = report_data["target_annual_t"]
    fleet = report_data["fleet"]

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], fontSize=20, fontName=font_bold)
    h2_style = ParagraphStyle("H2Custom", parent=styles["Heading2"], spaceBefore=14, fontName=font_bold)
    body_style = ParagraphStyle("BodyCustom", parent=styles["BodyText"], fontName=font_regular)
    h4_style = ParagraphStyle("H4Custom", parent=styles["Heading4"], fontName=font_bold)

    def styled_table(data_rows, col_widths=None):
        t = Table(data_rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font_regular),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    story = []

    # --- Strona tytułowa ---
    story.append(Paragraph("5-Year Production Scale-Up Report", title_style))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}", body_style))
    story.append(Paragraph(f"Target annual production capacity (100%, Year 5+): {target_annual_t:,.0f} t/year", body_style))
    story.append(Paragraph(f"Currency: {waluta_report}", body_style))
    story.append(Spacer(1, 0.8 * cm))

    # --- Wykres 1: tonaż produkowany vs importowany vs cel, per rok ---
    fig1, ax1 = plt.subplots(figsize=(6.5, 3.2))
    yrs = [y["year"] for y in years]
    ax1.bar(yrs, [y["produced_t"] for y in years], label="Produced", color="#1f77b4")
    ax1.bar(yrs, [y["imported_t"] for y in years], bottom=[y["produced_t"] for y in years], label="Imported", color="#ff7f0e")
    ax1.axhline(target_annual_t, color="#d62728", linestyle="--", linewidth=1.2, label="Target (100%)")
    ax1.set_xlabel("Year")
    ax1.set_ylabel("Tonnage [t/year]")
    ax1.set_title("Production Volume: Own Production vs. Import")
    ax1.set_xticks(yrs)
    ax1.legend(fontsize=8)
    story.append(Image(io.BytesIO(_mpl_fig_to_png_bytes(fig1)), width=16 * cm, height=7.9 * cm))
    plt.close(fig1)

    # --- Wykres 2: stan magazynowy w czasie (jeśli dostępny) ---
    stock_df = st.session_state.get("stock_simulation_df")
    if stock_df is not None and not stock_df.empty:
        fig2, ax2 = plt.subplots(figsize=(6.5, 3.0))
        ax2.bar(stock_df["Miesiąc"], stock_df["Stan magazynowy [pal]"], color="#1f77b4", width=0.8)
        ax2.set_xlabel("Month (1-60)")
        ax2.set_ylabel("Pallet positions")
        ax2.set_title("Finished-Goods Warehouse Stock Level (Monthly)")
        story.append(Spacer(1, 0.3 * cm))
        story.append(Image(io.BytesIO(_mpl_fig_to_png_bytes(fig2)), width=16 * cm, height=7.4 * cm))
        plt.close(fig2)

    # --- Wykres 3: koszt energii per rok, w podziale ---
    if years[0].get("energy"):
        fig3, ax3 = plt.subplots(figsize=(6.5, 3.0))
        heat_vals = [y["energy"].get("Ogrzewanie [waluta/rok]", 0) for y in years]
        proc_vals = [y["energy"].get("Elektryczność - proces (w tym chłodzenie) [waluta/rok]", 0) for y in years]
        fac_vals = [y["energy"].get("Elektryczność - pozaprodukcyjne (stałe) [waluta/rok]", 0) for y in years]
        ax3.bar(yrs, heat_vals, label="Heating", color="#d62728")
        ax3.bar(yrs, proc_vals, bottom=heat_vals, label="Electricity - Process (incl. cooling)", color="#1f77b4")
        bottom2 = [h + p for h, p in zip(heat_vals, proc_vals)]
        ax3.bar(yrs, fac_vals, bottom=bottom2, label="Electricity - Facility (fixed)", color="#2ca02c")
        ax3.set_xlabel("Year")
        ax3.set_ylabel(f"Energy cost [{waluta_report}/year]")
        ax3.set_title("Energy Cost Breakdown by Year")
        ax3.set_xticks(yrs)
        ax3.legend(fontsize=7)
        story.append(Spacer(1, 0.3 * cm))
        story.append(Image(io.BytesIO(_mpl_fig_to_png_bytes(fig3)), width=16 * cm, height=7.4 * cm))
        plt.close(fig3)

    story.append(PageBreak())

    # --- Flota (kotły / mieszalniki) - raz, bo nie zmienia się per rok ---
    story.append(Paragraph("Production Fleet (Kettles / Mixers)", h2_style))
    fleet_table_data = [["Tag", "Product Family", "Capacity [m3]", "Batch Mass [kg]", "Cycle [h]"]]
    for m in fleet:
        fleet_table_data.append([m["tag"], m["product_family"], f'{m["capacity_m3"]:.1f}',
                                  f'{m["mass_per_batch"]:,.0f}', f'{m["cycle_h"]:.2f}'])
    story.append(styled_table(fleet_table_data))

    # --- Sekcje per rok ---
    for y in years:
        story.append(PageBreak())
        story.append(Paragraph(f"Year {y['year']} — {y['target_pct']:.0f}% of Target", h2_style))

        summary_data = [
            ["Metric", "Value"],
            ["Total volume (target x rampup)", f"{y['total_t']:,.0f} t"],
            ["Produced in-house", f"{y['produced_t']:,.0f} t"],
            ["Imported", f"{y['imported_t']:,.0f} t"],
            ["Warehouse utilization (FG, year-end)", f"{y['wh_util_pct']:.0f}%" if y["wh_util_pct"] is not None else "n/a"],
            ["Warehouse stock value (year-end)", f"{y['wh_value']:,.0f} {waluta_report}" if y["wh_value"] is not None else "n/a"],
        ]
        story.append(styled_table(summary_data, col_widths=[9 * cm, 7 * cm]))
        story.append(Spacer(1, 0.4 * cm))

        if y["energy"]:
            energy_label_en = {
                "Ogrzewanie [waluta/rok]": "Heating",
                "Elektryczność - proces (w tym chłodzenie) [waluta/rok]": "Electricity - Process (incl. cooling)",
                "Elektryczność - pozaprodukcyjne (stałe) [waluta/rok]": "Electricity - Facility (fixed)",
                "Energia razem [waluta/rok]": "Total Energy",
            }
            energy_data = [["Energy KPI", f"Value [{waluta_report}/year]"]]
            for k, v in y["energy"].items():
                if k == "Rok":
                    continue
                energy_data.append([energy_label_en.get(k, k), f"{v:,.0f}"])
            story.append(styled_table(energy_data, col_widths=[11 * cm, 5 * cm]))
            story.append(Spacer(1, 0.4 * cm))

        if y["products"]:
            story.append(Paragraph("Products (this year, top 12 by volume)", h4_style))
            prod_data = [["Product", "Group", "Mode", "Volume [kg]"]]
            for p in y["products"]:
                prod_data.append([p["product"], p["group"], p["mode"], f'{p["kg"]:,.0f}'])
            story.append(styled_table(prod_data, col_widths=[6 * cm, 4 * cm, 3 * cm, 3 * cm]))

    doc_buf = io.BytesIO()
    doc = SimpleDocTemplate(doc_buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    doc.build(story)
    doc_buf.seek(0)
    return doc_buf.getvalue()


def generate_excel_report_bytes(report_data, roi_rows, waluta_report):
    """
    Buduje skoroszyt Excel (openpyxl - już zależność tej apki, zero nowego ryzyka wdrożeniowego)
    z tym samym zestawem danych co raport PDF, ale w formie do dalszej pracy: arkusz Summary
    (5 lat: tonaż/produkcja/import/magazyn/energia/OPEX/ROI w jednej tabeli), Fleet (flota),
    Products (produkty per rok), oraz Charts z natywnymi, edytowalnymi wykresami Excela.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter

    years = report_data["years"]
    target_annual_t = report_data["target_annual_t"]
    fleet = report_data["fleet"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()

    # --- Arkusz Summary (lata jako KOLUMNY, metryki jako WIERSZE) ---
    ws = wb.active
    ws.title = "Summary"
    n_years = len(years)

    # Kolejność = kolejność wierszy w arkuszu; stałe indeksy wierszy używane też przy budowie
    # wykresów niżej, żeby nie odwoływać się do "magicznych" numerów.
    ROW_YEAR, ROW_TARGET_PCT, ROW_TOTAL_VOL, ROW_PRODUCED, ROW_IMPORTED = 1, 2, 3, 4, 5
    ROW_WH_UTIL, ROW_WH_VALUE = 6, 7
    ROW_HEATING, ROW_ELEC_PROCESS, ROW_ELEC_FACILITY, ROW_TOTAL_ENERGY = 8, 9, 10, 11
    ROW_OPEX, ROW_REVENUE, ROW_PROFIT, ROW_CUM_PROFIT, ROW_ROI = 12, 13, 14, 15, 16

    row_labels = {
        ROW_YEAR: "Metric", ROW_TARGET_PCT: "% of Target", ROW_TOTAL_VOL: "Total Volume [t]",
        ROW_PRODUCED: "Produced [t]", ROW_IMPORTED: "Imported [t]", ROW_WH_UTIL: "Warehouse Utilization [%]",
        ROW_WH_VALUE: f"Warehouse Stock Value [{waluta_report}]", ROW_HEATING: f"Heating [{waluta_report}/year]",
        ROW_ELEC_PROCESS: f"Electricity - Process, incl. cooling [{waluta_report}/year]",
        ROW_ELEC_FACILITY: f"Electricity - Facility, fixed [{waluta_report}/year]",
        ROW_TOTAL_ENERGY: f"Total Energy [{waluta_report}/year]", ROW_OPEX: f"OPEX [{waluta_report}/year]",
        ROW_REVENUE: f"Revenue [{waluta_report}/year]", ROW_PROFIT: f"Profit [{waluta_report}/year]",
        ROW_CUM_PROFIT: f"Cumulative Profit [{waluta_report}]", ROW_ROI: "ROI (this year) [%]",
    }
    for row_idx, label in row_labels.items():
        c = ws.cell(row=row_idx, column=1, value=label)
        c.font = header_font if row_idx == ROW_YEAR else Font(bold=True)
        if row_idx == ROW_YEAR:
            c.fill = header_fill

    for i, y in enumerate(years):
        col = 2 + i  # B, C, D, E, F = Year 1..5
        roi_row = roi_rows[i] if i < len(roi_rows) else {}
        energy = y.get("energy", {})
        col_letter = get_column_letter(col)

        header_cell = ws.cell(row=ROW_YEAR, column=col, value=f"Year {y['year']}")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = wrap_center

        ws.cell(row=ROW_TARGET_PCT, column=col, value=round(y["target_pct"], 1))
        ws.cell(row=ROW_TOTAL_VOL, column=col, value=round(y["total_t"], 1))
        ws.cell(row=ROW_PRODUCED, column=col, value=round(y["produced_t"], 1))
        ws.cell(row=ROW_IMPORTED, column=col, value=round(y["imported_t"], 1))
        ws.cell(row=ROW_WH_UTIL, column=col, value=round(y["wh_util_pct"], 1) if y["wh_util_pct"] is not None else None)
        ws.cell(row=ROW_WH_VALUE, column=col, value=round(y["wh_value"], 0) if y["wh_value"] is not None else None)
        ws.cell(row=ROW_HEATING, column=col, value=energy.get("Ogrzewanie [waluta/rok]"))
        ws.cell(row=ROW_ELEC_PROCESS, column=col, value=energy.get("Elektryczność - proces (w tym chłodzenie) [waluta/rok]"))
        ws.cell(row=ROW_ELEC_FACILITY, column=col, value=energy.get("Elektryczność - pozaprodukcyjne (stałe) [waluta/rok]"))
        ws.cell(row=ROW_TOTAL_ENERGY, column=col, value=energy.get("Energia razem [waluta/rok]"))
        ws.cell(row=ROW_OPEX, column=col, value=roi_row.get("OPEX roczny"))
        ws.cell(row=ROW_REVENUE, column=col, value=roi_row.get("Przychód roczny"))
        ws.cell(row=ROW_PROFIT, column=col, value=roi_row.get("Zysk roczny"))
        ws.cell(row=ROW_CUM_PROFIT, column=col, value=roi_row.get("Zysk skumulowany"))
        ws.cell(row=ROW_ROI, column=col, value=roi_row.get("ROI (ten rok) [%]"))

    ws.column_dimensions["A"].width = 42
    for i in range(n_years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14
    ws.freeze_panes = "B2"
    ws.cell(row=ROW_ROI + 2, column=1, value=f"Target annual capacity (100%, Year 5+): {target_annual_t:,.0f} t/year")
    ws.cell(row=ROW_ROI + 3, column=1, value=f"Currency: {waluta_report}")

    # --- Arkusz Fleet ---
    ws_fleet = wb.create_sheet("Fleet")
    fleet_headers = ["Tag", "Product Family", "Capacity [m3]", "Batch Mass [kg]", "Cycle [h]", "Batches/month (target)"]
    for col_idx, h in enumerate(fleet_headers, start=1):
        c = ws_fleet.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, m in enumerate(fleet, start=2):
        ws_fleet.cell(row=i, column=1, value=m["tag"])
        ws_fleet.cell(row=i, column=2, value=m["product_family"])
        ws_fleet.cell(row=i, column=3, value=round(m["capacity_m3"], 1))
        ws_fleet.cell(row=i, column=4, value=round(m["mass_per_batch"], 0))
        ws_fleet.cell(row=i, column=5, value=round(m["cycle_h"], 2))
        ws_fleet.cell(row=i, column=6, value=m["batches_count"])
    for col_idx in range(1, len(fleet_headers) + 1):
        ws_fleet.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- Arkusz Products ---
    ws_prod = wb.create_sheet("Products")
    prod_headers = ["Year", "Product", "Group", "Mode", "Volume [kg]"]
    for col_idx, h in enumerate(prod_headers, start=1):
        c = ws_prod.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
    row_ptr = 2
    for y in years:
        for p in y["products"]:
            ws_prod.cell(row=row_ptr, column=1, value=y["year"])
            ws_prod.cell(row=row_ptr, column=2, value=p["product"])
            ws_prod.cell(row=row_ptr, column=3, value=p["group"])
            ws_prod.cell(row=row_ptr, column=4, value=p["mode"])
            ws_prod.cell(row=row_ptr, column=5, value=round(p["kg"], 0))
            row_ptr += 1
    for col_idx, w in zip(range(1, 6), [8, 26, 18, 12, 14]):
        ws_prod.column_dimensions[get_column_letter(col_idx)].width = w

    # --- Arkusz Charts: natywne, edytowalne wykresy Excela zbudowane z danych w Summary ---
    ws_charts = wb.create_sheet("Charts")
    last_year_col = 1 + n_years  # kolumna F dla 5 lat (B..F)
    cats_row = Reference(ws, min_col=2, max_col=last_year_col, min_row=ROW_YEAR, max_row=ROW_YEAR)

    bar1 = BarChart()
    bar1.type = "col"
    bar1.grouping = "stacked"
    bar1.title = "Production Volume: Own Production vs. Import"
    bar1.y_axis.title = "Tonnage [t/year]"
    bar1.x_axis.title = "Year"
    data1 = Reference(ws, min_col=1, max_col=last_year_col, min_row=ROW_PRODUCED, max_row=ROW_IMPORTED)
    bar1.add_data(data1, titles_from_data=True, from_rows=True)
    bar1.set_categories(cats_row)
    ws_charts.add_chart(bar1, "A1")

    bar2 = BarChart()
    bar2.type = "col"
    bar2.grouping = "stacked"
    bar2.title = "Energy Cost Breakdown by Year"
    bar2.y_axis.title = f"Cost [{waluta_report}/year]"
    bar2.x_axis.title = "Year"
    data2 = Reference(ws, min_col=1, max_col=last_year_col, min_row=ROW_HEATING, max_row=ROW_ELEC_FACILITY)
    bar2.add_data(data2, titles_from_data=True, from_rows=True)
    bar2.set_categories(cats_row)
    ws_charts.add_chart(bar2, "A20")

    line1 = LineChart()
    line1.title = "Cumulative Profit vs. CAPEX Payback"
    line1.y_axis.title = f"[{waluta_report}]"
    line1.x_axis.title = "Year"
    data3 = Reference(ws, min_col=1, max_col=last_year_col, min_row=ROW_CUM_PROFIT, max_row=ROW_CUM_PROFIT)
    line1.add_data(data3, titles_from_data=True, from_rows=True)
    line1.set_categories(cats_row)
    ws_charts.add_chart(line1, "A39")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()



# ==========================================
# FUNKCJE POMOCNICZE (wydzielone z pętli UI, aby dało się je testować niezależnie)
# ==========================================

def reynolds_number(velocity_ms, diameter_m, viscosity_cst):
    """Liczba Reynoldsa dla przepływu w rurze kołowej."""
    viscosity_m2s = viscosity_cst * 1e-6
    if viscosity_m2s <= 0 or diameter_m <= 0:
        return 0.0
    return (velocity_ms * diameter_m) / viscosity_m2s


def friction_factor(re):
    """
    Współczynnik oporów liniowych.
    - Re <= 2320: przepływ laminarny, wzór Hagen-Poiseuille (64/Re).
    - Re > 2320: przepływ turbulentny, korelacja Blasiusa (ważna dla Re < ~1e5,
      co pokrywa typowe warunki przetłaczania olejów/emulsji w tej instalacji).
    """
    if re <= 0:
        return 0.0
    if re <= 2320:
        return 64.0 / re
    return 0.316 * (re ** -0.25)


def compute_hydraulics(q_m3h, pipe_dn_mm, pipe_length_m, delta_h_m, viscosity_cst,
                        density_kgm3, zeta_sum, pump_efficiency):
    """Zwraca (Re, opór całkowity [bar], moc pompy [kW], prędkość [m/s])."""
    q_m3s = q_m3h / 3600.0
    d_m = pipe_dn_mm / 1000.0
    area_m2 = math.pi * (d_m ** 2) / 4.0
    velocity = q_m3s / area_m2 if area_m2 > 0 else 0.0

    dynamic_pressure = density_kgm3 * (velocity ** 2) / 2.0
    p_hydrostatic = density_kgm3 * G_ACCEL * delta_h_m

    re = reynolds_number(velocity, d_m, viscosity_cst)
    lam = friction_factor(re)

    p_loss_lin = lam * (pipe_length_m / d_m) * dynamic_pressure if d_m > 0 else 0.0
    tot_p_pa = p_loss_lin + (zeta_sum * dynamic_pressure) + p_hydrostatic
    tot_p_bar = tot_p_pa / 100000.0
    power_kw = (q_m3s * tot_p_pa) / pump_efficiency / 1000.0 if pump_efficiency > 0 else 0.0

    return re, tot_p_bar, power_kw, velocity


def estimate_tank_surface_area_m2(capacity_m3):
    """
    Szacunkowa powierzchnia zewnętrzna zbiornika cylindrycznego, przy założeniu H≈D (typowe dla
    małych/średnich zbiorników magazynowych) - powierzchnia boczna + dwie podstawy.
    """
    if capacity_m3 <= 0:
        return 0.0
    d_m = (4.0 * capacity_m3 / math.pi) ** (1.0 / 3.0)  # H=D -> V = (pi/4)*D^3
    area_side = math.pi * d_m * d_m  # H=D, więc powierzchnia boczna = pi*D*H = pi*D^2
    area_ends = 2.0 * (math.pi / 4.0) * d_m ** 2
    return area_side + area_ends


def compute_tank_overall_u_wm2k(insulation_mm, k_insulation_wmk=0.04, h_inside_wm2k=50.0, h_outside_wm2k=10.0):
    """
    Całkowity współczynnik przenikania ciepła [W/(m2*K)] przez ściankę zbiornika - opór stali
    pominięty (bardzo dobry przewodnik, opór pomijalny wobec izolacji/konwekcji). Bez izolacji
    (0 mm) liczy się tylko z oporów konwekcji wewnątrz/na zewnątrz.
    """
    r_total = (1.0 / h_inside_wm2k) + (1.0 / h_outside_wm2k)
    if insulation_mm > 0:
        r_total += (insulation_mm / 1000.0) / k_insulation_wmk
    return 1.0 / r_total if r_total > 0 else 0.0


def compute_tank_cooling_curve(capacity_m3, insulation_mm, t_start_c, t_ambient_c,
                                density_kgm3=900.0, cp_j_kgk=2000.0, hours=range(0, 73, 4)):
    """
    Krzywa wychładzania zbiornika (prawo stygnięcia Newtona): T(t) = T_amb + (T0-T_amb)*exp(-t/tau),
    tau = (m*cp)/(U*A). Zwraca listę temperatur [°C] dla podanych godzin. Uproszczenie inżynierskie:
    stały współczynnik U (nie uwzględnia np. wiatru, promieniowania, zmiany lepkości z temperaturą).
    """
    area_m2 = estimate_tank_surface_area_m2(capacity_m3)
    mass_kg = capacity_m3 * TANK_SAFETY_FILL * density_kgm3
    u_val = compute_tank_overall_u_wm2k(insulation_mm)
    if area_m2 <= 0 or mass_kg <= 0 or u_val <= 0:
        return [t_start_c for _ in hours]
    tau_s = (mass_kg * cp_j_kgk) / (u_val * area_m2)
    return [t_ambient_c + (t_start_c - t_ambient_c) * math.exp(-(h * 3600.0) / tau_s) for h in hours]


def compute_tank_heating_power_kw(capacity_m3, insulation_mm, t_target_c, t_ambient_c, safety_margin=1.2):
    """
    Szacunkowa moc grzania [kW] potrzebna, żeby w stanie ustalonym utrzymać zbiornik w temperaturze
    docelowej wobec strat do otoczenia (Q = U*A*ΔT), z marginesem bezpieczeństwa na rozruch.
    """
    area_m2 = estimate_tank_surface_area_m2(capacity_m3)
    u_val = compute_tank_overall_u_wm2k(insulation_mm)
    delta_t = max(t_target_c - t_ambient_c, 0.0)
    q_watts = u_val * area_m2 * delta_t * safety_margin
    return q_watts / 1000.0


def check_coil_velocity(flow_value, flow_unit, medium_type, coil_pipe_dn_mm):
    """
    Sprawdza, czy podany przepływ medium grzewczego/chłodzącego przez rurę wężownicy/płaszcza o
    danej średnicy daje FIZYCZNIE SENSOWNĄ prędkość - inaczej łatwo o pomyłkę rzędu wielkości
    (np. wpisanie 500 zamiast 5 m³/h), która przy niewielkiej średnicy rury dałaby prędkość
    wielokrotnie przekraczającą prędkość dźwięku. Te same granice co dla rurociągu produktu
    (VELOCITY_MIN_MS - VELOCITY_MAX_MS), para pominięta (kondensat, inna fizyka przepływu).
    Zwraca (prędkość [m/s], ostrzeżenie: None/"za_wolno"/"za_szybko") lub (None, None) dla pary.
    """
    if MEDIA_PROCESOWE[medium_type]["steam"]:
        return None, None
    flow_m3h = flow_value if flow_unit == "m3/h" else flow_value * 60.0 / 1000.0
    area_m2 = math.pi * ((coil_pipe_dn_mm / 1000.0) / 2.0) ** 2
    if area_m2 <= 0:
        return None, None
    velocity_ms = (flow_m3h / 3600.0) / area_m2
    if velocity_ms < VELOCITY_MIN_MS:
        return velocity_ms, "za_wolno"
    if velocity_ms > VELOCITY_MAX_MS:
        return velocity_ms, "za_szybko"
    return velocity_ms, None


def compute_thermal_ntu(mass_kg, cp_product, t_initial, t_target, k_coeff_w_m2k, area_m2,
                          utility_type, flow_value, flow_unit, t_utility_in):
    """
    Bilans grzania/chłodzenia metodą NTU-efektywności (produkt dobrze wymieszany = jednolita
    temperatura w danej chwili; medium przepływa CIĄGLE przez płaszcz/wężownicę o stałym
    UA = k*F). W przeciwieństwie do uproszczonej metody "średnie ΔT", poprawnie oddaje, że
    temperatura produktu zbliża się do temperatury medium ASYMPTOTYCZNIE (szybciej na początku,
    wolniej pod koniec) - wymagany czas liczony jest z całki różniczkowej (rozwiązanie
    analityczne: eksponencjalne dążenie do temperatury medium), nie z jednej uśrednionej mocy.

    flow_unit: "m3/h" lub "L/min". Para nasycona traktowana specjalnie (kondensacja = stała
    temperatura = efektywnie nieskończona pojemność cieplna strumienia, efektywność = 1).

    Zwraca dict z kluczem "status" ("ok" / "brak_potrzeby" / "niewystarczajace_dt" /
    "niemozliwe_do_osiagniecia" / "brak_przeplywu") i przy "ok" dodatkowo: q_total_kj,
    effectiveness, required_time_h, power_start_kw.
    """
    if abs(t_target - t_initial) < 1e-9:
        return {"status": "brak_potrzeby"}
    is_heating = t_target > t_initial
    if is_heating and t_utility_in <= t_initial:
        return {"status": "niewystarczajace_dt", "message": "Medium grzewcze nie jest cieplejsze od produktu początkowego."}
    if not is_heating and t_utility_in >= t_initial:
        return {"status": "niewystarczajace_dt", "message": "Medium chłodzące nie jest zimniejsze od produktu początkowego."}

    media_cfg = MEDIA_PROCESOWE[utility_type]
    cp_medium = media_cfg["cp"]
    kF_kw_k = (k_coeff_w_m2k * area_m2) / 1000.0
    mass_cp_kj_k = mass_kg * cp_product
    q_total_kj = mass_cp_kj_k * abs(t_target - t_initial)

    if media_cfg["steam"]:
        # Kondensująca para: temperatura medium stała (nasycona) niezależnie od przepływu -
        # efektywnie nieskończona pojemność cieplna strumienia. Moc liczona wprost z kF i średniej
        # różnicy temperatur (uproszczenie uzasadnione fizycznie przez stałą temp. kondensacji,
        # nie arbitralne założenie jak w dawnym modelu dla cieczy).
        effectiveness = 1.0
        approx_dt = t_utility_in - ((t_initial + t_target) / 2.0)
        power_avg_kw = kF_kw_k * approx_dt
        if power_avg_kw <= 0:
            return {"status": "niewystarczajace_dt", "message": "Zbyt mała różnica temperatur względem pary."}
        required_time_h = q_total_kj / (power_avg_kw * 3600.0)
        power_start_kw = kF_kw_k * abs(t_utility_in - t_initial)
        return {"status": "ok", "q_total_kj": q_total_kj, "effectiveness": effectiveness,
                "required_time_h": required_time_h, "power_start_kw": power_start_kw}

    density = media_cfg["density_kg_m3"]
    flow_m3h = flow_value if flow_unit == "m3/h" else flow_value * 60.0 / 1000.0
    flow_kg_h = flow_m3h * density
    w_kw_k = (flow_kg_h * cp_medium) / 3600.0
    if w_kw_k <= 0:
        return {"status": "brak_przeplywu"}

    ntu = kF_kw_k / w_kw_k
    effectiveness = 1.0 - math.exp(-ntu)
    if effectiveness <= 0:
        return {"status": "brak_wymiany", "q_total_kj": q_total_kj, "effectiveness": effectiveness}

    ratio = (t_target - t_utility_in) / (t_initial - t_utility_in)
    if ratio <= 0:
        return {"status": "niemozliwe_do_osiagniecia", "q_total_kj": q_total_kj, "effectiveness": effectiveness}

    time_constant_s = mass_cp_kj_k / (effectiveness * w_kw_k)
    required_time_h = (-time_constant_s * math.log(ratio)) / 3600.0
    power_start_kw = effectiveness * w_kw_k * abs(t_utility_in - t_initial)

    return {"status": "ok", "q_total_kj": q_total_kj, "effectiveness": effectiveness,
            "required_time_h": required_time_h, "power_start_kw": power_start_kw}


def compute_agitator_power(agitator_type, rpm, impeller_d_m, density_kgm3, viscosity_cst):
    """
    Szacunkowa moc mieszania na podstawie liczby Reynoldsa mieszania.
    Reżim laminarny:  P = C * mu * N^2 * D^3
    Reżim turbulentny: P = Ne * rho * N^3 * D^5
    Zwraca (Re_mix, reżim, moc [kW]).
    """
    cfg = AGITATOR_TYPES.get(agitator_type)
    if cfg is None or rpm <= 0 or impeller_d_m <= 0:
        return 0.0, "brak danych", 0.0

    n_rps = rpm / 60.0
    mu_pas = (viscosity_cst * 1e-6) * density_kgm3  # cSt -> m2/s -> Pa*s

    re_mix = (density_kgm3 * n_rps * (impeller_d_m ** 2)) / mu_pas if mu_pas > 0 else 1e9

    if re_mix <= 10:
        regime = "laminarny"
        power_w = cfg["laminar_C"] * mu_pas * (n_rps ** 2) * (impeller_d_m ** 3)
    else:
        regime = "turbulentny"
        power_w = cfg["turbulent_Ne"] * density_kgm3 * (n_rps ** 3) * (impeller_d_m ** 5)

    return re_mix, regime, power_w / 1000.0


# Ściągawka średnic rur stalowych (średnica wewnętrzna [m]) do wyboru DN w module pary.
STEAM_PIPE_DN_REFERENCE_M = {
    "DN80": 0.0800, "DN100": 0.1053, "DN125": 0.1300, "DN150": 0.1586, "DN200": 0.2073,
}


def compute_vent_line_scenario(mass_flow_kgs, rho_steam, pipe_length_m, pipe_diameter_m, lambda_friction):
    """
    Bilans hydrauliczny linii odpowietrzenia/zrzutu pary dla jednego scenariusza pracy,
    metodą Darcy-Weisbacha - analogicznie do compute_hydraulics dla cieczy, ale tu strumień
    wejściowy to zadany strumień masowy pary [kg/s] (nie przepływ objętościowy pompy), a opory
    liczone są jako czysto liniowe (bez lokalnych zeta), zgodnie z modelem referencyjnym
    użytkownika (ISO 28300 / API 2000).
    Zwraca (objętość [m3/s], prędkość [m/s], opory [Pa], opory [bar]).
    """
    if rho_steam <= 0 or pipe_diameter_m <= 0:
        return 0.0, 0.0, 0.0, 0.0
    volumetric_flow_m3s = mass_flow_kgs / rho_steam
    area_m2 = 3.14159265 * (pipe_diameter_m ** 2) / 4.0
    velocity_ms = volumetric_flow_m3s / area_m2 if area_m2 > 0 else 0.0
    dp_pa = lambda_friction * (pipe_length_m / pipe_diameter_m) * (rho_steam * velocity_ms ** 2 / 2.0) if pipe_diameter_m > 0 else 0.0
    dp_bar = dp_pa / 100000.0
    return volumetric_flow_m3s, velocity_ms, dp_pa, dp_bar


# --- 2. INICJALIZACJA STRUKTUR W SESJI ---
if "active_portfolio" not in st.session_state:
    # Startowo = 7 generycznych grup produktowych (GENERIC_PORTFOLIO); wgrana receptura
    # (Zakładka 1) odświeża gęstość/cykl/materiał danej grupy jej własnymi wartościami.
    st.session_state.active_portfolio = {k: dict(v) for k, v in GENERIC_PORTFOLIO.items()}

if "prod_dict" not in st.session_state:
    st.session_state.prod_dict = {
        k: {"roczna": 1200000, "user_vol_m3": 15.0, "skus": 1, "num_tanks": 1, "tank_volumes": [15.0]} for k in st.session_state.active_portfolio.keys()
    }

if "confirmed_mixers" not in st.session_state:
    st.session_state.confirmed_mixers = []

if "calculated_times" not in st.session_state:
    st.session_state.calculated_times = {}

if "mixer_tech_advanced_details" not in st.session_state:
    st.session_state.mixer_tech_advanced_details = {}

if "batch_time_components" not in st.session_state:
    st.session_state.batch_time_components = {}

if "recipes_df" not in st.session_state:
    st.session_state.recipes_df = None  # DataFrame wgranych receptur (produkt, roczne zapotrzebowanie, % surowców)

if "recipe_raw_material_consumption" not in st.session_state:
    st.session_state.recipe_raw_material_consumption = None  # dict: surowiec -> t/rok, liczone z recipes_df

if "pack_configs" not in st.session_state:
    # Startowo = wbudowane wartości domyślne (PACK_CONFIGS); po wgraniu arkusza 'Opakowania'
    # w Zakładce 1 są nadpisywane/uzupełniane, a dalej pozostają w pełni edytowalne w apce.
    st.session_state.pack_configs = {k: dict(v) for k, v in PACK_CONFIGS.items()}

if "qc_tests_by_product" not in st.session_state:
    # dict: nazwa produktu -> lista nazw testów QC (z arkusza 'Badania Laboratoryjne', jeśli
    # wgrany) - ma pierwszeństwo nad panelem zwolnienia per linia (Zakładka 6, VSM).
    st.session_state.qc_tests_by_product = {}

if "direct_raw_materials" not in st.session_state:
    # lista dictów {"group","material","annual_t"} - surowce zadeklarowane wprost (bez pełnej
    # receptury), z arkusza 'Zużycie Surowców (bez receptury)', jeśli wgrany. Dolicza się do
    # zużycia liczonego z receptur, nie zastępuje go.
    st.session_state.direct_raw_materials = []

if "qc_equipment_count_override" not in st.session_state:
    # dict: nazwa aparatu -> liczba fizycznych sztuk w laboratorium (domyślnie 1 z katalogu,
    # ale niektóre testy mogą mieć 2+ identycznych aparatów - to bezpośrednio podwaja
    # przepustowość tego konkretnego testu w całym zakładzie, nie tylko dla jednej szarży).
    st.session_state.qc_equipment_count_override = {}

if "supplier_splits" not in st.session_state:
    # lista dictów {"group","category","supplier","pct","tech","tank_id"} - rozbicie kategorii
    # surowca na konkretnych dostawców/bazy, z arkusza 'Rozbicie Dostawców Surowców', jeśli
    # wgrany. Pozwala pokazać kilka mniejszych zbiorników zamiast jednego dużego per kategoria.
    st.session_state.supplier_splits = []

if "custom_qc_tests" not in st.session_state:
    # dict: nazwa testu -> {"duration_min","equipment","count"} - testy DOPISANE przez
    # użytkownika w arkuszu 'Badania Laboratoryjne' (własne kolumny 'Sprzęt'/'Czas [min]'),
    # których nie ma w wbudowanym katalogu QC_TEST_CATALOG. Scalane z wbudowanym katalogiem
    # wszędzie przez get_full_qc_catalog() - nie jesteś ograniczony do wbudowanej listy testów.
    st.session_state.custom_qc_tests = {}

if "equipment_df" not in st.session_state:
    st.session_state.equipment_df = None  # DataFrame cennika standardowej instalacji (Zakładka 5)

if "equipment_install_counts" not in st.session_state:
    st.session_state.equipment_install_counts = {}  # dict: Grupa Produktowa -> liczba planowanych instalacji

if "recipe_groups_seen" not in st.session_state:
    st.session_state.recipe_groups_seen = set()  # grupy produktowe z receptury już raz zsynchronizowane z flotą

if "_last_recipe_group_signature" not in st.session_state:
    st.session_state._last_recipe_group_signature = None

if "shared_pumps" not in st.session_state:
    # Pompy współdzielone przez kilka zbiorników (jedna fizyczna pompa obsługująca kilka
    # mieszalników na przemian) - jedno źródło prawdy dla przepływu/sprawności/MTBF/MTTR
    # takiej pompy, niezależnie od tego, ile zbiorników ją współdzieli. Klucz = ID pompy
    # nadane przez użytkownika (np. "P-01"), wartość = dict z parametrami.
    st.session_state.shared_pumps = {}

if "confirmed_rm_tanks" not in st.session_state:
    st.session_state.confirmed_rm_tanks = []  # zatwierdzone zbiorniki RM (Zakładka 2) - tag/materiał/pojemność

if "rm_tank_tech_details" not in st.session_state:
    st.session_state.rm_tank_tech_details = {}  # per RM tank tag: pompa (dedyk./współdz.), DN/długość/wysokość/zawory

if "confirmed_fg_buffer_tanks" not in st.session_state:
    # Zbiorniki buforowe GOTOWEGO PRODUKTU (nie surowca) - ten sam mechanizm co zbiorniki RM,
    # tag/recipe_product/pojemność. Służą do (a) uwolnienia mieszalnika zaraz po szarży (produkt
    # przechodzi do bufora, mieszalnik startuje kolejną szarżę) i (b) bezpośredniego wydania
    # produktu w cysternie z tego zbiornika, bez pośredniego składowania na paletach.
    st.session_state.confirmed_fg_buffer_tanks = []

if "fg_buffer_tank_tech_details" not in st.session_state:
    st.session_state.fg_buffer_tank_tech_details = {}  # ta sama struktura co rm_tank_tech_details

if "rampup_global_pct" not in st.session_state:
    # Domyślna krzywa rozruchu (% docelowej produkcji osiągane w kolejnych latach) - flota i
    # magazyn są wymiarowane od razu pod docelową (100%) produkcję; ta krzywa opisuje jedynie
    # jak rośnie ich WYKORZYSTANIE w pierwszych RAMPUP_YEARS latach.
    st.session_state.rampup_global_pct = [40.0, 60.0, 80.0, 95.0, 100.0][:RAMPUP_YEARS]
if "rampup_differentiate" not in st.session_state:
    st.session_state.rampup_differentiate = False
if "rampup_per_line_pct" not in st.session_state:
    st.session_state.rampup_per_line_pct = {}  # dict: linia -> lista % per rok (tylko gdy differentiate=True)
if "rampup_start_pct" not in st.session_state:
    # Punkt startowy krzywej PRZED Rokiem 1 (t=0) - domyślnie 0% (start od zera, czysty rozruch).
    # Parametr celowo elastyczny: w przyszłości można tu wpisać np. 20%, żeby zasymulować
    # rozruch od fazy z istniejącą bazą klientów/zapasem, zamiast zawsze zaczynać od zera.
    st.session_state.rampup_start_pct = 0.0
if "import_follows_rampup" not in st.session_state:
    # Domyślnie: wolumen produktu JESZCZE importowanego rośnie tą samą krzywą rozruchu co
    # produkcja własna (założenie: popyt klienta rośnie niezależnie od tego, kto go zaspokaja).
    # Jeśli False: import liczony zawsze jako pełny docelowy wolumen (100%), niezależnie od roku -
    # np. gdy realny kontrakt importowy nie jest powiązany z krzywą rozruchu Twojej fabryki.
    st.session_state.import_follows_rampup = True


def _rampup_control_points(product_family):
    """Punkty kontrolne krzywej rozruchu w % (indeks 0 = start przed Rokiem 1, indeks i = koniec
    Roku i), używane do interpolacji liniowej ciągłej w czasie - patrz get_rampup_fraction_continuous."""
    if st.session_state.get("rampup_differentiate") and product_family in st.session_state.get("rampup_per_line_pct", {}):
        pct_list = st.session_state.rampup_per_line_pct[product_family]
    else:
        pct_list = st.session_state.get("rampup_global_pct", [100.0] * RAMPUP_YEARS)
    return [st.session_state.get("rampup_start_pct", 0.0)] + list(pct_list)


def get_rampup_fraction_continuous(product_family, t_years):
    """
    Ułamek (0-1) docelowej produkcji w CIĄGŁYM czasie t_years (0 = start Roku 1, 1 = koniec
    Roku 1 / start Roku 2, itd.) - interpolacja LINIOWA między punktami kontrolnymi (punkt
    startowy + wartości Rok 1..N), zamiast skokowej zmiany na granicy lat. Rzeczywistość rzadko
    jest liniowa "schodkami" - produkcja rośnie płynnie w trakcie roku, nie skacze 1 stycznia.
    """
    points = _rampup_control_points(product_family)
    t_clamped = max(0.0, min(t_years, float(len(points) - 1)))
    idx_lo = int(t_clamped)
    idx_hi = min(idx_lo + 1, len(points) - 1)
    frac_within = t_clamped - idx_lo
    pct_lo = max(0.0, min(float(points[idx_lo]), 100.0))
    pct_hi = max(0.0, min(float(points[idx_hi]), 100.0))
    pct = pct_lo + (pct_hi - pct_lo) * frac_within
    return pct / 100.0


def get_rampup_fraction_month(product_family, year_idx, month_idx):
    """Ułamek (0-1) w połowie danego miesiąca (month_idx 0-11) danego roku symulacji (year_idx
    0-4) - do symulacji magazynowej i wykresów min/max/średnia w obrębie roku (Dashboard)."""
    t_years = year_idx + (month_idx + 0.5) / MONTHS_PER_YEAR
    return get_rampup_fraction_continuous(product_family, t_years)


def get_rampup_fraction(product_family, year_idx):
    """
    Ułamek (0-1) docelowej produkcji osiąganej w danym roku symulacji rozruchu, dla danej
    linii produktowej. Jedno źródło prawdy używane zarówno w Zakładce 2 (utylizacja floty)
    jak i Zakładce 2 (wykorzystanie magazynu FG+RM) - patrz sync_recipes_into_fleet_defaults
    dla analogicznego wzorca "jedno miejsce prawdy, wiele zakładek czyta".

    Zwraca ŚREDNIĄ z całego roku (nie wartość na koniec roku) - dokładną, bo interpolacja jest
    liniowa (odcinek pomiędzy dwoma punktami kontrolnymi ma stałe nachylenie, więc średnia
    odcinka to dokładnie średnia jego dwóch końców, bez potrzeby próbkowania miesiąc po miesiącu).
    Dzięki temu każde dotychczasowe miejsce wywołania (roczny tonaż/koszt/zużycie surowców)
    automatycznie odzwierciedla płynny rozruch w trakcie roku, bez zmiany logiki wywołującej.
    """
    if year_idx == RAMPUP_YEAR_TARGET_SENTINEL:
        return 1.0
    points = _rampup_control_points(product_family)
    if year_idx < 0 or year_idx >= len(points) - 1:
        return max(0.0, min(float(points[-1]), 100.0)) / 100.0
    frac_start = get_rampup_fraction_continuous(product_family, float(year_idx))
    frac_end = get_rampup_fraction_continuous(product_family, float(year_idx + 1))
    return (frac_start + frac_end) / 2.0


def check_fleet_staleness_warning():
    """
    Flota (confirmed_mixers) to RĘCZNA migawka - zapisywana tylko kiedy użytkownik kliknie
    'Zatwierdź' w Zakładce 1. Jeśli receptura zostanie potem zmieniona (np. nowy produkt,
    inna roczna produkcja, zmiana sposobu pozyskania), migawka NIE aktualizuje się sama -
    a każda inna zakładka (Karta Maszyn, Logistyka, Finanse, Dashboard) i tak czyta ze starej
    migawki, dając liczby niezgodne z tym, co Zakładka 1 pokazuje na żywo. Ta funkcja wykrywa
    taką rozbieżność i zwraca gotowy komunikat ostrzegawczy do wyświetlenia (st.error) na
    początku każdej zależnej zakładki - albo None, jeśli flota jest aktualna.

    WAŻNE: porównuje z LIVE_FLEET_ANNUAL_KG_ROUNDED (ta sama formuła zaokrąglania szarż w górę
    co przy zatwierdzaniu), NIE z dokładnym tonażem z receptury wprost - flota fizycznie NIE
    MOŻE dokładnie równać się dokładnemu celowi (nie da się zrobić ułamka szarży), więc
    porównanie z dokładną wartością dawało fałszywe ostrzeżenie nawet zaraz po zatwierdzeniu.
    """
    if not st.session_state.get("confirmed_mixers"):
        return None
    if "live_fleet_annual_kg_rounded" not in st.session_state:
        return None  # Zakładka 1 jeszcze nie wykonała się w tej sesji - nie ma z czym porównać
    live_target_t = st.session_state["live_fleet_annual_kg_rounded"] / 1000.0
    confirmed_target_t = sum(m["annual_volume"] for m in st.session_state.confirmed_mixers) / 1000.0
    if live_target_t <= 0:
        return None
    diff_t = abs(live_target_t - confirmed_target_t)
    if diff_t > max(1.0, confirmed_target_t * 0.01):
        return (f"⚠️ **FLOTA JEST NIEAKTUALNA** — Zakładka 1 pokazuje obecnie **{live_target_t:,.0f} t/rok** "
                f"(po zaokrągleniu szarż), ale zatwierdzona flota (której używa ta zakładka) to "
                f"**{confirmed_target_t:,.0f} t/rok** — różnica **{diff_t:,.0f} t**. Receptura została zmieniona PO "
                "ostatnim zatwierdzeniu floty. Wszystkie liczby tutaj (tonaż, utylizacja, koszty, magazyn) są "
                "policzone ze STAREJ floty. Wróć do **Zakładki 1** i kliknij **'📥 Zatwierdź i wyślij konfigurację'**, "
                "żeby zsynchronizować.")
    return None


def sync_recipes_into_fleet_defaults():
    """
    Spina Zakładkę 1 (Receptury) z Zakładką 2 (Flota) NA POZIOMIE GRUPY PRODUKTOWEJ - tak jak
    w pliku Excel (Cleaners/Engine Oils/Glycols/Greases/Hydraulic Oils/Watermiscibles/Waxes),
    a nie per pojedynczy produkt czy stara szczegółowa lista marek. Każda grupa
    obecna w recepturze staje się JEDNĄ pozycją do wyboru w panelu bocznym.

    W ramach grupy, wiersze (produkty) dzielą się na "sloty" zbiornikowe wg RECIPE_TANK_ID_COL:
    - Puste ID -> każdy produkt dostaje swój własny, dedykowany zbiornik (jak dotychczas).
    - To samo niepuste ID w kilku wierszach -> te produkty WSPÓŁDZIELĄ jeden fizyczny
      mieszalnik (produkcja kampanijna). Pojemność takiego zbiornika = max spośród zadanych
      (musi pomieścić największą recepturę), a wykorzystanie czasowe liczone jest jako SUMA
      (szarże_i x cykl_i) po wszystkich produktach tego zbiornika - patrz tank_members,
      zużywane w Zakładce 2 do właściwego (nie uproszczonego) przeliczenia utylizacji.

    active_portfolio odświeża się przy każdym uruchomieniu (bezpieczne - nic w Zakładce 2 tego
    ręcznie nie edytuje). prod_dict dla danej grupy ustawiany jest z receptury TYLKO przy jej
    pierwszym pojawieniu się - późniejsze ręczne poprawki w Zakładce 2 nie są nadpisywane przy
    kolejnych edycjach receptury. Wybór w panelu bocznym jest resetowany do dokładnie grup z
    receptury TYLKO gdy zmieni się sam ZESTAW grup w recepturze (nowa/usunięta grupa) - między
    takimi zmianami można swobodnie dopisać ręcznie inne linie bez ich utraty.
    """
    df = st.session_state.get("recipes_df")
    if df is None or df.empty:
        return

    groups_in_recipe = sorted(df[RECIPE_GROUP_COL].dropna().unique().tolist())

    for group_name in groups_in_recipe:
        group_rows_all = df[df[RECIPE_GROUP_COL] == group_name]

        # Produkty "Import - Nigdy" (stały import) nigdy nie dostają własnego mieszalnika -
        # nawet w widoku docelowym/100% pozostają importowane, więc wykluczamy je z sizingu
        # floty. Produkty "Import" z ustalonym rokiem przejścia ZOSTAJĄ w sizingu, bo flota
        # jest wymiarowana pod docelową (w pełni dojrzałą, czyli już produkowaną) zdolność.
        if RECIPE_SOURCING_COL in group_rows_all.columns and RECIPE_IMPORT_TRANSITION_COL in group_rows_all.columns:
            permanent_import_mask = (group_rows_all[RECIPE_SOURCING_COL] == "Import") & \
                                     (group_rows_all[RECIPE_IMPORT_TRANSITION_COL].isin(["Nigdy (stały import)", "Nigdy (bufor)"]))
            group_rows = group_rows_all[~permanent_import_mask]
        else:
            group_rows = group_rows_all

        if group_rows.empty:
            continue  # cała grupa to stały import - brak mieszalnika dla tej grupy

        defaults = GROUP_PHYSICAL_DEFAULTS.get(group_name, GROUP_PHYSICAL_DEFAULTS["Hydraulic Oils"])

        weights = group_rows[RECIPE_ANNUAL_COL].clip(lower=0.001)
        avg_density = float((group_rows[RECIPE_DENSITY_COL] * weights).sum() / weights.sum()) if weights.sum() > 0 else 0.88
        total_annual_kg = float(group_rows[RECIPE_ANNUAL_COL].sum()) * 1000.0

        # Podział wierszy grupy na sloty zbiornikowe wg ID Zbiornika (puste = własny slot).
        tank_id_col_vals = group_rows[RECIPE_TANK_ID_COL].fillna("").astype(str).str.strip() if RECIPE_TANK_ID_COL in group_rows.columns else pd.Series([""] * len(group_rows), index=group_rows.index)
        slots, seen_ids = [], {}
        for idx, tid in tank_id_col_vals.items():
            if tid:
                if tid not in seen_ids:
                    seen_ids[tid] = len(slots)
                    slots.append([idx])
                else:
                    slots[seen_ids[tid]].append(idx)
            else:
                slots.append([idx])

        tank_volumes, tank_cycles, tank_products, tank_members = [], [], [], []
        for slot_idxs in slots:
            slot_rows = group_rows.loc[slot_idxs]
            members, slot_vols = [], []
            for _, r in slot_rows.iterrows():
                v = r.get(RECIPE_MIXER_VOL_COL, 0) or 0
                v = float(v) if v > 0 else 15.0
                slot_vols.append(v)
                c = r.get(RECIPE_CYCLE_COL, 0) or 0
                c = float(c) if c > 0 else defaults["cycle_h"]
                dens = float(r.get(RECIPE_DENSITY_COL, 0) or 0.88)
                annual_kg_i = float(r.get(RECIPE_ANNUAL_COL, 0) or 0) * 1000.0
                members.append({"product": str(r[RECIPE_PRODUCT_COL]), "annual_kg": annual_kg_i, "density": dens, "cycle_h": c})

            slot_vol = max(slot_vols) if slot_vols else 15.0
            total_w = sum(m["annual_kg"] for m in members) or 1.0
            weighted_cycle = sum(m["cycle_h"] * m["annual_kg"] for m in members) / total_w

            tank_volumes.append(slot_vol)
            tank_cycles.append(weighted_cycle)
            tank_products.append(members[0]["product"] if len(members) == 1 else None)
            tank_members.append(members)

        avg_vol = sum(tank_volumes) / len(tank_volumes) if tank_volumes else 15.0
        avg_cycle = sum(tank_cycles) / len(tank_cycles) if tank_cycles else defaults["cycle_h"]
        n_tanks = len(slots)
        n_skus = len(group_rows)

        # active_portfolio - bezpieczne do odświeżania co przebieg (nic tego ręcznie nie edytuje).
        st.session_state.active_portfolio[group_name] = {
            "material": defaults["material"], "density": avg_density, "cycle_h": avg_cycle,
            "cp": defaults["cp"], "oil_group": defaults["oil_group"], "water_content": defaults["water_content"],
        }

        if group_name not in st.session_state.recipe_groups_seen:
            st.session_state.prod_dict[group_name] = {
                "roczna": total_annual_kg, "user_vol_m3": avg_vol, "skus": n_skus, "num_tanks": n_tanks,
                "tank_volumes": tank_volumes, "tank_cycles": tank_cycles, "cycle_h_base": avg_cycle,
                "tank_products": tank_products, "tank_members": tank_members,
            }
            st.session_state.recipe_groups_seen.add(group_name)
        elif "tank_members" not in st.session_state.prod_dict[group_name]:
            # Zgodność wsteczna: grupa zsynchronizowana starszą wersją tej funkcji.
            st.session_state.prod_dict[group_name]["tank_products"] = tank_products
            st.session_state.prod_dict[group_name]["tank_members"] = tank_members
        else:
            # Grupa była już wcześniej zsynchronizowana - sprawdź, czy ZESTAW produktów tej
            # grupy w recepturze (nazwa, sposób pozyskania, rok przejścia, ID zbiornika) zmienił
            # się od ostatniej synchronizacji. Bez tego sprawdzenia zmiana np. "Sposób
            # Pozyskania" produktu z "Produkcja własna" na "Import" po ponownym wgraniu
            # receptury NIGDY by się nie przebiła do floty (Zakładka 2), bo ta gałąź kodu
            # istnieje właśnie po to, żeby NIE nadpisywać ręcznych edycji użytkownika w
            # Zakładce 2 przy każdym przebiegu - ale musi się jednak odświeżyć, gdy dane
            # źródłowe faktycznie się zmieniły, a nie tylko przy pierwszym pojawieniu się grupy.
            current_signature = tuple(sorted(
                (str(r[RECIPE_PRODUCT_COL]), str(r.get(RECIPE_SOURCING_COL, "") or ""),
                 str(r.get(RECIPE_IMPORT_TRANSITION_COL, "") or ""), str(r.get(RECIPE_TANK_ID_COL, "") or ""))
                for _, r in group_rows.iterrows()
            ))
            st.session_state.setdefault("recipe_group_product_signature", {})
            stored_signature = st.session_state.recipe_group_product_signature.get(group_name)
            if stored_signature != current_signature:
                st.session_state.prod_dict[group_name] = {
                    "roczna": total_annual_kg, "user_vol_m3": avg_vol, "skus": n_skus, "num_tanks": n_tanks,
                    "tank_volumes": tank_volumes, "tank_cycles": tank_cycles, "cycle_h_base": avg_cycle,
                    "tank_products": tank_products, "tank_members": tank_members,
                }
            st.session_state.recipe_group_product_signature[group_name] = current_signature


    group_signature = tuple(groups_in_recipe)
    if st.session_state._last_recipe_group_signature != group_signature:
        st.session_state["wybrane_kategorie_ms"] = groups_in_recipe
        st.session_state._last_recipe_group_signature = group_signature


sync_recipes_into_fleet_defaults()

if st.session_state.recipes_df is None:
    st.info("👋 **Zacznij od Zakładki 1 (Receptury Produktów)** — pobierz szablon, wgraj recepturę produktów "
            "(z wolumenem produkcji, opcjonalnie wielkością mieszalnika i rozbiciem na opakowania) i wróć tu, "
            "żeby dalej skonfigurować flotę i instalację. Możesz też pominąć ten krok i pracować w pełni ręcznie "
            "w kolejnych zakładkach.")
else:
    st.success("✅ Receptura wgrana — grupy produktowe pojawiły się automatycznie w Zakładce 2 (panel boczny, "
               "Krok 1: Wybór Rodzin), każda z liczbą zbiorników = liczbie produktów tej grupy w recepturze. "
               "Możesz je tam dalej edytować.")
    st.caption("⚠️ **Jeśli to ponowne wgranie już wcześniej używanej receptury** (np. zmieniłeś '{}' albo dane "
               "importu dla produktu): zmiana na poziomie produktu (kto importuje, kto produkuje, kiedy) trafi do "
               "Zakładki 2, ale **flota widoczna w Zakładce 3/6 aktualizuje się dopiero po ponownym kliknięciu "
               "'📥 Zatwierdź i wyślij konfigurację do kolejnych kroków' w Zakładce 2** — to jest jawny krok "
               "zatwierdzenia, nie dzieje się automatycznie.".format(RECIPE_SOURCING_COL))

# ==========================================
# PANEL BOCZNY (Wybór Rodzin i Opakowań)
# ==========================================
st.sidebar.header("📋 KROK 1: Wybór Rodzin")
_default_lines = (sorted(st.session_state.recipe_groups_seen) if st.session_state.recipe_groups_seen
                   else ["Hydraulic Oils", "Engine Oils", "Watermiscibles"])
wybrane_kategorie = st.sidebar.multiselect(
    "Wybierz aktywne grupy produktowe:",
    list(st.session_state.active_portfolio.keys()),
    default=_default_lines,
    key="wybrane_kategorie_ms"
)
st.session_state["wybrane_kategorie_snapshot"] = wybrane_kategorie

st.sidebar.markdown("---")

st.sidebar.header("⏱️ KROK 2: Założenia Czasu Pracy")
liczba_zmian = st.sidebar.slider("Liczba zmian produkcyjnych:", min_value=1.0, max_value=3.0, value=1.0, step=0.5)
godziny_na_zmiane = st.sidebar.slider("Liczba godzin na jedną zmianę:", min_value=4.0, max_value=12.0, value=8.0, step=0.5)

godziny_dziennie = liczba_zmian * godziny_na_zmiane
AVAILABLE_HOURS_MONTH = (WORKING_DAYS_YEAR * godziny_dziennie) / MONTHS_PER_YEAR

st.sidebar.markdown("---")

st.sidebar.header("📈 KROK 3: Symulacja Rozruchu (5 lat)")
st.sidebar.caption("Flota i magazyn budowane są od razu pod docelową (100%) produkcję — to poniżej steruje "
                    "tylko tym, jak WYKORZYSTANIE tej floty i magazynu rośnie w pierwszych 5 latach. Krzywa jest "
                    "teraz interpolowana LINIOWO w czasie (płynny wzrost w trakcie roku, nie skok 1 stycznia) — "
                    "wpływa na wszystkie zakładki na żywo.")
st.session_state.rampup_start_pct = st.sidebar.slider(
    "Punkt startowy przed Rokiem 1 [%]", min_value=0.0, max_value=100.0,
    value=float(st.session_state.rampup_start_pct), step=5.0, key="rampup_start_pct_slider",
    help="Domyślnie 0% — czysty rozruch od zera 1 stycznia Roku 1. Podnieś, jeśli symulujesz start z "
         "istniejącej bazy (np. przejęty zakład, istniejący portfel klientów)."
)
st.session_state.rampup_differentiate = st.sidebar.checkbox(
    "🔧 Zróżnicuj tempo rozruchu per linia produktowa", value=st.session_state.rampup_differentiate,
    help="Domyślnie jedna wspólna krzywa dla całej fabryki. Włącz, jeśli np. Engine Oils ma ruszyć "
         "szybciej niż Greases."
)
st.session_state.import_follows_rampup = st.sidebar.checkbox(
    "📦 Import rośnie razem z krzywą rozruchu", value=st.session_state.import_follows_rampup,
    help="Włączone (domyślnie): wolumen produktu jeszcze importowanego skaluje się tą samą krzywą co "
         "produkcja własna (założenie: popyt klienta rośnie niezależnie od tego, kto go zaspokaja w danym "
         "roku). Wyłącz, jeśli Twój import ma WŁASNY harmonogram (np. stały kontrakt niezależny od tempa "
         "rozruchu fabryki) — wtedy import liczony jest zawsze jako pełny docelowy wolumen, bez skalowania."
)
year_labels = [f"Rok {i+1}" for i in range(RAMPUP_YEARS)]
if not st.session_state.rampup_differentiate:
    for i in range(RAMPUP_YEARS):
        st.session_state.rampup_global_pct[i] = st.sidebar.slider(
            f"{year_labels[i]} [%]", min_value=0.0, max_value=100.0,
            value=float(st.session_state.rampup_global_pct[i]), step=5.0, key=f"rampup_global_{i}"
        )
else:
    for kat in wybrane_kategorie:
        st.session_state.rampup_per_line_pct.setdefault(kat, list(st.session_state.rampup_global_pct))
        with st.sidebar.expander(f"{kat}", expanded=False):
            for i in range(RAMPUP_YEARS):
                st.session_state.rampup_per_line_pct[kat][i] = st.slider(
                    f"{year_labels[i]} [%]", min_value=0.0, max_value=100.0,
                    value=float(st.session_state.rampup_per_line_pct[kat][i]), step=5.0,
                    key=f"rampup_line_{kat}_{i}"
                )

st.sidebar.markdown("---")

# --- STRUKTURA INTERFEJSU ---
tab1, tab5, tab2, tab3, tab4, tab6, tab8 = st.tabs([
    "📋 1. Receptury Produktów i Flota",
    "🛢️ 2. Magazynowanie (Surowce i Park Zbiorników)",
    "📐 3. Karta Maszyn, Kocioł i Zasilanie",
    "📦 4. Logistyka i Czas Rozlewu",
    "💰 5. Analiza Finansowa, CAPEX i ROI",
    "🧵 6. Mapa Strumienia Wartości (VSM)",
    "🏠 7. Dashboard"
])

# ==========================================
# ZAKŁADKA 1: FLOTA MIESZALNIKÓW (tab1)
# ==========================================
# ==========================================
# ==========================================
# RECEPTURY PRODUKTÓW (import z Excela) - scalone z Flotą w Zakładce 1 (tab1)
# ==========================================
with tab1:
    st.header("📋 Receptury Produktów: Import z Excela")

    st.markdown("### 💾 Zapisz / Wczytaj Projekt")
    st.caption("Cała konfiguracja (receptura, flota, zbiorniki, ceny, rozruch, testy QC) w jednym pliku — pobierz go, "
               "żeby nie uzupełniać wszystkiego od nowa przy następnej sesji, albo żeby zarządzać kilkoma zakładami "
               "naraz (każdy jako osobny plik projektu).")
    col_save1, col_save2 = st.columns(2)
    with col_save1:
        if st.session_state.get("recipes_df") is not None:
            st.download_button(
                "⬇️ Pobierz projekt (.json)", data=export_project_bytes(),
                file_name=f"projekt_{datetime.date.today().isoformat()}.json", mime="application/json",
                key="download_project_btn"
            )
        else:
            st.caption("Wgraj/skonfiguruj recepturę poniżej, żeby móc zapisać projekt.")
    with col_save2:
        uploaded_project_file = st.file_uploader("⬆️ Wczytaj projekt (.json)", type=["json"], key="upload_project_file")
        if uploaded_project_file is not None:
            success, msg = import_project_bytes(uploaded_project_file.getvalue())
            if success:
                st.success(f"✅ {msg}")
                st.rerun()
            else:
                st.error(f"❌ {msg}")
    st.caption("⚠️ Obejmuje receptury, flotę, zbiorniki, ceny, ustawienia rozruchu i testów QC — **nie** obejmuje "
               "wygenerowanych raportów PDF/Excel (te przeliczą się same z wczytanej konfiguracji).")
    st.markdown("---")

    st.caption("Wgraj plik Excel z listą produktów (przypisanych do grupy produktowej), rocznym zapotrzebowaniem "
               "i dozowaniem surowców [kg/t] (bazy olejowe, dodatki, pakiety, zagęszczacze, smary stałe, woda DEMI, "
               "biocyd). Dane z tej zakładki zasilają dodatkowo wymiarowanie silosów **per surowiec** w Zakładce 4 "
               "oraz podpowiedź, dla których surowców opłaca się dedykowany zbiornik, a które lepiej zostawić "
               "w beczkach/IBC/workach.")

    st.markdown("### 📥 Krok 1: Pobierz szablon")
    st.caption("Szablon zawiera dokładną strukturę kolumn wymaganą przez aplikację (nie zmieniaj nazw/kolejności "
               "kolumn), listę rozwijaną grup produktowych, dwa przykładowe wiersze pokazujące format oraz formuły "
               "kontrolne — 'Suma Udziałów Składników' podświetla się na czerwono, jeśli odbiega od 1000 kg/t o "
               "więcej niż tolerancja, a 'Roczne Zapotrzebowanie Surowcowe' wylicza się z uwzględnieniem strat "
               "procesowych. Dodatkowy arkusz 'Opakowania' pozwala predefiniować typy opakowań i ich pojemności — "
               "opcjonalny, wypełniony domyślnymi wartościami z aplikacji, można dopisać nowe wiersze lub zostawić bez zmian.")

    template_bytes = generate_recipe_template_bytes()
    st.download_button(
        label="⬇️ Pobierz szablon Excel (Receptury_Szablon.xlsx)",
        data=template_bytes,
        file_name="Receptury_Szablon.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="btn_download_recipe_template"
    )

    st.markdown("---")
    st.markdown("### 📤 Krok 2: Wgraj uzupełniony plik")
    uploaded_recipe_file = st.file_uploader(
        "Wybierz plik .xlsx z recepturami:", type=["xlsx"], key="recipe_uploader"
    )

    if uploaded_recipe_file is not None:
        uploaded_recipe_file.seek(0)
        parsed_df, parse_errors = parse_recipe_excel(uploaded_recipe_file)

        if parse_errors:
            for err in parse_errors:
                st.warning(f"⚠️ {err}")

        if parsed_df is not None and not parsed_df.empty:
            st.session_state.recipes_df = parsed_df
            st.success(f"✅ Wczytano {len(parsed_df)} poprawnych receptur produktowych.")
        elif parsed_df is None:
            st.error("❌ Nie udało się wczytać żadnych poprawnych receptur z tego pliku — popraw błędy powyżej i wgraj ponownie.")

        # Opcjonalny arkusz 'Opakowania' w tym samym pliku - jeśli obecny, nadpisuje/uzupełnia
        # domyślne typy opakowań i ich pojemności (Zakładki 1 i 3), z zachowaniem edycji w apce.
        uploaded_recipe_file.seek(0)
        packaging_result, packaging_errors = parse_packaging_excel(uploaded_recipe_file)
        for err in packaging_errors:
            st.warning(f"⚠️ {err}")
        if packaging_result is not None:
            st.session_state.pack_configs.update(packaging_result["pack_configs"])
            if "filling_lines_config" not in st.session_state:
                st.session_state.filling_lines_config = {}
            st.session_state.filling_lines_config.update(packaging_result["filling_defaults"])
            st.success(f"✅ Wczytano/zaktualizowano {len(packaging_result['pack_configs'])} typów opakowań z arkusza '{PACKAGING_SHEET_NAME}'.")

        # Opcjonalny arkusz 'Badania Laboratoryjne' - testy jako wiersze, produkty jako kolumny
        # (patrz parse_qc_tests_excel). Ma PIERWSZEŃSTWO nad panelem zwolnienia per linia
        # (Zakładka 6/VSM) wszędzie, gdzie liczba badań QC jest liczona per konkretny produkt.
        # Testy spoza wbudowanego katalogu, ale z wypełnionym własnym 'Sprzęt'/'Czas [min]',
        # rejestrują się jako NOWE testy (custom_qc_tests) - nie musisz ograniczać się do
        # wbudowanej listy ~32 testów.
        uploaded_recipe_file.seek(0)
        qc_tests_result, new_qc_tests, qc_tests_errors = parse_qc_tests_excel(uploaded_recipe_file)
        for err in qc_tests_errors:
            st.warning(f"⚠️ {err}")
        if new_qc_tests:
            st.session_state.custom_qc_tests.update(new_qc_tests)
            st.success(f"✅ Zarejestrowano {len(new_qc_tests)} nowych, własnych testów QC (spoza wbudowanego katalogu): "
                       f"{', '.join(new_qc_tests.keys())}.")
        if qc_tests_result is not None:
            st.session_state.qc_tests_by_product = qc_tests_result
            st.success(f"✅ Wczytano przypisanie badań laboratoryjnych dla {len(qc_tests_result)} produktów z arkusza '{QC_SHEET_NAME}'.")

        # Opcjonalny arkusz 'Zużycie Surowców (bez receptury)' - dla surowców, których zużycie
        # znasz wprost, bez potrzeby ujawniania pełnej receptury (ochrona know-how). Dolicza się
        # do zużycia wynikającego z pełnych receptur, nie zastępuje go.
        uploaded_recipe_file.seek(0)
        direct_rm_entries, direct_rm_errors = parse_direct_raw_materials_excel(uploaded_recipe_file)
        for err in direct_rm_errors:
            st.warning(f"⚠️ {err}")
        st.session_state.direct_raw_materials = direct_rm_entries
        if direct_rm_entries:
            st.success(f"✅ Wczytano {len(direct_rm_entries)} bezpośrednich deklaracji zużycia surowców (bez receptury) "
                       f"z arkusza '{DIRECT_RM_SHEET_NAME}'.")

        # Opcjonalny arkusz 'Rozbicie Dostawców Surowców' - dzieli zużycie kategorii surowca
        # (np. Base Oil Group II) na kilku konkretnych dostawców/baz, żeby pokazać kilka
        # mniejszych zbiorników zamiast jednego dużego, wspólnego.
        uploaded_recipe_file.seek(0)
        supplier_split_entries, supplier_split_errors = parse_supplier_split_excel(uploaded_recipe_file)
        for err in supplier_split_errors:
            st.warning(f"⚠️ {err}")
        st.session_state.supplier_splits = supplier_split_entries
        if supplier_split_entries:
            st.success(f"✅ Wczytano rozbicie dostawców dla {len(set(e['category'] for e in supplier_split_entries))} "
                       f"kategorii surowców z arkusza '{SUPPLIER_SPLIT_SHEET_NAME}'.")

    st.markdown("---")

    if st.session_state.recipes_df is not None and not st.session_state.recipes_df.empty:
        st.markdown("### 📊 Krok 3: Wczytane receptury (edytowalne)")
        st.caption("Możesz jeszcze poprawić wartości tutaj przed przeliczeniem — zmiany nie są zapisywane z powrotem "
                   "do pliku Excel, tylko do tej sesji aplikacji. 'Suma Udziałów Składników' i 'Roczne Zapotrzebowanie "
                   "Surowcowe' przeliczają się na żywo pod tabelą.")

        edited_recipes_df = st.data_editor(
            st.session_state.recipes_df,
            hide_index=True,
            use_container_width=True,
            num_rows="dynamic",
            key="recipes_data_editor",
            column_config={
                RECIPE_GROUP_COL: st.column_config.SelectboxColumn(options=RECIPE_PRODUCT_GROUPS),
                RECIPE_SOURCING_COL: st.column_config.SelectboxColumn(options=RECIPE_SOURCING_OPTIONS),
                RECIPE_IMPORT_TRANSITION_COL: st.column_config.SelectboxColumn(options=[""] + RECIPE_IMPORT_TRANSITION_OPTIONS),
            }
        )
        st.session_state.recipes_df = edited_recipes_df

        # Przeliczenie kolumn wyliczanych na żywo po edycji w tabeli (niezależnie od tego,
        # co ewentualnie zostało wpisane ręcznie w tych kolumnach w Excelu).
        edited_recipes_df[RECIPE_SUM_COL] = edited_recipes_df[RECIPE_RAW_MATERIALS].sum(axis=1)
        loss_safe = edited_recipes_df[RECIPE_LOSS_COL].clip(lower=0, upper=99.9)
        edited_recipes_df[RECIPE_RAW_DEMAND_COL] = edited_recipes_df[RECIPE_ANNUAL_COL] / (1.0 - loss_safe / 100.0)

        edited_recipes_df[RECIPE_BATCH_MASS_COL] = edited_recipes_df[RECIPE_MIXER_VOL_COL] * 1000.0 * edited_recipes_df[RECIPE_DENSITY_COL]
        edited_recipes_df[RECIPE_BATCHES_YEAR_COL] = 0.0
        has_bm = edited_recipes_df[RECIPE_BATCH_MASS_COL] > 0
        edited_recipes_df.loc[has_bm, RECIPE_BATCHES_YEAR_COL] = (
            edited_recipes_df.loc[has_bm, RECIPE_ANNUAL_COL] * 1000.0 / edited_recipes_df.loc[has_bm, RECIPE_BATCH_MASS_COL])
        edited_recipes_df[RECIPE_UTILIZATION_COL] = 0.0
        has_ah = edited_recipes_df[RECIPE_AVAIL_HOURS_COL] > 0
        edited_recipes_df.loc[has_ah, RECIPE_UTILIZATION_COL] = (
            edited_recipes_df.loc[has_ah, RECIPE_BATCHES_YEAR_COL] * edited_recipes_df.loc[has_ah, RECIPE_CYCLE_COL]
            / edited_recipes_df.loc[has_ah, RECIPE_AVAIL_HOURS_COL] * 100.0)

        bad_sum_live_mask = (edited_recipes_df[RECIPE_SUM_COL] - RECIPE_TARGET_SUM_KG).abs() > RECIPE_SUM_TOLERANCE_KG
        if RECIPE_SOURCING_COL in edited_recipes_df.columns and RECIPE_IMPORT_TRANSITION_COL in edited_recipes_df.columns:
            permanent_import_exempt_live = (edited_recipes_df[RECIPE_SOURCING_COL] == "Import") & \
                                            (edited_recipes_df[RECIPE_IMPORT_TRANSITION_COL].isin(["Nigdy (stały import)", "Nigdy (bufor)"]))
            bad_sum_live_mask = bad_sum_live_mask & ~permanent_import_exempt_live
        bad_sum_live = edited_recipes_df[bad_sum_live_mask]
        if not bad_sum_live.empty:
            zle_produkty = ", ".join(f"{p} ({s:.0f} kg/t)" for p, s in zip(bad_sum_live[RECIPE_PRODUCT_COL], bad_sum_live[RECIPE_SUM_COL]))
            st.error(f"❌ Suma dozowania surowców odbiega od 1000 kg/t (± {RECIPE_SUM_TOLERANCE_KG:.0f} kg) dla: "
                     f"{zle_produkty}. Popraw przed dalszą analizą.")

        overloaded_live = edited_recipes_df[edited_recipes_df[RECIPE_UTILIZATION_COL] > RECIPE_UTILIZATION_WARN_PCT]
        if not overloaded_live.empty:
            zle_mieszalniki = ", ".join(f"{p} ({u:.0f}%)" for p, u in zip(overloaded_live[RECIPE_PRODUCT_COL], overloaded_live[RECIPE_UTILIZATION_COL]))
            st.warning(f"⚠️ Wykorzystanie mieszalnika powyżej {RECIPE_UTILIZATION_WARN_PCT:.0f}% dla: {zle_mieszalniki} — "
                       f"rozważ większy zbiornik, krótszy cykl, więcej zmian roboczych, lub kolejny mieszalnik.")

        st.markdown("---")
        st.info("💡 Zagregowane zużycie surowców per materiał i wymiarowanie silosów znajdziesz w **Zakładce 2 "
                "(Magazynowanie)**. Roczne zapotrzebowanie per grupa produktowa zobaczysz w tabeli floty poniżej. "
                "Typy opakowań i ich rozbicie procentowe skonfigurujesz w **Zakładce 4 (Logistyka i Czas Rozlewu)**.")

    else:
        st.info("💡 Wgraj plik z recepturami powyżej, aby zobaczyć tu zagregowane zużycie surowców.")


st.markdown("---")
st.markdown("## 📊 Flota i Utylizacja")
st.caption("Ciąg dalszy - dobór mieszalników na bazie receptur wgranych powyżej (albo w pełni ręcznie).")

with tab1:
    st.header(f"Zintegrowane Zestawienie Parametrów Procesowych")

    if "mixer_fill_factor" not in st.session_state:
        # Domyślnie 92,5% - realne napełnienie robocze reaktora mieszającego (nie 100% pojemności
        # nominalnej - potrzeba miejsca na mieszanie, pienienie, zapas bezpieczeństwa). Typowo 90-95%.
        st.session_state.mixer_fill_factor = 0.925

    if wybrane_kategorie:
        st.session_state.mixer_fill_factor = st.slider(
            "🧪 Współczynnik napełnienia mieszalnika (robocze % pojemności nominalnej):",
            min_value=0.70, max_value=1.00, value=float(st.session_state.mixer_fill_factor), step=0.025,
            help="Reaktor nigdy nie jest napełniany do samej góry — potrzeba miejsca na mieszanie, "
                 "pienienie i zapas bezpieczeństwa. Typowo 90-95% pojemności nominalnej. Wpływa na masę "
                 "szarży dla WSZYSTKICH mieszalników, wszędzie w aplikacji."
        )
        st.markdown("##### 📥 Krok A: Parametryzacja Tonażu, Pojemności Mieszalnika oraz SKUs")
        st.caption("Wybierz linię z listy, aby błyskawicznie i płynnie zmienić jej parametry. Wyniki w tabeli poniżej przeliczą się natychmiast.")

        selected_family_to_edit = st.selectbox("Wybierz linię produktową do modyfikacji:", wybrane_kategorie)

        # "Roczna produkcja" synchronizuje się z receptury TYLKO przy pierwszym pojawieniu się danej
        # linii (żeby nie nadpisywać Twoich ręcznych poprawek przy każdej edycji pliku) - ale to
        # oznacza, że jeśli WGRASZ NOWY plik ze zmienioną roczną produkcją dla linii, która już
        # istniała, wartość tutaj CICHO zostanie stara. Wykrywamy to i dajemy jawny wybór, zamiast
        # zostawiać mylącą, nieaktualną liczbę.
        recipes_df_check = st.session_state.get("recipes_df")
        if recipes_df_check is not None and not recipes_df_check.empty and RECIPE_GROUP_COL in recipes_df_check.columns:
            group_rows_check = recipes_df_check[recipes_df_check[RECIPE_GROUP_COL] == selected_family_to_edit]
            if not group_rows_check.empty:
                recipe_roczna_kg = float(group_rows_check[RECIPE_ANNUAL_COL].sum()) * 1000.0
                stored_roczna_kg = float(st.session_state.prod_dict[selected_family_to_edit]["roczna"])
                if abs(recipe_roczna_kg - stored_roczna_kg) > max(1.0, stored_roczna_kg * 0.01):
                    st.warning(f"⚠️ **Receptura mówi {recipe_roczna_kg:,.0f} kg/rok** dla '{selected_family_to_edit}', ale "
                               f"tu wciąż zapisane jest **{stored_roczna_kg:,.0f} kg** — to STARA wartość (być może z "
                               "poprzedniego pliku, sprzed ręcznej poprawki, albo z domyślnych ustawień). Nie nadpisujemy "
                               "tego automatycznie, żeby nie stracić Twoich ręcznych zmian.")
                    if st.button(f"🔄 Użyj wartości z receptury ({recipe_roczna_kg:,.0f} kg) dla '{selected_family_to_edit}'", key=f"sync_roczna_{selected_family_to_edit}"):
                        st.session_state.prod_dict[selected_family_to_edit]["roczna"] = recipe_roczna_kg
                        st.rerun()

        c_ed1, c_ed2, c_ed3, c_ed4 = st.columns(4)
        with c_ed1:
            st.session_state.prod_dict[selected_family_to_edit]["roczna"] = st.number_input(
                "Roczna produkcja [kg]:", min_value=0, value=int(st.session_state.prod_dict[selected_family_to_edit]["roczna"]), step=50000,
                key=f"roczna_{selected_family_to_edit}"
            )
        with c_ed2:
            st.session_state.prod_dict[selected_family_to_edit]["user_vol_m3"] = st.number_input(
                "Pojemność Mieszalnika (bazowa) [m³]:", min_value=0.5, value=float(st.session_state.prod_dict[selected_family_to_edit]["user_vol_m3"]), step=0.5,
                help="Używana bezpośrednio, gdy linia ma jeden zbiornik. Przy kilku zbiornikach to wartość domyślna dla nowo dodanych — każdy można potem zróżnicować poniżej.",
                key=f"pojemnosc_baza_{selected_family_to_edit}"
            )
        with c_ed3:
            st.session_state.prod_dict[selected_family_to_edit].setdefault("cycle_h_base", st.session_state.active_portfolio[selected_family_to_edit]["cycle_h"])
            st.session_state.prod_dict[selected_family_to_edit]["cycle_h_base"] = st.number_input(
                "Cykl Procesowy (bazowy, szacunkowy) [h]:", min_value=0.5, value=float(st.session_state.prod_dict[selected_family_to_edit]["cycle_h_base"]), step=0.5,
                help="Szacunkowy czas cyklu jednej szarży (dozowanie + grzanie + homogenizacja + chłodzenie + rozlew), do wstępnego wymiarowania floty — "
                     "różne receptury/wielkości szarży realnie różnią się czasem cyklu. Po skonfigurowaniu inżynierii w Zakładce 2/7 zobaczysz obok "
                     "rzeczywisty, policzony czas cyklu do porównania.",
                key=f"cykl_baza_{selected_family_to_edit}"
            )
        with c_ed4:
            st.session_state.prod_dict[selected_family_to_edit]["skus"] = st.number_input(
                "Liczba aktywnych SKUs:", min_value=1, value=int(st.session_state.prod_dict[selected_family_to_edit]["skus"]), step=1,
                key=f"skus_{selected_family_to_edit}"
            )
        if st.session_state.prod_dict[selected_family_to_edit].get("num_tanks", 1) > 1:
            st.caption(f"ℹ️ **'Pojemność Mieszalnika (bazowa)' NIE wpływa na już skonfigurowane zbiorniki tej linii** — "
                       f"{selected_family_to_edit} ma {st.session_state.prod_dict[selected_family_to_edit]['num_tanks']} "
                       "zbiorniki, każdy z własną pojemnością w sekcji 'Pojemności i Cykle Poszczególnych Zbiorników' "
                       "poniżej. To pole zadziała dopiero, gdy zwiększysz liczbę zbiorników (jako wartość startowa dla "
                       "nowego).")

        current_skus = st.session_state.prod_dict[selected_family_to_edit]["skus"]
        _tank_members_current = st.session_state.prod_dict[selected_family_to_edit].get("tank_members", [])
        _has_recipe_products = any(len(m) > 0 for m in _tank_members_current)

        if current_skus > 1:
            st.markdown("---")
            st.session_state.prod_dict[selected_family_to_edit]["num_tanks"] = st.number_input(
                f"🏭 **Wielkość floty dla {selected_family_to_edit}**: Na ile osobnych mieszalników chcesz rozbić produkcję tych {current_skus} SKUs?",
                min_value=1, max_value=int(current_skus), value=min(int(st.session_state.prod_dict[selected_family_to_edit].get("num_tanks", 1)), int(current_skus)),
                key=f"num_tanks_{selected_family_to_edit}"
            )
        else:
            st.session_state.prod_dict[selected_family_to_edit]["num_tanks"] = 1

        tanks_count_now = st.session_state.prod_dict[selected_family_to_edit]["num_tanks"]

        # --- Przypisanie produktów z receptury do konkretnych zbiorników (bez wracania do Excela) ---
        if _has_recipe_products:
            st.markdown("---")
            st.markdown("###### 🔀 Przypisanie Produktów do Zbiorników")
            st.caption("Zmień, do którego zbiornika trafia dany produkt z receptury — np. gdy zbiornik się przeciąża "
                       "i chcesz rozdzielić współdzielone produkty na osobne mieszalniki. Działa od razu, bez "
                       "ponownego wgrywania pliku Excel. (Jeśli chcesz *trwale* zmienić przypisanie na przyszłość, "
                       "warto też poprawić 'ID Zbiornika' w źródłowym pliku — inaczej kolejne wgranie receptury "
                       "przywróci oryginalny podział).")

            all_members_flat = [(slot_idx, mem) for slot_idx, slot in enumerate(_tank_members_current) for mem in slot]
            new_assignment = {}
            for slot_idx, mem in all_members_flat:
                target_options = list(range(1, tanks_count_now + 1))
                default_idx = min(slot_idx, tanks_count_now - 1)
                chosen = st.selectbox(
                    f"{mem['product']} → Zbiornik nr:", target_options, index=default_idx,
                    key=f"tank_assign_{selected_family_to_edit}_{mem['product']}"
                )
                new_assignment.setdefault(chosen - 1, []).append(mem)

            rebuilt_members = [new_assignment.get(i, []) for i in range(tanks_count_now)]
            st.session_state.prod_dict[selected_family_to_edit]["tank_members"] = rebuilt_members

            # Pojemności bazowe nowo powstałych/opróżnionych zbiorników - jeśli slot ma produkty
            # z receptury, podpowiedz pojemność jako max ich zadanych wielkości mieszalnika,
            # żeby nie trzeba było ręcznie ustawiać po każdym przeniesieniu.
            tank_volumes_now = st.session_state.prod_dict[selected_family_to_edit].setdefault("tank_volumes", [15.0])
            while len(tank_volumes_now) < tanks_count_now:
                tank_volumes_now.append(15.0)
            del tank_volumes_now[tanks_count_now:]

        # --- Pojemności i cykle poszczególnych zbiorników, gdy linia jest rozbita na kilka mieszalników ---
        # Domyślnie każdy nowy zbiornik dziedziczy wartości bazowe powyżej, ale można je zróżnicować per
        # zbiornik — np. mniejsza szarża o krótszym cyklu dla niskowolumenowego SKU.
        tanks_count_selected = st.session_state.prod_dict[selected_family_to_edit]["num_tanks"]
        base_vol = st.session_state.prod_dict[selected_family_to_edit]["user_vol_m3"]
        base_cycle = st.session_state.prod_dict[selected_family_to_edit]["cycle_h_base"]
        tank_volumes_sel = st.session_state.prod_dict[selected_family_to_edit].setdefault("tank_volumes", [base_vol])
        tank_cycles_sel = st.session_state.prod_dict[selected_family_to_edit].setdefault("tank_cycles", [base_cycle])
        if len(tank_volumes_sel) < tanks_count_selected:
            tank_volumes_sel.extend([base_vol] * (tanks_count_selected - len(tank_volumes_sel)))
        elif len(tank_volumes_sel) > tanks_count_selected:
            del tank_volumes_sel[tanks_count_selected:]
        if len(tank_cycles_sel) < tanks_count_selected:
            tank_cycles_sel.extend([base_cycle] * (tanks_count_selected - len(tank_cycles_sel)))
        elif len(tank_cycles_sel) > tanks_count_selected:
            del tank_cycles_sel[tanks_count_selected:]

        if tanks_count_selected == 1:
            tank_volumes_sel[0] = base_vol  # przy jednym zbiorniku pola bazowe są jedynym źródłem prawdy
            tank_cycles_sel[0] = base_cycle
        else:
            st.markdown("###### 🧪 Pojemności i Cykle Poszczególnych Zbiorników")
            st.caption("Domyślnie każdy zbiornik dziedziczy wartości bazowe powyżej — edytuj poniżej, jeśli zbiorniki różnią się wielkością szarży "
                       "i/lub czasem cyklu (np. mała szarża niszowego SKU z krótszym cyklem vs. duża szarża wysokowolumenowego SKU z dłuższym). "
                       "Roczna produkcja rodziny jest wtedy dzielona między zbiorniki proporcjonalnie do ich pojemności, a nie po równo.")
            cols_tanks = st.columns(min(tanks_count_selected, 4))
            for i in range(tanks_count_selected):
                with cols_tanks[i % len(cols_tanks)]:
                    tank_volumes_sel[i] = st.number_input(
                        f"Zbiornik #{i + 1} — Pojemność [m³]:", min_value=0.5, value=float(tank_volumes_sel[i]), step=0.5,
                        key=f"tankvol_{selected_family_to_edit}_{i}"
                    )
                    tank_cycles_sel[i] = st.number_input(
                        f"Zbiornik #{i + 1} — Cykl [h]:", min_value=0.5, value=float(tank_cycles_sel[i]), step=0.5,
                        key=f"tankcyc_{selected_family_to_edit}_{i}"
                    )

        final_fleet_rows = []
        real_cycle_reference_rows = []
        tag_counter = 1
        st.session_state.tag_to_recipe_product = {}
        st.session_state.tag_to_shared_members = {}  # dla zbiorników kampanijnych (>1 produkt): lista {"product","annual_kg","cycle_h","density"}

        for kat in wybrane_kategorie:
            m_annual = st.session_state.prod_dict[kat]["roczna"]
            tanks_count = st.session_state.prod_dict[kat].get("num_tanks", 1)
            base_vol_kat = st.session_state.prod_dict[kat]["user_vol_m3"]
            base_cycle_kat = st.session_state.prod_dict[kat].setdefault("cycle_h_base", st.session_state.active_portfolio[kat]["cycle_h"])

            tank_volumes = st.session_state.prod_dict[kat].setdefault("tank_volumes", [base_vol_kat])
            tank_cycles = st.session_state.prod_dict[kat].setdefault("tank_cycles", [base_cycle_kat])
            if len(tank_volumes) < tanks_count:
                tank_volumes.extend([base_vol_kat] * (tanks_count - len(tank_volumes)))
            elif len(tank_volumes) > tanks_count:
                del tank_volumes[tanks_count:]
            if len(tank_cycles) < tanks_count:
                tank_cycles.extend([base_cycle_kat] * (tanks_count - len(tank_cycles)))
            elif len(tank_cycles) > tanks_count:
                del tank_cycles[tanks_count:]
            if tanks_count == 1:
                tank_volumes[0] = base_vol_kat
                tank_cycles[0] = base_cycle_kat

            rho_product = st.session_state.active_portfolio[kat]["density"]
            total_capacity = sum(tank_volumes)

            for t_idx, v_tank_user in enumerate(tank_volumes):
                cyc_h = tank_cycles[t_idx]
                tank_members_list = st.session_state.prod_dict[kat].get("tank_members", [])
                members = tank_members_list[t_idx] if t_idx < len(tank_members_list) else None

                if members is not None and len(members) > 1:
                    # Zbiornik współdzielony przez kilka produktów (produkcja kampanijna) -
                    # każdy produkt ma WŁASNĄ masę szarży (ta sama pojemność zbiornika, ale
                    # własna gęstość) i własny cykl; sumujemy szarże i faktyczny czas zajętości
                    # zbiornika, zamiast zgadywać jednym uśrednionym cyklem.
                    total_batches, total_hours, mass_per_batch_list = 0, 0.0, []
                    for mem in members:
                        mass_pb_i = v_tank_user * mem["density"] * 1000.0 * st.session_state.mixer_fill_factor
                        mass_per_batch_list.append(mass_pb_i)
                        monthly_i = mem["annual_kg"] / MONTHS_PER_YEAR
                        batches_i = math.ceil(monthly_i / mass_pb_i) if mass_pb_i > 0 else 0
                        total_batches += batches_i
                        total_hours += batches_i * mem["cycle_h"]
                    batches_per_tank = total_batches
                    mass_per_batch = max(mass_per_batch_list) if mass_per_batch_list else 0.0
                    real_utilization = (total_hours / AVAILABLE_HOURS_MONTH * 100.0) if AVAILABLE_HOURS_MONTH > 0 else 0.0
                elif members is not None and len(members) == 1:
                    # Zbiornik dedykowany DOKŁADNIE jednemu produktowi z receptury - liczymy
                    # wprost z jego własnych danych (gęstość/roczna/cykl), bez zgadywania przez
                    # podział proporcjonalny do pojemności (mamy dokładne dane, więc ich używamy).
                    mem = members[0]
                    mass_per_batch = v_tank_user * mem["density"] * 1000.0 * st.session_state.mixer_fill_factor
                    monthly_mass = mem["annual_kg"] / MONTHS_PER_YEAR
                    batches_per_tank = math.ceil(monthly_mass / mass_per_batch) if mass_per_batch > 0 else 0
                    real_utilization = (batches_per_tank * mem["cycle_h"]) / AVAILABLE_HOURS_MONTH * 100.0 if AVAILABLE_HOURS_MONTH > 0 else 0.0
                elif members is not None and len(members) == 0:
                    # Zbiornik jawnie opróżniony (przeniesiono z niego wszystkie produkty w
                    # panelu przypisania powyżej) - nic w nim nie produkujemy.
                    mass_per_batch, batches_per_tank, real_utilization = 0.0, 0, 0.0
                else:
                    # Zbiornik niepowiązany z recepturą (dodany ręcznie, tryb w pełni manualny) —
                    # podział rocznej produkcji rodziny między zbiorniki PROPORCJONALNIE do ich
                    # pojemności, jeśli zbiorniki mają różne wielkości, większy zbiornik przejmuje
                    # większą część wolumenu zamiast wymuszania tej samej liczby szarż co na małym.
                    capacity_share = (v_tank_user / total_capacity) if total_capacity > 0 else (1.0 / tanks_count)
                    annual_per_tank = m_annual * capacity_share
                    monthly_per_tank = annual_per_tank / MONTHS_PER_YEAR

                    mass_per_batch = v_tank_user * rho_product * 1000.0 * st.session_state.mixer_fill_factor
                    batches_per_tank = math.ceil(monthly_per_tank / mass_per_batch) if mass_per_batch > 0 else 0
                    real_utilization = (batches_per_tank * cyc_h) / AVAILABLE_HOURS_MONTH * 100.0 if AVAILABLE_HOURS_MONTH > 0 else 0.0

                tag_id = f"MT-{tag_counter}" + (f"-Z{t_idx+1}" if tanks_count > 1 else "")
                if members is not None and len(members) == 1:
                    st.session_state.tag_to_recipe_product[tag_id] = members[0]["product"]
                elif members is not None and len(members) > 1:
                    st.session_state.tag_to_shared_members[tag_id] = members
                status_txt = "🟢 Optymalna" if real_utilization <= MAX_TANK_UTILIZATION_PCT else "⚠️ Przeciążenie (>85%)"
                if v_tank_user < MIN_TANK_VOLUME_M3:
                    status_txt = "❌ Poniżej min. fabryki (<5 m³)"

                # Rzeczywisty, policzony czas cyklu (dozowanie+grzanie+homog.+pompowanie+chłodzenie) z
                # Zakładki 2/7 — pokazywany OSOBNO poniżej (nie w tej samej edytowalnej tabeli!), bo ta
                # wartość zmienia się za każdym razem, gdy cokolwiek zostanie skonfigurowane na INNEJ
                # zakładce. Trzymanie jej w tej samej tabeli co edytowalna flota powodowało, że
                # st.data_editor dostawał na każdym przebiegu inną zawartość i potrafił zresetować
                # ręczne zmiany użytkownika (np. usunięte wiersze) — stąd "niekontrolowane odświeżanie".
                real_cycle_txt = "—"
                ct = st.session_state.calculated_times.get(tag_id)
                bt = st.session_state.batch_time_components.get(tag_id)
                if ct is not None:
                    real_cycle_h = ct.get("heating", 0.0) + ct.get("pumping", 0.0) + ct.get("cooling_h", 0.0)
                    if bt is not None:
                        real_cycle_h += bt.get("dosing", 0.0) + bt.get("homog", 0.0)
                    real_cycle_txt = f"{real_cycle_h:.2f}"
                real_cycle_reference_rows.append({
                    "ID Urządzenia": tag_id, "Linia": kat,
                    "Cykl Szacowany [h]": round(cyc_h, 2), "Cykl Rzeczywisty [h]": real_cycle_txt,
                })

                if members and len(members) > 1:
                    produkty_txt = f"{len(members)} produktów (współdzielony)"
                elif members and len(members) == 1:
                    produkty_txt = members[0]["product"]
                else:
                    produkty_txt = "—"

                final_fleet_rows.append({
                    "ID Urządzenia": tag_id,
                    "Przypisana Linia": kat,
                    "Produkty": produkty_txt,
                    "Pojemność [m³]": round(v_tank_user, 1),
                    "Masa Szarży [kg]": int(mass_per_batch),
                    "Cykl Szacowany [h]": round(cyc_h, 2),
                    "Szarż / miesiąc (per aparat)": int(batches_per_tank),
                    "Utylizacja Czasowa": f"{real_utilization:.1f}%",
                    "Status": status_txt
                })

            tag_counter += 1

        st.markdown("### 📊 Aktualne Zestawienie Floty Produkcyjnej (Możesz usuwać wiersze)")
        st.caption("💡 **Instrukcja:** Aby usunąć zbiornik, zaznacz pole wyboru po lewej stronie wiersza i naciśnij `Delete` na klawiaturze (lub użyj ikony kosza). "
                    "Kolumnę **Przypisana Linia** można edytować tylko na wartości z aktywnie wybranych linii produktowych - inne wartości zostaną odrzucone przy zatwierdzaniu. "
                    "**Pojemność, Masa Szarży, Cykl, Szarże i Utylizacja są tylko do odczytu** w tej tabeli - to są wyliczone wartości; żeby je zmienić, edytuj pola "
                    "'Pojemność Mieszalnika (bazowa)' / 'Zbiornik #N — Pojemność/Cykl' powyżej. Jeśli chcesz rozdzielić produkty ze **wspólnego zbiornika** "
                    "(receptura ze wspólnym ID Zbiornika) na osobne fizyczne mieszalniki, zrób to w pliku Excel (usuń/zmień ID Zbiornika dla jednego z produktów) "
                    "i wgraj recepturę ponownie w Zakładce 1 - samo zwiększenie liczby zbiorników tutaj dodaje pusty zbiornik, ale nie wie, który produkt ma do niego przenieść.")

        df_fleet = pd.DataFrame(final_fleet_rows)

        # Klucz data_editora zależy od parametrów, które faktycznie napędzają flotę (roczna
        # produkcja, pojemności/cykle bazowe i per-zbiornik, liczba zbiorników) - dzięki temu
        # KAŻDA zmiana tych wartości (np. pojemności bazowej w Kroku A) wymusza pełne
        # odświeżenie tabeli, zamiast pozwalać Streamlitowi zatrzymać nieaktualny stan komórek
        # z poprzedniego przebiegu. Między takimi zmianami klucz jest stabilny, więc ręczne
        # usuwanie wierszy w tabeli nadal działa normalnie do czasu kliknięcia "Zatwierdź".
        fleet_signature = hash(tuple(
            (kat, st.session_state.prod_dict[kat]["roczna"], st.session_state.prod_dict[kat]["user_vol_m3"],
             st.session_state.prod_dict[kat].get("num_tanks", 1),
             tuple(st.session_state.prod_dict[kat].get("tank_volumes", [])),
             tuple(st.session_state.prod_dict[kat].get("tank_cycles", [])))
            for kat in wybrane_kategorie
        ))

        edited_df = st.data_editor(
            df_fleet,
            hide_index=True,
            use_container_width=True,
            num_rows="dynamic",
            key=f"fleet_data_editor_v4_{fleet_signature}",
            disabled=["ID Urządzenia", "Produkty", "Pojemność [m³]", "Masa Szarży [kg]",
                      "Cykl Szacowany [h]", "Szarż / miesiąc (per aparat)", "Utylizacja Czasowa", "Status"]
        )

        if real_cycle_reference_rows:
            with st.expander("📊 Rzeczywisty czas cyklu (referencja z Zakładki 2/7 — informacyjnie, nieedytowalne)", expanded=False):
                st.caption("Ta tabela aktualizuje się automatycznie w miarę konfigurowania hydrauliki/bilansu cieplnego (Zakładka 2) "
                           "i dozowania/homogenizacji (Zakładka 6) — nie wpływa na flotę powyżej i nie da się jej edytować.")
                st.dataframe(pd.DataFrame(real_cycle_reference_rows), hide_index=True, use_container_width=True)

        if not edited_df.empty:
            total_annual_production_edited = sum(st.session_state.prod_dict[kat]["roczna"] for kat in wybrane_kategorie)
            total_batches_edited = pd.to_numeric(edited_df["Szarż / miesiąc (per aparat)"], errors="coerce").fillna(0).astype(int).sum()
            total_volume_edited = pd.to_numeric(edited_df["Pojemność [m³]"], errors="coerce").fillna(0.0).astype(float).sum()
            # Ta sama formuła (masa szarży x szarż/mies x 12) co przy zatwierdzaniu floty - PO
            # zaokrągleniu liczby szarż w górę do pełnych sztuk. Fleet ZAWSZE wychodzi trochę
            # WIĘKSZA niż dokładny cel z receptury (nie da się zrobić np. 17,3 szarży) - porównanie
            # "aktualności" floty musi więc odbywać się względem TEJ (też zaokrąglonej) liczby, nie
            # względem dokładnego celu, inaczej ostrzeżenie o nieaktualności nigdy by nie znikało,
            # nawet zaraz po zatwierdzeniu.
            mass_col_live = pd.to_numeric(edited_df["Masa Szarży [kg]"], errors="coerce").fillna(0.0)
            batches_col_live = pd.to_numeric(edited_df["Szarż / miesiąc (per aparat)"], errors="coerce").fillna(0).astype(int)
            st.session_state["live_fleet_annual_kg_rounded"] = float((mass_col_live * batches_col_live * MONTHS_PER_YEAR).sum())
        else:
            total_annual_production_edited = 0
            total_batches_edited = 0
            total_volume_edited = 0.0
            st.session_state["live_fleet_annual_kg_rounded"] = 0.0

        st.markdown("<br>", unsafe_allow_html=True)
        sum_col1, sum_col2, sum_col3 = st.columns(3)
        with sum_col1: st.metric(label="📈 Sumaryczny tonaż roczny zakładu", value=f"{total_annual_production_edited:,} kg")
        with sum_col2: st.metric(label="🔄 Suma szarż floty / miesiąc", value=f"{total_batches_edited} szarż")
        with sum_col3: st.metric(label="📐 Całkowita kubatura floty", value=f"{total_volume_edited:.1f} m³")

        recipes_df_check2 = st.session_state.get("recipes_df")
        if recipes_df_check2 is not None and not recipes_df_check2.empty and RECIPE_GROUP_COL in recipes_df_check2.columns:
            mismatched_lines = []
            for kat in wybrane_kategorie:
                group_rows_check2 = recipes_df_check2[recipes_df_check2[RECIPE_GROUP_COL] == kat]
                if group_rows_check2.empty:
                    continue
                recipe_roczna_kg2 = float(group_rows_check2[RECIPE_ANNUAL_COL].sum()) * 1000.0
                stored_roczna_kg2 = float(st.session_state.prod_dict[kat]["roczna"])
                if abs(recipe_roczna_kg2 - stored_roczna_kg2) > max(1.0, stored_roczna_kg2 * 0.01):
                    mismatched_lines.append(kat)
            if mismatched_lines:
                st.warning(f"⚠️ **Powyższy tonaż może być NIEAKTUALNY** dla: {', '.join(mismatched_lines)} — wybierz każdą z "
                           "tych linii w Kroku A powyżej, żeby zobaczyć i zsynchronizować dokładną różnicę.")

        st.markdown("---")

        if st.button("📥 Zatwierdź i wyślij konfigurację do kolejnych kroków", type="primary", use_container_width=True, key="btn_zatwierdz_flote_v3"):
            if edited_df.empty:
                st.error("❌ Flota nie może być pusta!")
            else:
                # Walidacja wierszy dodanych/edytowanych ręcznie w data_editor, aby uniknąć
                # KeyError przy próbie odczytu nieistniejącej linii z st.session_state.active_portfolio.
                invalid_rows = edited_df[~edited_df["Przypisana Linia"].isin(st.session_state.active_portfolio.keys())]
                numeric_cols = ["Pojemność [m³]", "Masa Szarży [kg]", "Cykl Szacowany [h]", "Szarż / miesiąc (per aparat)"]
                bad_numeric = edited_df[numeric_cols].apply(pd.to_numeric, errors="coerce").isna().any(axis=1)

                if not invalid_rows.empty:
                    zle_linie = sorted(set(invalid_rows["Przypisana Linia"].astype(str)))
                    st.error(f"❌ Wiersze wskazują na nieznane linie produktowe: {', '.join(zle_linie)}. "
                             f"Popraw kolumnę 'Przypisana Linia' na jedną z aktywnie wybranych linii i spróbuj ponownie.")
                elif bad_numeric.any():
                    st.error("❌ Niektóre wiersze zawierają niepoprawne (nienumeryczne) wartości w kolumnach liczbowych. Popraw dane i spróbuj ponownie.")
                else:
                    confirmed_mixers_blueprint = []
                    for _, row in edited_df.iterrows():
                        kat = row["Przypisana Linia"]
                        confirmed_mixers_blueprint.append({
                            "tag": row["ID Urządzenia"],
                            "product_family": kat,
                            "capacity_m3": float(row["Pojemność [m³]"]),
                            "material": st.session_state.active_portfolio[kat]["material"],
                            "batches_count": int(row["Szarż / miesiąc (per aparat)"]),
                            "mass_per_batch": int(row["Masa Szarży [kg]"]),
                            "cycle_h": float(row["Cykl Szacowany [h]"]),
                            "annual_volume": int(row["Masa Szarży [kg]"]) * int(row["Szarż / miesiąc (per aparat)"]) * MONTHS_PER_YEAR,
                            "recipe_product": st.session_state.tag_to_recipe_product.get(row["ID Urządzenia"]),
                            "shared_members": st.session_state.get("tag_to_shared_members", {}).get(row["ID Urządzenia"]),
                        })

                    st.session_state.confirmed_mixers = confirmed_mixers_blueprint
                    st.success(f"🎉 Zapisano strukturę floty ({len(confirmed_mixers_blueprint)} urządzeń).")
    else:
        st.info("💡 Wybierz co najmniej jedną linię produktową w panelu bocznym, aby rozpocząć.")


with tab2:
    st.header("📐 Karta Maszyn: Zaawansowane Projektowanie Procesowe")
    if not st.session_state.confirmed_mixers:
        st.warning("⚠️ Brak danych o flocie. Skonfiguruj i zatwierdź flotę w Zakładce 1, aby odblokować ten krok.")
    else:
        _stale_msg = check_fleet_staleness_warning()
        if _stale_msg:
            st.error(_stale_msg)
        summary_combined_rows = []

        # --- KROK 1: Inicjalizacja domyślnych parametrów (bez widgetów) dla każdego urządzenia ---
        for mixer in st.session_state.confirmed_mixers:
            m_id = mixer["tag"]
            kat = mixer["product_family"]

            if m_id not in st.session_state.mixer_tech_advanced_details:
                st.session_state.mixer_tech_advanced_details[m_id] = {}

            p = st.session_state.mixer_tech_advanced_details[m_id]

            # Domyślna średnica mieszadła SKALUJE SIĘ z pojemnością zbiornika (typowy stosunek
            # mieszadło:zbiornik ~1:3, przy założeniu geometrii H≈D) - wcześniej był tu sztywny
            # 0,6 m dla KAŻDEGO zbiornika niezależnie od wielkości, co przy mocy skalującej się
            # z D^5 dawało drastyczne zaniżenie mocy silnika dla dużych mieszalników (nawet
            # >200x dla zbiornika 120 m³) - to źródło "bardzo małych wartości dla silników".
            _tank_diameter_est_m = (4.0 * mixer["capacity_m3"] / 3.14159265) ** (1.0 / 3.0)
            _default_impeller_d_m = round(_tank_diameter_est_m / 3.0, 2)

            defaults = {
                "pump_flow_m3h": 15.0,
                "pipe_dn": 80,
                "pipe_length_m": 25.0,
                "delta_h_m": 5.0,
                "viscosity_min_cst": 30.0,
                "viscosity_max_cst": 300.0,
                "density_kg_m3": st.session_state.active_portfolio[kat]["density"] * 1000.0,
                "count_elbows_90": 4,
                "count_tees": 2,
                "count_valves": 3,
                "pump_efficiency": 0.65,
                "cp_product": 2.10,
                "t_product_in": 20.0,
                "t_product_out": 70.0,
                "k_coeff_grzania": 450.0,
                "tank_mass": 1200.0,
                "cp_steel": 0.46,
                "t_discharge_c": 30.0,
                "exchange_area_m2": 4.5,
                "flow_heat_value": 5.0,
                "flow_heat_unit": "m3/h",
                "flow_cool_value": 5.0,
                "flow_cool_unit": "m3/h",
                "coil_pipe_dn_mm": 50.0,
                "utility_type_heat": "Woda technologiczna",
                "utility_type_cool": "Woda technologiczna",
                "t_utility_heat_in": 95.0,
                "t_utility_cool_in": 12.0,
                "k_coeff": 350.0,
                "agitator_type": "Turbinowe (Rushton)",
                "agitator_rpm": 90.0,
                "agitator_diameter_m": _default_impeller_d_m,
                "pump_mode": "Dedykowana (dla tego zbiornika)",
                "shared_pump_id": "",
                "pump_mtbf_h": 2000.0,
                "pump_mttr_h": 8.0,
                "reactor_mtbf_h": 4000.0,
                "reactor_mttr_h": 12.0,
            }
            for key, val in defaults.items():
                if key not in p:
                    p[key] = val

        # --- KROK 2: Parametryzatory (widgety) — MUSZĄ wykonać się PRZED przeliczeniem poniżej. ---
        # POPRAWKA: wcześniej te widgety renderowały się PO pętli obliczeniowej, więc każda
        # zmiana wartości (np. DN rury) była widoczna w tabeli zbiorczej dopiero po KOLEJNEJ
        # interakcji — do tego czasu tabela pokazywała wynik dla poprzedniej wartości. Stąd
        # pozornie "odwrócona" fizyka (większe DN pokazujące większą prędkość) — to był efekt
        # przestarzałych danych, nie błąd wzoru.
        st.markdown("### ⚙️ Konfiguracja Techniczna Mieszalników (wejścia)")
        st.caption("**Białe pola edytujesz bezpośrednio w tabeli**; kolumny 🔧/🌀/🔥/❄️ to grupy (hydraulika/pompa, "
                   "mieszadło, grzanie, chłodzenie). Tryb pompy (dedykowana/współdzielona) i typ procesu "
                   "(zwykły/Smar-Wax z parą) ustawiasz osobno niżej, bo od nich zależy, które pola mają znaczenie.")

        mixer_table_rows = []
        for mixer in st.session_state.confirmed_mixers:
            m_id = mixer["tag"]
            p = st.session_state.mixer_tech_advanced_details[m_id]
            mixer_table_rows.append({
                "Tag": m_id, "Linia": mixer["product_family"],
                "🔧 DN rury": p["pipe_dn"], "🔧 Długość [m]": p["pipe_length_m"], "🔧 Δh [m]": p["delta_h_m"],
                "🔧 Lepkość MIN [cSt]": p["viscosity_min_cst"], "🔧 Lepkość MAX [cSt]": p["viscosity_max_cst"],
                "🔧 Kolana 90°": p["count_elbows_90"], "🔧 Zawory": p["count_valves"],
                "🔧 Przepływ pompy [m³/h]": p["pump_flow_m3h"],
                "🌀 Typ mieszadła": p["agitator_type"], "🌀 Obroty [obr/min]": p["agitator_rpm"],
                "🌀 Śr. mieszadła [m]": p["agitator_diameter_m"],
                "🔥 Medium grzewcze": p["utility_type_heat"], "🔥 Temp. zasil. grzew. [°C]": p["t_utility_heat_in"],
                "🔥 k grzania [W/m²K]": p["k_coeff_grzania"], "🔥 Przepływ medium grzew. [m³/h]": p["flow_heat_value"],
                "❄️ Medium chłodzące": p["utility_type_cool"], "❄️ Temp. wody chłodz. [°C]": p["t_utility_cool_in"],
                "❄️ k chłodzenia [W/m²K]": p["k_coeff"], "❄️ Przepływ medium chłodz. [m³/h]": p["flow_cool_value"],
                "🔧 DN wężownicy [mm]": p["coil_pipe_dn_mm"],
                "🔥 Powierzchnia wymiany [m²]": p["exchange_area_m2"], "🔥 Temp. pocz. [°C]": p["t_product_in"],
                "🔥 Temp. procesu [°C]": p["t_product_out"], "🔥 Temp. rozlewu [°C]": p["t_discharge_c"],
                "🔧 MTBF reaktora [h]": p["reactor_mtbf_h"], "🔧 MTTR reaktora [h]": p["reactor_mttr_h"],
            })

        edited_mixer_table = st.data_editor(
            pd.DataFrame(mixer_table_rows), hide_index=True, use_container_width=True, key="mixer_master_editor",
            disabled=["Tag", "Linia"],
            column_config={
                "🌀 Typ mieszadła": st.column_config.SelectboxColumn(options=list(AGITATOR_TYPES.keys())),
                "🔥 Medium grzewcze": st.column_config.SelectboxColumn(options=list(MEDIA_PROCESOWE.keys())),
                "❄️ Medium chłodzące": st.column_config.SelectboxColumn(options=list(MEDIA_PROCESOWE.keys())),
            }
        )
        _mixer_field_map = {
            "🔧 DN rury": "pipe_dn", "🔧 Długość [m]": "pipe_length_m", "🔧 Δh [m]": "delta_h_m",
            "🔧 Lepkość MIN [cSt]": "viscosity_min_cst", "🔧 Lepkość MAX [cSt]": "viscosity_max_cst",
            "🔧 Kolana 90°": "count_elbows_90", "🔧 Zawory": "count_valves",
            "🔧 Przepływ pompy [m³/h]": "pump_flow_m3h", "🌀 Typ mieszadła": "agitator_type",
            "🌀 Obroty [obr/min]": "agitator_rpm", "🌀 Śr. mieszadła [m]": "agitator_diameter_m",
            "🔥 Medium grzewcze": "utility_type_heat", "🔥 Temp. zasil. grzew. [°C]": "t_utility_heat_in",
            "🔥 k grzania [W/m²K]": "k_coeff_grzania", "🔥 Przepływ medium grzew. [m³/h]": "flow_heat_value",
            "❄️ Medium chłodzące": "utility_type_cool", "❄️ Temp. wody chłodz. [°C]": "t_utility_cool_in",
            "❄️ k chłodzenia [W/m²K]": "k_coeff", "❄️ Przepływ medium chłodz. [m³/h]": "flow_cool_value",
            "🔧 DN wężownicy [mm]": "coil_pipe_dn_mm",
            "🔥 Powierzchnia wymiany [m²]": "exchange_area_m2", "🔥 Temp. pocz. [°C]": "t_product_in",
            "🔥 Temp. procesu [°C]": "t_product_out", "🔥 Temp. rozlewu [°C]": "t_discharge_c",
            "🔧 MTBF reaktora [h]": "reactor_mtbf_h", "🔧 MTTR reaktora [h]": "reactor_mttr_h",
        }
        for _, row in edited_mixer_table.iterrows():
            p = st.session_state.mixer_tech_advanced_details[row["Tag"]]
            for col_name, field_key in _mixer_field_map.items():
                p[field_key] = row[col_name]
            if MEDIA_PROCESOWE[p["utility_type_heat"]].get("steam"):
                st.caption(f"ℹ️ {row['Tag']}: para nasycona jako medium grzewcze — bilans liczony przez ciepło "
                           "skraplania, nie cp·ΔT (ΔT medium grzewczego z tabeli nie dotyczy pary).")

        # Rezerwujemy tu miejsce na wyniki (wypełniamy je NIŻEJ, po tym jak selektor szczegółów i
        # pompy współdzielone ustawią swoje wartości - w Streamlit obliczenia MUSZĄ wykonać się PO
        # widżetach, żeby nie wrócił błąd "wyniki o jeden krok w tyle", który już raz naprawialiśmy;
        # ale kontener pozwala mimo to WYŚWIETLIĆ wynik zaraz pod tabelą konfiguracyjną).
        mixer_results_placeholder = st.container()

        st.markdown("###### 🔍 Porównanie mieszalników")
        st.caption("Wybierz 2-3 mieszalniki, żeby zobaczyć konfigurację (tryb pompy, typ procesu) i wyniki obok "
                   "siebie. Pola te zależą od siebie warunkowo, dlatego nie mieszczą się w tabeli powyżej.")
        _all_mixer_tags = [m["tag"] for m in st.session_state.confirmed_mixers]
        compare_mixer_tags = st.multiselect(
            "Wybierz mieszalniki (max 3):", _all_mixer_tags,
            default=_all_mixer_tags[:min(2, len(_all_mixer_tags))], key="mixer_compare_select"
        )
        if len(compare_mixer_tags) > 3:
            st.warning("⚠️ Wybierz maksymalnie 3 — pokazuję pierwsze 3 z wybranych.")
            compare_mixer_tags = compare_mixer_tags[:3]

        if not compare_mixer_tags:
            st.info("ℹ️ Wybierz przynajmniej jeden mieszalnik powyżej, żeby zobaczyć jego szczegóły.")
        else:
            if "tanker_capacity_t" not in st.session_state:
                st.session_state.tanker_capacity_t = 24.0  # typowa cysterna drogowa ~24 t
            st.session_state.tanker_capacity_t = st.number_input(
                "🚛 Pojemność cysterny (dostawy RM / wysyłki FG) [t]:", min_value=1.0,
                value=float(st.session_state.tanker_capacity_t), step=1.0, key="tanker_capacity_input",
                help="Wspólna dla wszystkich porównywanych mieszalników poniżej — typowa cysterna drogowa to ok. 24 t."
            )

            compare_cols = st.columns(len(compare_mixer_tags))
            for col, selected_mixer_tag in zip(compare_cols, compare_mixer_tags):
                mixer = next(m for m in st.session_state.confirmed_mixers if m["tag"] == selected_mixer_tag)
                p = st.session_state.mixer_tech_advanced_details[selected_mixer_tag]
                with col:
                    st.markdown(f"**🔧 {selected_mixer_tag}** ({mixer['product_family']})")

                    batches_month_cmp = mixer["batches_count"]
                    batches_year_cmp = batches_month_cmp * MONTHS_PER_YEAR
                    mass_per_batch_t = mixer["mass_per_batch"] / 1000.0
                    monthly_mass_t = batches_month_cmp * mass_per_batch_t
                    annual_mass_t = batches_year_cmp * mass_per_batch_t

                    bm1, bm2 = st.columns(2)
                    with bm1: st.metric("📦 Szarż/miesiąc", f"{batches_month_cmp}")
                    with bm2: st.metric("📦 Szarż/rok", f"{batches_year_cmp}")

                    # --- Rozbicie na surowce, ten sam uklad co plik uzytkownika: Surowiec | %
                    # dozowania | Tony/szarze | Zuzycie/miesiac / rok [t]. ---
                    recipes_df_cmp = st.session_state.get("recipes_df")
                    recipe_product_cmp = mixer.get("recipe_product")
                    rm_bulk_month_t_by_material = {}  # tylko materialy luzem (dostawa cysterna)
                    if recipes_df_cmp is not None and not recipes_df_cmp.empty and recipe_product_cmp:
                        match_cmp = recipes_df_cmp[recipes_df_cmp[RECIPE_PRODUCT_COL] == recipe_product_cmp]
                        if not match_cmp.empty:
                            row_cmp = match_cmp.iloc[0]
                            rm_storage_override_cmp = st.session_state.get("rm_storage_method_override", {})
                            material_rows_cmp = []
                            for mat in RECIPE_RAW_MATERIALS:
                                dozowanie_kg_t = float(row_cmp.get(mat, 0) or 0)
                                if dozowanie_kg_t <= 0:
                                    continue
                                mat_month_t = dozowanie_kg_t / 1000.0 * monthly_mass_t
                                is_bulk_mat = rm_storage_override_cmp.get(mat) == "Zbiornik (luzem)"
                                if is_bulk_mat:
                                    rm_bulk_month_t_by_material[mat] = mat_month_t
                                material_rows_cmp.append({
                                    "Surowiec": mat.replace(" [kg/t]", ""),
                                    "% dozowania": round_visible(dozowanie_kg_t / 10.0),
                                    "t/szarżę": round_visible(dozowanie_kg_t / 1000.0 * mass_per_batch_t),
                                    "t/miesiąc": round_visible(mat_month_t),
                                    "t/rok": round_visible(mat_month_t * MONTHS_PER_YEAR),
                                    "Dostawa": "🚚 Cysterna (luzem)" if is_bulk_mat else "📦 Beczki/IBC/worki",
                                })
                            if material_rows_cmp:
                                st.dataframe(pd.DataFrame(material_rows_cmp), hide_index=True, use_container_width=True)
                        else:
                            st.caption(f"⚠️ Nie znaleziono '{recipe_product_cmp}' w recepturze (Zakładka 1).")
                    else:
                        st.caption("Ten mieszalnik nie ma przypisanego konkretnego produktu z receptury (wspólny/kampanijny zbiornik).")

                    # --- Badania laboratoryjne: PRIORYTET 1 - arkusz 'Badania Laboratoryjne'
                    # (testy jako wiersze, produkty jako kolumny, patrz parse_qc_tests_excel);
                    # PRIORYTET 2 - kolumny 'QC: {test} [x]' wprost w arkuszu Receptury;
                    # PRIORYTET 3 (fallback) - panel zwolnienia per LINIA w Zakładce 6 (VSM). ---
                    qc_tests_used, qc_source_label = get_qc_tests_for_mixer(mixer)
                    n_tests_per_batch = len(qc_tests_used)
                    time_min_per_batch = sum(get_full_qc_catalog().get(t, {}).get("duration_min", 0) for t in qc_tests_used)
                    lab_tests_month = n_tests_per_batch * batches_month_cmp
                    lab_tests_year = n_tests_per_batch * batches_year_cmp
                    time_h_month = (time_min_per_batch * batches_month_cmp) / 60.0
                    time_h_year = (time_min_per_batch * batches_year_cmp) / 60.0
                    bl1, bl2 = st.columns(2)
                    with bl1: st.metric("🧪 Badań QC/miesiąc", f"{lab_tests_month}",
                                        help=f"{n_tests_per_batch} testów/szarżę ({qc_source_label}) × {batches_month_cmp} szarż/mies.")
                    with bl2: st.metric("🧪 Badań QC/rok", f"{lab_tests_year}")
                    bl3, bl4 = st.columns(2)
                    with bl3: st.metric("⏱️ Czas QC/miesiąc", f"{time_h_month:,.1f} h",
                                        help=f"{time_min_per_batch:.0f} min/szarżę (suma czasów wszystkich testów) × {batches_month_cmp} szarż/mies. "
                                             "Zakłada testy WYKONYWANE SEKWENCYJNIE (jeden po drugim) - jeśli część robicie równolegle, realny czas będzie krótszy.")
                    with bl4: st.metric("⏱️ Czas QC/rok", f"{time_h_year:,.0f} h")

                    # --- Cysterny: dostawy surowców (RM) - ROZBITE PER SUROWIEC (nie łącznie),
                    # i TYLKO dla materiałów faktycznie dostarczanych luzem cysterną (zbiornik
                    # dedykowany) - beczkowane/IBC jadą zwykłym transportem, nie cysterną. ---
                    st.markdown("**🚚 Cysterny — dostawy RM (per surowiec)**")
                    if rm_bulk_month_t_by_material:
                        rm_tanker_rows = []
                        for mat, mat_month_t in rm_bulk_month_t_by_material.items():
                            tankers_month_mat = math.ceil(mat_month_t / st.session_state.tanker_capacity_t) if st.session_state.tanker_capacity_t > 0 else 0
                            tankers_year_mat = math.ceil((mat_month_t * MONTHS_PER_YEAR) / st.session_state.tanker_capacity_t) if st.session_state.tanker_capacity_t > 0 else 0
                            rm_tanker_rows.append({
                                "Surowiec": mat.replace(" [kg/t]", ""),
                                "t/miesiąc": round_visible(mat_month_t),
                                "Cystern/miesiąc": tankers_month_mat,
                                "Cystern/rok": tankers_year_mat,
                            })
                        st.dataframe(pd.DataFrame(rm_tanker_rows), hide_index=True, use_container_width=True)
                        st.caption("Każdy surowiec liczony osobno (własna cysterna, własny dostawca) — dokładne do "
                                   "planowania harmonogramu dostaw, nie łączna suma masy.")
                    else:
                        st.caption("Brak surowców dostarczanych luzem (cysterną) dla tego produktu — wszystkie w "
                                   "beczkach/IBC/workach, albo sposób magazynowania nie został jeszcze ustawiony "
                                   "w Zakładce 2.")

                    fg_tankers_month = math.ceil(monthly_mass_t / st.session_state.tanker_capacity_t) if st.session_state.tanker_capacity_t > 0 else 0
                    fg_tankers_year = math.ceil(annual_mass_t / st.session_state.tanker_capacity_t) if st.session_state.tanker_capacity_t > 0 else 0

                    st.markdown("**🚚 Cysterny — wysyłki FG**")
                    bt3, bt4 = st.columns(2)
                    with bt3: st.metric("Wysyłki/miesiąc", f"{fg_tankers_month}")
                    with bt4: st.metric("Wysyłki/rok", f"{fg_tankers_year}")
                    st.caption("Dotyczy TYLKO produktu wysyłanego luzem w cysternie — jeśli ten produkt jest "
                               "pakowany (beczki/kanistry/palety), zignoruj tę metrykę (masz ją w Zakładce 4).")

                    p["pump_mode"] = st.selectbox(
                        "Tryb pompy:", ["Dedykowana (dla tego zbiornika)", "Współdzielona (kilka zbiorników)"],
                        index=["Dedykowana (dla tego zbiornika)", "Współdzielona (kilka zbiorników)"].index(p["pump_mode"]),
                        key=f"pump_mode_{selected_mixer_tag}",
                        help="Jedna fizyczna pompa może obsługiwać kilka zbiorników na przemian — wybierz "
                             "'Współdzielona' i podaj ten sam ID pompy dla wszystkich zbiorników, które ją dzielą."
                    )
                    if p["pump_mode"] == "Współdzielona (kilka zbiorników)":
                        p["shared_pump_id"] = st.text_input(
                            "ID pompy współdzielonej:", value=p["shared_pump_id"] or "P-01", key=f"shared_pump_id_{selected_mixer_tag}"
                        )
                        st.caption("Przepływ/sprawność/MTBF/MTTR tej pompy w tabeli '🔧 Pompy Współdzielone' niżej.")
                        shared = st.session_state.shared_pumps.get(p["shared_pump_id"], {})
                        pump_mtbf_disp = shared.get("mtbf_h", 2000.0)
                        pump_mttr_disp = shared.get("mttr_h", 8.0)
                        avail_pump_preview = pump_mtbf_disp / (pump_mtbf_disp + pump_mttr_disp) * 100.0
                    else:
                        p["shared_pump_id"] = ""
                        p["pump_mtbf_h"] = st.number_input("MTBF pompy [h]:", min_value=1.0, value=float(p["pump_mtbf_h"]), key=f"pump_mtbf_{selected_mixer_tag}")
                        p["pump_mttr_h"] = st.number_input("MTTR pompy [h]:", min_value=0.1, value=float(p["pump_mttr_h"]), key=f"pump_mttr_{selected_mixer_tag}")
                        avail_pump_preview = p["pump_mtbf_h"] / (p["pump_mtbf_h"] + p["pump_mttr_h"]) * 100.0
                    avail_reactor_preview = p["reactor_mtbf_h"] / (p["reactor_mtbf_h"] + p["reactor_mttr_h"]) * 100.0
                    avail_combined_preview = (avail_pump_preview / 100.0) * (avail_reactor_preview / 100.0) * 100.0
                    st.metric("Dostępność łączna", f"{avail_combined_preview:.1f}%")

                    p.setdefault("process_type", "Ciecz (mieszanie/blending)")
                    p["process_type"] = st.selectbox(
                        "Typ procesu:", ["Ciecz (mieszanie/blending)", "Smar/Wax (gotowanie z odparowaniem)"],
                        index=["Ciecz (mieszanie/blending)", "Smar/Wax (gotowanie z odparowaniem)"].index(p["process_type"]),
                        key=f"proc_type_{selected_mixer_tag}",
                        help="Wybierz 'Smar/Wax', jeśli ten reaktor gotuje z intensywnym odparowaniem — zbiorczy "
                             "rurociąg zrzutowy policzy się niżej, dla wszystkich reaktorów tego typu naraz."
                    )
                    if p["process_type"] == "Smar/Wax (gotowanie z odparowaniem)":
                        p.setdefault("steam_avg_flow", 0.0185)
                        p.setdefault("steam_max_process", 0.037)
                        p.setdefault("steam_max_decompress", 0.089)
                        p["steam_avg_flow"] = st.number_input("Średni strumień odwadniania [kg/s]:", min_value=0.0, value=float(p["steam_avg_flow"]), step=0.001, format="%.4f", key=f"steam_avg_{selected_mixer_tag}")
                        p["steam_max_process"] = st.number_input("Maks. strumień procesowy [kg/s]:", min_value=0.0, value=float(p["steam_max_process"]), step=0.001, format="%.4f", key=f"steam_proc_{selected_mixer_tag}")
                        p["steam_max_decompress"] = st.number_input("Maks. strumień dekompresji [kg/s]:", min_value=0.0, value=float(p["steam_max_decompress"]), step=0.001, format="%.4f", key=f"steam_decomp_{selected_mixer_tag}")

                    st.markdown("**⚡ Energetyczne KPI**")
                    try:
                        _visc_avg_kpi = (p["viscosity_min_cst"] + p["viscosity_max_cst"]) / 2.0
                        _, _, agitator_power_kw_kpi = compute_agitator_power(
                            p["agitator_type"], p["agitator_rpm"], p["agitator_diameter_m"], p["density_kg_m3"], _visc_avg_kpi
                        )
                        _heat_res_kpi = compute_thermal_ntu(
                            mixer["mass_per_batch"], p["cp_product"], p["t_product_in"], p["t_product_out"],
                            p["k_coeff_grzania"], p["exchange_area_m2"], p["utility_type_heat"],
                            p["flow_heat_value"], p["flow_heat_unit"], p["t_utility_heat_in"])
                        _cool_res_kpi = compute_thermal_ntu(
                            mixer["mass_per_batch"], p["cp_product"], p["t_product_out"], p["t_discharge_c"],
                            p["k_coeff"], p["exchange_area_m2"], p["utility_type_cool"],
                            p["flow_cool_value"], p["flow_cool_unit"], p["t_utility_cool_in"])

                        heating_kwh_batch = (_heat_res_kpi["q_total_kj"] * 0.2778 / 1000.0) if _heat_res_kpi["status"] == "ok" else 0.0
                        cooling_kwh_batch = (_cool_res_kpi["q_total_kj"] * 0.2778 / 1000.0) if _cool_res_kpi["status"] == "ok" else 0.0
                        mixing_kwh_batch = agitator_power_kw_kpi * mixer.get("cycle_h", 4.0)
                        batches_month_kpi = mixer.get("batches_count", 0)
                        total_kwh_month = (heating_kwh_batch + cooling_kwh_batch + mixing_kwh_batch) * batches_month_kpi

                        st.metric("Grzanie / szarżę", f"{heating_kwh_batch:.1f} kWh" if _heat_res_kpi["status"] == "ok" else "⚠️ N/A")
                        st.metric("Chłodzenie / szarżę", f"{cooling_kwh_batch:.1f} kWh" if _cool_res_kpi["status"] == "ok" else "⚠️ N/A")
                        st.metric("Mieszanie / szarżę", f"{mixing_kwh_batch:.1f} kWh")
                        st.metric("Razem / miesiąc", f"{total_kwh_month:,.0f} kWh")
                    except Exception as _kpi_exc:
                        st.caption(f"⚠️ Nie udało się policzyć KPI energetycznego: {_kpi_exc}")

        # --- Pompy współdzielone: jedno miejsce edycji przepływu/sprawności/MTBF/MTTR, ---
        # wspólne dla wszystkich zbiorników, które przypisano do tej samej pompy powyżej.
        shared_pump_ids_in_use = sorted(set(
            st.session_state.mixer_tech_advanced_details[m["tag"]]["shared_pump_id"]
            for m in st.session_state.confirmed_mixers
            if st.session_state.mixer_tech_advanced_details.get(m["tag"], {}).get("pump_mode") == "Współdzielona (kilka zbiorników)"
            and st.session_state.mixer_tech_advanced_details[m["tag"]]["shared_pump_id"]
        ))

        if shared_pump_ids_in_use:
            st.markdown("### 🔧 Pompy Współdzielone — Przepływ, Sprawność, MTBF/MTTR")
            st.caption("Jeden wiersz = jedna fizyczna pompa obsługująca kilka zbiorników na przemian. Zmiana tutaj "
                       "dotyczy od razu wszystkich zbiorników przypisanych do tej pompy powyżej.")
            for pid in shared_pump_ids_in_use:
                st.session_state.shared_pumps.setdefault(pid, {
                    "flow_m3h": 15.0, "efficiency": 0.65, "mtbf_h": 2000.0, "mttr_h": 8.0,
                })

            shared_pump_rows = [
                {"ID Pompy": pid, "Przepływ [m³/h]": cfg["flow_m3h"], "Sprawność [-]": cfg["efficiency"],
                 "MTBF [h]": cfg["mtbf_h"], "MTTR [h]": cfg["mttr_h"],
                 "Zbiorniki": ", ".join(m["tag"] for m in st.session_state.confirmed_mixers
                                        if st.session_state.mixer_tech_advanced_details.get(m["tag"], {}).get("shared_pump_id") == pid)}
                for pid, cfg in st.session_state.shared_pumps.items() if pid in shared_pump_ids_in_use
            ]
            edited_shared_pumps = st.data_editor(
                pd.DataFrame(shared_pump_rows), hide_index=True, use_container_width=True,
                disabled=["ID Pompy", "Zbiorniki"], key="shared_pumps_editor"
            )
            for _, row in edited_shared_pumps.iterrows():
                st.session_state.shared_pumps[row["ID Pompy"]] = {
                    "flow_m3h": float(row["Przepływ [m³/h]"]), "efficiency": float(row["Sprawność [-]"]),
                    "mtbf_h": float(row["MTBF [h]"]), "mttr_h": float(row["MTTR [h]"]),
                }

        st.markdown("---")

        # --- KROK 3: Przeliczenie hydrauliki/bilansu cieplnego — TERAZ z aktualnymi wartościami z KROKU 2. ---
        for mixer in st.session_state.confirmed_mixers:
            m_id = mixer["tag"]
            kat = mixer["product_family"]
            p = st.session_state.mixer_tech_advanced_details[m_id]

            try:
                # --- 0. Rozwiązanie parametrów pompy (dedykowana vs. współdzielona) i niezawodności ---
                if p["pump_mode"] == "Współdzielona (kilka zbiorników)" and p["shared_pump_id"] in st.session_state.shared_pumps:
                    shared_pump_cfg = st.session_state.shared_pumps[p["shared_pump_id"]]
                    effective_pump_flow_m3h = shared_pump_cfg["flow_m3h"]
                    effective_pump_efficiency = shared_pump_cfg["efficiency"]
                    pump_mtbf_h = shared_pump_cfg["mtbf_h"]
                    pump_mttr_h = shared_pump_cfg["mttr_h"]
                else:
                    effective_pump_flow_m3h = p["pump_flow_m3h"]
                    effective_pump_efficiency = p["pump_efficiency"]
                    pump_mtbf_h = p["pump_mtbf_h"]
                    pump_mttr_h = p["pump_mttr_h"]

                availability_pump_pct = pump_mtbf_h / (pump_mtbf_h + pump_mttr_h) * 100.0 if (pump_mtbf_h + pump_mttr_h) > 0 else 100.0
                availability_reactor_pct = p["reactor_mtbf_h"] / (p["reactor_mtbf_h"] + p["reactor_mttr_h"]) * 100.0 \
                    if (p["reactor_mtbf_h"] + p["reactor_mttr_h"]) > 0 else 100.0
                availability_combined_pct = (availability_pump_pct / 100.0) * (availability_reactor_pct / 100.0) * 100.0

                # --- 1. HYDRAULIKA POMPY (Re / opór / moc), po 3 punktach lepkości ---
                visc_min = p["viscosity_min_cst"]
                visc_max = p["viscosity_max_cst"]
                visc_avg = (visc_min + visc_max) / 2.0

                zeta_sum_calculated = (p["count_elbows_90"] * 0.5) + (p["count_tees"] * 1.5) + (p["count_valves"] * 0.2)

                re_min, p_bar_min, power_kw_min, velocity = compute_hydraulics(
                    effective_pump_flow_m3h, p["pipe_dn"], p["pipe_length_m"], p["delta_h_m"],
                    visc_min, p["density_kg_m3"], zeta_sum_calculated, effective_pump_efficiency)
                re_avg, p_bar_avg, power_kw_avg, _ = compute_hydraulics(
                    effective_pump_flow_m3h, p["pipe_dn"], p["pipe_length_m"], p["delta_h_m"],
                    visc_avg, p["density_kg_m3"], zeta_sum_calculated, effective_pump_efficiency)
                re_max, p_bar_max, power_kw_max, _ = compute_hydraulics(
                    effective_pump_flow_m3h, p["pipe_dn"], p["pipe_length_m"], p["delta_h_m"],
                    visc_max, p["density_kg_m3"], zeta_sum_calculated, effective_pump_efficiency)

                # --- 2. MOC MIESZANIA (dawniej zdefiniowane, ale nigdy nie używane) ---
                re_mix, mix_regime, agitator_power_kw = compute_agitator_power(
                    p["agitator_type"], p["agitator_rpm"], p["agitator_diameter_m"],
                    p["density_kg_m3"], visc_avg)

                # --- 3. BILANS CIEPLNY: GRZANIE i CHŁODZENIE, metodą NTU-efektywności (produkt
                # dobrze wymieszany, medium przepływa ciągle przez płaszcz/wężownicę o stałym UA).
                mass_product = mixer["mass_per_batch"]
                heat_res = compute_thermal_ntu(
                    mass_product, p["cp_product"], p["t_product_in"], p["t_product_out"],
                    p["k_coeff_grzania"], p["exchange_area_m2"], p["utility_type_heat"],
                    p["flow_heat_value"], p["flow_heat_unit"], p["t_utility_heat_in"])

                # --- 4. CHŁODZENIE DO ROZLEWU ---
                cool_res = compute_thermal_ntu(
                    mass_product, p["cp_product"], p["t_product_out"], p["t_discharge_c"],
                    p["k_coeff"], p["exchange_area_m2"], p["utility_type_cool"],
                    p["flow_cool_value"], p["flow_cool_unit"], p["t_utility_cool_in"])

                # --- 5. Zapis wyników z powrotem do stanu sesji, aby Zakładka 2 mogła z nich realnie korzystać ---
                # Czas pompowania: objętość szarży / wydajność pompy.
                pumping_time_h = (mass_product / p["density_kg_m3"]) / effective_pump_flow_m3h if effective_pump_flow_m3h > 0 else 0.0
                flow_heat_kg_h = (p["flow_heat_value"] if p["flow_heat_unit"] == "m3/h" else p["flow_heat_value"] * 60.0 / 1000.0) * MEDIA_PROCESOWE[p["utility_type_heat"]]["density_kg_m3"] if not MEDIA_PROCESOWE[p["utility_type_heat"]]["steam"] else 0.0
                flow_cool_kg_h = (p["flow_cool_value"] if p["flow_cool_unit"] == "m3/h" else p["flow_cool_value"] * 60.0 / 1000.0) * MEDIA_PROCESOWE[p["utility_type_cool"]]["density_kg_m3"]

                st.session_state.calculated_times[m_id] = {
                    "power_mix_kw": agitator_power_kw,
                    "power_pump_kw": power_kw_avg,
                    "heating": heat_res["required_time_h"] if heat_res["status"] == "ok" else 0.0,
                    "pumping": pumping_time_h,
                    "t_max_mix": p["t_product_out"],
                    "t_rozlew": p["t_discharge_c"],
                    "cooling_h": cool_res["required_time_h"] if cool_res["status"] == "ok" else 0.0,
                    "power_heating_kw": heat_res.get("power_start_kw", 0.0),
                    "power_cooling_kw": cool_res.get("power_start_kw", 0.0),
                    "flow_heating_kg_h": flow_heat_kg_h,
                    "flow_cooling_kg_h": flow_cool_kg_h,
                    "medium_grz": p["utility_type_heat"],
                    "medium_chl": p["utility_type_cool"],
                    "is_steam": MEDIA_PROCESOWE[p["utility_type_heat"]]["steam"],
                    "availability_pct": availability_combined_pct,
                    "availability_pump_pct": availability_pump_pct,
                    "availability_reactor_pct": availability_reactor_pct,
                }

                cooling_txt = f"{cool_res['required_time_h']:.2f}" if cool_res["status"] == "ok" else ("—" if cool_res["status"] == "brak_potrzeby" else "⚠️ N/A")
                heating_txt = f"{heat_res['required_time_h']:.2f}" if heat_res["status"] == "ok" else "⚠️ N/A"

                # Sanity-check prędkości medium grzewczego/chłodzącego przez wężownicę - łapie
                # rzędowe pomyłki we wpisanym przepływie (np. 500 zamiast 5 m³/h), które inaczej
                # dałyby fizycznie niemożliwe prędkości (patrz przykład z Excela: 52 000 m³/h przez
                # typową rurę = 21x prędkość dźwięku).
                v_heat_coil, v_heat_flag = check_coil_velocity(p["flow_heat_value"], p["flow_heat_unit"], p["utility_type_heat"], p["coil_pipe_dn_mm"])
                v_cool_coil, v_cool_flag = check_coil_velocity(p["flow_cool_value"], p["flow_cool_unit"], p["utility_type_cool"], p["coil_pipe_dn_mm"])
                coil_velocity_txt = " / ".join(filter(None, [
                    f"grz: {v_heat_coil:.1f} m/s" if v_heat_coil is not None else None,
                    f"chł: {v_cool_coil:.1f} m/s" if v_cool_coil is not None else None,
                ])) or "—"

                summary_combined_rows.append({
                    "ID Urządzenia": m_id,
                    "Linia": kat,
                    "Prędkość [m/s]": round(velocity, 2),
                    "Opór [bar] (Min/Śr/Max)": f"{p_bar_min:.2f}/{p_bar_avg:.2f}/{p_bar_max:.2f}",
                    "Moc Pompy [kW] (Min/Śr/Max)": f"{power_kw_min:.2f}/{power_kw_avg:.2f}/{power_kw_max:.2f}",
                    "Moc Mieszania [kW]": round(agitator_power_kw, 2),
                    "Reżim mieszania": mix_regime,
                    "Moc Grzania (pocz.) [kW]": round(heat_res.get("power_start_kw", 0.0), 1),
                    "Przepływ medium grzewczego [kg/h]": round(flow_heat_kg_h, 1),
                    "Czas Grzania [h]": heating_txt,
                    "Efektywność wymiennika (grzanie) [%]": round(heat_res.get("effectiveness", 0.0) * 100.0, 1) if heat_res["status"] == "ok" else "—",
                    "Moc Chłodzenia (pocz.) [kW]": round(cool_res.get("power_start_kw", 0.0), 1),
                    "Przepływ medium chłodzącego [kg/h]": round(flow_cool_kg_h, 1),
                    "Czas chłodzenia [h]": cooling_txt,
                    "Prędkość w wężownicy": coil_velocity_txt,
                    "Dostępność (MTBF/MTTR) [%]": round(availability_combined_pct, 1),
                    "_velocity_val": velocity,
                    "_low_effectiveness_heat": heat_res.get("effectiveness", 1.0) < 0.3 if heat_res["status"] == "ok" else False,
                    "_coil_velocity_flag": v_heat_flag == "za_szybko" or v_cool_flag == "za_szybko",
                    "_cooling_status": cool_res["status"],
                    "_heating_status": heat_res["status"],
                })
            except Exception as exc:
                st.error(f"⚠️ Błąd obliczeń dla urządzenia {m_id}: {exc}. Sprawdź parametry w sekcji poniżej.")
                continue

        with mixer_results_placeholder:
            st.markdown("### 📋 Zbiorcza Specyfikacja Techniczna Maszyn, Pompy i Mieszania")
            st.info("💡 **Kryteria inżynieryjne:** Czerwonym kolorem podświetlane są **wyłącznie komórki**, które wykraczają poza normy "
                    f"(Prędkość poza przedziałem **{VELOCITY_MIN_MS} - {VELOCITY_MAX_MS} m/s**, niewystarczające ΔT grzania/chłodzenia "
                    "względem medium). Żółtym - niska efektywność wymiennika (<30%, metoda NTU) - sygnał, że przepływ medium lub "
                    "powierzchnia wymiany mogą być za małe względem celu czasowego.")

            if summary_combined_rows:
                df_summary = pd.DataFrame(summary_combined_rows)
                columns_to_show = [c for c in df_summary.columns if not c.startswith('_')]

                def style_basic_with_alerts(df_data):
                    style_matrix = pd.DataFrame('', index=df_data.index, columns=df_data.columns)
                    for idx, row in df_data.iterrows():
                        v = df_summary.loc[idx, "_velocity_val"]
                        if v < VELOCITY_MIN_MS or v > VELOCITY_MAX_MS:
                            if "Prędkość [m/s]" in style_matrix.columns:
                                style_matrix.loc[idx, "Prędkość [m/s]"] = 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'

                        low_eff_flag = df_summary.loc[idx, "_low_effectiveness_heat"]
                        if low_eff_flag:
                            if "Efektywność wymiennika (grzanie) [%]" in style_matrix.columns:
                                style_matrix.loc[idx, "Efektywność wymiennika (grzanie) [%]"] = 'background-color: #FFF2CC; color: #7F6000;'

                        if df_summary.loc[idx, "_coil_velocity_flag"]:
                            if "Prędkość w wężownicy" in style_matrix.columns:
                                style_matrix.loc[idx, "Prędkość w wężownicy"] = 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'

                        if df_summary.loc[idx, "_cooling_status"] == "niewystarczajace_dt":
                            if "Czas chłodzenia [h]" in style_matrix.columns:
                                style_matrix.loc[idx, "Czas chłodzenia [h]"] = 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'

                        if df_summary.loc[idx, "_heating_status"] == "niewystarczajace_dt":
                            if "Czas Grzania [h]" in style_matrix.columns:
                                style_matrix.loc[idx, "Czas Grzania [h]"] = 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'
                    return style_matrix

                df_filtered = df_summary[columns_to_show]
                styled_grid = df_filtered.style.apply(style_basic_with_alerts, axis=None)

                st.dataframe(styled_grid, hide_index=True, use_container_width=True)
            else:
                df_filtered = pd.DataFrame()
                st.warning("Brak poprawnie policzonych urządzeń — sprawdź komunikaty o błędach powyżej.")

        st.markdown("---")
        st.markdown("### 🛢️ Zbiorniki RM — Konfiguracja")
        confirmed_rm_tanks = st.session_state.get("confirmed_rm_tanks", [])
        total_rm_pump_power = 0.0
        total_rm_heating_power = 0.0
        if not confirmed_rm_tanks:
            st.info("ℹ️ Brak zatwierdzonych zbiorników RM — zadeklaruj je w **Zakładce 2 (Magazynowanie)**, sekcja "
                    "'✅ Zatwierdź Zbiorniki RM', żeby skonfigurować tu ich pompy, rurociąg i grzanie.")
        else:
            st.caption("Zbiorniki zatwierdzone w Zakładce 2. **Edytujesz bezpośrednio w tabeli** (🔧 hydraulika/"
                       "pompa, 🌡️ grzanie/izolacja). Tryb pompy (dedykowana/współdzielona) ustawiasz osobno niżej, "
                       "w rozwijanym panelu każdego zbiornika — zależy od niego, które pola mają znaczenie.")

            for rm_tank in confirmed_rm_tanks:
                st.session_state.rm_tank_tech_details.setdefault(rm_tank["tag"], {
                    "pump_mode": "Dedykowana (dla tego zbiornika)", "shared_pump_id": "",
                    "pump_flow_m3h": 10.0, "pipe_dn": 65, "pipe_length_m": 15.0, "delta_h_m": 4.0,
                    "viscosity_cst": 50.0, "density_kg_m3": 900.0, "count_elbows_90": 3, "count_valves": 2,
                    "pump_efficiency": 0.6, "heated": False, "target_temp_c": 40.0, "insulation_mm": 50,
                    "ambient_temp_c": 10.0, "specific_heat_j_kgk": 2000.0,
                })

            rm_config_rows = [
                {
                    "Tag": rm_tank["tag"], "Surowiec": rm_tank["material"], "Pojemność [m³]": rm_tank["capacity_m3"],
                    "🔧 DN rury": (d := st.session_state.rm_tank_tech_details[rm_tank["tag"]])["pipe_dn"],
                    "🔧 Długość [m]": d["pipe_length_m"], "🔧 Δh [m]": d["delta_h_m"],
                    "🔧 Lepkość [cSt]": d["viscosity_cst"], "🔧 Gęstość [kg/m³]": d["density_kg_m3"],
                    "🔧 Kolana 90°": d["count_elbows_90"], "🔧 Zawory": d["count_valves"],
                    "🔧 Przepływ pompy [m³/h]": d["pump_flow_m3h"], "🔧 Sprawność pompy": d["pump_efficiency"],
                    "🌡️ Grzany": d["heated"], "🌡️ Temp. docelowa [°C]": d["target_temp_c"],
                    "🌡️ Izolacja [mm]": d["insulation_mm"], "🌡️ Temp. otoczenia [°C]": d["ambient_temp_c"],
                }
                for rm_tank in confirmed_rm_tanks
            ]
            edited_rm_table = st.data_editor(
                pd.DataFrame(rm_config_rows), hide_index=True, use_container_width=True, key="rm_tank_master_editor",
                disabled=["Tag", "Surowiec", "Pojemność [m³]"],
                column_config={
                    "🔧 DN rury": st.column_config.NumberColumn(min_value=15, step=5),
                    "🔧 Długość [m]": st.column_config.NumberColumn(min_value=0.5, step=0.5),
                    "🔧 Δh [m]": st.column_config.NumberColumn(min_value=0.0, step=0.5),
                    "🔧 Lepkość [cSt]": st.column_config.NumberColumn(min_value=0.5, step=1.0),
                    "🔧 Gęstość [kg/m³]": st.column_config.NumberColumn(min_value=500.0, step=10.0),
                    "🔧 Kolana 90°": st.column_config.NumberColumn(min_value=0, step=1),
                    "🔧 Zawory": st.column_config.NumberColumn(min_value=0, step=1),
                    "🔧 Przepływ pompy [m³/h]": st.column_config.NumberColumn(min_value=0.5, step=0.5),
                    "🔧 Sprawność pompy": st.column_config.NumberColumn(min_value=0.1, max_value=1.0, step=0.05),
                    "🌡️ Grzany": st.column_config.CheckboxColumn(),
                    "🌡️ Temp. docelowa [°C]": st.column_config.NumberColumn(min_value=0.0, max_value=120.0, step=1.0),
                    "🌡️ Izolacja [mm]": st.column_config.SelectboxColumn(options=[0, 50, 100]),
                    "🌡️ Temp. otoczenia [°C]": st.column_config.NumberColumn(min_value=-30.0, max_value=45.0, step=1.0),
                }
            )
            for _, row in edited_rm_table.iterrows():
                d = st.session_state.rm_tank_tech_details.setdefault(row["Tag"], {})
                d["pipe_dn"] = row["🔧 DN rury"]; d["pipe_length_m"] = row["🔧 Długość [m]"]
                d["delta_h_m"] = row["🔧 Δh [m]"]; d["viscosity_cst"] = row["🔧 Lepkość [cSt]"]
                d["density_kg_m3"] = row["🔧 Gęstość [kg/m³]"]; d["count_elbows_90"] = row["🔧 Kolana 90°"]
                d["count_valves"] = row["🔧 Zawory"]; d["pump_flow_m3h"] = row["🔧 Przepływ pompy [m³/h]"]
                d["pump_efficiency"] = row["🔧 Sprawność pompy"]; d["heated"] = row["🌡️ Grzany"]
                d["target_temp_c"] = row["🌡️ Temp. docelowa [°C]"]; d["insulation_mm"] = row["🌡️ Izolacja [mm]"]
                d["ambient_temp_c"] = row["🌡️ Temp. otoczenia [°C]"]

            # Rezerwujemy miejsce na wyniki zaraz pod tabelą konfiguracyjną (jak u mieszalników) -
            # wypełniamy je NIŻEJ, po tym jak panele "tryb pompy" ustawią swoje wartości.
            rm_results_placeholder = st.container()

            st.markdown("###### 🔍 Porównanie zbiorników")
            st.caption("Wybierz 2-3 zbiorniki, żeby zobaczyć konfigurację (tryb pompy) i wyniki obok siebie. "
                       "Tryb pompy i wychładzanie po utracie grzania zależą od siebie warunkowo.")
            _all_rm_tags = [t["tag"] for t in confirmed_rm_tanks]
            compare_rm_tags = st.multiselect(
                "Wybierz zbiorniki (max 3):", _all_rm_tags,
                default=_all_rm_tags[:min(2, len(_all_rm_tags))], key="rm_compare_select"
            )
            if len(compare_rm_tags) > 3:
                st.warning("⚠️ Wybierz maksymalnie 3 — pokazuję pierwsze 3 z wybranych.")
                compare_rm_tags = compare_rm_tags[:3]

            if not compare_rm_tags:
                st.info("ℹ️ Wybierz przynajmniej jeden zbiornik powyżej, żeby zobaczyć jego szczegóły.")
            else:
                compare_rm_cols = st.columns(len(compare_rm_tags))
                for col, selected_rm_tag in zip(compare_rm_cols, compare_rm_tags):
                    rm_tank = next(t for t in confirmed_rm_tanks if t["tag"] == selected_rm_tag)
                    rm_sel_defaults = st.session_state.rm_tank_tech_details[selected_rm_tag]
                    with col:
                        st.markdown(f"**🔧 {selected_rm_tag}** — {rm_tank['material']}")
                        rm_sel_defaults["pump_mode"] = st.selectbox(
                            "Tryb pompy:", ["Dedykowana (dla tego zbiornika)", "Współdzielona (kilka zbiorników)"],
                            index=["Dedykowana (dla tego zbiornika)", "Współdzielona (kilka zbiorników)"].index(rm_sel_defaults["pump_mode"]),
                            key=f"rm_pump_mode_{selected_rm_tag}"
                        )
                        if rm_sel_defaults["pump_mode"] == "Współdzielona (kilka zbiorników)":
                            rm_sel_defaults["shared_pump_id"] = st.text_input(
                                "ID pompy współdzielonej:", value=rm_sel_defaults["shared_pump_id"] or "P-RM-01",
                                key=f"rm_shared_pump_id_{selected_rm_tag}",
                                help="Może to być ta sama pula co pompy mieszalników — wpisz identyczny ID, jeśli "
                                     "fizycznie ma to być ta sama pompa."
                            )
                        else:
                            rm_sel_defaults["shared_pump_id"] = ""

                        if rm_sel_defaults["heated"]:
                            t_start_for_curve = rm_sel_defaults["target_temp_c"]
                        else:
                            t_start_for_curve = max(rm_sel_defaults["ambient_temp_c"] + 20.0, rm_sel_defaults["ambient_temp_c"])
                        checkpoint_hours = [4, 12, 24, 48, 72]
                        cooling_table_rows = []
                        for insul_label, insul_mm in [("Brak izolacji", 0), ("Izolacja 50 mm", 50), ("Izolacja 100 mm", 100)]:
                            temps_at_checkpoints = compute_tank_cooling_curve(
                                rm_tank["capacity_m3"], insul_mm, t_start_for_curve, rm_sel_defaults["ambient_temp_c"],
                                rm_sel_defaults["density_kg_m3"], rm_sel_defaults["specific_heat_j_kgk"], checkpoint_hours
                            )
                            crow = {"Izolacja": insul_label}
                            for h, t in zip(checkpoint_hours, temps_at_checkpoints):
                                crow[f"Po {h}h"] = f"{t:.1f}°C"
                            cooling_table_rows.append(crow)
                        st.caption(f"Wychładzanie bez dogrzewania, start {t_start_for_curve:.0f}°C, otoczenie "
                                   f"{rm_sel_defaults['ambient_temp_c']:.0f}°C.")
                        st.dataframe(pd.DataFrame(cooling_table_rows), hide_index=True, use_container_width=True)

                        st.markdown("**⚡ Energetyczne KPI**")
                        if rm_sel_defaults["heated"]:
                            heating_power_kw_detail = compute_tank_heating_power_kw(
                                rm_tank["capacity_m3"], rm_sel_defaults["insulation_mm"],
                                rm_sel_defaults["target_temp_c"], rm_sel_defaults["ambient_temp_c"]
                            )
                            st.metric("Moc grzania (stan ustalony)", f"{heating_power_kw_detail:.2f} kW")
                            st.metric("Energia grzania — dziennie", f"{heating_power_kw_detail * 24:.1f} kWh/dzień")
                            st.metric("Energia grzania — miesięcznie", f"{heating_power_kw_detail * 24 * 30.4:,.0f} kWh/mies.")
                        else:
                            # Te same 3 metryki co dla grzanego (z myślnikiem zamiast wartości), żeby
                            # wysokość kolumny nie "rozjeżdżała się" względem porównywanych obok
                            # zbiorników grzanych - inaczej tabele "Strata ciepła" niżej nie są wyrównane.
                            st.metric("Moc grzania (stan ustalony)", "— (niegrzany)")
                            st.metric("Energia grzania — dziennie", "—")
                            st.metric("Energia grzania — miesięcznie", "—")

                        st.markdown("**🌡️ Strata ciepła po 24h [kWh]**")
                        mass_kg_kpi = rm_tank["capacity_m3"] * TANK_SAFETY_FILL * rm_sel_defaults["density_kg_m3"]
                        heat_loss_rows = []
                        for insul_label, insul_mm in [("Brak izolacji", 0), ("Izolacja 50 mm", 50), ("Izolacja 100 mm", 100)]:
                            temp_after_24h = compute_tank_cooling_curve(
                                rm_tank["capacity_m3"], insul_mm, t_start_for_curve, rm_sel_defaults["ambient_temp_c"],
                                rm_sel_defaults["density_kg_m3"], rm_sel_defaults["specific_heat_j_kgk"], [24]
                            )[0]
                            heat_lost_kwh = (mass_kg_kpi * rm_sel_defaults["specific_heat_j_kgk"] * (t_start_for_curve - temp_after_24h)) / 3_600_000.0
                            heat_loss_rows.append({"Izolacja": insul_label, "Strata [kWh]": round(heat_lost_kwh, 1)})
                        st.dataframe(pd.DataFrame(heat_loss_rows), hide_index=True, use_container_width=True)

            # --- Wyniki: policzone TERAZ, z aktualnym trybem pompy z paneli powyżej ---
            rm_results_rows = []
            for rm_tank in confirmed_rm_tanks:
                rm_tag = rm_tank["tag"]
                rm_defaults = st.session_state.rm_tank_tech_details[rm_tag]
                if rm_defaults["pump_mode"] == "Współdzielona (kilka zbiorników)" and rm_defaults["shared_pump_id"] in st.session_state.shared_pumps:
                    shared_cfg = st.session_state.shared_pumps[rm_defaults["shared_pump_id"]]
                    eff_flow, eff_eff = shared_cfg["flow_m3h"], shared_cfg["efficiency"]
                else:
                    eff_flow, eff_eff = rm_defaults["pump_flow_m3h"], rm_defaults["pump_efficiency"]

                zeta_sum_rm = (rm_defaults["count_elbows_90"] * 0.5) + (rm_defaults["count_valves"] * 0.2)
                _, p_bar_rm, power_kw_rm, velocity_rm = compute_hydraulics(
                    eff_flow, rm_defaults["pipe_dn"], rm_defaults["pipe_length_m"], rm_defaults["delta_h_m"],
                    rm_defaults["viscosity_cst"], rm_defaults["density_kg_m3"], zeta_sum_rm, eff_eff
                )
                total_rm_pump_power += power_kw_rm

                heating_power_kw = 0.0
                if rm_defaults["heated"]:
                    heating_power_kw = compute_tank_heating_power_kw(
                        rm_tank["capacity_m3"], rm_defaults["insulation_mm"],
                        rm_defaults["target_temp_c"], rm_defaults["ambient_temp_c"]
                    )
                    total_rm_heating_power += heating_power_kw

                rm_results_rows.append({
                    "Tag": rm_tag, "Surowiec": rm_tank["material"],
                    "📊 Prędkość [m/s]": round(velocity_rm, 2), "📊 Opór [bar]": round(p_bar_rm, 2),
                    "📊 Moc pompy [kW]": round(power_kw_rm, 2), "📊 Moc grzania [kW]": round(heating_power_kw, 2),
                })

            with rm_results_placeholder:
                st.markdown("### 📋 Zbiorniki RM — Wyniki")
                st.dataframe(pd.DataFrame(rm_results_rows), hide_index=True, use_container_width=True)
                st.metric("⚡ Moc pomp RM razem (doliczana do bilansu elektrycznego niżej)", f"{total_rm_pump_power:.2f} kW")
                if total_rm_heating_power > 0:
                    st.metric("🔥 Moc grzania zbiorników RM razem", f"{total_rm_heating_power:.2f} kW")

        st.markdown("---")

        # ============================================================
        # BILANS PARY I ODPOWIETRZENIA (agregacja dla reaktorów typu Smar/Wax)
        # ============================================================
        grease_mixers_p = [
            (m, st.session_state.mixer_tech_advanced_details.get(m["tag"], {}))
            for m in st.session_state.confirmed_mixers
        ]
        grease_mixers_p = [(m, p) for m, p in grease_mixers_p if p.get("process_type") == "Smar/Wax (gotowanie z odparowaniem)"]

        if grease_mixers_p:
            st.markdown("### 🌫️ Bilans Pary i Odpowietrzenie — Zbiorczy Rurociąg Zrzutowy")
            st.caption("Agreguje strumienie masowe pary wpisane powyżej dla reaktorów oznaczonych jako "
                       "'Smar/Wax' i liczy hydraulikę wspólnego rurociągu zrzutowego metodą Darcy-Weisbacha, "
                       "względem progów bezpieczeństwa (ISO 28300 / API 2000).")
            st.warning("⚠️ Wspomaganie wstępnego oszacowania — wymaga przeglądu przez uprawnionego inżyniera "
                       "ds. bezpieczeństwa procesowego (HAZOP/SIL) przed wdrożeniem.")

            c_gv1, c_gv2, c_gv3, c_gv4 = st.columns(4)
            with c_gv1:
                gv_pipe_length = st.number_input("Długość rurociągu zbiorczego (L) [m]:", min_value=0.1, value=10.0, step=0.5, key="tab2_steam_L")
            with c_gv2:
                gv_dn_choice = st.selectbox("Średnica wg DN (lub 'Własna'):", list(STEAM_PIPE_DN_REFERENCE_M.keys()) + ["Własna"], index=0, key="tab2_steam_dn")
                if gv_dn_choice == "Własna":
                    gv_pipe_d = st.number_input("Średnica wewnętrzna (d) [m]:", min_value=0.01, value=0.0889, step=0.001, format="%.4f", key="tab2_steam_d_custom")
                else:
                    gv_pipe_d = STEAM_PIPE_DN_REFERENCE_M[gv_dn_choice]
            with c_gv3:
                gv_lambda = st.number_input("Współczynnik tarcia (λ) [-]:", min_value=0.001, value=0.025, step=0.001, format="%.3f", key="tab2_steam_lambda")
            with c_gv4:
                gv_rho = st.number_input("Gęstość pary (ρ) [kg/m³]:", min_value=0.01, value=0.598, step=0.01, key="tab2_steam_rho")

            c_ge1, c_ge2 = st.columns(2)
            with c_ge1:
                gv_emergency = st.number_input("Maks. strumień zrzutu awaryjnego (PRV) [kg/s]:", min_value=0.0, value=2.0, step=0.1, key="tab2_steam_emergency")
                gv_nastawa = st.number_input("Nastawa zaworu bezpieczeństwa [bar]:", min_value=0.1, value=10.0, step=0.5, key="tab2_steam_nastawa")
            with c_ge2:
                gv_prog_ab = st.number_input("Próg prędkości - scenariusze A/B [m/s]:", min_value=1.0, value=35.0, step=1.0, key="tab2_steam_prog_ab")
                gv_prog_c = st.number_input("Próg prędkości - scenariusz C [m/s]:", min_value=1.0, value=60.0, step=1.0, key="tab2_steam_prog_c")

            sum_avg_flow = sum(p.get("steam_avg_flow", 0.0) for _, p in grease_mixers_p)
            max_single_decompress = max((p.get("steam_max_decompress", 0.0) for _, p in grease_mixers_p), default=0.0)
            sum_max_process = sum(p.get("steam_max_process", 0.0) for _, p in grease_mixers_p)
            prog_dp_bar_gv = gv_nastawa * 0.10

            gv_scenarios = [
                {"name": f"A. Normalna praca średnia ({len(grease_mixers_p)} reaktor(y) Smar/Wax)", "mass_flow": sum_avg_flow, "check": "velocity", "prog": gv_prog_ab,
                 "msg_bad": "ZAGROŻENIE: Rura za ciasna nawet dla wartości średnich!",
                 "msg_ok": "Teoretycznie bezpiecznie, ale ignoruje szczyty chwilowe (dm/dt)."},
                {"name": "B. Szczyt odparowania (najgorszy pojedynczy reaktor, faza Flash)", "mass_flow": max_single_decompress, "check": "velocity", "prog": gv_prog_ab,
                 "msg_bad": "OSTRZEŻENIE: Pojedynczy zrzut dławi rurę. Spadek efektu flash.",
                 "msg_ok": "Prędkość optymalna dla jednej maszyny."},
                {"name": "C. Najgorszy szczyt roboczy (suma maks. procesowych, koincydencja)", "mass_flow": sum_max_process, "check": "velocity", "prog": gv_prog_c,
                 "msg_bad": f"KRYTYCZNE DŁAWIENIE! Prędkość > {gv_prog_c:.0f} m/s. Blokada zrzutu pary, ryzyko cofki produktu! Rozważ większą średnicę.",
                 "msg_ok": "ZGODNIE Z NORMĄ (15-30 m/s): Średnica dobrana poprawnie pod szczyt obciążenia."},
                {"name": f"D. Scenariusz awaryjny (otwarcie zaworu - {gv_nastawa:.0f} bar)", "mass_flow": gv_emergency, "check": "pressure", "prog": prog_dp_bar_gv,
                 "msg_bad": f"KATASTROFA API 2000! Opory przekraczają 10% nastawy ({prog_dp_bar_gv:.2f} bar). Zawór ulegnie ZABLOKOWANIU!",
                 "msg_ok": "ZGODNIE Z API 2000: Ciśnienie wsteczne poniżej 10% nastawy. Zawór zadziała prawidłowo."},
            ]

            gv_result_rows = []
            for sc in gv_scenarios:
                vol_flow, velocity, dp_pa, dp_bar = compute_vent_line_scenario(sc["mass_flow"], gv_rho, gv_pipe_length, gv_pipe_d, gv_lambda)
                is_bad = (velocity > sc["prog"]) if sc["check"] == "velocity" else (dp_bar > sc["prog"])
                gv_result_rows.append({
                    "Scenariusz Pracy Instalacji": sc["name"], "Masa pary [kg/s]": round(sc["mass_flow"], 4),
                    "Prędkość pary [m/s]": round(velocity, 2), "Opory liniowe [bar]": round(dp_bar, 4),
                    "Weryfikacja": sc["msg_bad"] if is_bad else sc["msg_ok"], "_is_bad": is_bad,
                })

            df_gv = pd.DataFrame(gv_result_rows)

            def style_gv_alerts(df_data):
                sm = pd.DataFrame('', index=df_data.index, columns=df_data.columns)
                for idx in df_data.index:
                    if df_gv.loc[idx, "_is_bad"]:
                        for col in ["Prędkość pary [m/s]", "Opory liniowe [bar]", "Weryfikacja"]:
                            sm.loc[idx, col] = 'background-color: #FFC7CE; color: #9C0006; font-weight: bold;'
                    else:
                        sm.loc[idx, "Weryfikacja"] = 'background-color: #C6EFCE; color: #006100;'
                return sm

            st.dataframe(df_gv.drop(columns=["_is_bad"]).style.apply(style_gv_alerts, axis=None), hide_index=True, use_container_width=True)
            n_gv_alerts = int(df_gv["_is_bad"].sum())
            if n_gv_alerts > 0:
                st.error(f"🔴 {n_gv_alerts} z {len(df_gv)} scenariuszy przekracza próg bezpieczeństwa.")
            else:
                st.success("🟢 Wszystkie scenariusze mieszczą się w przyjętych progach bezpieczeństwa.")

            st.markdown("---")

        # ============================================================
        # DOBÓR KOTŁA GRZEWCZEGO (agregacja zapotrzebowania cieplnego floty)
        # ============================================================
        st.markdown("### 🔥 Dobór Kotła Grzewczego i Instalacji Grzewczej")
        st.caption("Sumuje moc grzania I przepływ medium grzewczego (ta zakładka) po wszystkich mieszalnikach — moc dobiera kocioł, "
                   "przepływ dobiera pompy obiegowe i średnicę rurociągu rozdzielczego.")

        heat_by_medium = {}
        for m in st.session_state.confirmed_mixers:
            ct = st.session_state.calculated_times.get(m["tag"])
            if ct is None:
                continue
            medium = ct.get("medium_grz", "Woda technologiczna")
            heat_by_medium.setdefault(medium, {"power_kw": 0.0, "flow_kg_h": 0.0, "is_steam": ct.get("is_steam", False),
                                                "daily_energy_kwh": 0.0})
            heat_by_medium[medium]["power_kw"] += ct.get("power_heating_kw", 0.0)
            heat_by_medium[medium]["flow_kg_h"] += ct.get("flow_heating_kg_h", 0.0)
            batches_per_day = m["batches_count"] / (WORKING_DAYS_YEAR / MONTHS_PER_YEAR)
            heat_by_medium[medium]["daily_energy_kwh"] += ct.get("power_heating_kw", 0.0) * ct.get("heating", 0.0) * batches_per_day

        if not heat_by_medium:
            st.info("ℹ️ Skonfiguruj bilans cieplny mieszalników powyżej, aby dobrać kocioł.")
            total_heating_power_installed = 0.0
        else:
            c_b1, c_b2 = st.columns(2)
            with c_b1:
                wspolczynnik_jednoczesnosci_cieplny = st.slider(
                    "Współczynnik jednoczesności grzania [%]:", min_value=20, max_value=100, value=70,
                    help="Odsetek zainstalowanej mocy grzania, który realnie występuje jednocześnie w szczycie "
                         "(rzadko wszystkie reaktory grzeją w tej samej minucie). Używany, gdy NIE stosujemy bufora ciepła.") / 100.0
            with c_b2:
                margines_kotla = st.slider("Margines bezpieczeństwa kotła [%]:", min_value=0, max_value=50, value=20) / 100.0

            st.markdown("##### 🛢️ Bufor Ciepła (Zbiornik Akumulacyjny)")
            st.caption("Bufor ciepła gromadzi energię w spokojniejszych okresach i oddaje ją w szczycie — dzięki temu kocioł nie musi "
                       "być dobrany na chwilowy szczyt mocy, tylko na uśrednione, ciągłe zapotrzebowanie w ciągu dnia (mniejszy, tańszy "
                       "kocioł pracujący w sposób ciągły zamiast dużego kotła 'z doskoku').")
            stosuj_bufor = st.checkbox("Zastosuj bufor ciepła w instalacji", value=False, key="stosuj_bufor_cieplny")
            if stosuj_bufor:
                c_buf1, c_buf2 = st.columns(2)
                with c_buf1:
                    autonomia_bufora_h = st.number_input(
                        "Docelowa autonomia bufora [h]:", min_value=0.25, value=1.0, step=0.25, key="autonomia_bufora",
                        help="Ile godzin różnicy między szczytem a średnią ma pokryć bufor — typowo 0.5-2h dla procesów wsadowych."
                    )
                with c_buf2:
                    delta_t_bufora = st.number_input(
                        "Użyteczny wahań ΔT bufora [K]:", min_value=5.0, value=20.0, step=1.0, key="dt_bufora",
                        help="Różnica między maks. a min. temperaturą wody w zbiorniku akumulacyjnym, jaką realnie wykorzystujemy."
                    )

            boiler_rows = []
            total_heating_power_installed = 0.0
            total_buffer_kwh_needed = 0.0
            for medium, data in heat_by_medium.items():
                installed = data["power_kw"]
                total_heating_power_installed += installed
                needed_peak = installed * wspolczynnik_jednoczesnosci_cieplny * (1 + margines_kotla)

                if stosuj_bufor:
                    needed_avg = (data["daily_energy_kwh"] / godziny_dziennie) * (1 + margines_kotla) if godziny_dziennie > 0 else needed_peak
                    needed = min(needed_avg, needed_peak)  # bufor nigdy nie każe dobierać WIĘKSZEGO kotła niż szczyt
                    buffer_kwh = max(needed_peak - needed, 0.0) * autonomia_bufora_h
                    total_buffer_kwh_needed += buffer_kwh
                else:
                    needed = needed_peak

                STANDARD_BOILER_SIZES_KW = [50, 75, 100, 150, 200, 250, 300, 400, 500, 600, 800, 1000, 1250, 1600, 2000]
                recommended = next((s for s in STANDARD_BOILER_SIZES_KW if s >= needed), needed)
                row = {
                    "Medium": medium,
                    "Moc zainstalowana (szczyt) [kW]": round(installed, 1),
                    "Moc wymagana [kW]": round(needed, 1),
                    "Zalecana moc kotła [kW]": round(recommended, 1),
                    "Przepływ medium (szczyt) [kg/h]": round(data["flow_kg_h"], 1),
                }
                if data["is_steam"]:
                    row["Wydajność pary [kg/h]"] = round(recommended * 3600 / STEAM_LATENT_HEAT_KJKG, 1)
                boiler_rows.append(row)
            st.dataframe(pd.DataFrame(boiler_rows), hide_index=True, use_container_width=True)

            if stosuj_bufor and total_buffer_kwh_needed > 0:
                buffer_volume_m3 = (total_buffer_kwh_needed * 3600) / (4.19 * 1000 * delta_t_bufora)
                m_buf1, m_buf2 = st.columns(2)
                with m_buf1:
                    st.metric("🛢️ Szacowana wymagana pojemność cieplna bufora", f"{total_buffer_kwh_needed:.1f} kWh")
                with m_buf2:
                    st.metric("📐 Orientacyjna objętość zbiornika (woda)", f"{buffer_volume_m3:.1f} m³")
                st.caption("⚠️ To uproszczone oszacowanie (różnica szczyt/średnia × zadeklarowana autonomia), nie pełna symulacja "
                           "profilu obciążenia w czasie — realny dobór bufora warto zweryfikować analizą dynamiczną, szczególnie "
                           "przy nierównomiernym harmonogramie szarż.")

        st.markdown("##### ⛽ Typ Kotła i Koszt Paliwa")
        c_f1, c_f2, c_f3 = st.columns(3)
        with c_f1:
            typ_kotla = st.radio("Typ kotła centralnego:", ["Gazowy", "Elektryczny"], horizontal=True, key="typ_kotla")
        with c_f2:
            if typ_kotla == "Gazowy":
                cena_gazu_mwh = st.number_input("Cena gazu [PLN/MWh]:", min_value=1.0, value=250.0, key="cena_gazu")
                sprawnosc_kotla = st.number_input("Sprawność kotła gazowego [%]:", min_value=50.0, max_value=100.0, value=90.0, key="sprawnosc_gaz") / 100.0
            else:
                cena_gazu_mwh = None
                sprawnosc_kotla = st.number_input("Sprawność kotła elektrycznego [%]:", min_value=50.0, max_value=100.0, value=98.0, key="sprawnosc_el") / 100.0
        with c_f3:
            st.metric("Moc zainstalowana grzania (suma floty)", f"{total_heating_power_installed:.1f} kW")

        # Miesięczny koszt paliwa grzewczego — energia użyteczna (Q grzania) / sprawność kotła.
        total_heating_energy_mwh_month = 0.0
        for m in st.session_state.confirmed_mixers:
            ct = st.session_state.calculated_times.get(m["tag"])
            if ct is None:
                continue
            energy_kwh_batch = ct.get("power_heating_kw", 0.0) * ct.get("heating", 0.0)
            total_heating_energy_mwh_month += (energy_kwh_batch * m["batches_count"]) / 1000.0

        fuel_price_for_calc = cena_gazu_mwh if typ_kotla == "Gazowy" else st.session_state.get("cena_mwh_tab4", 750.0)
        koszt_paliwa_grzewczego_month = (total_heating_energy_mwh_month / sprawnosc_kotla) * fuel_price_for_calc if sprawnosc_kotla > 0 else 0.0

        st.session_state["koszt_paliwa_grzewczego_month"] = koszt_paliwa_grzewczego_month
        st.session_state["boiler_capacity_installed_kw"] = total_heating_power_installed
        st.session_state["sprawnosc_kotla_frac"] = sprawnosc_kotla
        st.session_state["cena_gazu_mwh"] = cena_gazu_mwh

        st.caption(f"Szacowany miesięczny koszt paliwa grzewczego: **{koszt_paliwa_grzewczego_month:,.2f}** "
                   f"({'gaz' if typ_kotla=='Gazowy' else 'energia elektryczna'}) — pozycja ta trafia teraz do "
                   f"Zakładki 5 (Analiza Finansowa) jako osobny koszt.")

        st.markdown("---")

        # ============================================================
        # DOBÓR UKŁADU CHŁODZENIA (moc + przepływ chłodziwa floty)
        # ============================================================
        st.markdown("### ❄️ Dobór Układu Chłodzenia")
        st.caption("Analogicznie do kotła grzewczego — suma mocy chłodzenia i przepływu chłodziwa po wszystkich mieszalnikach, "
                   "do doboru wydajności agregatu/wieży chłodniczej oraz pomp i rurociągu obiegu chłodzącego. Zapotrzebowanie "
                   "elektryczne chłodzenia (przez COP) jest już liczone osobno w sekcji ⚡ poniżej.")

        cool_by_medium = {}
        for m in st.session_state.confirmed_mixers:
            ct = st.session_state.calculated_times.get(m["tag"])
            if ct is None:
                continue
            medium = ct.get("medium_chl", "Woda technologiczna")
            cool_by_medium.setdefault(medium, {"power_kw": 0.0, "flow_kg_h": 0.0})
            cool_by_medium[medium]["power_kw"] += ct.get("power_cooling_kw", 0.0)
            cool_by_medium[medium]["flow_kg_h"] += ct.get("flow_cooling_kg_h", 0.0)

        if not cool_by_medium:
            st.info("ℹ️ Skonfiguruj bilans cieplny mieszalników powyżej, aby dobrać układ chłodzenia.")
        else:
            wspolczynnik_jednoczesnosci_chlodzenia = st.slider(
                "Współczynnik jednoczesności chłodzenia [%]:", min_value=20, max_value=100, value=70,
                help="Odsetek zainstalowanej mocy chłodzenia, który realnie występuje jednocześnie w szczycie.",
                key="wsp_jednocz_chlodzenie") / 100.0

            cooling_rows = []
            for medium, data in cool_by_medium.items():
                installed_cool = data["power_kw"]
                needed_cool = installed_cool * wspolczynnik_jednoczesnosci_chlodzenia
                cooling_rows.append({
                    "Medium": medium,
                    "Moc zainstalowana (szczyt) [kW]": round(installed_cool, 1),
                    "Moc wymagana (ze wsp. jednocz.) [kW]": round(needed_cool, 1),
                    "Przepływ chłodziwa (szczyt) [kg/h]": round(data["flow_kg_h"], 1),
                })
            st.dataframe(pd.DataFrame(cooling_rows), hide_index=True, use_container_width=True)

        st.markdown("---")


        st.markdown("---")
        st.markdown("### ⚡ Zapotrzebowanie na Moc Elektryczną i Dobór Transformatora")
        st.caption("Sumuje moc silników mieszadeł i pomp z floty (w tym pomp zbiorników RM powyżej), ewentualny "
                   "kocioł elektryczny oraz (opcjonalnie) szacunkowe zapotrzebowanie elektryczne chłodzenia (przez "
                   "współczynnik COP), z uwzględnieniem współczynnika jednoczesności i mocy transformatora w kVA.")

        total_mix_power = sum(st.session_state.calculated_times.get(m["tag"], {}).get("power_mix_kw", 0.0)
                               for m in st.session_state.confirmed_mixers)
        total_pump_power = sum(st.session_state.calculated_times.get(m["tag"], {}).get("power_pump_kw", 0.0)
                                for m in st.session_state.confirmed_mixers) + total_rm_pump_power
        total_cooling_duty = sum(st.session_state.calculated_times.get(m["tag"], {}).get("power_cooling_kw", 0.0)
                                  for m in st.session_state.confirmed_mixers)

        c_e1, c_e2, c_e3 = st.columns(3)
        with c_e1:
            uwzglednij_chlodzenie_el = st.checkbox("Uwzględnij elektryczne chłodzenie (agregaty/chillery)", value=True)
            cop_chlodzenia = st.number_input("COP chłodzenia [-]:", min_value=1.0, value=3.0,
                                              disabled=not uwzglednij_chlodzenie_el,
                                              help="Współczynnik wydajności chłodniczej — moc elektryczna = moc chłodzenia / COP.")
        with c_e2:
            wspolczynnik_jednoczesnosci_el = st.slider("Współczynnik jednoczesności elektrycznej [%]:",
                                                         min_value=20, max_value=100, value=65) / 100.0
            cos_phi = st.number_input("cos φ (współczynnik mocy):", min_value=0.5, max_value=1.0, value=0.90, step=0.01)
        with c_e3:
            margines_transformatora = st.slider("Margines bezpieczeństwa transformatora [%]:",
                                                  min_value=0, max_value=100, value=25) / 100.0

        electric_boiler_load = total_heating_power_installed if typ_kotla == "Elektryczny" else 0.0
        cooling_electric_load = (total_cooling_duty / cop_chlodzenia) if uwzglednij_chlodzenie_el and cop_chlodzenia > 0 else 0.0

        st.markdown("---")
        st.markdown("##### 🏢 Odbiory Pozaprodukcyjne (Serwery, Oświetlenie, Sprężarkownia, Inne)")
        st.caption("Odbiory niezwiązane bezpośrednio z reaktorami/pompami floty, ale realnie obciążające transformator "
                   "zakładu — zwykle pracują z wyższym współczynnikiem jednoczesności niż urządzenia procesowe "
                   "wsadowe (są 'włączone' niemal cały czas), więc liczone są z osobnym współczynnikiem poniżej.")

        c_f_1, c_f_2, c_f_3 = st.columns(3)
        with c_f_1:
            serwery_kw = st.number_input("Serwery / IT [kW]:", min_value=0.0, value=5.0, step=0.5, key="fac_servers_kw")
            hvac_kw = st.number_input("Wentylacja / HVAC (poza chłodzeniem procesowym) [kW]:", min_value=0.0, value=15.0, step=1.0, key="fac_hvac_kw")
        with c_f_2:
            powierzchnia_zakladu_m2 = st.number_input("Powierzchnia zakładu do oświetlenia [m²]:", min_value=0.0, value=2000.0, step=100.0, key="fac_area_m2")
            wskaznik_oswietlenia_w_m2 = st.number_input("Wskaźnik mocy oświetlenia [W/m²]:", min_value=1.0, value=8.0, step=0.5, key="fac_light_wm2",
                                                          help="Typowo 6-10 W/m² dla oświetlenia LED w halach przemysłowych.")
            oswietlenie_kw = (powierzchnia_zakladu_m2 * wskaznik_oswietlenia_w_m2) / 1000.0
            st.caption(f"Wyliczona moc oświetlenia: **{oswietlenie_kw:.1f} kW**")
        with c_f_3:
            sprezarkownia_kw = st.number_input("Sprężarkownia (sprężone powietrze) [kW]:", min_value=0.0, value=45.0, step=5.0, key="fac_compressed_air_kw")
            inne_odbiory_kw = st.number_input("Inne (biura, ładowarki wózków, warsztat UR) [kW]:", min_value=0.0, value=20.0, step=5.0, key="fac_other_kw")

        total_facility_load_kw = serwery_kw + hvac_kw + oswietlenie_kw + sprezarkownia_kw + inne_odbiory_kw
        wspolczynnik_jednoczesnosci_facility = st.slider(
            "Współczynnik jednoczesności odbiorów pozaprodukcyjnych [%]:", min_value=50, max_value=100, value=90,
            help="Serwery/oświetlenie/sprężarkownia pracują niemal ciągle, więc ten współczynnik jest zwykle wyższy niż dla floty procesowej.",
            key="fac_wsp_jednocz") / 100.0

        st.markdown("---")

        installed_electric_load_kw = total_mix_power + total_pump_power + electric_boiler_load + cooling_electric_load
        process_demand_kw = installed_electric_load_kw * wspolczynnik_jednoczesnosci_el
        facility_demand_kw = total_facility_load_kw * wspolczynnik_jednoczesnosci_facility
        demand_kw = process_demand_kw + facility_demand_kw
        demand_kva = (demand_kw / cos_phi) * (1 + margines_transformatora) if cos_phi > 0 else 0.0

        STANDARD_TRANSFORMER_SIZES_KVA = [100, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500]
        recommended_transformer = next((s for s in STANDARD_TRANSFORMER_SIZES_KVA if s >= demand_kva), demand_kva)

        load_rows = [
            {"Odbiornik": "Silniki mieszadeł (suma floty)", "Moc [kW]": round(total_mix_power, 1), "Kategoria": "Proces"},
            {"Odbiornik": "Silniki pomp (suma floty)", "Moc [kW]": round(total_pump_power, 1), "Kategoria": "Proces"},
            {"Odbiornik": "Kocioł elektryczny" if typ_kotla == "Elektryczny" else "Kocioł (gazowy — brak obciążenia elektrycznego)",
             "Moc [kW]": round(electric_boiler_load, 1), "Kategoria": "Proces"},
            {"Odbiornik": "Chłodzenie (agregaty, przez COP)" if uwzglednij_chlodzenie_el else "Chłodzenie (nieuwzględnione)",
             "Moc [kW]": round(cooling_electric_load, 1), "Kategoria": "Proces"},
            {"Odbiornik": "Serwery / IT", "Moc [kW]": round(serwery_kw, 1), "Kategoria": "Pozaprodukcyjne"},
            {"Odbiornik": "Wentylacja / HVAC", "Moc [kW]": round(hvac_kw, 1), "Kategoria": "Pozaprodukcyjne"},
            {"Odbiornik": "Oświetlenie zakładu", "Moc [kW]": round(oswietlenie_kw, 1), "Kategoria": "Pozaprodukcyjne"},
            {"Odbiornik": "Sprężarkownia", "Moc [kW]": round(sprezarkownia_kw, 1), "Kategoria": "Pozaprodukcyjne"},
            {"Odbiornik": "Inne (biura, ładowarki, UR)", "Moc [kW]": round(inne_odbiory_kw, 1), "Kategoria": "Pozaprodukcyjne"},
        ]
        st.dataframe(pd.DataFrame(load_rows), hide_index=True, use_container_width=True)

        m_e0a, m_e0b = st.columns(2)
        with m_e0a:
            st.metric("⚙️ Moc szczytowa — proces (ze wsp. jednocz.)", f"{process_demand_kw:.1f} kW")
        with m_e0b:
            st.metric("🏢 Moc szczytowa — pozaprodukcyjne (ze wsp. jednocz.)", f"{facility_demand_kw:.1f} kW")

        m_e1, m_e2, m_e3, m_e4 = st.columns(4)
        with m_e1: st.metric("Moc zainstalowana (razem)", f"{installed_electric_load_kw + total_facility_load_kw:.1f} kW")
        with m_e2: st.metric("Moc szczytowa (razem)", f"{demand_kw:.1f} kW")
        with m_e3: st.metric("Moc pozorna wymagana", f"{demand_kva:.1f} kVA")
        with m_e4: st.metric("🔌 Zalecany transformator", f"{recommended_transformer:.0f} kVA")

        # Zapis do session_state, żeby połączona Zakładka Analiza Finansowa mogła doliczyć
        # PEŁNY koszt energii elektrycznej (proces + odbiory pozaprodukcyjne), a nie tylko
        # mieszanie/pompowanie, jak dotychczas. Proces i odbiory pozaprodukcyjne zapisywane
        # OSOBNO, bo w symulacji rozruchu (ROI) tylko obciążenie procesowe skaluje się z
        # wolumenem produkcji — serwery/HVAC/oświetlenie/sprężarkownia działają praktycznie
        # niezależnie od tego, ile realnie produkujesz w danym roku.
        st.session_state["demand_kw_total"] = demand_kw
        st.session_state["process_demand_kw"] = process_demand_kw
        st.session_state["facility_demand_kw"] = facility_demand_kw

        st.caption("Uwaga: sekcja rozlewu/pakowania nie ma dziś modelowanych mocy silników (tylko wydajność kg/min), "
                   "więc nie jest tu ujęta — jeśli chcesz uwzględnić linie napełniające w bilansie elektrycznym, "
                   "podaj ich orientacyjną moc zainstalowaną do doliczenia ręcznego.")

        if not df_filtered.empty:
            csv_buffer = io.StringIO()
            df_filtered.to_csv(csv_buffer, index=False, sep=";")

            st.download_button(
                label="📊 Pobierz pełny raport procesowy (Format CSV)",
                data=csv_buffer.getvalue().encode("utf-8-sig"),  # BOM, żeby Excel poprawnie czytał polskie znaki
                file_name="Fuchs_Pelny_Model_Hydrauliczno_Procesowy.csv",
                mime="text/csv",
                use_container_width=True,
                key="btn_download_final_csv_v13"
            )

# ==========================================
# ZAKŁADKA 4: LOGISTYKA I OPALETOWANIE (tab3)
# ==========================================
with tab3:
    st.header("📦 Analiza Logistyczna, Czas Rozlewu i Gospodarka Paletowa")
    if not st.session_state.confirmed_mixers:
        st.info("💡 Najpierw zatwierdź konfigurację floty w Zakładce 1.")
    else:
        _stale_msg = check_fleet_staleness_warning()
        if _stale_msg:
            st.error(_stale_msg)
        mixers_fleet = st.session_state.confirmed_mixers

        rampup_year_options = ["Docelowa produkcja (100%)"] + [f"Rok {i+1}" for i in range(RAMPUP_YEARS)]
        selected_rampup_year_label = st.selectbox(
            "📈 Rok symulacji rozruchu:", rampup_year_options, index=0, key="tab3_rampup_year_select",
            help="Magazyn i flota są budowane od razu pod docelową (100%) produkcję — ten wybór tylko pokazuje, "
                 "jaki wolumen FG i RM realnie przepłynie przez fabrykę w danym roku rozruchu (krzywa z Zakładki 2), "
                 "i ile z gotowej powierzchni magazynowej to zajmie."
        )
        selected_rampup_year_idx = None if selected_rampup_year_label.startswith("Docelowa") else int(selected_rampup_year_label.split(" ")[1]) - 1

        st.markdown("---")

        st.markdown("##### 📦 Konfiguracja Opakowań i Głowic Rozlewniczych")
        st.caption("Jedna tabela na typ opakowania: pojemność i sztuk/paletę (domyślnie wbudowane, nadpisywane przez "
                   "opcjonalny arkusz 'Opakowania' z Zakładki 1) razem z parametrami głowic nalewaka tego opakowania. "
                   "Dodaj/usuń wiersz, aby dodać/usunąć typ opakowania.")

        if "filling_lines_config" not in st.session_state:
            st.session_state.filling_lines_config = {}
        for p in st.session_state.pack_configs.keys():
            if p not in st.session_state.filling_lines_config:
                st.session_state.filling_lines_config[p] = default_filling_line_config(p)

        pack_fill_editor_rows = [
            {"Nazwa Opakowania": name, "Pojemność [L]": cfg["size_l"], "Sztuk na Palecie": cfg["per_pallet"],
             "Głowice nalewaka [szt]": int(st.session_state.filling_lines_config.get(name, {"nozzles": 1})["nozzles"]),
             "Wydajność 1 głowicy [kg/min]": float(st.session_state.filling_lines_config.get(name, {"speed_kg_min": default_filling_speed_kg_min(name)})["speed_kg_min"])}
            for name, cfg in st.session_state.pack_configs.items()
        ]
        edited_pack_fill_df = st.data_editor(
            pd.DataFrame(pack_fill_editor_rows), hide_index=True, use_container_width=True,
            num_rows="dynamic", key="pack_fill_editor"
        )
        new_pack_configs = {}
        new_filling_config = {}
        for _, row in edited_pack_fill_df.iterrows():
            name = str(row["Nazwa Opakowania"])
            if not name or pd.isna(row["Pojemność [L]"]) or pd.isna(row["Sztuk na Palecie"]):
                continue
            existing = st.session_state.pack_configs.get(name, {"rate_szt_h": 0})
            new_pack_configs[name] = {
                "size_l": float(row["Pojemność [L]"]),
                "per_pallet": int(row["Sztuk na Palecie"]),
                "rate_szt_h": existing.get("rate_szt_h", 0),
            }
            new_filling_config[name] = {
                "nozzles": float(row["Głowice nalewaka [szt]"]) if pd.notna(row["Głowice nalewaka [szt]"]) else 1.0,
                "speed_kg_min": float(row["Wydajność 1 głowicy [kg/min]"]) if pd.notna(row["Wydajność 1 głowicy [kg/min]"]) else 30.0,
            }
        st.session_state.pack_configs = new_pack_configs
        st.session_state.filling_lines_config = new_filling_config

        aktywne_opakowania = set(st.session_state.pack_configs.keys())

        with st.expander("🧴 Rozbicie na Opakowania (ręczne, per linia) — używane, gdy receptura go nie określa", expanded=False):
            st.caption("Jeśli produkt ma w recepturze (Zakładka 1) wypełnione % opakowań, apka użyje ich wprost. "
                       "Dla pozostałych produktów obowiązuje rozbicie z tej tabeli, per linia produktowa.")

            opakowania_podzial = st.session_state.setdefault("opakowania_podzial", {})
            split_editor_rows = []
            for kat in wybrane_kategorie:
                for p in st.session_state.pack_configs.keys():
                    key_id = f"pct_{kat}_{p}"
                    split_editor_rows.append({"Linia": kat, "Opakowanie": p, "Udział [%]": opakowania_podzial.get(key_id, 0.0)})

            if split_editor_rows:
                edited_split_df = st.data_editor(
                    pd.DataFrame(split_editor_rows), hide_index=True, use_container_width=True,
                    disabled=["Linia", "Opakowanie"], key="opakowania_split_editor"
                )
                for _, row in edited_split_df.iterrows():
                    opakowania_podzial[f"pct_{row['Linia']}_{row['Opakowanie']}"] = float(row["Udział [%]"])

                sum_check = edited_split_df.groupby("Linia")["Udział [%]"].sum()
                bad_lines = sum_check[(sum_check > 0) & ((sum_check - 100).abs() > 0.5)]
                if not bad_lines.empty:
                    st.warning("⚠️ Suma % różni się od 100% dla: " + ", ".join(f"{k} ({v:.0f}%)" for k, v in bad_lines.items()) +
                               " — dotyczy tylko produktów bez własnego rozbicia w recepturze.")
            else:
                st.info("Wybierz aktywne linie produktowe w panelu bocznym, aby skonfigurować ręczny podział opakowań.")

        st.markdown("---")
        czas_skladowania_dni = st.number_input("Czas składowania palety (Rotacja) [dni]:", min_value=1, value=14)
        st.session_state["czas_skladowania_tab3"] = czas_skladowania_dni
        dni_robocze_miesiac = WORKING_DAYS_YEAR / MONTHS_PER_YEAR

        real_split_rows = []
        fg_positions_target_list = []  # miejsca magazynowe FG liczone przy 100% celu (stały rozmiar budynku)
        fg_pallets_target_list = []  # palet/mies. FG liczone przy 100% celu - do wykresu wysyłek na 5 lat
        recipes_df_lookup = st.session_state.get("recipes_df")
        pack_cols_in_recipe = []
        if recipes_df_lookup is not None and not recipes_df_lookup.empty:
            pack_cols_in_recipe = [c for c in recipes_df_lookup.columns if c.startswith("Opak: ") and c.endswith(" [%]")]

        effective_year_idx_for_import = selected_rampup_year_idx if selected_rampup_year_idx is not None else RAMPUP_YEAR_TARGET_SENTINEL
        products_imported_this_view = set()  # produkty (recipe_product) importowane w wybranym roku/widoku

        for m in mixers_fleet:
            kat = m["product_family"]
            recipe_product = m.get("recipe_product")

            # Produkt z DEDYKOWANYM zbiornikiem (recipe_product ustawiony), który w wybranym
            # roku/widoku jest jeszcze importowany (nie osiągnął swojego roku przejścia) - jego
            # BIEŻĄCA (ten rok) produkcja to 0, a zapotrzebowanie na magazyn liczy się osobno w
            # sekcji "Import" poniżej, z rytmu dostaw. Budynek (docelowe 100%) i tak MUSI
            # uwzględnić jego docelową (już produkowaną) wielkość, więc target liczy się zawsze,
            # niezależnie od statusu importu w wybranym roku. Zbiorniki współdzielone (bez
            # pojedynczego recipe_product) nie są dziś rozbijane per produkt - traktowane jak
            # dotychczas, w całości jako produkcja.
            is_imported_this_view = False
            if recipe_product and recipes_df_lookup is not None and RECIPE_SOURCING_COL in recipes_df_lookup.columns:
                match_src = recipes_df_lookup[recipes_df_lookup[RECIPE_PRODUCT_COL] == recipe_product]
                if not match_src.empty:
                    src_row = match_src.iloc[0]
                    if is_product_imported_in_year(src_row.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                                    src_row.get(RECIPE_IMPORT_TRANSITION_COL, ""),
                                                    effective_year_idx_for_import):
                        is_imported_this_view = True
                        products_imported_this_view.add(recipe_product)

            rampup_frac_fg = get_rampup_fraction(kat, selected_rampup_year_idx) if selected_rampup_year_idx is not None else 1.0
            target_monthly_mass = m["batches_count"] * m["mass_per_batch"]
            mixer_monthly_mass = 0.0 if is_imported_this_view else target_monthly_mass * rampup_frac_fg
            rho_linii = st.session_state.active_portfolio[kat]["density"]

            # Priorytet 1: rozbicie na opakowania WPROST z receptury tego konkretnego produktu
            # (Zakładka 1), jeśli podano i sumuje się w przybliżeniu do 100%.
            recipe_split = None
            if recipe_product and pack_cols_in_recipe:
                match = recipes_df_lookup[recipes_df_lookup[RECIPE_PRODUCT_COL] == recipe_product]
                if not match.empty:
                    row0 = match.iloc[0]
                    pack_sum = sum(row0.get(c, 0) or 0 for c in pack_cols_in_recipe)
                    if pack_sum > 0.5:
                        recipe_split = {c[len("Opak: "):-len(" [%]")]: (row0.get(c, 0) or 0) for c in pack_cols_in_recipe
                                         if (row0.get(c, 0) or 0) > 0}

            if recipe_split is not None:
                split_source = "receptura"
                pack_pcts = recipe_split
            else:
                # Priorytet 2 (fallback): ręczny podział per linia z tabeli powyżej.
                split_source = "ręczny"
                pack_pcts = {p: opakowania_podzial.get(f"pct_{kat}_{p}", 0.0) for p in st.session_state.pack_configs.keys()}

            # KRYTYCZNE: normalizacja do 100% - jeśli suma % opakowań w źródle (recepturze albo
            # ręcznym podziale) nie wynosi dokładnie 100% (np. wpisano tylko 67%, reszta pusta),
            # BEZ tej normalizacji brakująca masa po prostu ZNIKAŁA z obliczeń (nigdy nie trafiała
            # do żadnego opakowania) - produkcja/magazyn FG był wtedy cicho zaniżony. Skalujemy
            # proporcjonalnie, żeby cała masa zawsze się rozliczyła, i zgłaszamy niezgodność do
            # ostrzeżenia zbiorczego niżej, żeby dało się poprawić dane źródłowe.
            pack_pcts_sum = sum(pack_pcts.values())
            if pack_pcts_sum > 0.5 and abs(pack_pcts_sum - 100.0) > 2.0:
                packaging_mismatch_warnings.append(
                    f"{recipe_product or kat} ({split_source}): suma % opakowań = {pack_pcts_sum:.0f}% zamiast 100%"
                )
                pack_pcts = {p: v * (100.0 / pack_pcts_sum) for p, v in pack_pcts.items()}

            for p, udzial_pct in pack_pcts.items():
                if udzial_pct <= 0 or p not in st.session_state.pack_configs:
                    continue
                masa_opakowania_month = mixer_monthly_mass * (udzial_pct / 100.0)
                pack_capacity_kg = st.session_state.pack_configs[p]["size_l"] * rho_linii
                liczba_sztuk_month = math.ceil(masa_opakowania_month / pack_capacity_kg) if pack_capacity_kg > 0 else 0

                cfg_fill = st.session_state.filling_lines_config.get(p, default_filling_line_config(p))
                sekcja_nalewania_m3_h = (cfg_fill["nozzles"] * cfg_fill["speed_kg_min"] * 60.0) / (rho_linii * 1000.0)

                # Rzeczywisty przepływ pompy TEGO KONKRETNEGO mieszalnika z Zakładki 2 (nie
                # reprezentanta całej grupy jak poprzednio - każdy mieszalnik ma teraz własny wiersz).
                tech_details = st.session_state.get("mixer_tech_advanced_details", {}).get(m["tag"], {})
                q_pump_m3h = tech_details.get("pump_flow_m3h", 15.0)

                q_effective_flow_m3h = min(q_pump_m3h, sekcja_nalewania_m3_h)
                czas_rozlewu_h = (masa_opakowania_month / (rho_linii * 1000.0)) / q_effective_flow_m3h if q_effective_flow_m3h > 0 else 0.0

                is_tanker_pack = st.session_state.pack_configs[p]["per_pallet"] == 0
                if is_tanker_pack:
                    # Cysterna (luzem): "sztuka" = 1 pełny ładunek cysterny - nie ma palet ani
                    # miejsc magazynowych (wysyłane bezpośrednio, bez składowania na paletach).
                    liczba_palet_month = 0
                    miejsca_paletowe = 0
                    liczba_palet_month_target = 0
                    miejsca_paletowe_target = 0
                    cystern_month = liczba_sztuk_month
                else:
                    liczba_palet_month = math.ceil(liczba_sztuk_month / st.session_state.pack_configs[p]["per_pallet"])
                    miejsca_paletowe = math.ceil((liczba_palet_month / dni_robocze_miesiac) * czas_skladowania_dni)
                    cystern_month = 0

                # Miejsca magazynowe przy 100% celu (niezależnie od wybranego roku symulacji) -
                # to jest rozmiar BUDYNKU, który stawia się raz, pod docelową zdolność produkcyjną.
                masa_opakowania_month_target = target_monthly_mass * (udzial_pct / 100.0)
                liczba_sztuk_month_target = math.ceil(masa_opakowania_month_target / pack_capacity_kg) if pack_capacity_kg > 0 else 0
                if not is_tanker_pack:
                    liczba_palet_month_target = math.ceil(liczba_sztuk_month_target / st.session_state.pack_configs[p]["per_pallet"])
                    miejsca_paletowe_target = math.ceil((liczba_palet_month_target / dni_robocze_miesiac) * czas_skladowania_dni)
                fg_positions_target_list.append(miejsca_paletowe_target)
                fg_pallets_target_list.append(liczba_palet_month_target)

                if not is_imported_this_view:
                    real_split_rows.append({
                        "Typ": "FG", "Reaktor 🔒": m["tag"], "Linia 🔒": kat, "Opakowanie 📦": p, "Udział": f"{udzial_pct:.1f}%",
                        "Źródło %": split_source,
                        "Opakowań [/mies]": int(liczba_sztuk_month), "Palet [/mies] 🧱": int(liczba_palet_month),
                        "Cystern [/mies] 🚚": int(cystern_month),
                        "Miejsca magazynowe [szt] 📐": int(miejsca_paletowe), "Czas rozlewu strumienia [h] ⏱️": round(czas_rozlewu_h, 1),
                        "Wąskie gardło": "Pompa" if q_pump_m3h < sekcja_nalewania_m3_h else "Sekcja nalewania",
                        "_masa_kg_miesiac": masa_opakowania_month,
                    })

        st.session_state["logistics_results"] = real_split_rows

        # ============================================================
        # IMPORT — bufor magazynowy dla produktów jeszcze/na stałe importowanych, liczony z
        # rytmu dostaw (częstotliwość + wielkość), a nie z cyklu produkcji. Model zapasu
        # cyklicznego: szczytowy stan magazynu = wielkość 1 dostawy + bufor bezpieczeństwa
        # (bo tuż przed kolejną dostawą bufor bezpieczeństwa wciąż stoi, a nowa dostawa
        # dokłada się na wierzch) - to jest wartość, pod którą trzeba realnie zarezerwować
        # miejsce, nie średnia.
        # ============================================================
        import_pallet_mass_kg = st.session_state.get("import_pallet_mass_kg", 800.0)

        def compute_import_positions(year_idx_for_calc):
            rows, total_positions = [], 0
            if recipes_df_lookup is None or recipes_df_lookup.empty or RECIPE_SOURCING_COL not in recipes_df_lookup.columns:
                return rows, total_positions
            import_rows_df = recipes_df_lookup[
                (recipes_df_lookup[RECIPE_SOURCING_COL] == "Import") &
                (recipes_df_lookup.get(RECIPE_IMPORT_TRANSITION_COL, pd.Series(dtype=str)) != "Nigdy (bufor)")
            ]
            for _, r in import_rows_df.iterrows():
                if not is_product_imported_in_year(r.get(RECIPE_SOURCING_COL, "Produkcja własna"),
                                                     r.get(RECIPE_IMPORT_TRANSITION_COL, ""), year_idx_for_calc):
                    continue
                annual_t_target = float(r.get(RECIPE_ANNUAL_COL, 0) or 0)
                frac = get_import_volume_fraction(r[RECIPE_GROUP_COL], year_idx_for_calc)
                effective_annual_t = annual_t_target * frac
                daily_t = effective_annual_t / WORKING_DAYS_YEAR if WORKING_DAYS_YEAR > 0 else 0.0
                freq_days = float(r.get(RECIPE_IMPORT_FREQ_COL, 0) or 0)
                lot_t = float(r.get(RECIPE_IMPORT_LOT_COL, 0) or 0)
                safety_days = float(r.get(RECIPE_IMPORT_SAFETY_DAYS_COL, 0) or 0)
                safety_stock_t = safety_days * daily_t
                peak_stock_t = lot_t + safety_stock_t
                deliveries_per_year = math.ceil(effective_annual_t / lot_t) if lot_t > 0 else 0
                miejsca_paletowe_import = math.ceil((peak_stock_t * 1000.0) / import_pallet_mass_kg) if import_pallet_mass_kg > 0 else 0
                rows.append({
                    "Produkt": r[RECIPE_PRODUCT_COL], "Linia": r[RECIPE_GROUP_COL],
                    "Tryb": "Stały import" if r.get(RECIPE_IMPORT_TRANSITION_COL, "") == "Nigdy (stały import)" else r.get(RECIPE_IMPORT_TRANSITION_COL, ""),
                    "Wolumen [t/rok]": round(effective_annual_t, 1),
                    "Częstotliwość dostawy [dni]": freq_days if freq_days > 0 else "—",
                    "Wielkość dostawy [t]": lot_t,
                    "Dostaw/rok": deliveries_per_year,
                    "Bufor bezpieczeństwa [t]": round(safety_stock_t, 2),
                    "Szczytowy zapas [t]": round(peak_stock_t, 2),
                    "Miejsca magazynowe [szt]": int(miejsca_paletowe_import),
                })
                total_positions += miejsca_paletowe_import
            # UWAGA: total_positions liczony powyżej wiersz-po-wierszu MUSI dawać identyczny wynik
            # co wspólna funkcja compute_import_positions_for_year (używana też w Dashboardzie) -
            # nadpisujemy nim na wszelki wypadek, żeby te dwa miejsca nigdy nie mogły się rozjechać.
            return rows, compute_import_positions_for_year(year_idx_for_calc, import_pallet_mass_kg)

        import_warehouse_rows_view, total_import_positions = compute_import_positions(effective_year_idx_for_import)
        # Baseline "budynku" (target/100%) uwzględnia TYLKO stały import ("Nigdy") - produkty z
        # rokiem przejścia w widoku docelowym są już produkowane, więc ich miejsce liczy fg_positions_target_list.
        import_warehouse_rows_target, total_import_positions_target = compute_import_positions(RAMPUP_YEAR_TARGET_SENTINEL)

        rm_warehouse_rows = st.session_state.get("raw_material_warehouse_rows", [])

        if real_split_rows or rm_warehouse_rows or import_warehouse_rows_view:
            st.markdown("##### 🔀 Wyniki Symulacji Logistyczno-Magazynowej (Wyroby Gotowe, FG)")
            if real_split_rows:
                st.caption("Kolumna **Źródło %** pokazuje, czy rozbicie na opakowania tego wiersza pochodzi z receptury "
                           "(Zakładka 1, per produkt) czy z ręcznego podziału w panelu bocznym (per grupa, gdy receptura "
                           "nie precyzuje opakowań dla tego produktu). Kolumna **Wąskie gardło** pokazuje, czy czas rozlewu "
                           "jest dziś limitowany przez wydajność pompy TEGO reaktora (Zakładka 2), czy przez sekcję głowic nalewczych.")
                _df_fg_display = pd.DataFrame(real_split_rows)
                st.dataframe(_df_fg_display[[c for c in _df_fg_display.columns if not c.startswith("_")]],
                             hide_index=True, use_container_width=True)
            else:
                st.info("ℹ️ W wybranym roku/widoku wszystkie produkty tej floty są jeszcze importowane (patrz sekcja "
                        "'📦 Import' poniżej) — brak własnej produkcji FG do pokazania.")

            # ============================================================
            # IMPORT — produkty importowane w wybranym roku/widoku
            # ============================================================
            if import_warehouse_rows_view:
                st.markdown("##### 📦 Import — Bufor Magazynowy")
                st.caption("Produkty jeszcze (lub na stałe) importowane w wybranym roku/widoku — bufor liczony z rytmu "
                           "dostaw, nie z cyklu produkcji. **Szczytowy zapas** = wielkość 1 dostawy + bufor bezpieczeństwa "
                           "(bo tuż przed kolejną dostawą bufor bezpieczeństwa wciąż stoi w magazynie).")
                import_pallet_mass_kg = st.number_input(
                    "Masa 1 palety importowej [kg]:", min_value=1.0, value=float(import_pallet_mass_kg), step=50.0,
                    key="import_pallet_mass_kg_input",
                    help="Uproszczenie: jedna wspólna masa/paletę dla wszystkich produktów importowanych, do przeliczenia "
                         "szczytowego zapasu [t] na miejsca magazynowe [szt]."
                )
                st.session_state["import_pallet_mass_kg"] = import_pallet_mass_kg
                import_warehouse_rows_view, total_import_positions = compute_import_positions(effective_year_idx_for_import)
                import_warehouse_rows_target, total_import_positions_target = compute_import_positions(RAMPUP_YEAR_TARGET_SENTINEL)
                st.dataframe(pd.DataFrame(import_warehouse_rows_view), hide_index=True, use_container_width=True)
                st.metric("📦 Miejsca magazynowe — Import (ten rok)", f"{total_import_positions} szt.")
            else:
                # Sprawdzamy OSOBNO, czy "brak importu" wynika z tego, że wszystko, co zostało, to
                # produkty "Nigdy (bufor)" (celowo wykluczone stąd, bo idą do zbiornika w Zakładce 2,
                # nie na paletę) - inaczej komunikat myląco sugerowałby, że import w ogóle nie trwa.
                buffer_products_still_importing = []
                if st.session_state.recipes_df is not None and not st.session_state.recipes_df.empty and RECIPE_IMPORT_TRANSITION_COL in st.session_state.recipes_df.columns:
                    buffer_mask = (st.session_state.recipes_df[RECIPE_SOURCING_COL] == "Import") & \
                                  (st.session_state.recipes_df[RECIPE_IMPORT_TRANSITION_COL] == "Nigdy (bufor)")
                    buffer_products_still_importing = st.session_state.recipes_df.loc[buffer_mask, RECIPE_PRODUCT_COL].tolist()

                if buffer_products_still_importing:
                    st.info("ℹ️ Brak produktów w buforze paletowym importu w wybranym roku/widoku — ale "
                            f"**{', '.join(buffer_products_still_importing)}** to import na stałe ('Nigdy (bufor)'), "
                            "który trwa cały czas i ma własny zbiornik zamiast palety — zobacz sekcję "
                            "'🔵 Zbiorniki buforowe' w **Zakładce 2**.")
                else:
                    st.info("ℹ️ Brak produktów importowanych w wybranym roku/widoku (albo brak wgranych receptur z "
                            "'Sposób Pozyskania' = 'Import' w Zakładce 1).")

            # ============================================================
            # SUROWCE (RM) W BECZKACH/IBC/WORKACH — z Zakładki 4, jeśli policzone
            # ============================================================
            if rm_warehouse_rows:
                with st.expander("🧴 Surowce w Beczkach/IBC/Workach (RM) — z Zakładki 4", expanded=False):
                    st.caption("Surowce nietrafiające do zbiorników (Zakładka 4) — te też stoją w tym samym "
                               "magazynie co wyroby gotowe i wliczają się do łącznej powierzchni poniżej.")
                    st.dataframe(pd.DataFrame(rm_warehouse_rows), hide_index=True, use_container_width=True)
            else:
                st.info("ℹ️ Brak policzonych surowców w beczkach/IBC/workach — wgraj receptury (Zakładka 1) i "
                        "odwiedź Zakładkę 3, aby doliczyć ich miejsca magazynowe do bilansu poniżej.")

            # ============================================================
            # PODSUMOWANIE POWIERZCHNI MAGAZYNOWEJ — FG + RM RAZEM (suma miejsc paletowych -> m²)
            # ============================================================
            st.markdown("##### 📐 Podsumowanie Powierzchni Magazynowej (FG + RM, wspólny magazyn)")
            st.caption("Ten sam magazyn przechowuje zarówno wyroby gotowe (FG), jak i surowce nietrafiające do "
                       "zbiorników (RM) — wszystko, co nie stoi w silosie, musi stanąć tutaj. Powierzchnia na "
                       "jedno miejsce paletowe zależy od typu składowania — regały selektywne wymagają więcej "
                       "przestrzeni na alejki niż składowanie blokowe, ale za to (razem z liczbą poziomów poniżej) "
                       "pozwalają postawić więcej palet na tej samej powierzchni podłogi.")

            RACKING_PRESETS_AISLE_PCT = {
                # % powierzchni "straconej" na alejki i drogi transportowe względem netto powierzchni
                # palet - szacunek orientacyjny, wynikający z tego, jak gęsto dany typ składowania
                # pozwala ustawiać palety (regały selektywne potrzebują szerokich alejek na wjazd
                # wózka do KAŻDEGO rzędu; drive-in i block stacking - znacznie mniej, bo wózek
                # wjeżdża rzadziej / wcale nie wjeżdża między palety).
                "Składowanie blokowe (block stacking)": 8.0,
                "Regały wjezdne (drive-in)": 28.0,
                "Regały paletowe selektywne (standardowe)": 60.0,
            }
            POWIERZCHNIA_NETTO_PALETY_M2 = 1.2  # standardowa paleta EUR ~1.2 m x 1.0 m, z marginesem

            RACKING_PRESETS_M2 = {
                "Składowanie blokowe (block stacking)": {"m2": 1.3, "poziomy": 2},
                "Regały wjezdne (drive-in)": {"m2": 1.8, "poziomy": 3},
                "Regały paletowe selektywne (standardowe)": {"m2": 3.0, "poziomy": 4},
                "Własna wartość": {"m2": None, "poziomy": None},
            }
            c_wh1, c_wh2, c_wh3 = st.columns(3)
            with c_wh1:
                typ_skladowania = st.selectbox("Typ składowania:", list(RACKING_PRESETS_M2.keys()), index=2, key="typ_skladowania_tab3")
            with c_wh2:
                domyslna_powierzchnia = RACKING_PRESETS_M2[typ_skladowania]["m2"]
                if domyslna_powierzchnia is not None:
                    aisle_pct_preset = RACKING_PRESETS_AISLE_PCT[typ_skladowania]
                    powierzchnia_na_miejsce = domyslna_powierzchnia
                    st.metric("Powierzchnia / miejsce paletowe (1 poziom)", f"{powierzchnia_na_miejsce:.2f} m²",
                              help=f"Zawiera ~{aisle_pct_preset:.0f}% na alejki i drogi transportowe wózka widłowego "
                                   f"(szacunek dla tego typu składowania — {POWIERZCHNIA_NETTO_PALETY_M2:.1f} m² "
                                   f"netto/paletę ÷ (1 − {aisle_pct_preset:.0f}%)).")
                else:
                    st.caption("**Własna wartość** — rozbij poniżej na wymiary palety + szerokość alejki wózka "
                               "widłowego, żeby mieć pełną, fizyczną kontrolę nad tym, ile miejsca to zajmuje.")
                    aw1, aw2, aw3 = st.columns(3)
                    with aw1:
                        paleta_szerokosc_m = st.number_input(
                            "Szerokość palety [m]:", min_value=0.5, value=1.2, step=0.1, key="paleta_szerokosc_input",
                            help="Wymiar palety WZDŁUŻ alejki — standardowa paleta EUR ustawiona dłuższym bokiem "
                                 "do alejki to ok. 1,2 m."
                        )
                    with aw2:
                        paleta_glebokosc_m = st.number_input(
                            "Głębokość palety [m]:", min_value=0.5, value=1.0, step=0.1, key="paleta_glebokosc_input",
                            help="Wymiar palety W GŁĄB regału (prostopadle do alejki) — standardowa paleta EUR to ok. 0,8-1,0 m."
                        )
                    with aw3:
                        szerokosc_alejki_m = st.number_input(
                            "Szerokość alejki wózka widłowego [m]:", min_value=1.5, value=3.5, step=0.1, key="szerokosc_alejki_input",
                            help="Typowe wartości: ~3,5 m dla wózka czołowego/reach trucka (regały selektywne), "
                                 "~2,5 m dla wózka wąskoalejkowego (VNA), ~1,5 m minimalnie tam, gdzie wózek nie "
                                 "musi skręcać (block stacking/drive-in)."
                        )
                    # Geometria rzędu regałowego: alejka jest WSPÓLNA dla dwóch rzędów palet po obu
                    # stronach, więc każda pozycja "zajmuje" połowę szerokości alejki w głąb, na
                    # całej szerokości palety wzdłuż alejki.
                    powierzchnia_netto_wlasna = paleta_szerokosc_m * paleta_glebokosc_m
                    powierzchnia_na_miejsce = paleta_szerokosc_m * (paleta_glebokosc_m + szerokosc_alejki_m / 2.0)
                    aisle_pct_wlasna = (1.0 - powierzchnia_netto_wlasna / powierzchnia_na_miejsce) * 100.0
                    st.metric("→ Powierzchnia / miejsce paletowe (1 poziom, wynik)", f"{powierzchnia_na_miejsce:.2f} m²",
                              help=f"= szerokość palety × (głębokość palety + szerokość alejki ÷ 2) — alejka "
                                   f"współdzielona z rzędem po drugiej stronie. Odpowiada ~{aisle_pct_wlasna:.0f}% "
                                   "powierzchni na alejki.")
            with c_wh3:
                domyslne_poziomy = RACKING_PRESETS_M2[typ_skladowania]["poziomy"]
                liczba_poziomow = st.number_input(
                    "Liczba poziomów składowania:", min_value=1, max_value=10,
                    value=int(domyslne_poziomy) if domyslne_poziomy is not None else 3, step=1,
                    disabled=domyslne_poziomy is not None,
                    help="Ile palet w pionie mieści jedno miejsce podłogowe — dla block stackingu to fizyczne "
                         "piętrzenie palet jedna na drugiej (zwykle 2-3), dla regałów to liczba poziomów "
                         "regału (zwykle 3-5). Więcej poziomów = mniejsza wymagana powierzchnia podłogi przy "
                         "tej samej liczbie miejsc paletowych. Wybierz 'Własna wartość' powyżej, aby to odblokować."
                )
            st.session_state["powierzchnia_na_miejsce_report"] = powierzchnia_na_miejsce
            st.session_state["liczba_poziomow_report"] = liczba_poziomow

            total_fg_positions = sum(r["Miejsca magazynowe [szt] 📐"] for r in real_split_rows)
            # RM już liczone w Zakładce 2 dla WYBRANEGO tam roku/importu (nie doliczamy tu drugi
            # raz przybliżonego globalnego przelicznika rampupu - to prowadziło do podwójnego
            # skalowania i było niezależne od statusu importu per surowiec).
            total_rm_positions = sum(r["Miejsca magazynowe [szt]"] for r in rm_warehouse_rows)
            total_miejsca_magazynowe = total_fg_positions + total_rm_positions + total_import_positions

            # Budynek stawiany RAZ, pod docelową (100%) produkcję — niezależnie od wybranego roku
            # symulacji. RM w rm_warehouse_rows jest już liczone przy 100% (Zakładka 2 nie skaluje
            # rampupem), więc target = suma bez przeliczenia; FG target liczony osobno w pętli wyżej;
            # Import target = TYLKO produkty na stałe importowane ("Nigdy") - te potrzebują miejsca
            # w magazynie nawet w pełnej dojrzałości.
            total_fg_positions_target = sum(fg_positions_target_list)
            total_rm_positions_target = sum(r["Miejsca magazynowe [szt]"] for r in rm_warehouse_rows)
            total_miejsca_magazynowe_target = total_fg_positions_target + total_rm_positions_target + total_import_positions_target
            total_floor_slots_target = math.ceil(total_miejsca_magazynowe_target / liczba_poziomow) if liczba_poziomow > 0 else total_miejsca_magazynowe_target
            total_powierzchnia_m2 = total_floor_slots_target * powierzchnia_na_miejsce

            wykorzystanie_magazynu_pct = (total_miejsca_magazynowe / total_miejsca_magazynowe_target * 100.0) \
                if total_miejsca_magazynowe_target > 0 else 0.0

            st.session_state["total_miejsca_magazynowe_target_report"] = total_miejsca_magazynowe_target
            st.session_state["total_powierzchnia_m2_report"] = total_powierzchnia_m2

            m_wh1, m_wh2, m_wh3, m_wh4 = st.columns(4)
            with m_wh1: st.metric("📦 Miejsca paletowe — ten rok (FG + RM + Import)", f"{total_miejsca_magazynowe:,} szt.",
                                   help=f"FG: {total_fg_positions:,} szt. · RM: {total_rm_positions:,} szt. · Import: {total_import_positions:,} szt.")
            with m_wh2: st.metric("🎯 Miejsca paletowe — docelowe (100%)", f"{total_miejsca_magazynowe_target:,} szt.",
                                   help=f"W tym stały import (\"Nigdy\"): {total_import_positions_target:,} szt.")
            with m_wh3: st.metric("📐 Powierzchnia magazynu (budowana pod 100%)", f"{total_powierzchnia_m2:,.0f} m²")
            with m_wh4: st.metric("📊 Wykorzystanie magazynu w tym roku", f"{wykorzystanie_magazynu_pct:.0f}%")

            dostepna_powierzchnia_m2 = st.number_input(
                "🏢 Dostępna powierzchnia magazynowa (istniejący/planowany budynek) [m²] — opcjonalnie:",
                min_value=0.0, value=0.0, step=50.0, key="dostepna_powierzchnia_input",
                help="Podaj, jeśli masz już konkretną halę (istniejącą lub zaplanowaną) — porównamy z wyliczonym "
                     "zapotrzebowaniem powyżej. Zostaw 0, żeby pominąć to porównanie."
            )
            if dostepna_powierzchnia_m2 > 0:
                roznica_m2 = dostepna_powierzchnia_m2 - total_powierzchnia_m2
                if roznica_m2 >= 0:
                    st.success(f"✅ Dostępna hala ({dostepna_powierzchnia_m2:,.0f} m²) **wystarcza** — nadwyżka {roznica_m2:,.0f} m² "
                               f"({roznica_m2 / total_powierzchnia_m2 * 100.0:.0f}% zapasu) względem wyliczonego zapotrzebowania.")
                else:
                    st.error(f"❌ Dostępna hala ({dostepna_powierzchnia_m2:,.0f} m²) **NIE wystarcza** — brakuje {abs(roznica_m2):,.0f} m² "
                             f"({abs(roznica_m2) / total_powierzchnia_m2 * 100.0:.0f}% deficytu) względem wyliczonego zapotrzebowania.")

            # Jasna, bezwzględna ilość materiału w magazynie [t] I w paletach - nie tylko liczba
            # miejsc paletowych (które uwzględniają dni zapasu i poziomy składowania i nie mówią
            # wprost "ile mamy" komuś, kto nie myśli w tych kategoriach).
            total_fg_tony_month = sum(r.get("_masa_kg_miesiac", 0.0) for r in real_split_rows) / 1000.0
            total_fg_palety_month = sum(r["Palet [/mies] 🧱"] for r in real_split_rows)
            rm_consumption_this_view = compute_rm_consumption_for_year(effective_year_idx_for_import)

            # RM "w magazynie" (paletowym) = TYLKO beczki/IBC/worki - dokładnie te same surowce, co
            # w rm_warehouse_rows (z Zakładki 2), skąd też liczą się palety. Surowce w zbiornikach
            # (silosy, tank farm) NIE stoją w tym magazynie paletowym - mają osobne miejsce (silosy)
            # i pokazujemy je tu jako OSOBNĄ liczbę, żeby te dwie tonaże nigdy się nie myliły.
            total_rm_drummed_tony_month = (sum(r["Zużycie [t/rok] 🔒"] for r in rm_warehouse_rows) / MONTHS_PER_YEAR) if rm_warehouse_rows else 0.0
            total_rm_palety_month = sum(r["Palet [/mies]"] for r in rm_warehouse_rows) if rm_warehouse_rows else 0
            drummed_materials_set = {r["Surowiec 🔒"] for r in rm_warehouse_rows} if rm_warehouse_rows else set()
            total_rm_tank_tony_month = (sum(t for mat, t in rm_consumption_this_view.items() if mat not in drummed_materials_set)
                                         / MONTHS_PER_YEAR) if rm_consumption_this_view else 0.0

            m_qty1, m_qty2, m_qty3 = st.columns(3)
            with m_qty1: st.metric("🏷️ Wyroby Gotowe (FG) w magazynie", f"{total_fg_tony_month:,.1f} t",
                                    help=f"≈ {total_fg_palety_month:,.0f} palet/mies. Miesięczna produkcja FG w wybranym roku/widoku (przybliżenie stanu magazynowego).")
            with m_qty2: st.metric("🛢️ Surowce (RM) w magazynie (beczki/IBC)", f"{total_rm_drummed_tony_month:,.1f} t",
                                    help=f"≈ {total_rm_palety_month:,.0f} palet/mies. Tylko surowce w beczkach/IBC/workach — to jest to, co faktycznie zajmuje miejsce paletowe w tym magazynie.")
            with m_qty3: st.metric("🧱 Surowce (RM) w zbiornikach (tank farm)", f"{total_rm_tank_tony_month:,.1f} t",
                                    help="Surowce magazynowane luzem w silosach (Zakładka 2) — NIE zajmują miejsca paletowego, liczone osobno, poza tym magazynem.")

            st.caption("💡 Powierzchnia = ⌈(docelowe miejsca paletowe FG + RM + stały Import) / liczba poziomów⌉ × "
                       "powierzchnia/miejsce (1 poziom) — budynek stawiany RAZ, pod pełną (100%) zdolność. "
                       "Bufor **surowców w zbiornikach** (silosy) jest liczony i wymiarowany osobno w Zakładce 4. "
                       f"Powierzchnia budynku pozostaje wymiarowana pod 100% celu niezależnie od wybranego roku — "
                       f"zmienia się tylko pokazane wykorzystanie ({selected_rampup_year_label}).")

            # ============================================================
            # WYSYŁKI — ile trzeba wysyłać dziennie, żeby magazyn WG nie rósł ponad projektową
            # rotację, i co się dzieje, jeśli realne wysyłki są mniejsze od sugerowanych.
            # ============================================================
            st.markdown("---")
            st.markdown("##### 🚚 Wysyłki — Utrzymanie Optymalnego Poziomu Magazynu")
            st.caption("Magazyn jest zaprojektowany pod rotację **{:.0f} dni** (pole 'Czas składowania palety' powyżej) "
                       "— żeby jej dotrzymać, tyle samo palet WG musi dziennie WYJEŻDŻAĆ, ile średnio dziennie "
                       "PRZYBYWA z produkcji. Poniżej: sugerowane tempo wysyłek w tym roku/widoku, oraz co się "
                       "dzieje, jeśli realne wysyłki są od niego mniejsze.".format(czas_skladowania_dni))

            total_palety_month_fg = sum(r["Palet [/mies] 🧱"] for r in real_split_rows)
            suggested_pallets_per_day = (total_palety_month_fg / dni_robocze_miesiac) if dni_robocze_miesiac > 0 else 0.0

            c_ship1, c_ship2 = st.columns(2)
            with c_ship1:
                pallets_per_truck = st.number_input(
                    "Palet na 1 wysyłkę (naczepa/kontener) [szt]:", min_value=1, value=33, step=1, key="pallets_per_truck",
                    help="Typowa naczepa standardowa mieści ~33 palety EUR — dostosuj do realnego taboru."
                )
            with c_ship2:
                suggested_trucks_per_day = suggested_pallets_per_day / pallets_per_truck if pallets_per_truck > 0 else 0.0
                st.metric("📦 Sugerowane wysyłki / dzień (ten rok)", f"{suggested_trucks_per_day:.2f} wysyłki/dzień",
                          help=f"= {suggested_pallets_per_day:.1f} palet/dzień ({total_palety_month_fg:.0f} palet/mies. ÷ "
                               f"{dni_robocze_miesiac:.1f} dni roboczych/mies.)")

            actual_trucks_per_day = st.number_input(
                "Rzeczywiste/planowane wysyłki / dzień:", min_value=0.0,
                value=round(suggested_trucks_per_day, 2), step=0.5, key="actual_trucks_per_day",
                help="Domyślnie ustawione na wartość sugerowaną — zmień, żeby zobaczyć skutek wysyłania mniej (lub więcej) niż potrzeba."
            )
            actual_pallets_per_day = actual_trucks_per_day * pallets_per_truck
            gap_pallets_per_day = suggested_pallets_per_day - actual_pallets_per_day

            if actual_trucks_per_day <= 0:
                st.warning("⚠️ Przy zerowych wysyłkach magazyn WG napełni się od zera do pełnej pojemności w "
                           f"**{(total_fg_positions_target / suggested_pallets_per_day):.0f} dni roboczych** "
                           "(przy tempie produkcji z wybranego roku) — i dalej rosnąć, bez odpływu.")
            elif gap_pallets_per_day > 0.001:
                effective_rotation_days = czas_skladowania_dni * (suggested_pallets_per_day / actual_pallets_per_day)
                # Ile miejsc podłogowych zajmie magazyn WG, jeśli realnie rotacja wydłuża się do effective_rotation_days
                # zamiast projektowych czas_skladowania_dni - proporcjonalnie do dziennego niedoboru wysyłek.
                fg_positions_at_actual_rate = math.ceil(total_fg_positions_target * (effective_rotation_days / czas_skladowania_dni)) \
                    if czas_skladowania_dni > 0 else total_fg_positions_target
                extra_positions_needed = max(fg_positions_at_actual_rate - total_fg_positions_target, 0)
                spare_capacity = max(total_miejsca_magazynowe_target - total_miejsca_magazynowe, 0)
                days_to_overflow = (spare_capacity / gap_pallets_per_day) if gap_pallets_per_day > 0 else float("inf")

                st.error(f"🔴 Wysyłasz **{gap_pallets_per_day:.1f} palety/dzień mniej** niż sugerowane — magazyn WG "
                         f"rośnie ponad projekt. Efektywna rotacja wydłuża się z {czas_skladowania_dni:.0f} do "
                         f"~**{effective_rotation_days:.0f} dni**, co odpowiada ok. **{extra_positions_needed:,} dodatkowym "
                         f"miejscom paletowym** ponad dzisiejszy projekt budynku. Przy obecnym zapasie wolnej "
                         f"powierzchni ({spare_capacity:,} miejsc w tym roku/widoku) magazyn **przepełni się za "
                         f"~{days_to_overflow:.0f} dni roboczych**, jeśli nic się nie zmieni.")
            else:
                st.success(f"🟢 Planowane wysyłki ({actual_trucks_per_day:.2f}/dzień) pokrywają lub przewyższają "
                           f"sugerowane tempo — magazyn WG powinien utrzymać projektową rotację "
                           f"{czas_skladowania_dni:.0f} dni.")

            # ============================================================
            # SYMULACJA STANU MAGAZYNOWEGO W CZASIE (60 miesięcy = 5 lat rozruchu): jak zapas
            # WG będzie się zmieniał w zależności od PLANOWANEJ PRODUKCJI (krzywa rozruchu) przy
            # ZAŁOŻONYM TEMPIE WYSYŁEK (pole "Rzeczywiste/planowane wysyłki/dzień" powyżej) - sam
            # poziom wysyłek na wykresie POMIJAMY (jest już ujęty w komunikacie ostrzegawczym
            # powyżej) i pokazujemy tylko wynikowy stan magazynowy, razem z jego wartością.
            # ============================================================
            st.markdown("###### 📈 Symulacja Stanu Magazynowego — 5 lat")
            st.caption("Zgodnie z planowaną produkcją (krzywa rozruchu z Zakładki 3) i wysyłkami ustawionymi "
                       "powyżej: jak zmienia się stan magazynowy WG w palet oraz jego wartość, na tle projektowej "
                       "pojemności budynku.")

            total_palety_month_fg_target = sum(fg_pallets_target_list)
            st.session_state["total_palety_month_fg_target_report"] = total_palety_month_fg_target
            target_annual_t_ship = sum(m["annual_volume"] for m in mixers_fleet) / 1000.0
            # Stosunek rzeczywistych wysyłek do sugerowanych (dla WYBRANEGO powyżej roku/widoku) -
            # ten sam stosunek stosujemy do KAŻDEGO roku symulacji, żeby wysyłki rosły razem z
            # produkcją (a nie były jedną stałą liczbą przez wszystkie 5 lat - to dawało fałszywy
            # obraz "0% wykorzystania w latach 1-2, potem gwałtowny skok", bo stała liczba wysyłek
            # dobrana pod PÓŹNIEJSZY, wyższy rok automatycznie przewyższała niską produkcję z
            # wczesnych lat i zerowała magazyn).
            shipment_ratio = (actual_pallets_per_day / suggested_pallets_per_day) if suggested_pallets_per_day > 0 else 1.0
            fg_capacity_pallets = total_fg_positions_target

            # Przelicznik palety -> kg -> wartość SPRZEDAŻNA (Zakładka 5, Krok 1: cena sprzedaży per
            # grupa produktowa [waluta/L], przeliczona na [waluta/kg] przez gęstość) - to jest realna
            # wartość towaru leżącego w magazynie, nie tylko koszt wytworzenia. Ważona rzeczywistym
            # miksem produkcji tej floty; jeśli Zakładka 5 nie była jeszcze skonfigurowana, używane
            # są wartości domyślne.
            target_monthly_mass_kg_total = sum(m["batches_count"] * m["mass_per_batch"] for m in mixers_fleet)
            avg_kg_per_pallet = (target_monthly_mass_kg_total / total_palety_month_fg_target) if total_palety_month_fg_target > 0 else 0.0
            group_pricing_stock = st.session_state.get("group_pricing", {})
            total_annual_volume_kg = sum(m["annual_volume"] for m in mixers_fleet)

            def _selling_value_per_kg(grp):
                pricing = group_pricing_stock.get(grp)
                if not pricing:
                    return 2.54  # domyślna wartość, spójna z resztą apki, dopóki Krok 1 nie zostanie ustawiony
                density_kg_l = st.session_state.active_portfolio.get(grp, {}).get("density", 0.9)
                return pricing["sales_price"] / density_kg_l if density_kg_l > 0 else 2.54

            avg_selling_value_per_kg = (sum(m["annual_volume"] * _selling_value_per_kg(m["product_family"]) for m in mixers_fleet)
                                         / total_annual_volume_kg) if total_annual_volume_kg > 0 else 2.54

            waluta_stock = st.selectbox("Waluta wyceny zapasu:", ["PLN", "EUR", "USD"], key="waluta_stock_value")
            if not group_pricing_stock:
                st.caption(f"ℹ️ Cena sprzedaży per grupa nie była jeszcze ustawiana w Zakładce 5, Krok 1 — użyto "
                           f"wartości domyślnej ({avg_selling_value_per_kg:.2f} {waluta_stock}/kg). "
                           f"Ustaw ją w Zakładce 5, aby wycena była dokładniejsza.")

            initial_stock_pct = st.slider(
                "Początkowe zapełnienie magazynu (przed Rokiem 1) [% pojemności FG]:", min_value=0, max_value=100,
                value=0, step=5, key="initial_stock_pct_fg",
                help="Jeśli magazyn nie startuje pusty (np. zapas bezpieczeństwa zbudowany przed rozruchem, albo "
                     "przejęty istniejący zapas) — ustaw tu punkt startowy. Symulacja poniżej śledzi zmiany "
                     "WZGLĘDEM tego poziomu, a nie zawsze od zera."
            )
            initial_stock_pallets = fg_capacity_pallets * (initial_stock_pct / 100.0)
            if initial_stock_pct > 0:
                st.caption(f"➡️ Start: {initial_stock_pallets:,.0f} palet ({initial_stock_pct}% z {fg_capacity_pallets:,.0f} palet pojemności).")

            stock_rows = []
            stock_level = initial_stock_pallets
            for yi in range(RAMPUP_YEARS):
                for mi in range(1, 13):
                    # Ułamek MIESIĘCZNY (interpolowany liniowo), nie płaski roczny - dzięki temu
                    # symulacja magazynu faktycznie rośnie płynnie w trakcie roku (jak w
                    # rzeczywistości), zamiast robić skok na każdej granicy roku.
                    month_tonnage_t = sum((m["annual_volume"] / 1000.0) * get_rampup_fraction_month(m["product_family"], yi, mi - 1)
                                           for m in mixers_fleet)
                    frac_month = (month_tonnage_t / target_annual_t_ship) if target_annual_t_ship > 0 else 0.0
                    production_pallets_month = total_palety_month_fg_target * frac_month
                    shipped_pallets_month = production_pallets_month * shipment_ratio

                    stock_level = max(stock_level + production_pallets_month - shipped_pallets_month, 0.0)
                    wartosc_zapasu = stock_level * avg_kg_per_pallet * avg_selling_value_per_kg
                    stock_rows.append({
                        "Miesiąc": yi * 12 + mi, "Okres": f"Y{yi + 1}-{mi:02d}", "Rok": f"Rok {yi + 1}",
                        "Stan magazynowy [pal]": stock_level,
                        f"Wartość zapasu [{waluta_stock}]": wartosc_zapasu,
                    })

            df_stock = pd.DataFrame(stock_rows)
            wartosc_col = f"Wartość zapasu [{waluta_stock}]"
            st.session_state["stock_simulation_df"] = df_stock  # do raportu PDF (Zakładka 5)
            st.session_state["fg_capacity_pallets_report"] = fg_capacity_pallets

            st.caption(f"Wartość sprzedażna użyta do wyceny: {avg_selling_value_per_kg:.2f} {waluta_stock}/kg "
                       "(cena sprzedaży per grupa z Zakładki 5, Krok 1, ważona miksem produkcji floty). "
                       "Rosnące słupki w kolejnych latach pokazują rosnące wykorzystanie magazynu wraz z rozruchem produkcji.")

            st.markdown("**Stan magazynowy [palety]**")
            st.bar_chart(df_stock.set_index("Okres")[["Stan magazynowy [pal]"]])

            # Max/min/średnia liczba palet w symulacji + odpowiadający im metraż, tym samym
            # przelicznikiem co w podsumowaniu powierzchni magazynowej wyżej (poziomy składowania
            # + powierzchnia/miejsce). Min pomijamy jeśli i tak wynosi 0 (naturalny start symulacji).
            pal_max = df_stock["Stan magazynowy [pal]"].max()
            pal_min = df_stock["Stan magazynowy [pal]"].min()
            pal_avg = df_stock["Stan magazynowy [pal]"].mean()

            def _pal_to_m2(pal_count):
                floor_slots = math.ceil(pal_count / liczba_poziomow) if liczba_poziomow > 0 else pal_count
                return floor_slots * powierzchnia_na_miejsce

            p_c1, p_c2, p_c3 = st.columns(3)
            with p_c1:
                st.metric("📉 Min. stan magazynowy", f"{pal_min:,.0f} pal.", help=f"≈ {_pal_to_m2(pal_min):,.0f} m² potrzebnej powierzchni")
            with p_c2:
                st.metric("📊 Śr. stan magazynowy", f"{pal_avg:,.0f} pal.", help=f"≈ {_pal_to_m2(pal_avg):,.0f} m² potrzebnej powierzchni")
            with p_c3:
                st.metric("📈 Maks. stan magazynowy", f"{pal_max:,.0f} pal.", help=f"≈ {_pal_to_m2(pal_max):,.0f} m² potrzebnej powierzchni")
            st.caption(f"Metraż liczony tym samym przelicznikiem co powyżej ({powierzchnia_na_miejsce:.2f} m²/miejsce, "
                       f"{liczba_poziomow:.0f} poziomów składowania).")

            st.caption("Wartość zapasu na koniec każdego roku (patrz też metryki poniżej):")
            year_end_table = df_stock.iloc[[11, 23, 35, 47, 59]][["Rok", "Stan magazynowy [pal]", wartosc_col]].copy()
            year_end_table["Stan magazynowy [pal]"] = year_end_table["Stan magazynowy [pal]"].round(0).astype(int)
            year_end_table[wartosc_col] = year_end_table[wartosc_col].round(0)
            st.dataframe(year_end_table, hide_index=True, use_container_width=True)

            v_c1, v_c2, v_c3 = st.columns(3)
            with v_c1:
                st.metric("💰 Wartość zapasu — koniec Roku 1", f"{df_stock.iloc[11][f'Wartość zapasu [{waluta_stock}]']:,.0f} {waluta_stock}")
            with v_c2:
                st.metric("💰 Wartość zapasu — koniec Roku 5", f"{df_stock.iloc[-1][f'Wartość zapasu [{waluta_stock}]']:,.0f} {waluta_stock}")
            with v_c3:
                st.metric("📈 Szczytowa wartość zapasu", f"{df_stock[f'Wartość zapasu [{waluta_stock}]'].max():,.0f} {waluta_stock}")

            if stock_level > fg_capacity_pallets:
                st.error(f"🔴 Przy tych założeniach stan magazynowy po 5 latach ({stock_level:,.0f} palet) "
                         f"**przekracza** projektową pojemność FG ({fg_capacity_pallets:,.0f} palet) — wysyłki nie "
                         f"nadążają za rozruchem produkcji.")
            else:
                st.success(f"🟢 Przy tych założeniach stan magazynowy po 5 latach ({stock_level:,.0f} palet) mieści "
                           f"się w projektowej pojemności FG ({fg_capacity_pallets:,.0f} palet).")
        else:
            st.info("Brak skonfigurowanego podziału opakowań o niezerowym udziale — uzupełnij procenty w panelu bocznym, "
                    "albo (dla produktów importowanych) uzupełnij dane importu w Zakładce 1.")

# ==========================================
# ZAKŁADKA 5: ANALIZA FINANSOWA, CAPEX I ROI (tab4)
# ==========================================
with tab4:
    st.header("💰 Analiza Finansowa, CAPEX i ROI")
    if not st.session_state.confirmed_mixers:
        st.warning("⚠️ Najpierw zatwierdź flotę w Zakładce 1.")
    else:
        _stale_msg = check_fleet_staleness_warning()
        if _stale_msg:
            st.error(_stale_msg)
        waluta = st.selectbox("Wybierz walutę operacyjną:", ["PLN", "EUR", "USD"])

        # ==========================================
        # SYMULACJA ROZRUCHU (RAMPUP) — wejścia ustawiasz w panelu bocznym (KROK 3), tu tylko
        # wynikowe podsumowanie. Flota budowana od razu pod cel, ale wykorzystanie rośnie w
        # czasie; ta sama krzywa % jest reużywana w Zakładce 2 (magazyn FG+RM), żeby historia
        # "startujemy nisko, dochodzimy do celu" była spójna w całej aplikacji.
        # ==========================================
        st.markdown("---")
        with st.expander("📈 Symulacja Rozruchu (Rampup) — 5 lat — wyniki", expanded=False):
            st.caption("Ustawienia krzywej rozruchu zmieniasz w **panelu bocznym (KROK 3)** — działają na żywo "
                       "we wszystkich zakładkach. Tu tylko podsumowanie wynikowe.")

            year_labels = [f"Rok {i+1}" for i in range(RAMPUP_YEARS)]

            # --- Przeliczenie: tonaż roczny i średnia utylizacja floty per rok symulacji ---
            target_annual_t = sum(m["annual_volume"] for m in st.session_state.confirmed_mixers) / 1000.0
            target_batches_month_total = sum(m["batches_count"] for m in st.session_state.confirmed_mixers)
            rampup_summary_rows = []
            rampup_tonnage_chart = {"Rok": [], "Tonaż [t/rok]": [], "Cel [t/rok]": []}
            rampup_cost_revenue_chart = {"Rok": [], f"Koszt produkcji [{waluta}/rok]": [], f"Przychód sprzedaży [{waluta}/rok]": []}
            for i in range(RAMPUP_YEARS):
                year_tonnage_t = 0.0
                util_weighted, util_weight_sum = 0.0, 0.0
                util_ciagla_weighted = 0.0
                total_batches_month_yi = 0
                for m in st.session_state.confirmed_mixers:
                    frac = get_rampup_fraction(m["product_family"], i)
                    year_tonnage_t += (m["annual_volume"] / 1000.0) * frac

                    scaled_monthly_mass = (m["annual_volume"] / MONTHS_PER_YEAR) * frac
                    scaled_batches = math.ceil(scaled_monthly_mass / m["mass_per_batch"]) if m["mass_per_batch"] > 0 else 0
                    total_batches_month_yi += scaled_batches
                    scaled_util_pct = (scaled_batches * m["cycle_h"]) / AVAILABLE_HOURS_MONTH * 100.0 if AVAILABLE_HOURS_MONTH > 0 else 0.0
                    util_weighted += scaled_util_pct * m["capacity_m3"]
                    util_weight_sum += m["capacity_m3"]

                    # Wariant ciągły (bez zaokrąglania w górę do pełnych szarż) - czysto
                    # diagnostyczny, żeby pokazać płynny trend rosnącego zapotrzebowania nawet
                    # gdy realna (całkowita) liczba szarż jeszcze się nie zmienia między latami.
                    scaled_batches_ciagle = scaled_monthly_mass / m["mass_per_batch"] if m["mass_per_batch"] > 0 else 0.0
                    scaled_util_pct_ciagly = (scaled_batches_ciagle * m["cycle_h"]) / AVAILABLE_HOURS_MONTH * 100.0 if AVAILABLE_HOURS_MONTH > 0 else 0.0
                    util_ciagla_weighted += scaled_util_pct_ciagly * m["capacity_m3"]

                avg_util_pct = (util_weighted / util_weight_sum) if util_weight_sum > 0 else 0.0
                avg_util_ciagly_pct = (util_ciagla_weighted / util_weight_sum) if util_weight_sum > 0 else 0.0

                # Śr. miesięczne palety FG i RM (beczkowane) w tym roku - ten sam frac co tonaż,
                # zastosowany do docelowej (100%) stawki palet/mies. Zbiornikowe RM (silosy) NIE są
                # tu ujęte - to nie palety, tylko osobno wymiarowane zbiorniki (Zakładka 2).
                frac_year_blended = (year_tonnage_t / target_annual_t) if target_annual_t > 0 else 0.0
                fg_pallets_month_yi = st.session_state.get("total_palety_month_fg_target_report", 0.0) * frac_year_blended

                # Ta sama logika co Zakładka 4/Dashboard - liczona RAZ we wspólnej funkcji, żeby
                # te trzy miejsca nigdy nie mogły się rozjechać.
                rm_pallets_month_yi = sum(compute_rm_drummed_pallets_per_month(i).values())

                # Koszt produkcji vs przychód ze sprzedaży, ten sam rok - per grupa produktowa,
                # tą samą metodą co ROI (Krok 3) niżej: cena sprzedaży i koszty z Kroku 1, każda
                # grupa skalowana WŁASNĄ krzywą rozruchu (nie ogólną frac_year_blended powyżej).
                revenue_year_chart, cost_year_chart = 0.0, 0.0
                for grp in sorted(set(mm["product_family"] for mm in st.session_state.confirmed_mixers)):
                    density_kg_l_chart = st.session_state.active_portfolio[grp]["density"]
                    annual_volume_kg_grp = sum(mm["annual_volume"] for mm in st.session_state.confirmed_mixers if mm["product_family"] == grp)
                    annual_volume_l_grp = annual_volume_kg_grp / density_kg_l_chart if density_kg_l_chart > 0 else 0.0
                    frac_grp = get_rampup_fraction(grp, i)
                    vol_l_year_grp = annual_volume_l_grp * frac_grp
                    pricing_grp = st.session_state.get("group_pricing", {}).get(grp, {
                        "sales_price": 2.0, "raw_material_cost": 1.35, "labour": 0.07, "energy": 0.04,
                        "qc": 0.04, "maintenance": 0.04, "logistics": 0.04, "overhead": 0.04,
                    })
                    cost_per_l_grp = pricing_grp["raw_material_cost"] + sum(
                        pricing_grp[k] for k in ["labour", "energy", "qc", "maintenance", "logistics", "overhead"]
                    )
                    revenue_year_chart += vol_l_year_grp * pricing_grp["sales_price"]
                    cost_year_chart += vol_l_year_grp * cost_per_l_grp

                rampup_summary_rows.append({
                    "Rok": year_labels[i], "Tonaż [t/rok]": round(year_tonnage_t, 0),
                    "% Celu": f"{(year_tonnage_t / target_annual_t * 100.0) if target_annual_t > 0 else 0:.0f}%",
                    "Szarż / miesiąc (cała flota)": total_batches_month_yi,
                    "Szarż / rok (cała flota)": total_batches_month_yi * MONTHS_PER_YEAR,
                    "Śr. Utylizacja Floty [%] (pełne szarże)": round(avg_util_pct, 1),
                    "Śr. Utylizacja Floty [%] (ciągła)": round(avg_util_ciagly_pct, 1),
                    "Śr. palet FG/mies.": round(fg_pallets_month_yi, 0),
                    "Śr. palet RM (beczk.)/mies.": round(rm_pallets_month_yi, 0),
                    f"Koszt produkcji [{waluta}/rok]": round(cost_year_chart, 0),
                    f"Przychód sprzedaży [{waluta}/rok]": round(revenue_year_chart, 0),
                })
                rampup_tonnage_chart["Rok"].append(year_labels[i])
                rampup_tonnage_chart["Tonaż [t/rok]"].append(year_tonnage_t)
                rampup_tonnage_chart["Cel [t/rok]"].append(target_annual_t)
                rampup_cost_revenue_chart["Rok"].append(year_labels[i])
                rampup_cost_revenue_chart[f"Koszt produkcji [{waluta}/rok]"].append(cost_year_chart)
                rampup_cost_revenue_chart[f"Przychód sprzedaży [{waluta}/rok]"].append(revenue_year_chart)

            ch_c1, ch_c2 = st.columns(2)
            with ch_c1:
                st.markdown("**Tonaż [t/rok]**")
                chart_df = pd.DataFrame(rampup_tonnage_chart).set_index("Rok")
                st.line_chart(chart_df)
            with ch_c2:
                st.markdown(f"**Koszt produkcji vs Przychód sprzedaży [{waluta}/rok]**")
                st.caption("Z tych samych cen/kosztów per grupa co w Kroku 1 (Rentowność), każda grupa skalowana "
                           "WŁASNĄ krzywą rozruchu — to ten sam model, który zasila ROI w Kroku 3 niżej.")
                chart_cost_rev_df = pd.DataFrame(rampup_cost_revenue_chart).set_index("Rok")
                try:
                    st.bar_chart(chart_cost_rev_df, stack=False)  # słupki obok siebie, nie jeden na drugim
                except TypeError:
                    # Starsza wersja Streamlit bez parametru "stack" - fallback do domyślnego (skumulowanego).
                    st.bar_chart(chart_cost_rev_df)

            st.dataframe(pd.DataFrame(rampup_summary_rows), hide_index=True, use_container_width=True)
            st.caption("ℹ️ **Pełne szarże** — realna liczba szarż zaokrąglona w górę do liczb całkowitych (tak faktycznie "
                       "planuje się produkcję); może być identyczna w sąsiednich latach, jeśli wzrost popytu nie "
                       "przekroczył jeszcze progu kolejnej pełnej szarży na wystarczającej liczbie mieszalników. "
                       "**Ciągła** — ta sama utylizacja bez zaokrąglania, czysto diagnostyczna: pokazuje płynny trend "
                       "rosnącego zapotrzebowania nawet między takimi progami. **Palety FG/RM** — śr. miesięczna "
                       "liczba palet danego roku; RM obejmuje wyłącznie surowce w beczkach/IBC/workach — surowce "
                       "w zbiornikach (silosy) są wymiarowane osobno w Zakładce 2, nie liczą się w paletach.")
            st.caption(f"🎯 Docelowa produkcja (100%, jak wymiarowana jest flota): **{target_annual_t:,.0f} t/rok** "
                       f"= **{target_batches_month_total} szarż/miesiąc** (**{target_batches_month_total * MONTHS_PER_YEAR} szarż/rok**) "
                       "dla całej floty. Ta sama krzywa steruje wykorzystaniem magazynu w Zakładce 2 (wybór roku symulacji).")

            st.markdown("---")
            st.markdown("###### 🏭 Mieszalniki — tonaż i produkt przypisany, per rok")
            st.caption("Liczba i pojemność mieszalników jest stała (budowane raz, pod 100% celu) — poniżej widać, "
                       "**który z nich faktycznie pracuje i z jakim tonażem w danym roku**. 'Nieużywany' = w tym "
                       "roku popyt na ten produkt/linię jeszcze nie generuje ani jednej pełnej szarży.")
            # Mapa produkt -> (sposób pozyskania, rok przejścia), z receptur - żeby sprawdzić, czy
            # KONKRETNY produkt przypisany do tego mieszalnika jest jeszcze importowany w danym
            # roku (a nie tylko czy linia produktowa w ogóle ma tam rampup > 0 - to był błąd:
            # mieszalnik dedykowany pod produkt wciąż importowany pokazywał się jako aktywny).
            product_sourcing_lookup = {}
            if st.session_state.recipes_df is not None and not st.session_state.recipes_df.empty and RECIPE_SOURCING_COL in st.session_state.recipes_df.columns:
                for _, r in st.session_state.recipes_df.iterrows():
                    product_sourcing_lookup[r[RECIPE_PRODUCT_COL]] = (
                        r.get(RECIPE_SOURCING_COL, "Produkcja własna"), r.get(RECIPE_IMPORT_TRANSITION_COL, "")
                    )

            mixer_year_rows = []
            for i in range(RAMPUP_YEARS):
                for m in st.session_state.confirmed_mixers:
                    recipe_product = m.get("recipe_product")
                    is_still_imported = False
                    if recipe_product and recipe_product in product_sourcing_lookup:
                        sourcing, transition = product_sourcing_lookup[recipe_product]
                        is_still_imported = is_product_imported_in_year(sourcing, transition, i)

                    if is_still_imported:
                        tonnage_year_t, scaled_batches = 0.0, 0
                        status_txt = "🔵 Jeszcze importowany"
                    else:
                        frac = get_rampup_fraction(m["product_family"], i)
                        tonnage_year_t = (m["annual_volume"] / 1000.0) * frac
                        scaled_monthly_mass = (m["annual_volume"] / MONTHS_PER_YEAR) * frac
                        scaled_batches = math.ceil(scaled_monthly_mass / m["mass_per_batch"]) if m["mass_per_batch"] > 0 else 0
                        status_txt = "🟢 Aktywny" if scaled_batches > 0 else "⚪ Nieużywany (jeszcze)"

                    mixer_year_rows.append({
                        "Rok": year_labels[i], "Tag": m["tag"], "Linia": m["product_family"],
                        "Produkt": recipe_product or "—", "Pojemność [m³]": m["capacity_m3"],
                        "Tonaż ten rok [t]": round(tonnage_year_t, 1), "Szarż/miesiąc ten rok": scaled_batches,
                        "Status": status_txt,
                    })
            df_mixer_year = pd.DataFrame(mixer_year_rows)
            selected_year_view = st.selectbox("Pokaż rok:", year_labels, key="mixer_year_view_select")
            st.dataframe(df_mixer_year[df_mixer_year["Rok"] == selected_year_view].drop(columns=["Rok"]),
                         hide_index=True, use_container_width=True)
            n_active_this_year = (df_mixer_year[df_mixer_year["Rok"] == selected_year_view]["Status"] == "🟢 Aktywny").sum()
            st.caption(f"Aktywnych mieszalników w {selected_year_view}: {n_active_this_year} / {len(st.session_state.confirmed_mixers)}.")

            confirmed_rm_tanks_ramp = st.session_state.get("confirmed_rm_tanks", [])
            if confirmed_rm_tanks_ramp:
                st.markdown("###### 🛢️ Zbiorniki RM — zużycie surowca przypisane, per rok")
                st.caption("Liczba i pojemność zbiorników RM też jest stała (budowane raz, pod 100% zużycia) — "
                           "poniżej widać zużycie surowca w danym roku i jak duża część łącznej pojemności "
                           "przypisanych zbiorników jest realnie wykorzystywana.")
                rm_stock_days = st.session_state.get("days_of_stock_tab5", 14)
                tanks_by_material = {}
                for t in confirmed_rm_tanks_ramp:
                    tanks_by_material.setdefault(t["material"], []).append(t)

                rm_year_rows = []
                for i in range(RAMPUP_YEARS):
                    year_consumption = compute_rm_consumption_for_year(i)
                    for material, tanks_this_material in tanks_by_material.items():
                        year_t = year_consumption.get(material, 0.0)
                        total_capacity_m3 = sum(t["capacity_m3"] for t in tanks_this_material)
                        required_buffer_m3 = ((year_t / WORKING_DAYS_YEAR) * rm_stock_days) / OIL_FILL_FACTOR
                        utilization_pct = (required_buffer_m3 / total_capacity_m3 * 100.0) if total_capacity_m3 > 0 else 0.0
                        rm_year_rows.append({
                            "Rok": year_labels[i], "Surowiec": material, "Zbiorników (zatwierdzonych)": len(tanks_this_material),
                            "Łączna pojemność [m³]": round(total_capacity_m3, 0), "Zużycie ten rok [t/rok]": round(year_t, 1),
                            "Wykorzystanie pojemności": f"{utilization_pct:.0f}%",
                            "Status": "🟢 Aktywny" if year_t > 0 else "⚪ Nieużywany (jeszcze)",
                        })
                df_rm_year = pd.DataFrame(rm_year_rows)
                st.dataframe(df_rm_year[df_rm_year["Rok"] == selected_year_view].drop(columns=["Rok"]),
                             hide_index=True, use_container_width=True)

        st.markdown("### 💵 Krok 1: Rentowność per Grupa Produktowa (Cena − Koszty)")
        st.caption("Ustaw cenę sprzedaży, koszt surowców i pozostałe czynniki kosztotwórcze — wszystko w "
                   "[waluta]/L (jak w Twoim zestawieniu). Marża brutto i procent ceny, jaki stanowi każdy koszt, "
                   "liczą się automatycznie. **Białe pola edytujesz, szare (marża, %) to wynik.**")

        if "group_pricing" not in st.session_state:
            st.session_state.group_pricing = {}
        active_groups_fin = sorted(set(m["product_family"] for m in st.session_state.confirmed_mixers))

        cost_item_keys = ["labour", "energy", "qc", "maintenance", "logistics", "overhead"]
        cost_item_labels = {
            "labour": "Robocizna bezpośrednia", "energy": "Energia i media", "qc": "QC/lab/eksploatacyjne",
            "maintenance": "Utrzymanie ruchu", "logistics": "Logistyka wych.", "overhead": "Narzut fabryczny",
        }

        pricing_table_rows = []
        for grp in active_groups_fin:
            density_kg_l = st.session_state.active_portfolio[grp]["density"]
            annual_volume_kg = sum(m["annual_volume"] for m in st.session_state.confirmed_mixers if m["product_family"] == grp)
            annual_volume_l = annual_volume_kg / density_kg_l if density_kg_l > 0 else 0.0

            defaults = st.session_state.group_pricing.setdefault(grp, {
                "sales_price": 2.0, "raw_material_cost": 1.35, "labour": 0.07, "energy": 0.04,
                "qc": 0.04, "maintenance": 0.04, "logistics": 0.04, "overhead": 0.04,
            })
            total_cost_l = defaults["raw_material_cost"] + sum(defaults[k] for k in cost_item_keys)
            margin_l = defaults["sales_price"] - total_cost_l
            margin_pct = (margin_l / defaults["sales_price"] * 100.0) if defaults["sales_price"] > 0 else 0.0

            row = {
                "Grupa": grp, "Produkcja [L/rok]": round(annual_volume_l, 0),
                f"Cena sprzedaży [{waluta}/L]": defaults["sales_price"],
                f"Koszt surowców [{waluta}/L]": defaults["raw_material_cost"],
            }
            for k in cost_item_keys:
                row[f"{cost_item_labels[k]} [{waluta}/L]"] = defaults[k]
            row[f"Marża brutto [{waluta}/L]"] = round(margin_l, 3)
            row["Marża brutto [%]"] = round(margin_pct, 1)
            pricing_table_rows.append(row)

        cost_col_config = {
            f"Cena sprzedaży [{waluta}/L]": st.column_config.NumberColumn(min_value=0.01, step=0.01, format="%.2f"),
            f"Koszt surowców [{waluta}/L]": st.column_config.NumberColumn(min_value=0.0, step=0.01, format="%.2f"),
        }
        for k in cost_item_keys:
            cost_col_config[f"{cost_item_labels[k]} [{waluta}/L]"] = st.column_config.NumberColumn(min_value=0.0, step=0.01, format="%.2f")

        edited_pricing_df = st.data_editor(
            pd.DataFrame(pricing_table_rows), hide_index=True, use_container_width=True, key="group_pricing_editor",
            disabled=["Grupa", "Produkcja [L/rok]", f"Marża brutto [{waluta}/L]", "Marża brutto [%]"],
            column_config=cost_col_config
        )
        for _, row in edited_pricing_df.iterrows():
            grp = row["Grupa"]
            d = st.session_state.group_pricing.setdefault(grp, {})
            d["sales_price"] = row[f"Cena sprzedaży [{waluta}/L]"]
            d["raw_material_cost"] = row[f"Koszt surowców [{waluta}/L]"]
            for k in cost_item_keys:
                d[k] = row[f"{cost_item_labels[k]} [{waluta}/L]"]

        total_annual_revenue_target = sum(
            (sum(m["annual_volume"] for m in st.session_state.confirmed_mixers if m["product_family"] == grp)
             / st.session_state.active_portfolio[grp]["density"]) * st.session_state.group_pricing[grp]["sales_price"]
            for grp in active_groups_fin
        )
        st.metric(f"💰 Roczny przychód (100% celu, wszystkie grupy)", f"{total_annual_revenue_target:,.0f} {waluta}")

        with st.expander("🔍 Szczegółowe źródła kosztu energii (opcjonalnie, do porównania z % 'Energia i media' powyżej)", expanded=False):
            cena_mwh = st.number_input(f"Cena energii elektrycznej [{waluta}/MWh]:", min_value=1.0, value=750.0)
            st.session_state["cena_mwh_tab4"] = cena_mwh

            if not st.session_state.calculated_times:
                st.info("ℹ️ Skonfiguruj urządzenia w Zakładce 2, aby koszty energii odzwierciedlały rzeczywistą hydraulikę i bilans cieplny "
                        "(w przeciwnym razie poniżej używane są bezpieczne wartości domyślne).")

            heating_fuel_price = st.session_state.get("cena_mwh_tab4", cena_mwh)
            heating_fuel_efficiency = st.session_state.get("sprawnosc_kotla_frac", 0.98)
            if st.session_state.get("typ_kotla") == "Gazowy":
                heating_fuel_price = st.session_state.get("cena_gazu_mwh", 250.0)
            heating_fuel_price = heating_fuel_price if heating_fuel_price else cena_mwh
            heating_fuel_efficiency = heating_fuel_efficiency if heating_fuel_efficiency else 1.0

            total_monthly_saving_thermal = 0.0
            total_energy_cost_el = 0.0
            calculated_times = st.session_state.get("calculated_times", {})

            for mixer in st.session_state.confirmed_mixers:
                tag = mixer["tag"]
                kat = mixer["product_family"]
                prod_info = st.session_state.active_portfolio[kat]
                m_monthly_kg = mixer["annual_volume"] / MONTHS_PER_YEAR
                batches_per_month = mixer["batches_count"]
                m_data = calculated_times.get(tag, {"power_mix_kw": 5.5, "power_pump_kw": 1.5, "heating": 1.5, "pumping": 0.75, "t_max_mix": 60.0, "t_rozlew": 30.0})
                mixing_energy = m_data["power_mix_kw"] * mixer.get("cycle_h", prod_info["cycle_h"]) * batches_per_month
                pumping_energy = m_data["power_pump_kw"] * m_data["pumping"] * batches_per_month
                cost_el = ((mixing_energy + pumping_energy) / 1000.0) * cena_mwh
                total_energy_cost_el += cost_el
                if m_data["t_rozlew"] < m_data["t_max_mix"]:
                    total_monthly_saving_thermal += ((m_monthly_kg * prod_info["cp"] * (m_data["t_max_mix"] - m_data["t_rozlew"])) / 3_600_000.0) * heating_fuel_price / heating_fuel_efficiency

            koszt_paliwa_grzewczego = st.session_state.get("koszt_paliwa_grzewczego_month", 0.0)
            process_demand_kw = st.session_state.get("process_demand_kw", 0.0)
            facility_demand_kw = st.session_state.get("facility_demand_kw", 0.0)
            godziny_rocznie_zakladu = godziny_dziennie * WORKING_DAYS_YEAR
            koszt_energii_proces_month = ((process_demand_kw * godziny_rocznie_zakladu) / 1000.0 / MONTHS_PER_YEAR) * cena_mwh
            koszt_energii_facility_month = ((facility_demand_kw * godziny_rocznie_zakladu) / 1000.0 / MONTHS_PER_YEAR) * cena_mwh

            real_energy_cost_month = (total_energy_cost_el + koszt_paliwa_grzewczego + koszt_energii_proces_month
                                       + koszt_energii_facility_month - total_monthly_saving_thermal)
            real_energy_pct_of_revenue = (real_energy_cost_month * MONTHS_PER_YEAR / total_annual_revenue_target * 100.0) \
                if total_annual_revenue_target > 0 else 0.0
            st.metric("📐 Rzeczywisty koszt energii (z inżynierii) jako % przychodu",
                      f"{real_energy_pct_of_revenue:.1f}%",
                      help=f"Realny koszt energii: {real_energy_cost_month:,.0f} {waluta}/mies. Porównaj z sumą "
                           f"'% Energia i media' wpisaną w tabeli powyżej — jeśli mocno się różnią, dostosuj %.")

            fixed_monthly_opex = koszt_energii_facility_month

        st.markdown("---")
        st.markdown("### 🧰 Krok 2: CAPEX — Nakłady Inwestycyjne")
        st.caption("Wybierz sposób określenia CAPEX zależnie od zaawansowania prac projektowych: kwota całkowita "
                   "jako szybki szacunek na wczesnym etapie, albo szczegółowy cennik instalacji z Excela, gdy masz "
                   "już rozpisany projekt (P&ID, dobrane urządzenia, ceny jednostkowe).")

        capex_mode = st.radio(
            "Sposób określenia CAPEX:",
            ["Kwota całkowita (szacunek wstępny)", "Szczegółowy cennik z Excela (zaawansowany projekt)"],
            key="capex_mode", horizontal=True
        )

        total_capex = 0.0

        if capex_mode == "Kwota całkowita (szacunek wstępny)":
            capex_lump_sum = st.number_input(
                f"Całkowity CAPEX [{waluta}]:", min_value=0.0,
                value=float(st.session_state.get("capex_lump_sum", 0.0)), step=50_000.0, format="%.0f",
                key="capex_lump_sum_input",
                help="Jedna zbiorcza kwota — instalacje, budynek, infrastruktura itd. razem. Użyj tego trybu, "
                     "dopóki nie masz jeszcze szczegółowego rozbicia kosztów."
            )
            st.session_state["capex_lump_sum"] = capex_lump_sum
            total_capex = capex_lump_sum
            if total_capex > 0:
                st.metric("💰 Całkowity CAPEX", f"{total_capex:,.0f} {waluta}")
            else:
                st.info("ℹ️ Wpisz szacunkową kwotę CAPEX powyżej, aby policzyć ROI w Kroku 4.")

        else:
            st.caption("Zdefiniuj listę komponentów standardowej instalacji per grupa produktowa (pompy, czujniki, "
                       "zawory, elektrozawory itd.) wraz z cenami jednostkowymi — treść zwykle przepisana z Twojego "
                       "istniejącego P&ID danej instalacji standardowej. Podaj, ile takich instalacji planujesz, "
                       "a aplikacja przeliczy szacunkowy CAPEX. Cennik wgrywasz ponownie, gdy ceny się zmienią.")

            equipment_template_bytes = generate_equipment_template_bytes()
            st.download_button(
                label="⬇️ Pobierz szablon Excel (Cennik_Instalacji_Szablon.xlsx)",
                data=equipment_template_bytes,
                file_name="Cennik_Instalacji_Szablon.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_download_equipment_template"
            )

            uploaded_equipment_file = st.file_uploader("Wybierz plik .xlsx z cennikiem instalacji:", type=["xlsx"], key="equipment_uploader")

            if uploaded_equipment_file is not None:
                parsed_eq_df, eq_errors = parse_equipment_excel(uploaded_equipment_file)
                for err in eq_errors:
                    st.warning(f"⚠️ {err}")
                if parsed_eq_df is not None and not parsed_eq_df.empty:
                    st.session_state.equipment_df = parsed_eq_df
                    st.success(f"✅ Wczytano {len(parsed_eq_df)} pozycji cennika.")
                elif parsed_eq_df is None:
                    st.error("❌ Nie udało się wczytać żadnych poprawnych pozycji z tego pliku — popraw błędy powyżej i wgraj ponownie.")

            if st.session_state.equipment_df is not None and not st.session_state.equipment_df.empty:
                edited_eq_df = st.data_editor(
                    st.session_state.equipment_df, hide_index=True, use_container_width=True,
                    num_rows="dynamic", key="equipment_data_editor",
                    column_config={EQUIPMENT_GROUP_COL: st.column_config.SelectboxColumn(options=RECIPE_PRODUCT_GROUPS)}
                )
                edited_eq_df[EQUIPMENT_LINE_TOTAL_COL] = edited_eq_df[EQUIPMENT_QTY_COL] * edited_eq_df[EQUIPMENT_UNIT_PRICE_COL]
                st.session_state.equipment_df = edited_eq_df

                groups_in_price_list = sorted(edited_eq_df[EQUIPMENT_GROUP_COL].dropna().unique().tolist())

                # Domyślna liczba instalacji = liczba MIESZALNIKÓW tej grupy faktycznie
                # zatwierdzonych we flocie (Zakładka 2) - to jest realna liczba fizycznych
                # instalacji, którą CAPEX powinien odzwierciedlać (1 mieszalnik = 1 standardowa
                # instalacja tej grupy). Jeśli flota nie jest jeszcze zatwierdzona, spadamy do
                # liczby unikalnych produktów w recepturze jako przybliżenia, a przy jej braku - 1.
                fleet_counts_by_group = {}
                if st.session_state.confirmed_mixers:
                    for m in st.session_state.confirmed_mixers:
                        fleet_counts_by_group[m["product_family"]] = fleet_counts_by_group.get(m["product_family"], 0) + 1
                default_counts_from_recipes = {}
                if st.session_state.recipes_df is not None and not st.session_state.recipes_df.empty:
                    default_counts_from_recipes = st.session_state.recipes_df.groupby(RECIPE_GROUP_COL)[RECIPE_PRODUCT_COL].nunique().to_dict()

                c_sync1, c_sync2 = st.columns([3, 1])
                with c_sync1:
                    if fleet_counts_by_group:
                        st.caption("💡 Domyślna liczba instalacji = liczba mieszalników tej grupy zatwierdzonych "
                                   "we Zakładce 2 (1 mieszalnik = 1 instalacja). Możesz nadpisać ręcznie poniżej.")
                    else:
                        st.caption("ℹ️ Flota nie jest jeszcze zatwierdzona (Zakładka 2) — domyślna liczba instalacji "
                                   "pochodzi z liczby produktów w recepturze (przybliżenie), nie z realnej floty.")
                with c_sync2:
                    if st.button("🔄 Zsynchronizuj z flotą", key="btn_sync_capex_with_fleet",
                                  disabled=not fleet_counts_by_group, use_container_width=True,
                                  help="Nadpisuje liczbę instalacji dla wszystkich grup wartościami z aktualnie "
                                       "zatwierdzonej floty (Zakładka 2) - przydatne, gdy flota zmieniła się po "
                                       "pierwszym ustawieniu tych liczb."):
                        for grp in groups_in_price_list:
                            st.session_state.equipment_install_counts[grp] = fleet_counts_by_group.get(grp, 0)
                        st.rerun()

                cols_counts = st.columns(min(len(groups_in_price_list), 4)) if groups_in_price_list else []
                for i, grp in enumerate(groups_in_price_list):
                    with cols_counts[i % len(cols_counts)]:
                        default_val = int(st.session_state.equipment_install_counts.get(
                            grp, fleet_counts_by_group.get(grp, default_counts_from_recipes.get(grp, 1))))
                        st.session_state.equipment_install_counts[grp] = st.number_input(
                            f"{grp} — instalacji:", min_value=0, value=default_val, step=1, key=f"eq_count_{grp}"
                        )

                capex_rows = []
                for grp in groups_in_price_list:
                    n_install = st.session_state.equipment_install_counts.get(grp, 0)
                    grp_df = edited_eq_df[edited_eq_df[EQUIPMENT_GROUP_COL] == grp]
                    per_install_cost = grp_df[EQUIPMENT_LINE_TOTAL_COL].sum()
                    group_total = per_install_cost * n_install
                    total_capex += group_total
                    currencies = grp_df[EQUIPMENT_CURRENCY_COL].dropna().unique().tolist()
                    currency_label = currencies[0] if len(currencies) == 1 else "/".join(currencies) if currencies else "—"
                    capex_rows.append({
                        "Grupa Produktowa": grp, "Komponentów w cenniku": len(grp_df),
                        "Koszt 1 instalacji": round(per_install_cost, 2), "Liczba instalacji": n_install,
                        "CAPEX grupy": round(group_total, 2), "Waluta": currency_label,
                    })

                st.dataframe(pd.DataFrame(capex_rows), hide_index=True, use_container_width=True)
                st.metric("💰 Łączny szacowany CAPEX (wszystkie grupy)", f"{total_capex:,.0f}")

                with st.expander("📋 Szczegółowa lista komponentów per grupa", expanded=False):
                    for grp in groups_in_price_list:
                        st.markdown(f"**{grp}**")
                        st.dataframe(edited_eq_df[edited_eq_df[EQUIPMENT_GROUP_COL] == grp].drop(columns=[EQUIPMENT_GROUP_COL]),
                                     hide_index=True, use_container_width=True)
            else:
                st.info("💡 Wgraj cennik powyżej, aby zobaczyć tu podsumowanie CAPEX per grupa produktowa.")

        if total_capex > 0:
            st.markdown("---")
            contingency_pct = st.slider(
                "Rezerwa na nieprzewidziane [%]:", min_value=0, max_value=30, value=15, step=1,
                key="capex_contingency_pct",
                help="Standardowa praktyka przy szacowaniu CAPEX na wczesnym etapie projektu — bufor na "
                     "nieprzewidziane koszty (zwykle 10-20%)."
            )
            capex_before_contingency = total_capex
            total_capex = capex_before_contingency * (1.0 + contingency_pct / 100.0)
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                st.metric("CAPEX bazowy", f"{capex_before_contingency:,.0f} {waluta}")
            with cc2:
                st.metric(f"+ Rezerwa ({contingency_pct}%)", f"{capex_before_contingency * contingency_pct / 100.0:,.0f} {waluta}")
            with cc3:
                st.metric("💰 CAPEX razem (z rezerwą)", f"{total_capex:,.0f} {waluta}")

        st.markdown("---")
        st.markdown("### 📈 Krok 3: ROI (z uwzględnieniem krzywej rozruchu)")
        st.caption("Przychód i koszt liczone **per grupa produktowa** (ceny/koszty z Kroku 1), każda skalowana "
                   "WŁASNĄ krzywą rozruchu (panel boczny). Doliczana jest też stała energia pozaprodukcyjna "
                   "(z rozwijanego panelu w Kroku 1, jeśli skonfigurowana).")

        # Ta sama krzywa rozruchu co w panelu bocznym, per grupa - liczona tu od nowa na bazie
        # confirmed_mixers, żeby ROI nie zależało od tego, czy użytkownik odwiedził Zakładkę 2.
        target_annual_t_fin = sum(m["annual_volume"] for m in st.session_state.confirmed_mixers) / 1000.0
        roi_rows = []
        energy_kpi_rows = []  # do raportu PDF i wglądu tutaj: rozbicie energii per rok (szacunek uproszczony)
        cumulative_profit = 0.0
        payback_year_fraction = None
        for i in range(RAMPUP_YEARS):
            annual_revenue_year = 0.0
            annual_variable_cost_year = 0.0
            year_tonnage_t = 0.0
            for grp in active_groups_fin:
                density_kg_l = st.session_state.active_portfolio[grp]["density"]
                annual_volume_kg_target = sum(m["annual_volume"] for m in st.session_state.confirmed_mixers if m["product_family"] == grp)
                annual_volume_l_target = annual_volume_kg_target / density_kg_l if density_kg_l > 0 else 0.0
                frac = get_rampup_fraction(grp, i)
                vol_l_year = annual_volume_l_target * frac
                year_tonnage_t += (annual_volume_kg_target / 1000.0) * frac

                pricing = st.session_state.group_pricing.get(grp, {
                    "sales_price": 2.0, "raw_material_cost": 1.35, "labour": 0.07, "energy": 0.04,
                    "qc": 0.04, "maintenance": 0.04, "logistics": 0.04, "overhead": 0.04,
                })
                cost_per_l = pricing["raw_material_cost"] + sum(
                    pricing[k] for k in ["labour", "energy", "qc", "maintenance", "logistics", "overhead"]
                )
                annual_revenue_year += vol_l_year * pricing["sales_price"]
                annual_variable_cost_year += vol_l_year * cost_per_l

            frac_year_blended = (year_tonnage_t / target_annual_t_fin) if target_annual_t_fin > 0 else 1.0
            annual_fixed_opex = fixed_monthly_opex * MONTHS_PER_YEAR
            annual_opex_year = annual_variable_cost_year + annual_fixed_opex
            annual_profit_year = annual_revenue_year - annual_opex_year

            # Rozbicie energii per rok - uproszczony szacunek (do raportów/Dashboardu), skalowany
            # tą samą (ważoną wolumenem) krzywą co reszta OPEX-u zmiennego.
            heating_annual_year = koszt_paliwa_grzewczego * MONTHS_PER_YEAR * frac_year_blended
            electricity_process_annual_year = (total_energy_cost_el + koszt_energii_proces_month) * MONTHS_PER_YEAR * frac_year_blended
            electricity_facility_annual_year = fixed_monthly_opex * MONTHS_PER_YEAR
            energy_kpi_rows.append({
                "Rok": f"Rok {i + 1}",
                "Ogrzewanie [waluta/rok]": round(heating_annual_year, 0),
                "Elektryczność - proces (w tym chłodzenie) [waluta/rok]": round(electricity_process_annual_year, 0),
                "Elektryczność - pozaprodukcyjne (stałe) [waluta/rok]": round(electricity_facility_annual_year, 0),
                "Energia razem [waluta/rok]": round(heating_annual_year + electricity_process_annual_year + electricity_facility_annual_year, 0),
            })

            profit_before = cumulative_profit
            cumulative_profit += annual_profit_year
            if payback_year_fraction is None and total_capex > 0 and cumulative_profit >= total_capex:
                needed = total_capex - profit_before
                payback_year_fraction = i + (needed / annual_profit_year if annual_profit_year > 0 else 0.0)

            roi_rows.append({
                "Rok": f"Rok {i + 1}", "% Celu": f"{frac_year_blended * 100.0:.0f}%",
                "OPEX roczny": round(annual_opex_year, 0), "Przychód roczny": round(annual_revenue_year, 0),
                "Zysk roczny": round(annual_profit_year, 0), "Zysk skumulowany": round(cumulative_profit, 0),
                "ROI (ten rok) [%]": round((annual_profit_year / total_capex * 100.0), 1) if total_capex > 0 else None,
            })

        st.dataframe(pd.DataFrame(roi_rows), hide_index=True, use_container_width=True)
        st.session_state["roi_rows_report"] = roi_rows  # do raportu Excel/PDF
        st.session_state["energy_kpi_rows_report"] = energy_kpi_rows  # do raportu PDF

        r_c1, r_c2, r_c3, r_c4 = st.columns(4)
        with r_c1:
            st.metric("💰 OPEX w Roku 1 (rozruch)", f"{roi_rows[0]['OPEX roczny']:,.0f} {waluta}")
        with r_c2:
            st.metric("💵 Zysk w Roku 5 (pełna dojrzałość)", f"{roi_rows[-1]['Zysk roczny']:,.0f} {waluta}")
        with r_c3:
            roi_year5 = roi_rows[-1]["ROI (ten rok) [%]"]
            st.metric("🎯 ROI w Roku 5 (ustabilizowane)", f"{roi_year5:.1f}%" if roi_year5 is not None else "—")
        with r_c4:
            if total_capex <= 0:
                st.metric("⏳ Okres zwrotu (z rozruchem)", "—", help="Uzupełnij cennik CAPEX w Kroku 2 powyżej.")
            elif payback_year_fraction is not None:
                st.metric("⏳ Okres zwrotu (z rozruchem)", f"{payback_year_fraction:.1f} lat")
            else:
                st.metric("⏳ Okres zwrotu (z rozruchem)", f"> {RAMPUP_YEARS} lat",
                          help="Skumulowany zysk nie pokrywa CAPEX nawet w Roku 5 przy obecnych założeniach.")

        if total_capex <= 0:
            st.info("ℹ️ ROI wymaga policzonego CAPEX — uzupełnij cennik instalacji w Kroku 2 powyżej.")

        # Zapis do session_state, żeby nowa Zakładka Dashboard mogła pokazać te liczby na "pierwszy rzut oka".
        st.session_state["total_capex_report"] = total_capex
        st.session_state["payback_year_fraction_report"] = payback_year_fraction
        st.session_state["roi_year5_report"] = roi_rows[-1]["ROI (ten rok) [%]"] if roi_rows else None
        st.session_state["waluta_report"] = waluta

        st.markdown("---")
        st.markdown("### 📄 Krok 4: Raport 5-letni (PDF / Excel)")
        st.caption("Zbiera w jeden dokument to, co już policzone w aplikacji: skalę produkcji per rok, produkty, "
                   "produkcja własna vs import, flotę (mieszalniki), wykorzystanie magazynu i KPI energetyczne. "
                   "**PDF** — czytelny, sformatowany dokument (po angielsku) do pokazania/udostępnienia. **Excel** "
                   "— te same dane w formie tabel + natywne, edytowalne wykresy, do dalszej pracy (przestawianie, "
                   "własne zestawienia). Odwiedź Zakładkę 3 (wykres stanu magazynowego), żeby dane magazynowe w "
                   "raporcie były aktualne — inaczej te pola pokażą 'n/a'.")

        c_rep1, c_rep2 = st.columns(2)
        with c_rep1:
            if st.button("📄 Wygeneruj raport PDF", key="btn_generate_pdf_report", use_container_width=True):
                report_data = compute_pdf_report_year_data()
                if report_data is None:
                    st.error("❌ Brak zatwierdzonej floty (Zakładka 5) — nie ma czego raportować.")
                else:
                    try:
                        pdf_bytes = generate_pdf_report_bytes(report_data, waluta)
                        st.session_state["pdf_report_bytes"] = pdf_bytes
                        st.success("✅ Raport PDF wygenerowany — pobierz poniżej.")
                    except ImportError as exc:
                        st.error(f"❌ Brakująca biblioteka do generowania PDF: {exc}. Dopisz do `requirements.txt` na "
                                 f"Streamlit Cloud: `reportlab` oraz `matplotlib` (jeśli jeszcze nie ma), zapisz plik i "
                                 f"poczekaj na automatyczne ponowne wdrożenie aplikacji, potem spróbuj ponownie.")
                    except Exception as exc:
                        st.error(f"❌ Nie udało się wygenerować raportu PDF: {exc}")

            if st.session_state.get("pdf_report_bytes"):
                st.download_button(
                    label="⬇️ Pobierz raport PDF",
                    data=st.session_state["pdf_report_bytes"],
                    file_name="5_Year_Production_Scaleup_Report.pdf",
                    mime="application/pdf",
                    key="btn_download_pdf_report",
                    use_container_width=True
                )

        with c_rep2:
            if st.button("📊 Wygeneruj raport Excel", key="btn_generate_excel_report", use_container_width=True):
                report_data = compute_pdf_report_year_data()
                if report_data is None:
                    st.error("❌ Brak zatwierdzonej floty (Zakładka 5) — nie ma czego raportować.")
                else:
                    try:
                        roi_rows_for_report = st.session_state.get("roi_rows_report", [])
                        xlsx_bytes = generate_excel_report_bytes(report_data, roi_rows_for_report, waluta)
                        st.session_state["excel_report_bytes"] = xlsx_bytes
                        st.success("✅ Raport Excel wygenerowany — pobierz poniżej.")
                    except Exception as exc:
                        st.error(f"❌ Nie udało się wygenerować raportu Excel: {exc}")

            if st.session_state.get("excel_report_bytes"):
                st.download_button(
                    label="⬇️ Pobierz raport Excel",
                    data=st.session_state["excel_report_bytes"],
                    file_name="5_Year_Production_Scaleup_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_excel_report",
                    use_container_width=True
                )

# ==========================================
# ZAKŁADKA 2: PARK ZBIORNIKÓW (TANK FARM) (tab5)
# ==========================================
with tab5:
    st.header("🛢️ Logistyka Surowcowa i Grupy Magazynowe (Tank Farm)")
    if not st.session_state.confirmed_mixers:
        st.warning("⚠️ Brak danych technicznych. Uruchom konfigurację w Zakładce 1.")
    else:
        _stale_msg = check_fleet_staleness_warning()
        if _stale_msg:
            st.error(_stale_msg)
        days_of_stock = st.number_input("Wymagany zapas bezpieczeństwa surowca [dni]:", min_value=5, value=14)
        st.session_state["days_of_stock_tab5"] = days_of_stock
        max_single_tank_m3 = st.slider(
            "Maksymalna pojemność pojedynczego zbiornika [m³]:", min_value=10, max_value=250, value=100, step=5,
            help="Górny limit dla JEDNEGO fizycznego zbiornika. Dla każdego surowca aplikacja dobierze najmniejszy "
                 "standardowy rozmiar, który mieści cały wymagany bufor w jednym zbiorniku — a jeśli surowiec "
                 "potrzebuje więcej niż ten limit, zaproponuje kilka zbiorników o tej maksymalnej pojemności."
        )

        st.markdown("---")
        st.markdown("### 🏢 Wymiarowanie Silosów Magazynowych")

        recipe_consumption_target = collapse_shared_tank_materials(apply_supplier_splits(
            compute_rm_consumption_for_year(RAMPUP_YEAR_TARGET_SENTINEL), RAMPUP_YEAR_TARGET_SENTINEL))
        has_recipe_consumption = any(v > 0 for v in recipe_consumption_target.values())

        if has_recipe_consumption:
            # Ten sam rok co w Zakładce 3 (odczyt, nie osobny widget - Streamlit nie pozwala na
            # dwa widżety z tym samym kluczem w jednym przebiegu skryptu). Zakładka 3 renderuje
            # się wcześniej w kolejności zakładek, więc jej wybór jest tu już dostępny.
            selected_rm_year_label = st.session_state.get("tab3_rampup_year_select", "Docelowa produkcja (100%)")
            st.caption(f"📈 Rok symulacji rozruchu: **{selected_rm_year_label}** (ten sam wybór co w Zakładce 3 — "
                       "zmień go tam). Silosy (zbiorniki dedykowane) są dobierane RAZ, pod docelowe (100%) zużycie, "
                       "budynek/instalacja stawiana raz. Wybór roku wpływa na pokazane BIEŻĄCE zużycie i na liczbę "
                       "beczek/IBC/palet w danym roku — surowce potrzebne WYŁĄCZNIE do produktów jeszcze "
                       "importowanych w wybranym roku są pomijane.")
            selected_rm_year_idx = None if selected_rm_year_label.startswith("Docelowa") else int(selected_rm_year_label.split(" ")[1]) - 1
            effective_rm_year_idx = selected_rm_year_idx if selected_rm_year_idx is not None else RAMPUP_YEAR_TARGET_SENTINEL
            recipe_consumption = (recipe_consumption_target if selected_rm_year_idx is None
                                   else collapse_shared_tank_materials(apply_supplier_splits(
                                       compute_rm_consumption_for_year(selected_rm_year_idx), selected_rm_year_idx)))

            st.caption("Liczone **per pojedynczy surowiec** (np. osobno Base Oil II i Base Oil III) na podstawie "
                       "receptur wgranych w **Zakładce 1**. Dla każdego surowca sprawdzane jest też (a) czy fizycznie/"
                       "praktycznie nadaje się do magazynowania luzem w zbiorniku, oraz (b) czy roczne zużycie "
                       "przekracza próg opłacalności dedykowanego zbiornika — jeśli oba warunki są spełnione, "
                       "proponowany jest zbiornik o określonej pojemności; w przeciwnym razie beczki/IBC/worki. "
                       "**Silosy dobierane są zawsze pod docelowe (100%) zużycie** (budowane raz); poniższa tabela "
                       "pokazuje też, ile z tej pojemności realnie wykorzystujesz w wybranym roku/widoku.")

            c_thr1, c_thr2 = st.columns(2)
            with c_thr1:
                prog_zbiornika_t = st.number_input(
                    "Próg rocznego zużycia do zbiornika dedykowanego [t/rok]:", min_value=1.0, value=50.0, step=5.0,
                    key="prog_zbiornika_tab6",
                    help="Poniżej tego wolumenu zbiornik dedykowany zwykle się nie zwraca — surowiec zostaje w "
                         "beczkach/IBC, nawet jeśli fizycznie nadaje się do magazynowania luzem."
                )
            with c_thr2:
                st.caption(f"Zapas bezpieczeństwa i maks. pojemność zbiornika jak ustawione powyżej ({days_of_stock:.0f} dni, "
                           f"{max_single_tank_m3} m³).")

            STANDARD_SMALL_TANK_SIZES_M3 = [10, 15, 20, 25, 30, 40, 50, 60, 80, 100, 125, 150, 200, 250]

            with st.expander("ℹ️ Jak liczony jest wymagany bufor i dobór zbiornika?", expanded=False):
                st.markdown(
                    "1. **Dzienne zużycie** = roczne zużycie surowca ÷ dni robocze w roku.\n"
                    "2. **Wymagany bufor [m³]** = (dzienne zużycie × zapas bezpieczeństwa [dni]) ÷ współczynnik "
                    "napełnienia zbiornika (uwzględnia gęstość i margines na rozszerzalność cieplną).\n"
                    "3. **Dobór zbiornika** — jeśli wymagany bufor mieści się w jednym zbiorniku poniżej ustawionego "
                    "limitu maksymalnej pojemności, wybierany jest **najmniejszy standardowy rozmiar**, który go "
                    "pomieści (przy uwzględnieniu współczynnika bezpiecznego napełnienia — zbiornik nigdy nie jest "
                    "wypełniany do 100%). Jeśli bufor przekracza limit, aplikacja proponuje **kilka zbiorników o "
                    "maksymalnej dozwolonej pojemności** — stąd 'Liczba silosów' i 'Rekomendacja' zawsze opisują "
                    "TEN SAM, spójny wariant (wcześniej to były dwa niezależne, rozjeżdżające się wyliczenia)."
                )

            # --- Sposób magazynowania: automatyczna sugestia + RĘCZNE nadpisanie per surowiec ---
            # Algorytm sugeruje "Zbiornik" tylko gdy fizycznie się nadaje (bulk_eligible) I przekracza
            # próg opłacalności - ale to tylko sugestia. Materiał sypki, dostępny wyłącznie w
            # paczkach/workach, albo z innego powodu niemagazynowalny luzem MUSI dać się ręcznie
            # skierować do beczek/IBC/worków, niezależnie od sugestii.
            if "rm_storage_method_override" not in st.session_state:
                st.session_state.rm_storage_method_override = {}

            override_rows = []
            for material, annual_tony_target in sorted(recipe_consumption_target.items(), key=lambda x: -x[1]):
                if annual_tony_target <= 0:
                    continue
                info = RAW_MATERIAL_STORAGE_INFO.get(material, {"bulk_eligible": True, "note": "Brak danych - domyślnie traktowany jak ciecz magazynowalna luzem."})
                auto_suggestion = "Zbiornik (luzem)" if (info["bulk_eligible"] and annual_tony_target >= prog_zbiornika_t) else "Beczki / IBC / worki"
                current_choice = st.session_state.rm_storage_method_override.get(material, auto_suggestion)
                override_rows.append({
                    "Surowiec": material, "Zużycie docelowe [t/rok]": round(annual_tony_target, 1),
                    "Sugestia algorytmu": auto_suggestion, "Sposób magazynowania": current_choice,
                })

            st.markdown("##### 🔀 Sposób magazynowania per surowiec (możesz nadpisać sugestię)")
            st.caption("Np. materiał sypki, dostępny tylko w paczkach/workach, albo z innego powodu "
                       "niemagazynowalny luzem — zmień 'Sposób magazynowania' na 'Beczki / IBC / worki', a "
                       "przeliczy się to poprawnie do magazynu RM poniżej.")
            edited_override_df = st.data_editor(
                pd.DataFrame(override_rows), hide_index=True, use_container_width=True,
                disabled=["Surowiec", "Zużycie docelowe [t/rok]", "Sugestia algorytmu"], key="rm_storage_override_editor",
                column_config={
                    "Sposób magazynowania": st.column_config.SelectboxColumn(options=["Zbiornik (luzem)", "Beczki / IBC / worki"]),
                }
            )
            for _, r in edited_override_df.iterrows():
                st.session_state.rm_storage_method_override[r["Surowiec"]] = r["Sposób magazynowania"]

            recipe_silos_rows = []
            recipe_total_tanks = 0
            drummed_materials = []  # surowce NIE trafiające do zbiornika - potrzebują miejsca w magazynie
            dedicated_tank_candidates = []  # surowce KWALIFIKUJĄCE się do zbiornika dedykowanego - do zatwierdzenia niżej
            for material, annual_tony_target in sorted(recipe_consumption_target.items(), key=lambda x: -x[1]):
                if annual_tony_target <= 0:
                    continue
                annual_tony_year = recipe_consumption.get(material, 0.0)
                info = RAW_MATERIAL_STORAGE_INFO.get(material, {"bulk_eligible": True, "note": "Brak danych - domyślnie traktowany jak ciecz magazynowalna luzem."})
                bulk_ok = info["bulk_eligible"]
                recommend_tank = st.session_state.rm_storage_method_override.get(material) == "Zbiornik (luzem)"

                daily_t = annual_tony_target / WORKING_DAYS_YEAR
                is_manual_override = st.session_state.rm_storage_method_override.get(material) != (
                    "Zbiornik (luzem)" if (bulk_ok and annual_tony_target >= prog_zbiornika_t) else "Beczki / IBC / worki")
                if recommend_tank:
                    required_m3 = (daily_t * days_of_stock) / OIL_FILL_FACTOR
                    # Netto pojemność potrzebna PO uwzględnieniu współczynnika bezpiecznego napełnienia -
                    # jedna wspólna liczba, z której wynikają OBIE kolumny (liczba I pojemność), więc zawsze
                    # są ze sobą spójne.
                    required_m3_gross = required_m3 / TANK_SAFETY_FILL
                    if required_m3_gross <= max_single_tank_m3:
                        needed_tanks = 1
                        recommended_capacity = next((s for s in STANDARD_SMALL_TANK_SIZES_M3 if s >= required_m3_gross), max_single_tank_m3)
                    else:
                        recommended_capacity = max_single_tank_m3
                        needed_tanks = math.ceil(required_m3_gross / max_single_tank_m3)
                    recipe_total_tanks += needed_tanks
                    rekomendacja = f"🛢️ {needed_tanks}× zbiornik dedykowany ({recommended_capacity:.0f} m³)"
                    uzasadnienie = (f"Zużycie {annual_tony_target:.1f} t/rok (docelowo) ≥ próg {prog_zbiornika_t:.0f} t/rok, nadaje się do magazynowania luzem."
                                     if not is_manual_override else "Ręcznie wybrane powyżej (odbiega od sugestii algorytmu).")
                    bufor_txt = f"{required_m3:.1f}"
                    silosy_txt = f"{needed_tanks} szt."
                    wykorzystanie_txt = f"{(annual_tony_year / annual_tony_target * 100.0):.0f}%" if annual_tony_target > 0 else "—"
                    dedicated_tank_candidates.append({
                        "material": material, "needed_tanks": needed_tanks, "recommended_capacity": recommended_capacity,
                    })
                else:
                    rekomendacja = "🧴 Beczki / IBC / worki"
                    if is_manual_override:
                        uzasadnienie = "Ręcznie wybrane powyżej (odbiega od sugestii algorytmu) — np. materiał sypki/w paczkach."
                    else:
                        uzasadnienie = info["note"] if not bulk_ok else f"Zużycie {annual_tony_target:.1f} t/rok (docelowo) < próg {prog_zbiornika_t:.0f} t/rok — zbiornik się nie opłaca."
                    bufor_txt = "—"
                    silosy_txt = "—"
                    wykorzystanie_txt = "—"
                    if annual_tony_year > 0:
                        drummed_materials.append({"material": material, "annual_tony": annual_tony_year, "info": info})

                recipe_silos_rows.append({
                    "Surowiec": material, "Konsumpcja [t/rok] (docelowo)": round(annual_tony_target, 2),
                    "Konsumpcja [t/rok] (ten rok)": round(annual_tony_year, 2),
                    "Wymagany Bufor [m³]": bufor_txt, "Liczba silosów": silosy_txt,
                    "Wykorzystanie w tym roku": wykorzystanie_txt,
                    "Rekomendacja": rekomendacja, "Uzasadnienie": uzasadnienie,
                })

            st.dataframe(pd.DataFrame(recipe_silos_rows), hide_index=True, use_container_width=True)
            m_silo1, m_silo2 = st.columns(2)
            with m_silo1:
                st.metric("🧱 Całkowita liczba silosów (surowce w zbiornikach, docelowo)", f"{recipe_total_tanks} szt.")
            with m_silo2:
                n_drums = sum(1 for r in recipe_silos_rows if "Beczki" in r["Rekomendacja"])
                st.metric("🧴 Surowce zostające w beczkach/IBC", f"{n_drums} / {len(recipe_silos_rows)}")

            # --- Zbiorniki buforowe dla produktów IMPORTOWANYCH NA STAŁE, oznaczonych w Zakładce 1
            # jako "Nigdy (bufor)" - to gotowy produkt, nie surowiec (nie liczy się do zużycia RM,
            # nie dostaje mieszalnika), ale mimo to potrzebuje lokalnego zbiornika (np. żeby baza
            # się nie rozwarstwiła) zamiast zwykłych palet importowych. Ta sama logika wymiarowania
            # (dni zapasu × wolumen), tylko wolumen produktu zamiast zużycia surowca.
            buffer_tank_candidates = []
            if st.session_state.recipes_df is not None and not st.session_state.recipes_df.empty and RECIPE_IMPORT_TRANSITION_COL in st.session_state.recipes_df.columns:
                buffer_products_df = st.session_state.recipes_df[
                    (st.session_state.recipes_df[RECIPE_SOURCING_COL] == "Import") &
                    (st.session_state.recipes_df[RECIPE_IMPORT_TRANSITION_COL] == "Nigdy (bufor)")
                ]
                if not buffer_products_df.empty:
                    st.markdown("###### 🔵 Zbiorniki buforowe — produkty importowane na stałe, wymagające lokalnego bufora")
                    st.caption("Oznaczone w Zakładce 1 jako 'Nigdy (bufor)' — gotowy produkt, nie surowiec (nie liczy "
                               "się do zużycia RM, nie ma mieszalnika), ale dostaje własny zbiornik zamiast miejsca "
                               "paletowego w buforze importu. Trafią do tej samej tabeli zbiorników RM w Karcie "
                               "Maszyn — z pompą, hydrauliką i opcjonalnym mieszaniem, żeby zapobiec sedymentacji.")
                    buffer_rows_display = []
                    for _, r in buffer_products_df.iterrows():
                        product_name = r[RECIPE_PRODUCT_COL]
                        annual_t = float(r.get(RECIPE_ANNUAL_COL, 0) or 0)
                        if annual_t <= 0:
                            continue
                        daily_t_buf = annual_t / WORKING_DAYS_YEAR
                        required_m3_buf = (daily_t_buf * days_of_stock) / OIL_FILL_FACTOR
                        required_m3_gross_buf = required_m3_buf / TANK_SAFETY_FILL
                        if required_m3_gross_buf <= max_single_tank_m3:
                            needed_tanks_buf = 1
                            recommended_capacity_buf = next((s for s in STANDARD_SMALL_TANK_SIZES_M3 if s >= required_m3_gross_buf), max_single_tank_m3)
                        else:
                            recommended_capacity_buf = max_single_tank_m3
                            needed_tanks_buf = math.ceil(required_m3_gross_buf / max_single_tank_m3)
                        material_label = f"🔵 Produkt (bufor): {product_name}"
                        buffer_tank_candidates.append({
                            "material": material_label, "needed_tanks": needed_tanks_buf, "recommended_capacity": recommended_capacity_buf,
                        })
                        buffer_rows_display.append({
                            "Produkt": product_name, "Linia": r[RECIPE_GROUP_COL], "Wolumen [t/rok]": round(annual_t, 1),
                            "Wymagany bufor [m³]": round(required_m3_buf, 1), "Zbiorników": needed_tanks_buf,
                            "Pojemność 1 zbiornika [m³]": round(recommended_capacity_buf, 1),
                        })
                    st.dataframe(pd.DataFrame(buffer_rows_display), hide_index=True, use_container_width=True)
                    dedicated_tank_candidates.extend(buffer_tank_candidates)

            st.markdown("---")
            st.markdown("### ✅ Zatwierdź Zbiorniki RM (liczba i pojemność)")
            st.caption("Powyższe to sugestia — tutaj **deklarujesz ostateczną** liczbę zbiorników i ich pojemność per "
                       "surowiec (możesz nadpisać sugestię, np. jeśli wolisz mniej, większych zbiorników). Po "
                       "zatwierdzeniu ta lista pojawi się w **Karcie Maszyn**, obok mieszalników, żebyś mógł "
                       "przypisać im pompy (dedykowane/współdzielone) i rurociąg — dziś to była luka powodująca "
                       "rozbieżności w bilansie mocy pomp.")

            if dedicated_tank_candidates:
                rm_tank_editor_rows = [
                    {"Surowiec": c["material"], "Liczba zbiorników": c["needed_tanks"],
                     "Pojemność 1 zbiornika [m³]": round(c["recommended_capacity"], 1)}
                    for c in dedicated_tank_candidates
                ]
                edited_rm_tanks_df = st.data_editor(
                    pd.DataFrame(rm_tank_editor_rows), hide_index=True, use_container_width=True,
                    disabled=["Surowiec"], key="rm_tank_editor",
                    column_config={
                        "Liczba zbiorników": st.column_config.NumberColumn(min_value=1, step=1),
                        "Pojemność 1 zbiornika [m³]": st.column_config.NumberColumn(min_value=0.5, step=0.5),
                    }
                )
                if st.button("📥 Zatwierdź zbiorniki RM", type="primary", key="btn_confirm_rm_tanks"):
                    confirmed_rm_tanks = []
                    tag_counter = 1
                    for _, row in edited_rm_tanks_df.iterrows():
                        material_name = row["Surowiec"]
                        n_tanks_confirmed = int(row["Liczba zbiorników"])
                        cap_confirmed = float(row["Pojemność 1 zbiornika [m³]"])
                        for t_idx in range(n_tanks_confirmed):
                            confirmed_rm_tanks.append({
                                "tag": f"T-RM-{tag_counter}" + (f"-{t_idx+1}" if n_tanks_confirmed > 1 else ""),
                                "material": material_name, "capacity_m3": cap_confirmed,
                            })
                        tag_counter += 1
                    st.session_state["confirmed_rm_tanks"] = confirmed_rm_tanks
                    # Karta Maszyn (Zakładka 3) wykonuje się WCZEŚNIEJ w skrypcie niż ta zakładka
                    # (Magazynowanie) - bez wymuszenia ponownego przebiegu, w TYM renderze Karta
                    # Maszyn zdążyłaby już przeczytać STARĄ (sprzed kliknięcia) wartość. st.rerun()
                    # startuje świeży przebieg, w którym Karta Maszyn od razu widzi zaktualizowany stan.
                    # (Komunikat st.success() pominięty celowo - i tak nie zdążyłby się wyświetlić
                    # przed przebudową strony; potwierdzeniem jest podpis poniżej.)
                    st.rerun()

                if st.session_state.get("confirmed_rm_tanks"):
                    st.success(f"✅ Zatwierdzone: {len(st.session_state['confirmed_rm_tanks'])} zbiorników RM — widoczne w Karcie Maszyn (Zakładka 3).")
            else:
                st.info("ℹ️ Żaden surowiec nie kwalifikuje się dziś do zbiornika dedykowanego (wszystko w "
                        "beczkach/IBC/workach) — nic do zatwierdzenia.")
                st.session_state["confirmed_rm_tanks"] = []

            st.markdown("---")
            st.markdown("### 🛢️ Zbiorniki Buforowe FG (Gotowy Produkt)")
            st.caption("Dla produktów wysyłanych **cysterną luzem** ('Cysterna (luzem)' w rozbiciu na opakowania, "
                       "Zakładka 1) — zbiornik buforowy pozwala uwolnić mieszalnik zaraz po szarży (produkt "
                       "przechodzi do bufora, mieszalnik startuje kolejną szarżę) i jest bezpośrednim źródłem "
                       "wysyłki cysterną, bez pośredniego składowania na paletach.")

            recipes_df_fgbuf = st.session_state.get("recipes_df")
            fg_buffer_candidates = []
            if recipes_df_fgbuf is not None and not recipes_df_fgbuf.empty:
                tanker_pct_col = recipe_pack_pct_col("Cysterna (luzem)")
                if tanker_pct_col in recipes_df_fgbuf.columns:
                    for m in st.session_state.confirmed_mixers:
                        recipe_product_fgbuf = m.get("recipe_product")
                        if not recipe_product_fgbuf:
                            continue
                        match_fgbuf = recipes_df_fgbuf[recipes_df_fgbuf[RECIPE_PRODUCT_COL] == recipe_product_fgbuf]
                        if match_fgbuf.empty or float(match_fgbuf.iloc[0].get(tanker_pct_col, 0) or 0) <= 0:
                            continue
                        density_fgbuf = st.session_state.active_portfolio.get(m["product_family"], {}).get("density", 0.9)
                        # Sugestia: ~1,5x masy szarży, żeby zmieścić szarżę + margines na kolejną,
                        # zanim cysterna zdąży odebrać poprzednią.
                        suggested_capacity_m3 = round((m["mass_per_batch"] * 1.5 / 1000.0) / density_fgbuf, 1)
                        fg_buffer_candidates.append({
                            "recipe_product": recipe_product_fgbuf, "mixer_tag": m["tag"],
                            "suggested_capacity_m3": suggested_capacity_m3,
                        })

            if fg_buffer_candidates:
                fg_buf_editor_rows = [
                    {"Produkt": c["recipe_product"], "Mieszalnik": c["mixer_tag"],
                     "Pojemność zbiornika [m³]": c["suggested_capacity_m3"]}
                    for c in fg_buffer_candidates
                ]
                edited_fg_buf_df = st.data_editor(
                    pd.DataFrame(fg_buf_editor_rows), hide_index=True, use_container_width=True,
                    disabled=["Produkt", "Mieszalnik"], key="fg_buffer_tank_editor",
                    column_config={"Pojemność zbiornika [m³]": st.column_config.NumberColumn(min_value=0.5, step=0.5)}
                )
                if st.button("📥 Zatwierdź zbiorniki buforowe FG", type="primary", key="btn_confirm_fg_buffer_tanks"):
                    confirmed_fg_buffer_tanks = []
                    for idx, row in edited_fg_buf_df.iterrows():
                        confirmed_fg_buffer_tanks.append({
                            "tag": f"T-FG-{idx + 1}", "recipe_product": row["Produkt"],
                            "capacity_m3": float(row["Pojemność zbiornika [m³]"]),
                        })
                    st.session_state["confirmed_fg_buffer_tanks"] = confirmed_fg_buffer_tanks
                    st.rerun()

                if st.session_state.get("confirmed_fg_buffer_tanks"):
                    st.success(f"✅ Zatwierdzone: {len(st.session_state['confirmed_fg_buffer_tanks'])} zbiorników buforowych FG.")
            else:
                st.info("ℹ️ Żaden produkt nie ma dziś przypisanej 'Cysterna (luzem)' w rozbiciu na opakowania "
                        "(Zakładka 1) — nic do zatwierdzenia. Przypisz % do tego opakowania w recepturze, żeby "
                        "zobaczyć tu kandydatów.")
                st.session_state["confirmed_fg_buffer_tanks"] = []

            st.markdown("---")
            st.markdown("### 📦 Magazynowanie Surowców w Beczkach/IBC/Workach")
            st.caption("Surowce, które nie trafiają do zbiornika (powyżej), i tak muszą stanąć w magazynie — "
                       "przypisz każdemu typ pojemnika, a aplikacja przeliczy liczbę pojemników/palet/miejsc "
                       f"magazynowych **dla wybranego wyżej roku/widoku** ({selected_rm_year_label}). Wynik doliczy "
                       "się do **łącznej powierzchni magazynowej w Zakładce 3** razem z wyrobami gotowymi (FG) — "
                       "to jeden, wspólny magazyn.")

            if drummed_materials:
                if "rm_container_assignment" not in st.session_state:
                    st.session_state.rm_container_assignment = {}
                rm_container_rows = []
                for dm in drummed_materials:
                    mat, ann_t, info = dm["material"], dm["annual_tony"], dm["info"]
                    default_container = st.session_state.rm_container_assignment.get(
                        mat, default_rm_container_for(mat, info))
                    rm_container_rows.append({"Surowiec": mat, "Konsumpcja [t/rok]": round(ann_t, 2),
                                               "Typ pojemnika": default_container})

                edited_rm_containers = st.data_editor(
                    pd.DataFrame(rm_container_rows), hide_index=True, use_container_width=True,
                    disabled=["Surowiec", "Konsumpcja [t/rok]"], key="rm_container_editor",
                    column_config={"Typ pojemnika": st.column_config.SelectboxColumn(options=list(RM_CONTAINER_TYPES.keys()))}
                )
                for _, row in edited_rm_containers.iterrows():
                    st.session_state.rm_container_assignment[row["Surowiec"]] = row["Typ pojemnika"]

                dni_robocze_miesiac_rm = WORKING_DAYS_YEAR / MONTHS_PER_YEAR
                rm_warehouse_rows = []
                for dm in drummed_materials:
                    mat, ann_t = dm["material"], dm["annual_tony"]
                    container_name = st.session_state.rm_container_assignment.get(mat, "Beczka 200 kg (ciecz)")
                    container_cfg = RM_CONTAINER_TYPES[container_name]
                    monthly_kg = (ann_t * 1000.0) / MONTHS_PER_YEAR
                    n_containers_month = math.ceil(monthly_kg / container_cfg["capacity_kg"]) if container_cfg["capacity_kg"] > 0 else 0
                    n_pallets_month = math.ceil(n_containers_month / container_cfg["per_pallet"])
                    miejsca_paletowe_rm = math.ceil((n_pallets_month / dni_robocze_miesiac_rm) * days_of_stock)
                    rm_warehouse_rows.append({
                        "Surowiec 🔒": mat, "Typ pojemnika 🔒": container_name, "Zużycie [t/rok] 🔒": round(ann_t, 2),
                        "Pojemników [/mies]": int(n_containers_month), "Palet [/mies]": int(n_pallets_month),
                        "Miejsca magazynowe [szt]": int(miejsca_paletowe_rm),
                    })

                st.dataframe(pd.DataFrame(rm_warehouse_rows), hide_index=True, use_container_width=True)
                total_rm_pallet_positions = sum(r["Miejsca magazynowe [szt]"] for r in rm_warehouse_rows)
                st.metric("📦 Miejsca magazynowe surowców (RM)", f"{total_rm_pallet_positions} szt.")
                st.session_state["raw_material_warehouse_rows"] = rm_warehouse_rows
            else:
                st.info("Wszystkie surowce z receptur trafiają do zbiorników — brak surowców do magazynowania w beczkach/IBC/workach.")
                st.session_state["raw_material_warehouse_rows"] = []
        else:
            st.info("💡 Wgraj receptury produktów w **Zakładce 1**, aby uzyskać dokładniejsze wymiarowanie per "
                    "pojedynczy surowiec. Poniżej uproszczony szacunek grupowy (wg typu bazy z floty).")

            active_chemical_ratio = st.slider("Średni udział fazy ciekłej (baza + woda) w recepturze [%]:", 50, 95, 85) / 100.0
            silos_aggregation = {"Mineralne (Gr. I/II)": 0.0, "Syntetyczne (Gr. III/IV)": 0.0, "Woda Procesowa DEMI": 0.0, "Inne / Pakiety płynne": 0.0}
            raw_material_summary = []
            for mixer in st.session_state.confirmed_mixers:
                kat = mixer["product_family"]
                prod_info = st.session_state.active_portfolio[kat]
                total_liquid_tony = (mixer["annual_volume"] / 1000.0) * active_chemical_ratio
                water_annual = total_liquid_tony * prod_info["water_content"]
                oil_annual = total_liquid_tony * (1.0 - prod_info["water_content"]) if prod_info["oil_group"] != "Brak (Specjalistyczne)" else 0.0
                other_liquid = total_liquid_tony - water_annual - oil_annual
                silos_aggregation["Woda Procesowa DEMI"] += water_annual
                if oil_annual > 0:
                    silos_aggregation[prod_info["oil_group"]] += oil_annual
                silos_aggregation["Inne / Pakiety płynne"] += other_liquid
                raw_material_summary.append({
                    "ID Reaktora 🔒": mixer["tag"], "Linia 🔒": kat, "Typ Bazy": prod_info["oil_group"],
                    "Produkcja [t/rok]": round(mixer["annual_volume"] / 1000.0, 1),
                    "Baza Olejowa [t/rok]": round(oil_annual, 1), "Woda DEMI [t/rok]": round(water_annual, 1),
                })
            st.markdown("###### 📋 Zestawienie Surowcowe Floty (per Reaktor) — szacunek grupowy")
            st.dataframe(pd.DataFrame(raw_material_summary), hide_index=True, use_container_width=True)

            silos_rows = []
            total_tanks = 0
            for group_name, annual_tony in silos_aggregation.items():
                if annual_tony > 0:
                    daily_t = annual_tony / WORKING_DAYS_YEAR
                    fill_factor = WATER_FILL_FACTOR if "Woda" in group_name else OIL_FILL_FACTOR
                    required_m3 = (daily_t * days_of_stock) / fill_factor
                    required_m3_gross = required_m3 / TANK_SAFETY_FILL
                    if required_m3_gross <= max_single_tank_m3:
                        needed_tanks = 1
                        rec_cap = next((s for s in STANDARD_SMALL_TANK_SIZES_M3 if s >= required_m3_gross), max_single_tank_m3)
                    else:
                        rec_cap = max_single_tank_m3
                        needed_tanks = math.ceil(required_m3_gross / max_single_tank_m3)
                    total_tanks += needed_tanks
                    silos_rows.append({
                        "Grupa Surowcowa": group_name, "Konsumpcja [t/rok]": round(annual_tony, 1),
                        "Wymagany Bufor [m³]": round(required_m3, 1), "Liczba silosów": f"{needed_tanks} szt.",
                        "Pojemność 1 silosu": f"{rec_cap:.0f} m³",
                    })
            st.dataframe(pd.DataFrame(silos_rows), hide_index=True, use_container_width=True)
            st.metric("🧱 Całkowita wymagana liczba silosów surowcowych (szacunek grupowy)", f"{total_tanks} szt.")

# ==========================================
# ZAKŁADKA 6: MAPA STRUMIENIA WARTOŚCI (VSM) (tab6)
# ==========================================
with tab6:
    st.header("🧵 Mapa Strumienia Wartości (Value Stream Mapping)")
    st.caption("Ta zakładka **nie liczy niczego od nowa** — składa w jeden łańcuch czasy już policzone w Zakładkach 2-4 "
               "(hydraulika/bilans cieplny, rozlew, bufory magazynowe), więc automatycznie aktualizuje się razem z nimi.")

    if not st.session_state.confirmed_mixers:
        st.warning("⚠️ Najpierw zatwierdź flotę w Zakładce 1.")
    else:
        _stale_msg = check_fleet_staleness_warning()
        if _stale_msg:
            st.error(_stale_msg)
        rodziny_w_flocie = sorted(set(m["product_family"] for m in st.session_state.confirmed_mixers))
        selected_vsm_family = st.selectbox("Wybierz linię produktową do mapowania:", rodziny_w_flocie, key="vsm_family_select")

        # --- KONFIGURACJA PANELU ZWOLNIENIA QC ---
        st.markdown("##### 🧪 Panel testów QC do zwolnienia szarży")
        st.caption("Wybierz testy z katalogu laboratoryjnego wchodzące w standardowy panel zwolnienia dla tej linii. "
                   "Czasy trwania są edytowalnymi wartościami domyślnymi — popraw je na rzeczywiste, jeśli różnią się w Twoim laboratorium.")

        if "vsm_qc_config" not in st.session_state:
            st.session_state.vsm_qc_config = {}

        # Testy per KONKRETNY produkt (arkusz 'Badania Laboratoryjne', Zakładka 1) - sprawdzamy PRZED
        # ustawieniem domyślnej konfiguracji per linia, żeby przy PIERWSZYM pojawieniu się tej linii
        # w panelu od razu użyć testów z Excela (jeśli spójne dla wszystkich jej produktów), zamiast
        # ogólnego, generycznego fallbacku. Ta sama filozofia co "roczna produkcja" w Zakładce 1:
        # synchronizacja z recepturą TYLKO przy pierwszym pojawieniu się - późniejsze ręczne zmiany
        # tutaj nie są nadpisywane automatycznie.
        qc_tests_by_product_vsm = st.session_state.get("qc_tests_by_product", {})
        products_in_line = [m["recipe_product"] for m in st.session_state.confirmed_mixers
                             if m["product_family"] == selected_vsm_family and m.get("recipe_product") in qc_tests_by_product_vsm]
        distinct_test_sets = {tuple(sorted(qc_tests_by_product_vsm[p])) for p in products_in_line} if products_in_line else set()
        recipe_tests_for_line = list(distinct_test_sets.pop()) if len(distinct_test_sets) == 1 else None

        if selected_vsm_family not in st.session_state.vsm_qc_config and recipe_tests_for_line:
            initial_tests = recipe_tests_for_line
        else:
            initial_tests = ["Lepkość kinematyczna @40°C (półautomat)", "Barwa ASTM", "Temp. zapłonu - półautomat"]
        qc_cfg = st.session_state.vsm_qc_config.setdefault(selected_vsm_family, {
            "tests": initial_tests, "mode": "Sekwencyjnie (jeden technik, jedno stanowisko)", "custom_durations": {},
        })

        # Jeśli lista już istnieje (linia była wcześniej konfigurowana) i różni się od tego, co mówi
        # Excel - informujemy i dajemy przycisk, zamiast po cichu nadpisywać ręczne zmiany.
        if products_in_line:
            if recipe_tests_for_line is not None:
                if sorted(recipe_tests_for_line) != sorted(qc_cfg["tests"]):
                    st.info(f"📄 Arkusz 'Badania Laboratoryjne' definiuje dla tej linii: **{', '.join(recipe_tests_for_line)}** "
                            "— różni się od listy poniżej.")
                    if st.button(f"🔄 Użyj testów z Excela dla '{selected_vsm_family}'", key=f"sync_qc_{selected_vsm_family}"):
                        qc_cfg["tests"] = recipe_tests_for_line
                        st.rerun()
            else:
                st.warning("⚠️ Arkusz 'Badania Laboratoryjne' definiuje **RÓŻNE** testy dla różnych produktów tej linii "
                           f"({len(distinct_test_sets)} różnych zestawów) — panel poniżej to tylko uproszczenie "
                           "per LINIA (jedna wspólna lista). Dokładne, per-produktowe liczby badań QC znajdziesz w "
                           "widgecie porównawczym (Zakładka 2, Karta Maszyn).")

        full_qc_catalog = get_full_qc_catalog()
        c_qc1, c_qc2 = st.columns([2, 1])
        with c_qc1:
            qc_cfg["tests"] = st.multiselect(
                "Testy w panelu zwolnienia:", list(full_qc_catalog.keys()),
                default=[t for t in qc_cfg["tests"] if t in full_qc_catalog],
                key=f"qc_tests_{selected_vsm_family}"
            )
        with c_qc2:
            qc_cfg["mode"] = st.radio(
                "Sposób wykonania:",
                ["Sekwencyjnie (jeden technik, jedno stanowisko)", "Równolegle (kilku techników / aparatów)"],
                index=0 if qc_cfg["mode"].startswith("Sekw") else 1,
                key=f"qc_mode_{selected_vsm_family}"
            )

        if qc_cfg["tests"]:
            df_qc = pd.DataFrame([{
                "Test": t,
                "Czas [min]": qc_cfg["custom_durations"].get(t, full_qc_catalog[t]["duration_min"]),
                "Sprzęt": full_qc_catalog[t]["equipment"],
            } for t in qc_cfg["tests"]])

            edited_qc = st.data_editor(
                df_qc, hide_index=True, use_container_width=True,
                disabled=["Test", "Sprzęt"], key=f"qc_dur_editor_{selected_vsm_family}"
            )
            for _, r in edited_qc.iterrows():
                qc_cfg["custom_durations"][r["Test"]] = float(r["Czas [min]"])

            durations_min = [qc_cfg["custom_durations"][t] for t in qc_cfg["tests"]]
            qc_time_h = (sum(durations_min) if qc_cfg["mode"].startswith("Sekw") else max(durations_min)) / 60.0
        else:
            qc_time_h = 0.0
            st.info("Brak wybranych testów — czas zwolnienia QC przyjęto jako 0h.")

        # Kolejka laboratoryjna: czas oczekiwania próbki na wolne stanowisko/technika,
        # ODDZIELNY od czasu samego wykonania testów (qc_time_h powyżej).
        if "vsm_qc_queue_days" not in st.session_state:
            st.session_state.vsm_qc_queue_days = {}
        qc_queue_days = st.number_input(
            "⏳ Kolejka laboratoryjna przed rozpoczęciem testów [dni]:", min_value=0.0,
            value=float(st.session_state.vsm_qc_queue_days.get(selected_vsm_family, 0.0)), step=0.5,
            key=f"qc_queue_{selected_vsm_family}"
        )
        st.session_state.vsm_qc_queue_days[selected_vsm_family] = qc_queue_days

        st.markdown("---")

        # --- CZASY PROCESOWE: PER KONKRETNY MIESZALNIK, nie uśrednione po całej linii - inaczej
        # mieszalnik 31 m³ i 120 m³ (różny cykl, różna moc, różny czas rozlewu) wpadałyby do jednej
        # wspólnej, mylącej liczby. ---
        mixers_in_family = [m for m in st.session_state.confirmed_mixers if m["product_family"] == selected_vsm_family]
        mixer_tags_in_family = [m["tag"] for m in mixers_in_family]
        selected_vsm_mixer_tag = st.selectbox(
            "Wybierz mieszalnik do mapowania (w obrębie linii):", mixer_tags_in_family, key="vsm_mixer_select",
            help="Każdy mieszalnik ma własny cykl, moc i czas rozlewu — mapa strumienia wartości poniżej dotyczy "
                 "TEGO KONKRETNEGO urządzenia, nie uśrednionej linii."
        )
        selected_vsm_mixer = next(m for m in mixers_in_family if m["tag"] == selected_vsm_mixer_tag)

        # --- Zbiornik kampanijny (kilka produktów) - dotąd VSM pokazywał tylko zsumowaną/uśrednioną
        # wartość dla całego zbiornika, przez co np. dwa produkty o różnej gęstości/cyklu wpadały do
        # jednej mylącej liczby. Jeśli zbiornik jest współdzielony, wybierz KONKRETNY produkt. ---
        shared_members_vsm = selected_vsm_mixer.get("shared_members")
        if shared_members_vsm:
            product_names_vsm = [mem["product"] for mem in shared_members_vsm]
            selected_vsm_product_name = st.selectbox(
                f"🔀 {selected_vsm_mixer_tag} to zbiornik kampanijny ({len(shared_members_vsm)} produktów) — wybierz produkt:",
                product_names_vsm, key="vsm_shared_product_select"
            )
            selected_member = next(mem for mem in shared_members_vsm if mem["product"] == selected_vsm_product_name)
            mass_per_batch_vsm = selected_vsm_mixer["capacity_m3"] * selected_member["density"] * 1000.0 * st.session_state.mixer_fill_factor
            monthly_mass_vsm = selected_member["annual_kg"] / MONTHS_PER_YEAR
            batches_month_this_mixer = math.ceil(monthly_mass_vsm / mass_per_batch_vsm) if mass_per_batch_vsm > 0 else 0
        else:
            batches_month_this_mixer = selected_vsm_mixer.get("batches_count", 0)

        calc_times_selected = st.session_state.calculated_times.get(selected_vsm_mixer_tag)

        if calc_times_selected is None:
            st.info(f"ℹ️ Skonfiguruj hydraulikę i bilans cieplny dla {selected_vsm_mixer_tag} w Zakładce 2, aby uzyskać "
                    "rzeczywiste czasy grzania/pompowania/chłodzenia (poniżej użyto bezpiecznych wartości domyślnych).")
            heating_h, pumping_h, cooling_h = 1.5, 0.75, 0.5
        else:
            heating_h = calc_times_selected["heating"]
            pumping_h = calc_times_selected["pumping"]
            cooling_h = calc_times_selected.get("cooling_h", 0.0)

        # --- DOZOWANIE / HOMOGENIZACJA: konfiguracja per urządzenie, przeniesiona tu z Zakładki 6 ---
        # (dawniej wpisywana w Zakładce 2 i odczytywana przez klucz widgetu — jeśli nikt nie odwiedził
        # tamtej zakładki dla danej rodziny, VSM cicho używał wartości domyślnych; teraz to jest
        # jedyne miejsce konfiguracji tych czasów).
        st.markdown("##### ⏱️ Dozowanie i Homogenizacja (per urządzenie)")
        for mixer in mixers_in_family:
            tag = mixer["tag"]
            defaults_bt = st.session_state.batch_time_components.setdefault(tag, {"dosing": 1.0, "homog": 2.0})
            with st.expander(f"⏱️ Składniki czasu operacyjnego dla: {tag}", expanded=(tag == selected_vsm_mixer_tag)):
                defaults_bt["dosing"] = st.number_input(
                    "Dozowanie surowców [h]:", min_value=0.1, value=float(defaults_bt["dosing"]), key=f"vsm_tdos_{tag}")
                defaults_bt["homog"] = st.number_input(
                    "Homogenizacja właściwa [h]:", min_value=0.1, value=float(defaults_bt["homog"]), key=f"vsm_thom_{tag}")

        t_dosing = st.session_state.batch_time_components[selected_vsm_mixer_tag]["dosing"]
        t_homog = st.session_state.batch_time_components[selected_vsm_mixer_tag]["homog"]

        # Czas rozlewu — liczony WPROST z masy JEDNEJ SZARŻY (nie z miesięcznej wartości dzielonej
        # przez liczbę szarż) - prostsze, bardziej przejrzyste i niezależne od tego, czy Zakładka 4
        # była już odwiedzona. Miesiąc/rok pokazane niżej wynikają z tej samej wartości, pomnożonej
        # przez liczbę szarż - w tym kierunku, nie odwrotnie.
        recipe_product_for_logistics = selected_vsm_product_name if shared_members_vsm else selected_vsm_mixer.get("recipe_product")
        mass_per_batch_for_filling = mass_per_batch_vsm if shared_members_vsm else selected_vsm_mixer["mass_per_batch"]
        rho_linii_vsm = st.session_state.active_portfolio[selected_vsm_family]["density"]
        filling_h = compute_filling_time_h(
            mass_per_batch_for_filling, recipe_product_for_logistics, selected_vsm_family,
            selected_vsm_mixer_tag, rho_linii_vsm, opakowania_podzial=st.session_state.get("opakowania_podzial", {})
        )
        if filling_h == 0.0:
            st.info(f"ℹ️ Skonfiguruj podział opakowań w panelu bocznym i odwiedź Zakładkę 3, aby uzyskać rzeczywisty "
                    f"czas rozlewu dla {selected_vsm_mixer_tag}.")
        elif filling_h > 8.0:
            st.warning(f"⚠️ **Czas rozlewu ({filling_h:.1f} h) wygląda długo** dla jednej szarży — sprawdź liczbę "
                       "głowic nalewających i ich wydajność w **Zakładce 4** (sekcja opakowań) — domyślnie 1-2 "
                       "głowice mogą być za mało dla dużej szarży. Duży mieszalnik potrzebuje zwykle szybszej "
                       "linii nalewającej (więcej głowic), żeby nie stał bezczynnie podczas rozlewu.")

        # Bufory magazynowe — założenia globalne z Zakładek 3 i 5 (te wejścia nie są dziś różnicowane per rodzina).
        raw_material_buffer_days = st.session_state.get("days_of_stock_tab5", 14)
        fg_storage_days = st.session_state.get("czas_skladowania_tab3", 14)

        HOURS_PER_DAY = 24.0
        raw_material_buffer_h = raw_material_buffer_days * HOURS_PER_DAY
        fg_storage_h = fg_storage_days * HOURS_PER_DAY
        qc_queue_h = qc_queue_days * HOURS_PER_DAY

        process_steps = [
            {"name": "Dozowanie", "hours": t_dosing, "value_added": True},
            {"name": "Grzanie", "hours": heating_h, "value_added": True},
            {"name": "Homogenizacja", "hours": t_homog, "value_added": True},
            {"name": "Zwolnienie QC", "hours": qc_time_h, "value_added": False, "extra_wait_h": qc_queue_h},
            {"name": "Pompowanie", "hours": pumping_h, "value_added": True},
            {"name": "Chłodzenie", "hours": cooling_h, "value_added": True},
            {"name": "Rozlew", "hours": filling_h, "value_added": True},
        ]
        for s in process_steps:
            s.setdefault("extra_wait_h", 0.0)

        # --- OEE: C/O, Uptime, Dostępność, Pass rate per etap (edytowalne, domyślnie neutralne) ---
        st.markdown("##### ⚙️ Zmiana, Dostępność i Jakość per Etap (OEE)")
        st.caption("Domyślnie C/O=0h i Uptime/Dostępność/Pass=100% (brak strat) — popraw na wartości rzeczywiste tam, "
                   "gdzie mają znaczenie (typowo: Grzanie/Homogenizacja przy zmianie produktu w reaktorze, oraz Rozlew "
                   "przy zmianie SKU na linii pakującej). **C/O jest wliczane do Lead Time** — zajmuje realny czas "
                   "na reaktorze/linii, nawet jeśli księgowane jest jako strata, a nie czas procesu.")

        if "vsm_oee" not in st.session_state:
            st.session_state.vsm_oee = {}
        oee_cfg = st.session_state.vsm_oee.setdefault(selected_vsm_family, {})
        for s in process_steps:
            oee_cfg.setdefault(s["name"], {"co_h": 0.0, "uptime_pct": 100.0, "availability_pct": 100.0, "pass_pct": 100.0})

        # Dostępność wyliczona z MTBF/MTTR (Zakładka 6, karta maszyn) — średnia dla mieszalników
        # tej rodziny, ważona liczbą szarż/miesiąc (urządzenia bardziej obciążone mają większy
        # wpływ na realną dostępność linii). Dotyczy etapów, w których fizycznie bierze udział
        # reaktor i/lub pompa (Dozowanie/Grzanie/Homogenizacja/Pompowanie/Chłodzenie) — Zwolnienie
        # QC i Rozlew korzystają z innych zasobów (laboratorium, linia pakująca), nieujętych tu.
        reactor_pump_steps = ["Dozowanie", "Grzanie", "Homogenizacja", "Pompowanie", "Chłodzenie"]
        mtbf_weighted_avail, mtbf_weight_sum = 0.0, 0.0
        for mx in mixers_in_family:
            ct_mx = st.session_state.calculated_times.get(mx["tag"])
            if ct_mx is not None and "availability_pct" in ct_mx:
                w = max(mx.get("batches_count", 1), 1)
                mtbf_weighted_avail += ct_mx["availability_pct"] * w
                mtbf_weight_sum += w
        mtbf_derived_availability_pct = (mtbf_weighted_avail / mtbf_weight_sum) if mtbf_weight_sum > 0 else None

        c_avail1, c_avail2 = st.columns([3, 1])
        with c_avail1:
            if mtbf_derived_availability_pct is not None:
                st.caption(f"📟 Dostępność wyliczona z MTBF/MTTR (Zakładka 6, ważona liczbą szarż) dla tej rodziny: "
                           f"**{mtbf_derived_availability_pct:.1f}%**. Dotyczy etapów: {', '.join(reactor_pump_steps)}.")
            else:
                st.caption("ℹ️ Skonfiguruj MTBF/MTTR w Zakładce 6 (karta maszyn, sekcja 🔧 Niezawodność), aby móc "
                           "podstawić tu wyliczoną dostępność zamiast wpisywać ją ręcznie.")
        with c_avail2:
            if st.button("🔄 Zastosuj wyliczoną Dostępność", key=f"apply_mtbf_avail_{selected_vsm_family}",
                          disabled=mtbf_derived_availability_pct is None, use_container_width=True):
                for step_name in reactor_pump_steps:
                    oee_cfg[step_name]["availability_pct"] = round(mtbf_derived_availability_pct, 1)
                st.rerun()

        df_oee_in = pd.DataFrame([{
            "Etap": s["name"],
            "C/T [h]": round(s["hours"], 2),
            "C/O [h]": oee_cfg[s["name"]]["co_h"],
            "Uptime [%]": oee_cfg[s["name"]]["uptime_pct"],
            "Dostępność [%]": oee_cfg[s["name"]]["availability_pct"],
            "Pass [%]": oee_cfg[s["name"]]["pass_pct"],
        } for s in process_steps])

        edited_oee = st.data_editor(
            df_oee_in, hide_index=True, use_container_width=True,
            disabled=["Etap", "C/T [h]"], key=f"oee_editor_{selected_vsm_family}",
            column_config={
                "Uptime [%]": st.column_config.NumberColumn(min_value=0.0, max_value=100.0),
                "Dostępność [%]": st.column_config.NumberColumn(min_value=0.0, max_value=100.0),
                "Pass [%]": st.column_config.NumberColumn(min_value=0.0, max_value=100.0),
                "C/O [h]": st.column_config.NumberColumn(min_value=0.0),
            }
        )
        for _, r in edited_oee.iterrows():
            oee_cfg[r["Etap"]] = {
                "co_h": float(r["C/O [h]"]),
                "uptime_pct": float(r["Uptime [%]"]),
                "availability_pct": float(r["Dostępność [%]"]),
                "pass_pct": float(r["Pass [%]"]),
            }

        for s in process_steps:
            o = oee_cfg[s["name"]]
            s["co_h"] = o["co_h"]
            s["oee_pct"] = (o["uptime_pct"] / 100.0) * (o["availability_pct"] / 100.0) * (o["pass_pct"] / 100.0) * 100.0

        st.markdown("---")

        # --- WIDOK: szczegółowy (7 etapów) lub zbiorczy (4 etapy, jak w dashboardzie referencyjnym) ---
        widok = st.radio(
            "Widok:", ["Szczegółowy (7 etapów)", "Zbiorczy (4 etapy, jak w dashboardzie referencyjnym)"],
            horizontal=True, key=f"vsm_widok_{selected_vsm_family}"
        )

        if widok.startswith("Szczegółowy"):
            display_steps = [{
                "name": s["name"], "hours": s["hours"], "value_added": s["value_added"],
                "co_h": s["co_h"], "wait_h": s["extra_wait_h"], "oee_pct": s["oee_pct"],
            } for s in process_steps]
            lead_start_wait = ("Bufor surowców", raw_material_buffer_h)
            lead_end_wait = ("Bufor wyrobów gotowych", fg_storage_h)
        else:
            # Grupowanie 4 makro-etapów zgodnie z układem dashboardu referencyjnego.
            def group_stats(names):
                members = [s for s in process_steps if s["name"] in names]
                ct_h = sum(m["hours"] for m in members)
                co_h = sum(m["co_h"] for m in members)
                wait_h = sum(m["extra_wait_h"] for m in members)
                va = any(m["value_added"] for m in members)
                if ct_h > 0:
                    uptime = sum(oee_cfg[m["name"]]["uptime_pct"] * m["hours"] for m in members) / ct_h
                    avail = sum(oee_cfg[m["name"]]["availability_pct"] * m["hours"] for m in members) / ct_h
                else:
                    uptime = avail = 100.0
                pass_combined = 100.0
                for m in members:
                    pass_combined *= oee_cfg[m["name"]]["pass_pct"] / 100.0
                pass_combined *= 100.0
                oee_pct = (uptime / 100.0) * (avail / 100.0) * (pass_combined / 100.0) * 100.0
                return {"hours": ct_h, "co_h": co_h, "wait_h": wait_h, "value_added": va, "oee_pct": oee_pct}

            blending = group_stats(["Dozowanie", "Grzanie", "Homogenizacja", "Pompowanie", "Chłodzenie"])
            qc_group = group_stats(["Zwolnienie QC"])
            filling_group = group_stats(["Rozlew"])

            display_steps = [
                {"name": "Blending/Cooking", **blending},
                {"name": "QC", **qc_group},
                {"name": "Filling/Packing", **filling_group},
            ]
            lead_start_wait = ("Receiving & Staging", raw_material_buffer_h)
            lead_end_wait = (None, 0.0)  # bufor WG dołączony do Filling/Packing, jak w dashboardzie referencyjnym
            display_steps[-1]["wait_h"] += fg_storage_h

        total_process_h = sum(s["hours"] for s in display_steps)
        total_co_h = sum(s["co_h"] for s in display_steps)
        total_wait_h = lead_start_wait[1] + lead_end_wait[1] + sum(s["wait_h"] for s in display_steps)
        value_added_h = sum(s["hours"] for s in display_steps if s["value_added"])
        total_lead_time_h = total_process_h + total_co_h + total_wait_h
        pce_pct = (value_added_h / total_lead_time_h * 100.0) if total_lead_time_h > 0 else 0.0

        st.markdown("### 📈 Kluczowe Metryki Strumienia Wartości")
        m1, m2, m3, m4, m5 = st.columns(5)
        with m1: st.metric("⏳ Całkowity Lead Time", f"{total_lead_time_h / 24.0:.1f} dni")
        with m2: st.metric("⚙️ Czas przetwarzania (VA)", f"{value_added_h:.1f} h")
        with m3: st.metric("🔄 Suma C/O", f"{total_co_h:.1f} h")
        with m4: st.metric("🎯 Process Cycle Efficiency", f"{pce_pct:.1f}%")
        with m5: st.metric("🧪 Czas zwolnienia QC", f"{qc_time_h:.2f} h")

        if pce_pct < 10.0 and total_lead_time_h > 0:
            st.warning("⚠️ PCE poniżej 10% jest typowe dla procesów wsadowych z dużym buforowaniem magazynowym — "
                       "największa dźwignia poprawy leży zwykle w skróceniu dni bufora surowców/wyrobów gotowych "
                       "lub kolejki laboratoryjnej, a nie w przyspieszaniu samego procesu w reaktorze.")

        # --- DIAGRAM VSM: proste boksy/strzałki w HTML+CSS (bez zależności od graphviz) ---
        st.markdown("### 🗺️ Diagram Strumienia Wartości")

        def render_box(title, ct_txt, co_h=0.0, oee_pct=None, va=True):
            color = "#2E7D32" if va else "#B45309"
            bg = "#E8F5E9" if va else "#FEF3C7"
            extra = f'<div style="font-size:10px; color:#555;">C/O {co_h:.2f}h</div>' if co_h > 0 else ""
            oee_line = f'<div style="font-size:10px; color:#555;">OEE {oee_pct:.0f}%</div>' if oee_pct is not None else ""
            return (f'<div style="border:2px solid {color}; border-radius:6px; padding:8px 10px; min-width:118px; '
                    f'text-align:center; background:{bg}; flex-shrink:0;">'
                    f'<div style="font-size:12px; font-weight:600; color:#111;">{title}</div>'
                    f'<div style="font-size:14px; font-weight:700; color:{color};">{ct_txt}</div>{extra}{oee_line}</div>')

        def render_triangle(label, days_txt):
            return (f'<div style="display:flex; flex-direction:column; align-items:center; flex-shrink:0; margin:0 4px;">'
                    f'<div style="width:0; height:0; border-left:20px solid transparent; border-right:20px solid transparent; '
                    f'border-bottom:34px solid #FDE68A;"></div>'
                    f'<div style="font-size:11px; font-weight:600; margin-top:2px; white-space:nowrap;">{label} {days_txt}</div></div>')

        def render_arrow():
            return '<div style="display:flex; align-items:center; padding:0 2px; flex-shrink:0; color:#555; font-size:18px;">➜</div>'

        pieces = [render_triangle(lead_start_wait[0], f"{lead_start_wait[1] / 24.0:.0f} dni"), render_arrow()]
        for s in display_steps:
            pieces.append(render_box(s["name"], f"{s['hours']:.2f} h", co_h=s["co_h"], oee_pct=s.get("oee_pct"), va=s["value_added"]))
            if s.get("wait_h", 0.0) > 0:
                pieces.append(render_arrow())
                pieces.append(render_triangle("oczekiwanie", f"{s['wait_h'] / 24.0:.1f} dni"))
            pieces.append(render_arrow())
        if lead_end_wait[0] is not None:
            pieces.append(render_triangle(lead_end_wait[0], f"{lead_end_wait[1] / 24.0:.0f} dni"))
        else:
            pieces.pop()  # usuń ostatnią, niepotrzebną strzałkę, gdy nie ma końcowego bufora

        diagram_html = f'<div style="display:flex; align-items:center; overflow-x:auto; padding:14px 4px;">{"".join(pieces)}</div>'
        st.markdown(diagram_html, unsafe_allow_html=True)

        st.caption("🟢 Zielone pola = czas dodający wartość (przetwarzanie produktu). 🟠 Pomarańczowe pola = czas "
                   "niedodający wartości bezpośrednio produktowi (kontrola jakości, magazynowanie) — często konieczny "
                   "operacyjnie, ale to właśnie tu zwykle leży potencjał skrócenia lead time. **C/O** i **OEE** "
                   "pokazane pod nazwą etapu, gdy dotyczy.")

        # --- WSKAŹNIKI: DZIEŃ / MIESIĄC / ROK - produkcja, laboratorium, logistyka razem, dla
        # WYBRANEGO wyżej mieszalnika (i konkretnego produktu, jeśli zbiornik kampanijny). ---
        st.markdown("### 📊 Wskaźniki: Dzień / Miesiąc / Rok")
        dni_robocze_miesiac_vsm = WORKING_DAYS_YEAR / MONTHS_PER_YEAR
        batches_year_this = batches_month_this_mixer * MONTHS_PER_YEAR
        batches_day_this = batches_month_this_mixer / dni_robocze_miesiac_vsm if dni_robocze_miesiac_vsm > 0 else 0.0

        n_tests_per_batch_vsm = len(qc_cfg.get("tests", []))
        qc_day = n_tests_per_batch_vsm * batches_day_this
        qc_month = n_tests_per_batch_vsm * batches_month_this_mixer
        qc_year = n_tests_per_batch_vsm * batches_year_this

        wp1, wp2, wp3 = st.columns(3)
        with wp1: st.metric("🏭 Szarż — dzień", f"{batches_day_this:.2f}")
        with wp2: st.metric("🏭 Szarż — miesiąc", f"{batches_month_this_mixer}")
        with wp3: st.metric("🏭 Szarż — rok", f"{batches_year_this}")

        wr1, wr2, wr3 = st.columns(3)
        with wr1: st.metric("🚿 Rozlew — na szarżę", f"{filling_h:.2f} h")
        with wr2: st.metric("🚿 Rozlew — miesiąc", f"{filling_h * batches_month_this_mixer:.1f} h",
                             help=f"{filling_h:.2f} h/szarżę × {batches_month_this_mixer} szarż/mies.")
        with wr3: st.metric("🚿 Rozlew — rok", f"{filling_h * batches_year_this:.0f} h",
                             help=f"{filling_h:.2f} h/szarżę × {batches_year_this} szarż/rok")

        wq1, wq2, wq3 = st.columns(3)
        with wq1: st.metric("🧪 Badań QC — dzień", f"{qc_day:.2f}")
        with wq2: st.metric("🧪 Badań QC — miesiąc", f"{qc_month}")
        with wq3: st.metric("🧪 Badań QC — rok", f"{qc_year}")

        # Logistyka: dostawy RM (cysterny per surowiec) i wysyłki FG (jeśli "Cysterna (luzem)")
        # dla konkretnego produktu tego mieszalnika - ta sama logika co widget porównawczy
        # (Zakładka 2, Karta Maszyn), przeliczona tu na dzień/miesiąc/rok.
        recipes_df_vsm = st.session_state.get("recipes_df")
        rm_tankers_month_total, fg_tankers_month = 0, 0
        if recipes_df_vsm is not None and not recipes_df_vsm.empty and recipe_product_for_logistics:
            match_vsm = recipes_df_vsm[recipes_df_vsm[RECIPE_PRODUCT_COL] == recipe_product_for_logistics]
            if not match_vsm.empty:
                row_vsm = match_vsm.iloc[0]
                mass_per_batch_for_log = (mass_per_batch_vsm if shared_members_vsm else selected_vsm_mixer["mass_per_batch"])
                monthly_mass_for_log = mass_per_batch_for_log * batches_month_this_mixer
                rm_storage_override_vsm = st.session_state.get("rm_storage_method_override", {})
                for mat in RECIPE_RAW_MATERIALS:
                    dozowanie_kg_t = float(row_vsm.get(mat, 0) or 0)
                    if dozowanie_kg_t <= 0 or rm_storage_override_vsm.get(mat) == "Zbiornik (luzem)":
                        continue
                    mat_month_t = dozowanie_kg_t / 1000.0 * (monthly_mass_for_log / 1000.0)
                    rm_tankers_month_total += math.ceil(mat_month_t / st.session_state.tanker_capacity_t) if st.session_state.tanker_capacity_t > 0 else 0
                tanker_col_vsm = recipe_pack_pct_col("Cysterna (luzem)")
                if tanker_col_vsm in recipes_df_vsm.columns and float(row_vsm.get(tanker_col_vsm, 0) or 0) > 0:
                    fg_tankers_month = math.ceil((monthly_mass_for_log / 1000.0) / st.session_state.tanker_capacity_t) if st.session_state.tanker_capacity_t > 0 else 0

        wl1, wl2, wl3 = st.columns(3)
        with wl1: st.metric("🚚 Cystern RM — dzień", f"{(rm_tankers_month_total / dni_robocze_miesiac_vsm) if dni_robocze_miesiac_vsm > 0 else 0:.2f}")
        with wl2: st.metric("🚚 Cystern RM — miesiąc", f"{rm_tankers_month_total}")
        with wl3: st.metric("🚚 Cystern RM — rok", f"{rm_tankers_month_total * MONTHS_PER_YEAR}")
        if fg_tankers_month > 0:
            wf1, wf2, wf3 = st.columns(3)
            with wf1: st.metric("📦 Wysyłek FG (cysterna) — dzień", f"{(fg_tankers_month / dni_robocze_miesiac_vsm) if dni_robocze_miesiac_vsm > 0 else 0:.2f}")
            with wf2: st.metric("📦 Wysyłek FG (cysterna) — miesiąc", f"{fg_tankers_month}")
            with wf3: st.metric("📦 Wysyłek FG (cysterna) — rok", f"{fg_tankers_month * MONTHS_PER_YEAR}")

        # --- POJEMNOŚĆ LABORATORIUM: obciążenie każdego aparatu w CAŁYM ZAKŁADZIE (nie tylko dla
        # wybranego mieszalnika) - niektóre testy mogą mieć WIĘCEJ NIŻ JEDEN identyczny aparat w
        # laboratorium (np. 2 aparaty do pienienia), co bezpośrednio podwaja przepustowość tego
        # konkretnego testu na cały zakład, nie tylko dla jednej szarży/mieszalnika. ---
        st.markdown("---")
        st.markdown("### 🔬 Pojemność Laboratorium (cały zakład)")
        st.caption("Zużycie czasu każdego aparatu, zsumowane po WSZYSTKICH mieszalnikach i produktach na raz — "
                   "jeśli laboratorium ma więcej niż 1 sztukę danego aparatu, wpisz to poniżej: bezpośrednio "
                   "podwaja (lub więcej) przepustowość tego konkretnego testu w całym zakładzie.")

        equipment_demand_min_month = {}  # nazwa aparatu -> suma [min/mies] potrzebna w calym zakladzie
        equipment_test_names = {}  # nazwa aparatu -> zbior nazw testow ktore go uzywaja (do etykiety)
        for m_lab in st.session_state.confirmed_mixers:
            tests_lab, _ = get_qc_tests_for_mixer(m_lab)
            batches_month_lab = m_lab.get("batches_count", 0)
            for t_name in tests_lab:
                t_info = get_full_qc_catalog().get(t_name)
                if not t_info:
                    continue
                equip_name = t_info["equipment"]
                equipment_demand_min_month[equip_name] = equipment_demand_min_month.get(equip_name, 0.0) + t_info["duration_min"] * batches_month_lab
                equipment_test_names.setdefault(equip_name, set()).add(t_name)

        if not equipment_demand_min_month:
            st.info("ℹ️ Brak jeszcze skonfigurowanych badań QC dla żadnego mieszalnika.")
        else:
            lab_rows = []
            for equip_name, demand_min in sorted(equipment_demand_min_month.items(), key=lambda x: -x[1]):
                count = st.session_state.qc_equipment_count_override.get(equip_name, 1)
                available_min_month = count * WORKING_DAYS_YEAR / MONTHS_PER_YEAR * 8 * 60  # 1 zmiana laboratoryjna, 8h/dzień
                utilization_pct = (demand_min / available_min_month * 100.0) if available_min_month > 0 else 0.0
                lab_rows.append({
                    "Aparat": equip_name, "Używany do": ", ".join(sorted(equipment_test_names[equip_name])),
                    "Liczba sztuk": count, "Zapotrzebowanie [h/mies]": round(demand_min / 60.0, 1),
                    "Wykorzystanie [%]": round(utilization_pct, 0),
                })
            lab_df = pd.DataFrame(lab_rows)
            edited_lab_df = st.data_editor(
                lab_df, hide_index=True, use_container_width=True, key="lab_capacity_editor",
                disabled=["Aparat", "Używany do", "Zapotrzebowanie [h/mies]", "Wykorzystanie [%]"],
                column_config={"Liczba sztuk": st.column_config.NumberColumn(min_value=1, step=1)}
            )
            for _, row_lab in edited_lab_df.iterrows():
                st.session_state.qc_equipment_count_override[row_lab["Aparat"]] = int(row_lab["Liczba sztuk"])

            bottleneck_rows = edited_lab_df[edited_lab_df["Wykorzystanie [%]"] > 85]
            if not bottleneck_rows.empty:
                st.warning("⚠️ **Możliwe wąskie gardło laboratorium** (wykorzystanie >85%, przyjmując 1 zmianę, 8h/dzień): "
                           f"{', '.join(bottleneck_rows['Aparat'].tolist())}. Rozważ dodatkowy aparat, drugą zmianę "
                           "w laboratorium, albo sprawdź, czy któryś test naprawdę musi być wykonywany tak często.")
            st.caption("Wykorzystanie liczone przy założeniu 1 zmiany laboratoryjnej (8h/dzień roboczy) na aparat — "
                       "zmień 'Liczba sztuk', jeśli laboratorium ma więcej fizycznych egzemplarzy danego aparatu.")

        # --- DRABINKA CZASU: pełny rozkład lead time vs. czas przetwarzania ---
        st.markdown("### ⏱️ Drabinka Czasu (Lead Time vs. Czas Przetwarzania)")
        ladder_rows = [{"Etap": lead_start_wait[0], "Waiting [dni]": round(lead_start_wait[1] / 24.0, 2),
                        "C/T [h]": 0.0, "C/O [h]": 0.0, "Typ": "Magazynowanie"}]
        for s in display_steps:
            ladder_rows.append({
                "Etap": s["name"], "Waiting [dni]": round(s.get("wait_h", 0.0) / 24.0, 2),
                "C/T [h]": round(s["hours"], 2), "C/O [h]": round(s["co_h"], 2),
                "Typ": "Wartość dodana" if s["value_added"] else "Kontrola / Oczekiwanie"
            })
        if lead_end_wait[0] is not None:
            ladder_rows.append({"Etap": lead_end_wait[0], "Waiting [dni]": round(lead_end_wait[1] / 24.0, 2),
                                 "C/T [h]": 0.0, "C/O [h]": 0.0, "Typ": "Magazynowanie"})

        df_ladder = pd.DataFrame(ladder_rows)
        st.dataframe(df_ladder, hide_index=True, use_container_width=True)

        total_waiting_days = df_ladder["Waiting [dni]"].sum()
        total_ct_days = df_ladder["C/T [h]"].sum() / 24.0
        total_co_days = df_ladder["C/O [h]"].sum() / 24.0
        st.markdown(
            f"**TOTAL** — Waiting: `{total_waiting_days:.2f} dni` · C/T: `{total_ct_days:.2f} dni` · "
            f"C/O: `{total_co_days:.2f} dni` · **Lead Time: `{(total_waiting_days + total_ct_days + total_co_days):.2f} dni`** · "
            f"**VA ratio: `{pce_pct:.1f}%`**"
        )

# ==========================================
# ZAKŁADKA 7: DASHBOARD (tab8) — podsumowanie na pierwszy rzut oka
# ==========================================
with tab8:
    st.header("🏠 Dashboard — Podsumowanie Projektu")
    st.caption("Zbiera w jednym miejscu najważniejsze liczby z całej aplikacji. Pola pokazujące '—' oznaczają, że "
               "dana zakładka nie została jeszcze skonfigurowana/odwiedzona w tej sesji — apka nie liczy niczego "
               "od nowa tutaj, tylko czyta to, co już policzone gdzie indziej.")

    if not st.session_state.confirmed_mixers:
        st.warning("⚠️ Flota nie jest jeszcze zatwierdzona. Zacznij od **Zakładki 1 (Receptury)** lub od razu skonfiguruj "
                   "flotę w **Zakładce 1 (Receptury Produktów i Flota)**.")
    else:
        _stale_msg = check_fleet_staleness_warning()
        if _stale_msg:
            st.error(_stale_msg)
        # Docelowy tonaż liczony NA ŻYWO z receptury (dokładnie tak jak "Sumaryczny tonaż roczny
        # zakładu" w Zakładce 1) - NIE ze starej migawki confirmed_mixers, żeby ta liczba nigdy nie
        # mogła pokazywać czegoś innego niż to, co widać w Zakładce 1 w danym momencie.
        target_annual_t_dash = sum(
            st.session_state.prod_dict.get(kat, {}).get("roczna", 0)
            for kat in st.session_state.get("wybrane_kategorie_snapshot", [])
        ) / 1000.0
        n_mixers = len(st.session_state.confirmed_mixers)
        groups_active = sorted(set(m["product_family"] for m in st.session_state.confirmed_mixers))

        st.markdown("### 🏭 Produkcja i Flota")
        st.metric("🎯 Docelowa produkcja własna (100%, Rok 5+)", f"{target_annual_t_dash:,.0f} t/rok — {n_mixers} mieszalników, {len(groups_active)} grup produktowych")
        st.caption("ℹ️ **Utylizacja = czas pracy ÷ dostępny czas pracy w miesiącu** (szarże × cykl [h] ÷ godziny "
                   "dostępne) — ta sama logika co Zakładka 1 ('Utylizacja Czasowa'). To realny wskaźnik: nawet przy "
                   "100% docelowej produkcji mieszalnik zwykle **nie** osiąga 100% (zbiornik ma margines na "
                   "przezbrojenia, konserwację, wahania popytu) — 100% oznaczałoby zerowy zapas czasu na cokolwiek "
                   "poza produkcją, co w praktyce się nie zdarza.")

        recipes_df_dash = st.session_state.get("recipes_df")
        product_sourcing_lookup_dash = {}
        if recipes_df_dash is not None and not recipes_df_dash.empty and RECIPE_SOURCING_COL in recipes_df_dash.columns:
            for _, r in recipes_df_dash.iterrows():
                product_sourcing_lookup_dash[r[RECIPE_PRODUCT_COL]] = (
                    r.get(RECIPE_SOURCING_COL, "Produkcja własna"), r.get(RECIPE_IMPORT_TRANSITION_COL, "")
                )

        year_cols_prod = st.columns(RAMPUP_YEARS)
        for year_idx, col in enumerate(year_cols_prod):
            active_n, inactive_n = 0, 0
            utilizations = []  # (tag, %)
            own_tonnage_y, import_tonnage_y = 0.0, 0.0
            import_breakdown_y = []  # (produkt, t) - do przejrzystości, co dokładnie się sumuje
            for m in st.session_state.confirmed_mixers:
                recipe_product = m.get("recipe_product")
                is_imported = False
                if recipe_product and recipe_product in product_sourcing_lookup_dash:
                    sourcing, transition = product_sourcing_lookup_dash[recipe_product]
                    is_imported = is_product_imported_in_year(sourcing, transition, year_idx)
                if is_imported:
                    inactive_n += 1
                    import_frac = get_import_volume_fraction(m["product_family"], year_idx)
                    import_t_this = (m["annual_volume"] / 1000.0) * import_frac
                    import_tonnage_y += import_t_this
                    import_breakdown_y.append((recipe_product or m["tag"], import_t_this))
                    continue
                frac = get_rampup_fraction(m["product_family"], year_idx)
                scaled_monthly_mass = (m["annual_volume"] / MONTHS_PER_YEAR) * frac
                scaled_batches = math.ceil(scaled_monthly_mass / m["mass_per_batch"]) if m["mass_per_batch"] > 0 else 0
                # Ta sama logika co Zakładka 1 ("Utylizacja Czasowa") - CZAS pracy (szarże × cykl)
                # względem DOSTĘPNYCH godzin pracy w miesiącu, NIE stosunek liczby szarż do celu
                # (ten drugi z definicji zawsze wychodzi 100% przy pełnym rozruchu, bo licznik i
                # mianownik są tą samą wielkością - nie mówi nic o REALNYM wykorzystaniu czasu).
                util_pct = (scaled_batches * m["cycle_h"] / AVAILABLE_HOURS_MONTH * 100.0) if AVAILABLE_HOURS_MONTH > 0 else 0.0
                utilizations.append((m["tag"], util_pct))
                if scaled_batches > 0:
                    active_n += 1
                else:
                    inactive_n += 1
                own_tonnage_y += (m["annual_volume"] / 1000.0) * frac

            with col:
                st.markdown(f"**Rok {year_idx + 1}**")
                st.metric("🟢 Aktywne", f"{active_n} szt.")
                st.metric("⚪ Nieaktywne", f"{inactive_n} szt.")
                if utilizations:
                    tag_min, util_min = min(utilizations, key=lambda x: x[1])
                    tag_max, util_max = max(utilizations, key=lambda x: x[1])
                    util_avg = sum(u for _, u in utilizations) / len(utilizations)
                    st.metric("📉 Min. utylizacja", f"{util_min:.0f}%", help=f"Mieszalnik: {tag_min}")
                    st.metric("📈 Maks. utylizacja", f"{util_max:.0f}%", help=f"Mieszalnik: {tag_max}")
                    st.metric("📊 Śr. utylizacja", f"{util_avg:.0f}%")
                else:
                    st.caption("Brak aktywnych mieszalników w tym roku.")
                st.metric("🏭 Produkcja własna", f"{own_tonnage_y:,.0f} t")
                if import_breakdown_y:
                    breakdown_txt = " · ".join(f"{name}: {t:,.0f} t" for name, t in import_breakdown_y)
                    st.metric("📦 Import", f"{import_tonnage_y:,.0f} t",
                              help=f"{len(import_breakdown_y)} produkt(ów) jeszcze importowanych w tym roku: {breakdown_txt}")
                else:
                    st.metric("📦 Import", f"{import_tonnage_y:,.0f} t")

        st.markdown("---")
        st.markdown("### 🗺️ Porównanie VSM (Mapa Strumienia Wartości)")
        st.caption("Porównaj 2-3 mieszalniki obok siebie — np. 31 m³ vs 120 m³ — żeby zobaczyć, jak bardzo różnią "
                   "się ich rzeczywiste czasy cyklu, nie tylko pojemność.")
        all_mixer_tags_vsm_cmp = [m["tag"] for m in st.session_state.confirmed_mixers]
        compare_vsm_tags = st.multiselect(
            "Wybierz mieszalniki (max 3):", all_mixer_tags_vsm_cmp,
            default=all_mixer_tags_vsm_cmp[:min(2, len(all_mixer_tags_vsm_cmp))], key="vsm_dashboard_compare_select"
        )
        if len(compare_vsm_tags) > 3:
            st.warning("⚠️ Wybierz maksymalnie 3 — pokazuję pierwsze 3 z wybranych.")
            compare_vsm_tags = compare_vsm_tags[:3]

        if not compare_vsm_tags:
            st.info("ℹ️ Wybierz przynajmniej jeden mieszalnik powyżej, żeby zobaczyć jego mapę strumienia wartości.")
        else:
            vsm_cmp_cols = st.columns(len(compare_vsm_tags))
            for col, tag_cmp in zip(vsm_cmp_cols, compare_vsm_tags):
                m_cmp = next(m for m in st.session_state.confirmed_mixers if m["tag"] == tag_cmp)
                with col:
                    st.markdown(f"**🔧 {tag_cmp}** ({m_cmp['product_family']}, {m_cmp['capacity_m3']:.0f} m³)")
                    ct_cmp = st.session_state.get("calculated_times", {}).get(tag_cmp, {"heating": 1.5, "pumping": 0.75, "cooling_h": 0.5})
                    bt_cmp = st.session_state.get("batch_time_components", {}).get(tag_cmp, {"dosing": 1.0, "homog": 2.0})
                    qc_tests_cmp_dash, _ = get_qc_tests_for_mixer(m_cmp)
                    durations_min_cmp = [get_full_qc_catalog().get(t, {}).get("duration_min", 0) for t in qc_tests_cmp_dash]
                    qc_mode_cmp = st.session_state.get("vsm_qc_config", {}).get(m_cmp["product_family"], {}).get("mode", "Sekwencyjnie (jeden technik, jedno stanowisko)")
                    qc_time_h_cmp = (sum(durations_min_cmp) if qc_mode_cmp.startswith("Sekw") else max(durations_min_cmp, default=0)) / 60.0

                    rho_cmp = st.session_state.active_portfolio.get(m_cmp["product_family"], {}).get("density", 0.9)
                    filling_h_cmp = compute_filling_time_h(
                        m_cmp["mass_per_batch"], m_cmp.get("recipe_product"), m_cmp["product_family"], tag_cmp,
                        rho_cmp, opakowania_podzial=st.session_state.get("opakowania_podzial", {})
                    )

                    steps_cmp = [
                        ("Dozowanie", bt_cmp["dosing"]), ("Grzanie", ct_cmp["heating"]), ("Homogenizacja", bt_cmp["homog"]),
                        ("Zwolnienie QC", qc_time_h_cmp), ("Pompowanie", ct_cmp["pumping"]),
                        ("Chłodzenie", ct_cmp.get("cooling_h", 0.0)), ("Rozlew", filling_h_cmp),
                    ]
                    total_cycle_h_cmp = sum(v for _, v in steps_cmp)
                    st.dataframe(pd.DataFrame(steps_cmp, columns=["Etap", "Czas [h]"]).round(2), hide_index=True, use_container_width=True)
                    st.metric("⏱️ Łączny cykl (bez buforów magazynowych)", f"{total_cycle_h_cmp:.1f} h")

        st.markdown("---")
        st.markdown("### 🛢️ Zbiorniki RM")
        confirmed_rm_tanks_dash2 = st.session_state.get("confirmed_rm_tanks", [])
        if not confirmed_rm_tanks_dash2:
            st.info("ℹ️ Brak zatwierdzonych zbiorników RM — odwiedź Zakładkę 2 (Magazynowanie), sekcję "
                    "'✅ Zatwierdź Zbiorniki RM'.")
        else:
            tanks_by_material_dash2 = {}
            for t in confirmed_rm_tanks_dash2:
                tanks_by_material_dash2.setdefault(t["material"], []).append(t)
            rm_tank_tech_details_dash2 = st.session_state.get("rm_tank_tech_details", {})
            year_cols_tanks = st.columns(RAMPUP_YEARS)
            for year_idx, col in enumerate(year_cols_tanks):
                active_tanks_n, inactive_tanks_n = 0, 0
                total_material_t_year = 0.0
                total_refills_year = 0
                year_consumption = compute_rm_consumption_for_year(year_idx)
                for material, tanks_this_material in tanks_by_material_dash2.items():
                    consumption_t = year_consumption.get(material, 0.0)
                    is_active = consumption_t > 0
                    if is_active:
                        active_tanks_n += len(tanks_this_material)
                        total_material_t_year += consumption_t
                        # Szacowana liczba uzupełnień/rok = roczne zużycie ÷ użyteczna pojemność
                        # WSZYSTKICH zbiorników tego materiału razem (bufor bezpiecznego napełnienia
                        # uwzględniony) - przybliżenie do porównania skali, nie dokładny harmonogram dostaw.
                        total_capacity_m3 = sum(t["capacity_m3"] for t in tanks_this_material)
                        density_kg_m3 = rm_tank_tech_details_dash2.get(tanks_this_material[0]["tag"], {}).get("density_kg_m3", 900.0)
                        usable_t = total_capacity_m3 * TANK_SAFETY_FILL * density_kg_m3 / 1000.0
                        if usable_t > 0:
                            total_refills_year += math.ceil(consumption_t / usable_t)
                    else:
                        inactive_tanks_n += len(tanks_this_material)
                with col:
                    st.markdown(f"**Rok {year_idx + 1}**")
                    st.metric("🟢 Aktywne", f"{active_tanks_n} szt.")
                    st.metric("⚪ Nieaktywne", f"{inactive_tanks_n} szt.")
                    st.metric("📦 Materiał w zbiornikach", f"{total_material_t_year:,.0f} t/rok")
                    st.metric("🔄 Uzupełnienia (szac.)", f"{total_refills_year}/rok",
                              help="Szacunkowa liczba dostaw/napełnień rocznie: zużycie ÷ użyteczna pojemność "
                                   "zbiorników danego surowca — przybliżenie, nie harmonogram dostaw.")

        st.markdown("---")
        st.markdown("### 💰 CAPEX i ROI")
        total_capex_dash = st.session_state.get("total_capex_report")
        payback_dash = st.session_state.get("payback_year_fraction_report")
        roi_year5_dash = st.session_state.get("roi_year5_report")
        waluta_dash = st.session_state.get("waluta_report", "PLN")

        e1, e2, e3 = st.columns(3)
        with e1:
            st.metric("💰 CAPEX (z rezerwą)", f"{total_capex_dash:,.0f} {waluta_dash}" if total_capex_dash else "—",
                      help=None if total_capex_dash else "Skonfiguruj CAPEX w Zakładce 5, Krok 3.")
        with e2:
            if payback_dash is not None:
                st.metric("⏳ Okres zwrotu", f"{payback_dash:.1f} lat")
            elif total_capex_dash:
                st.metric("⏳ Okres zwrotu", f"> {RAMPUP_YEARS} lat")
            else:
                st.metric("⏳ Okres zwrotu", "—", help="Skonfiguruj CAPEX i ROI w Zakładce 5.")
        with e3:
            st.metric("🎯 ROI w Roku 5", f"{roi_year5_dash:.1f}%" if roi_year5_dash is not None else "—",
                      help=None if roi_year5_dash is not None else "Odwiedź Krok 4 (ROI) w Zakładce 5.")

        st.markdown("---")
        st.markdown("### 🏢 Magazyn — FG / RM (beczki) / Import, osobno per rok")
        st.caption("Wszystkie trzy liczby poniżej liczone TYM SAMYM modelem — bufor dni zapasu (Zakładka 4, "
                   "'Wymagany zapas bezpieczeństwa') — więc są ze sobą w pełni spójne i sumują się do liczby "
                   "docelowej poniżej.")
        total_wh_target_dash = st.session_state.get("total_miejsca_magazynowe_target_report")
        total_wh_m2_dash = st.session_state.get("total_powierzchnia_m2_report")
        if total_wh_target_dash:
            st.metric("🎯 Docelowa pojemność magazynu (FG + RM w beczkach + stały import, razem)",
                      f"{total_wh_target_dash:,.0f} szt." + (f" · {total_wh_m2_dash:,.0f} m²" if total_wh_m2_dash else ""))

        fg_capacity_dash = st.session_state.get("fg_capacity_pallets_report")
        powierzchnia_na_miejsce_dash = st.session_state.get("powierzchnia_na_miejsce_report")
        liczba_poziomow_dash = st.session_state.get("liczba_poziomow_report")
        import_pallet_mass_kg_dash = st.session_state.get("import_pallet_mass_kg", 800.0)
        recipes_df_dash2 = st.session_state.get("recipes_df")

        if not fg_capacity_dash:
            st.info("ℹ️ Odwiedź Zakładkę 4 (Logistyka), sekcję symulacji stanu magazynowego, aby zobaczyć wykorzystanie per rok.")
        else:
            days_of_stock_dash = st.session_state.get("days_of_stock_tab5", 14)

            def _buffer_import_volume_for_year(year_idx_calc):
                """Wolumen [t/rok] produktów importowanych NA STAŁE do zbiornika buforowego ('Nigdy
                bufor') - import trwa ZAWSZE dla tych produktów (nie ma przejścia na produkcję
                własną), więc to NIE JEST zero w żadnym roku, tylko inna forma (zbiornik, nie paleta)."""
                if recipes_df_dash2 is None or recipes_df_dash2.empty or RECIPE_IMPORT_TRANSITION_COL not in recipes_df_dash2.columns:
                    return 0.0, []
                buffer_rows = recipes_df_dash2[
                    (recipes_df_dash2[RECIPE_SOURCING_COL] == "Import") &
                    (recipes_df_dash2[RECIPE_IMPORT_TRANSITION_COL] == "Nigdy (bufor)")
                ]
                total_t, names = 0.0, []
                for _, r in buffer_rows.iterrows():
                    annual_t_target = float(r.get(RECIPE_ANNUAL_COL, 0) or 0)
                    frac = get_import_volume_fraction(r[RECIPE_GROUP_COL], year_idx_calc)
                    total_t += annual_t_target * frac
                    names.append(r[RECIPE_PRODUCT_COL])
                return total_t, names

            year_cols_wh = st.columns(RAMPUP_YEARS)
            for year_idx, col in enumerate(year_cols_wh):
                # Wszystkie trzy - te same wspólne funkcje (compute_fg_buffer_positions_for_year,
                # compute_rm_drummed_positions_for_year, compute_import_positions_for_year), ten
                # sam model bufora dni zapasu - żeby te liczby NIGDY nie mogły się rozjechać.
                fg_used_pal = compute_fg_buffer_positions_for_year(year_idx, days_of_stock_dash)
                rm_used_pal = compute_rm_drummed_positions_for_year(year_idx, days_of_stock_dash)
                import_used_pal = compute_import_positions_for_year(year_idx, import_pallet_mass_kg_dash)
                buffer_import_t, buffer_import_names = _buffer_import_volume_for_year(year_idx)
                with col:
                    st.markdown(f"**Rok {year_idx + 1}**")
                    st.metric("🏷️ FG — miejsca magazynowe", f"{fg_used_pal:,.0f} pal.")
                    st.metric("🛢️ RM (beczki) — miejsca magazynowe", f"{rm_used_pal:,.0f} pal.")
                    st.metric("📦 Import (palety) — miejsca magazynowe", f"{import_used_pal:,.0f} pal.")
                    st.metric("🔵 Import (zbiornik bufor.)", f"{buffer_import_t:,.0f} t/rok",
                              help=(f"{', '.join(buffer_import_names)} — import na stałe, trwa cały czas, "
                                    "magazynowany w dedykowanym zbiorniku (Zakładka 2), nie na palecie."
                                    if buffer_import_names else "Brak produktów importowanych na stałe do zbiornika."))
                    total_used = fg_used_pal + rm_used_pal + import_used_pal
                    if total_wh_target_dash:
                        st.metric("📊 Wykorzystanie (FG+RM+Import)", f"{(total_used / total_wh_target_dash * 100.0):.0f}%")

        st.markdown("---")
        st.markdown("### ⚡ Energia — CAŁKOWITE zużycie, per rok")
        st.caption("Fizyczne zużycie [kWh], nie koszt — obejmuje WSZYSTKO: mieszanie/pompowanie, grzanie i "
                   "chłodzenie produktu, pompy i grzanie zbiorników RM, ORAZ energię pozaprodukcyjną (HVAC, "
                   "oświetlenie, serwerownia — działa niezależnie od tego, ile akurat produkujesz). Pominięta "
                   "jest jedynie 'moc szczytowa procesu' z bilansu elektrycznego (Karta Maszyn) — to moc "
                   "instalowana do doboru transformatora, nie realne zużycie, więc doliczenie jej zawyżałoby "
                   "wynik i podwajało to, co już liczone jest tu wprost per szarżę.")
        recipes_df_dash3 = st.session_state.get("recipes_df")
        product_sourcing_lookup_dash3 = {}
        if recipes_df_dash3 is not None and not recipes_df_dash3.empty and RECIPE_SOURCING_COL in recipes_df_dash3.columns:
            for _, r in recipes_df_dash3.iterrows():
                product_sourcing_lookup_dash3[r[RECIPE_PRODUCT_COL]] = (
                    r.get(RECIPE_SOURCING_COL, "Produkcja własna"), r.get(RECIPE_IMPORT_TRANSITION_COL, "")
                )
        calculated_times_dash3 = st.session_state.get("calculated_times", {})
        mixer_tech_dash3 = st.session_state.get("mixer_tech_advanced_details", {})
        confirmed_rm_tanks_dash3 = st.session_state.get("confirmed_rm_tanks", [])
        rm_tank_tech_dash3 = st.session_state.get("rm_tank_tech_details", {})
        facility_demand_kw_dash = st.session_state.get("facility_demand_kw", 0.0)
        hours_year_dash = AVAILABLE_HOURS_MONTH * MONTHS_PER_YEAR

        year_cols_energy = st.columns(RAMPUP_YEARS)
        for year_idx, col in enumerate(year_cols_energy):
            electrical_kwh_year, thermal_kwh_year, produced_kg_year = 0.0, 0.0, 0.0
            for m in st.session_state.confirmed_mixers:
                recipe_product = m.get("recipe_product")
                is_imported = False
                if recipe_product and recipe_product in product_sourcing_lookup_dash3:
                    sourcing, transition = product_sourcing_lookup_dash3[recipe_product]
                    is_imported = is_product_imported_in_year(sourcing, transition, year_idx)
                if is_imported:
                    continue
                frac = get_rampup_fraction(m["product_family"], year_idx)
                month_mass_kg = (m["annual_volume"] / MONTHS_PER_YEAR) * frac
                scaled_batches_month = math.ceil(month_mass_kg / m["mass_per_batch"]) if m["mass_per_batch"] > 0 else 0
                scaled_batches_year = scaled_batches_month * MONTHS_PER_YEAR
                produced_kg_year += month_mass_kg * MONTHS_PER_YEAR

                m_data = calculated_times_dash3.get(m["tag"], {"power_mix_kw": 5.5, "power_pump_kw": 1.5})
                cycle_h_e = m.get("cycle_h", 4.0)
                electrical_kwh_year += (m_data["power_mix_kw"] * cycle_h_e + m_data["power_pump_kw"] * 0.75) * scaled_batches_year

                p_e = mixer_tech_dash3.get(m["tag"])
                if p_e:
                    try:
                        heat_res_e = compute_thermal_ntu(
                            m["mass_per_batch"], p_e["cp_product"], p_e["t_product_in"], p_e["t_product_out"],
                            p_e["k_coeff_grzania"], p_e["exchange_area_m2"], p_e["utility_type_heat"],
                            p_e["flow_heat_value"], p_e["flow_heat_unit"], p_e["t_utility_heat_in"])
                        if heat_res_e["status"] == "ok":
                            thermal_kwh_year += (heat_res_e["q_total_kj"] / 3600.0) * scaled_batches_year
                        cool_res_e = compute_thermal_ntu(
                            m["mass_per_batch"], p_e["cp_product"], p_e["t_product_out"], p_e["t_discharge_c"],
                            p_e["k_coeff"], p_e["exchange_area_m2"], p_e["utility_type_cool"],
                            p_e["flow_cool_value"], p_e["flow_cool_unit"], p_e["t_utility_cool_in"])
                        if cool_res_e["status"] == "ok":
                            # Chłodzenie liczone jako energia ELEKTRYCZNA (agregat/COP), nie cieplna.
                            electrical_kwh_year += (cool_res_e["q_total_kj"] / 3600.0) * scaled_batches_year
                    except Exception:
                        pass

            # Zbiorniki RM: pompy (elektryczna) i grzanie (cieplna) - TYLKO aktywne w danym roku
            # (te same warunki co w sekcji '🛢️ Zbiorniki RM' wyżej).
            year_consumption_rm = compute_rm_consumption_for_year(year_idx)
            for t in confirmed_rm_tanks_dash3:
                if year_consumption_rm.get(t["material"], 0.0) <= 0:
                    continue
                rd = rm_tank_tech_dash3.get(t["tag"])
                if not rd:
                    continue
                try:
                    eff_flow = rd["pump_flow_m3h"]
                    if rd["pump_mode"] == "Współdzielona (kilka zbiorników)" and rd["shared_pump_id"] in st.session_state.get("shared_pumps", {}):
                        shared_cfg = st.session_state.shared_pumps[rd["shared_pump_id"]]
                        eff_flow, eff_eff = shared_cfg["flow_m3h"], shared_cfg["efficiency"]
                    else:
                        eff_eff = rd["pump_efficiency"]
                    zeta_sum_rm_e = (rd["count_elbows_90"] * 0.5) + (rd["count_valves"] * 0.2)
                    _, _, power_kw_rm_e, _ = compute_hydraulics(
                        eff_flow, rd["pipe_dn"], rd["pipe_length_m"], rd["delta_h_m"],
                        rd["viscosity_cst"], rd["density_kg_m3"], zeta_sum_rm_e, eff_eff)
                    electrical_kwh_year += power_kw_rm_e * hours_year_dash
                    if rd["heated"]:
                        heating_power_kw_e = compute_tank_heating_power_kw(
                            t["capacity_m3"], rd["insulation_mm"], rd["target_temp_c"], rd["ambient_temp_c"])
                        thermal_kwh_year += heating_power_kw_e * hours_year_dash
                except Exception:
                    pass

            # Pozaprodukcyjne (HVAC/oświetlenie/serwerownia) - stałe, niezależne od wolumenu produkcji.
            facility_kwh_year = facility_demand_kw_dash * hours_year_dash
            electrical_kwh_year += facility_kwh_year
            total_kwh_year = electrical_kwh_year + thermal_kwh_year

            with col:
                st.markdown(f"**Rok {year_idx + 1}**")
                st.metric("⚡ Elektryczna — razem", f"{electrical_kwh_year:,.0f} kWh/rok",
                          help=f"W tym pozaprodukcyjne (HVAC/oświetlenie): {facility_kwh_year:,.0f} kWh/rok (stałe).")
                st.metric("⚡ Elektryczna — na kg", f"{(electrical_kwh_year / produced_kg_year):.3f} kWh/kg" if produced_kg_year > 0 else "—")
                st.metric("🔥 Cieplna — razem", f"{thermal_kwh_year:,.0f} kWh/rok")
                st.metric("🔥 Cieplna — na kg", f"{(thermal_kwh_year / produced_kg_year):.3f} kWh/kg" if produced_kg_year > 0 else "—")
                st.metric("⚡🔥 Razem — na kg", f"{(total_kwh_year / produced_kg_year):.3f} kWh/kg" if produced_kg_year > 0 else "—")

        st.markdown("---")
        st.caption("📄 Pełny raport (PDF/Excel) ze wszystkimi tymi liczbami znajdziesz w Zakładce 5, Krok 5.")
